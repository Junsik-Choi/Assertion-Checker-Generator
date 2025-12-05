#!/usr/bin/env python3
"""
Check Excel Signal Assignments to see if hierarchy/clocks/resets are there
"""
from openpyxl import load_workbook
from pathlib import Path

excel_path = Path(__file__).parent / "out/sessions/blur_scaler-20251203_141618/blur_scaler.xlsx"

print("=" * 70)
print("EXCEL SIGNAL ASSIGNMENTS INSPECTION")
print("=" * 70)

if excel_path.exists():
    wb = load_workbook(str(excel_path))
    
    # Check "Assertion Sheet" or "Signal Assignments"
    sheet_name = None
    for name in wb.sheetnames:
        if 'signal' in name.lower() or 'assignment' in name.lower():
            sheet_name = name
            break
    
    if not sheet_name and 'Assertion Sheet' in wb.sheetnames:
        sheet_name = 'Assertion Sheet'
    
    print(f"\nAll sheets: {wb.sheetnames}")
    
    if sheet_name:
        ws = wb[sheet_name]
        print(f"\nUsing sheet: {sheet_name}")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        print("\nFirst 20 rows:")
        for row_idx in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(8, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(val) if val else "")
            print(f"Row {row_idx}: {row_data}")
    else:
        print("\nNo 'Signal Assignments' sheet found")
        print("Available sheets:", wb.sheetnames)
    
    wb.close()
else:
    print(f"Excel not found: {excel_path}")

print("\n" + "=" * 70)
