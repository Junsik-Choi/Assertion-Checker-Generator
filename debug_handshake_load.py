#!/usr/bin/env python3
"""
Debug handshake loading
"""
from openpyxl import load_workbook
from pathlib import Path

excel_path = Path(__file__).parent / "out/sessions/ram_ctrl-20251202_190649/ram_ctrl.xlsx"

print("=" * 70)
print("DEBUG: Handshake loading")
print("=" * 70)

wb = load_workbook(str(excel_path))

# Find handshake sheet
handshake_sheet = None
for name in wb.sheetnames:
    if name.lower() == 'handshake':
        handshake_sheet = name
        break

print(f"\nHandshake sheet: {handshake_sheet}")

if handshake_sheet:
    ws = wb[handshake_sheet]
    print(f"Max row: {ws.max_row}")
    
    print("\nReading Handshake data (Row 7+):")
    for row_idx in range(7, min(15, ws.max_row + 1)):
        phase_type = ws.cell(row=row_idx, column=3).value  # Column C
        sender = ws.cell(row=row_idx, column=4).value       # Column D
        receiver = ws.cell(row=row_idx, column=5).value     # Column E
        
        print(f"Row {row_idx}: phase={phase_type}, sender={sender}, receiver={receiver}")
        
        # Check if reading process would skip this
        if not phase_type or str(phase_type).strip() == '':
            print(f"  -> Would BREAK (empty phase_type)")
            break
        
        # Check if both sender and receiver are empty
        if not sender and not receiver:
            print(f"  -> Would SKIP (both empty)")
            continue
        
        print(f"  -> Would INCLUDE")

wb.close()
print("\n" + "=" * 70)
