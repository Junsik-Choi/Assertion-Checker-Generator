#!/usr/bin/env python3
from pathlib import Path
from openpyxl import load_workbook

excel_path = 'out/sessions/blur_scaler-20251203_141618/blur_scaler.xlsx'
wb = load_workbook(excel_path)

print("Sheet names:", wb.sheetnames)
print()

# Find Counter sheet
counter_sheet = None
for name in wb.sheetnames:
    if name.lower() == 'counter':
        counter_sheet = name
        print(f"Found Counter sheet: {counter_sheet}")
        break

if counter_sheet:
    ws = wb[counter_sheet]
    print(f"Counter sheet max_row: {ws.max_row}")
    
    # Check rows 8-15
    for row_idx in range(8, min(15, ws.max_row + 1)):
        target = ws.cell(row=row_idx, column=2).value
        plus_con = ws.cell(row=row_idx, column=3).value
        reset_con = ws.cell(row=row_idx, column=4).value
        trigger_con = ws.cell(row=row_idx, column=5).value
        exp_cnt_val = ws.cell(row=row_idx, column=6).value
        
        print(f"Row {row_idx}: target={target}, plus={plus_con}, reset={reset_con}, trigger={trigger_con}, exp={exp_cnt_val}")
        
    print()
    print("Now trying Row 9 specifically:")
    target = ws.cell(row=9, column=2).value
    print(f"Row 9 target: {target}")
    print(f"Row 9 target type: {type(target)}")
    print(f"Row 9 target stripped: '{str(target).strip()}'")
    print(f"Is it 'abc'?: {str(target).strip().lower() == 'abc'}")
