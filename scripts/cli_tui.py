#!/usr/bin/env python3
"""
Full-screen terminal TUI for the Assertion Builder.

Features:
- Fills terminal and adapts to resize.
- Fixed-width grouped panels showing module/IP info, clocks/resets, and ports.
- Command hint line and an input prompt at the bottom.
- Commands: help/h, quit/q, set rtl/module/excel/out, scan, fill, json, sv.
- Integrates with scripts/rtl_parser.py and scripts/assertion_builder.py.

Windows note: requires the 'windows-curses' package: pip install windows-curses
"""

from __future__ import annotations

import os
import sys
import json
import shutil
import subprocess
from dataclasses import dataclass, field
from textwrap import wrap as _textwrap
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union
from datetime import datetime

# Ensure we can import local scripts
_THIS_DIR = Path(__file__).resolve().parent
if str(_THIS_DIR) not in sys.path:
    sys.path.insert(0, str(_THIS_DIR))

try:
    import curses  # type: ignore
except Exception as _curses_err:  # pragma: no cover - import-time environment dependent
    # Provide a clearer error for Windows users
    msg = (
        "Failed to import curses. On Windows, install 'windows-curses' first: "
        "pip install windows-curses\nOriginal error: %r" % (_curses_err,)
    )
    print(msg)
    raise

# Import RTL parsing helpers
from rtl_parser import (  # type: ignore
    discover_files,
    find_rtl_root_from,
    build_modules_db,
    find_top_modules,
    find_occurrences_of_target,
    find_module_instances_by_file,
    compute_env_for_occurrence,
    resolve_ports_with_params,
    classify_groups,
)
from assertion_builder import fill_define_excel_if_needed  # type: ignore
try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None  # type: ignore


def _remove_readonly(func, path, excinfo):
    """Error handler for Windows readonly files"""
    import stat
    import os
    os.chmod(path, stat.S_IWRITE)
    func(path)


def _force_delete_folder(folder_path: Path) -> Tuple[bool, str]:
    """
    Force delete folder with 3-tier strategy for Windows.
    Returns (success, error_message)
    """
    import time
    import subprocess
    
    if not folder_path.exists():
        return True, ""
    
    # Tier 1: Try normal deletion with readonly handler
    try:
        shutil.rmtree(folder_path, onerror=_remove_readonly)
        if not folder_path.exists():
            return True, ""
    except Exception as e:
        pass
    
    # Tier 2: Wait and retry
    try:
        time.sleep(0.1)
        shutil.rmtree(folder_path, onerror=_remove_readonly)
        if not folder_path.exists():
            return True, ""
    except Exception as e:
        pass
    
    # Tier 3: System command (Windows: rd /s /q, Unix: rm -rf)
    try:
        import platform
        if platform.system() == "Windows":
            result = subprocess.run(
                ['cmd', '/c', 'rd', '/s', '/q', str(folder_path)],
                capture_output=True,
                text=True,
                timeout=10
            )
            if not folder_path.exists():
                return True, ""
            else:
                return False, f"System command failed: {result.stderr or 'Unknown error'}"
        else:
            result = subprocess.run(
                ['rm', '-rf', str(folder_path)],
                capture_output=True,
                text=True,
                timeout=10
            )
            if not folder_path.exists():
                return True, ""
            else:
                return False, f"System command failed: {result.stderr or 'Unknown error'}"
    except Exception as e:
        return False, f"Force delete failed: {str(e)}"


def _robust_copy(src: Path, dst: Path) -> Path:
    """Copy file src->dst robustly. If dst exists or copy fails, try a unique name and raw copy."""
    target = dst
    if target.exists():
        stem = dst.stem
        suffix = dst.suffix
        for i in range(1, 1000):
            cand = dst.with_name(f"{stem}-{i}{suffix}")
            if not cand.exists():
                target = cand
                break
    try:
        shutil.copy2(src, target)
        return target
    except Exception:
        # Fallback: raw read/write
        try:
            with open(src, "rb") as rf, open(target, "wb") as wf:
                while True:
                    chunk = rf.read(1024 * 1024)
                    if not chunk:
                        break
                    wf.write(chunk)
            return target
        except Exception as e:
            raise RuntimeError(f"Copy failed: {e}")


def _create_session_excel_and_fill(state: AppState) -> Tuple[bool, str]:
    """
    Create session Excel in out/sessions/<module>-<timestamp>/, 
    verify Define sheet, run fill_define.py. Returns (ok, err).
    """
    # Debug logging to file
    debug_log = _THIS_DIR.parent / "out" / "session_creation_debug.log"
    debug_log.parent.mkdir(parents=True, exist_ok=True)
    
    def log_debug(msg):
        with open(debug_log, "a", encoding="utf-8") as f:
            from datetime import datetime
            f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
    
    try:
        log_debug("=== _create_session_excel_and_fill() CALLED ===")
        log_debug(f"state.excel_path = {state.excel_path}")
        log_debug(f"state.target_module = {state.target_module}")
        log_debug(f"state.module_info.module = {state.module_info.module}")
        
        if not state.excel_path or not Path(state.excel_path).exists():
            log_debug("ERROR: Reference Excel not set or doesn't exist")
            return False, "Reference Excel not set"
        
        # 세션 디렉터리를 out/sessions/<module>-<timestamp>/ 형태로 생성
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        mod = state.target_module or (state.module_info.module or "module")
        session_name = f"{mod}-{ts}"
        sess_dir = (_THIS_DIR.parent / "out" / "sessions" / session_name).resolve()
        
        log_debug(f"Creating session directory: {sess_dir}")
        
        sess_dir.mkdir(parents=True, exist_ok=True)
        log_debug(f"Session directory created successfully")
        
        # Create assertions output directory within session
        assertions_dir = sess_dir / "assertions"
        assertions_dir.mkdir(parents=True, exist_ok=True)
        log_debug(f"Assertions directory created: {assertions_dir}")
        
        # Update state.out_dir to point to session-specific assertions folder
        state.out_dir = assertions_dir
        
        # 엑셀 파일을 세션 폴더로 복사
        new_xlsx = sess_dir / f"{mod}.xlsx"
        log_debug(f"Copying Excel from {state.excel_path} to {new_xlsx}")
        
        try:
            new_xlsx = _robust_copy(Path(state.excel_path), new_xlsx)
            state.session_excel_path = new_xlsx
            log_debug(f"Excel copied successfully to {new_xlsx}")
            log_debug(f"state.session_excel_path SET TO: {state.session_excel_path}")
        except Exception as copy_err:
            log_debug(f"ERROR during copy: {copy_err}")
            return False, f"Failed to copy Excel: {copy_err}"
        
        # Verify Define sheet
        log_debug("Verifying Define sheet...")
        if not load_workbook:
            log_debug("ERROR: openpyxl missing")
            return False, "openpyxl missing"
        try:
            wb = load_workbook(str(new_xlsx))
            if "Define" not in wb.sheetnames:
                wb.close()
                log_debug("ERROR: Define sheet missing")
                return False, "Define sheet missing in reference Excel"
            wb.close()
            log_debug("Define sheet verified OK")
        except Exception as wb_err:
            log_debug(f"ERROR opening workbook: {wb_err}")
            return False, f"Failed to open workbook: {wb_err}"
        
        # Define JSON을 같은 세션 폴더에 생성
        log_debug("Creating define JSON...")
        
        # Ensure hierarchy is set (use module name as default if not set)
        hierarchy = state.module_info.module_hierarchy or state.module_info.module or ""
        
        define_json = fill_define_excel_if_needed(new_xlsx, {
            "module": state.module_info.module,
            "rtl_file_path": state.module_info.rtl_file_path,  # Internal use only
            "rtl_hierarchy": hierarchy,  # Display path (use module as fallback)
            "clocks": state.module_info.clocks,
            "resets": state.module_info.resets,
            "inputs": state.module_info.inputs,
            "outputs": state.module_info.outputs,
            "inouts": state.module_info.inouts,
            "parameters": state.module_info.parameters,
            "conditions": state.conditions,  # Add condition signals
        }, sess_dir)
        log_debug(f"Define JSON created: {define_json}")
        
        # Run fill_define.py
        log_debug("Running fill_define.py...")
        fill_script = _THIS_DIR / "fill_define.py"
        if not fill_script.exists():
            log_debug("ERROR: fill_define.py not found")
            return False, "fill_define.py not found"
        
        # Set environment to use UTF-8 encoding
        import os
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        
        result = subprocess.run(
            [sys.executable, str(fill_script), str(new_xlsx), str(define_json)], 
            check=False,
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            env=env,
            cwd=_THIS_DIR.parent  # Set working directory to project root
        )
        
        log_debug(f"fill_define.py returned: {result.returncode}")
        if result.stdout:
            log_debug(f"fill_define.py stdout: {result.stdout}")
        if result.stderr:
            log_debug(f"fill_define.py stderr: {result.stderr}")
        
        if result.returncode != 0:
            log_debug("ERROR: fill_define.py failed")
            # Don't fail the session creation just because fill_define failed
            # The Excel and folder are already created successfully
            log_debug("WARNING: Continuing despite fill_define failure (Excel already created)")
        
        log_debug(f"SUCCESS! Session created at: {sess_dir}")
        log_debug(f"Final state.session_excel_path: {state.session_excel_path}")
        
        msg = f"✓ Session created: {_sanitize_path_for_display(str(sess_dir))}"
        msg += f"\n✓ Session Excel: {_sanitize_path_for_display(str(new_xlsx))}"
        return True, msg
    except Exception as e:
        import traceback
        err_msg = f"Session creation error: {str(e)}\n{traceback.format_exc()}"
        log_debug(f"EXCEPTION: {err_msg}")
        return False, err_msg


def _update_define_sheet(state: AppState) -> None:
    """
    Update Define sheet in session Excel with current module_info.
    Called after scan or when module info changes.
    """
    if not state.session_excel_path or not state.session_excel_path.exists():
        raise RuntimeError("Session Excel not found")
    
    # Ensure hierarchy is set (use module name as default if not set)
    hierarchy = state.module_info.module_hierarchy or state.module_info.module or ""
    
    # Create define JSON in session directory
    sess_dir = state.session_excel_path.parent
    define_json = fill_define_excel_if_needed(state.session_excel_path, {
        "module": state.module_info.module,
        "rtl_file_path": state.module_info.rtl_file_path,  # Internal use only
        "rtl_hierarchy": hierarchy,  # Display path (use module as fallback)
        "clocks": state.module_info.clocks,
        "resets": state.module_info.resets,
        "inputs": state.module_info.inputs,
        "outputs": state.module_info.outputs,
        "inouts": state.module_info.inouts,
        "parameters": state.module_info.parameters,
        "conditions": state.conditions,  # Add condition signals
    }, sess_dir)
    
    # Run fill_define.py
    fill_script = _THIS_DIR / "fill_define.py"
    if not fill_script.exists():
        raise RuntimeError("fill_define.py not found")
    
    # Set environment to use UTF-8 encoding
    import os
    env = os.environ.copy()
    env['PYTHONIOENCODING'] = 'utf-8'
    
    rc = subprocess.run(
        [sys.executable, str(fill_script), str(state.session_excel_path), str(define_json)],
        check=False,
        capture_output=True,
        text=True,
        encoding='utf-8',
        errors='replace',
        env=env
    ).returncode
    
    if rc != 0:
        raise RuntimeError("fill_define.py failed")

_APP_VERSION = "v1.0"

@dataclass
class ModuleInfo:
    module: str = ""
    rtl_file_path: str = ""  # Full path to .v file containing the module (internal use)
    module_hierarchy: str = ""  # User-provided module hierarchy (e.g., top.dut.abc.u_abc)
    inputs: List[Dict[str, Any]] = field(default_factory=list)
    outputs: List[Dict[str, Any]] = field(default_factory=list)
    inouts: List[Dict[str, Any]] = field(default_factory=list)
    clocks: List[Dict[str, Any]] = field(default_factory=list)
    resets: List[Dict[str, Any]] = field(default_factory=list)
    parameters: List[Dict[str, Any]] = field(default_factory=list)


@dataclass
class AppState:
    rtl_start: Optional[Path] = None
    target_module: Optional[str] = None
    excel_path: Optional[Path] = None
    out_dir: Path = Path("out/assertions")
    modules_db: Dict[str, Any] = field(default_factory=dict)
    module_info: ModuleInfo = field(default_factory=ModuleInfo)
    occs: List[Any] = field(default_factory=list)
    messages: List[str] = field(default_factory=list)
    # Session persistence id
    session_id: Optional[str] = None
    # User-defined condition signals
    conditions: List[Dict[str, Any]] = field(default_factory=list)
    # Pending ms command awaiting bit width
    pending_ms_command: Optional[Dict[str, Any]] = None
    # Ports filter substring
    port_filter: Optional[str] = None
    # Onboarding wizard state
    onboarding_active: bool = False
    onboarding_stage: Optional[str] = None  # 'rtl' | 'module' | 'hierarchy' | 'excel' | None
    onboarding_filter: str = ""
    onboarding_page: int = 0
    onboarding_modules: List[str] = field(default_factory=list)  # Now stores "inst_name (module_type)" format
    onboarding_instances: List[Dict[str, Any]] = field(default_factory=list)  # Stores full instance info
    selected_instance: Optional[Dict[str, Any]] = None  # Selected instance metadata for hierarchy building
    onboarding_excel_autofound: Optional[Path] = None
    onboarding_compl_index: int = 0
    onboarding_cand_page: int = 0
    onboarding_cand_visible: bool = False
    onboarding_cand_h: int = 0
    # Session Excel generated for this run
    session_excel_path: Optional[Path] = None
    excel_error: Optional[str] = None
    # Created assertions list
    assertions: List[Dict[str, Any]] = field(default_factory=list)
    # Assertion creation wizard state
    assertion_wizard_active: bool = False
    assertion_wizard_stage: str = ""  # 'select_type' | 'input_data' | 'confirm'
    assertion_selected_type: Optional[str] = None
    assertion_input_data: Dict[str, Any] = field(default_factory=dict)
    # New: Store port_dict for each signal field (field_name -> port_dict with calculated_bit_width)
    assertion_signal_ports: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    # New: Track current field being edited and cursor position
    assertion_current_field_idx: int = 0
    # New: Preview for assertion (right side)
    assertion_preview_lines: List[str] = field(default_factory=list)
    # New: Signal selection map (index -> (signal_name, port_dict)) for wizard
    assertion_signal_map: Dict[int, Tuple[str, Dict[str, Any]]] = field(default_factory=dict)
    # New: Signal list pagination for wizard signal selection
    assertion_signal_page: int = 0  # Current page in signal selection list
    assertion_signal_list: List[Tuple[int, str, str, Dict[str, Any]]] = field(default_factory=list)  # (idx, name, type, port_dict)
    # New: Track when waiting for custom number input (for exp_cnt_val [0] option)
    assertion_waiting_custom_number: bool = False
    
    # File generation wizard state
    gen_wizard_active: bool = False
    gen_wizard_stage: str = ""  # 'filename' | 'file_type' | 'data_source' | 'preview' | 'confirm'
    gen_filename: str = ""  # Output filename without extension
    gen_file_type: Optional[int] = None  # 1=interface, 2=instance, 3=both
    gen_data_source: Optional[str] = None  # 'assertions' | 'signals' | 'both'
    gen_preview_lines: List[str] = field(default_factory=list)  # Preview of generated code
    gen_preview_page: int = 0  # Current page in preview (for pagination)
    gen_preview_file_idx: int = 0  # 0=interface, 1=instance (for "both" mode)

    def log(self, msg: str) -> None:
        self.messages.append(msg)
        # Limit message history
        self.messages = self.messages[-200:]


HELP_TEXT = (
    """
Type 'help' to open the full-screen overlay. Use Tab/Shift-Tab to switch pages, Esc/q to close.
"""
).strip("\n")


def _safe_str(o: Any) -> str:
    try:
        return str(o)
    except Exception:
        return repr(o)


def _get_port_width(port: Dict[str, Any]) -> str:
    """Extract bit width from port and return as string like '[7:0]' or ''."""
    width = port.get("width") or port.get("msb_lsb") or ""
    if isinstance(width, dict):
        msb = width.get('msb', '')
        lsb = width.get('lsb', '')
        if msb != '' and lsb != '':
            return f"[{msb}:{lsb}]"
    elif isinstance(width, str) and width.strip():
        return width
    return ""


def _get_port_param_info(port: Dict[str, Any]) -> Tuple[bool, str, int]:
    """
    Extract parameterization info from port.
    Returns: (is_parameterized, params_str, calculated_bit_width)
    Example: (True, "WEIGHT_WIDTH", 4)
    """
    is_param = port.get("is_parameterized", False)
    params_list = port.get("params_used", [])
    params_str = ",".join(params_list) if params_list else ""
    bit_width = port.get("calculated_bit_width", 0)
    return is_param, params_str, bit_width


def _format_port_with_width(port: Dict[str, Any], index: int) -> Tuple[str, bool]:
    """
    Format port as '[idx] name [width]'.
    Returns: (text, is_parameterized)
    For parameterized ports, display calculated bit width (e.g., [7:0]) not parameter name.
    Example: "[1] i_w1_cap [3:0]"  or  "[1] i_data [DATA_WIDTH-1:0] (8bits)" if unresolved
    """
    name = port.get('name', '?')
    width = _get_port_width(port)  # Original width from RTL (may contain parameters)
    is_param, params_str, bit_width = _get_port_param_info(port)
    
    # Check if width has unresolved parameters
    def has_param_expr(w: str) -> bool:
        import re
        return bool(re.search(r'[A-Za-z_]\w*', w)) if w else False
    
    if is_param:
        # Parameterized signal
        if bit_width > 0:
            # ✅ Show calculated bit width (e.g., [7:0]) instead of parameter expression
            formatted_width = f"[{bit_width-1}:0]"
            text = f"[{index+1}] {name} {formatted_width}"
        else:
            # ⚠️ Couldn't calculate - show parameter expression with warning
            text = f"[{index+1}] {name} {width}"  # e.g., [WEIGHT_WIDTH-1:0]
    elif width:
        text = f"[{index+1}] {name} {width}"
    else:
        text = f"[{index+1}] {name}"
    
    return text, is_param


def _draw_ports_two_columns(win: "curses._CursesWindow", ports: List[Dict[str, Any]], start_row: int = 1) -> None:
    """Draw ports in 2-column layout with bit width information.
    Only the bit width part [N:0] is shown in BLUE color for parameterized signals.
    """
    import re
    
    max_y, max_x = win.getmaxyx()
    usable_h = max_y - start_row - 1
    usable_w = max_x - 2
    
    # Split into two columns
    col_w = usable_w // 2
    col1_w = col_w - 1  # Space for separator
    col2_w = usable_w - col_w
    
    def draw_text_with_colored_width(win, row, col, text, max_w, is_param):
        """Draw text with only the width part [N:0] in blue if parameterized."""
        try:
            text = _truncate(text, max_w)
            
            # Find the bit width pattern [N:0] at the end
            match = re.search(r'\s+\[(\d+):0\]$', text)
            
            if is_param and match:
                # Split into: "[idx] name" and "[N:0]"
                width_start = match.start()
                prefix = text[:width_start]
                width_part = text[width_start:]
                
                # Draw prefix in normal color
                win.addstr(row, col, prefix, curses.A_NORMAL)
                
                # Draw width part in blue
                blue_pair = _PAIR_BY_NAME.get("blue", 0)
                win.addstr(row, col + len(prefix), width_part, curses.color_pair(blue_pair))
            else:
                # No parameterized width, draw all in normal color
                win.addstr(row, col, text, curses.A_NORMAL)
        except curses.error:
            pass
    
    row = start_row
    for i in range(0, len(ports), 2):
        if row >= max_y - 1:
            break
        
        # Left column
        if i < len(ports):
            left_text, is_param_left = _format_port_with_width(ports[i], i)
            draw_text_with_colored_width(win, row, 1, left_text, col1_w, is_param_left)
        
        # Right column
        if i + 1 < len(ports):
            right_text, is_param_right = _format_port_with_width(ports[i + 1], i + 1)
            draw_text_with_colored_width(win, row, col_w + 1, right_text, col2_w - 1, is_param_right)
        
        row += 1


def _truncate(text: str, width: int) -> str:
    if width <= 0:
        return ""
    if len(text) <= width:
        return text
    if width <= 1:
        return text[:width]
    return text[: width - 1] + "\N{HORIZONTAL ELLIPSIS}"


def build_context_from_rtl(rtl_start: Path, target_module: Optional[str]) -> Tuple[Dict[str, Any], ModuleInfo, List[Any]]:
    exts = [".v", ".sv"]
    # If a single file was provided, parse only that file; else use root + scope dir
    if rtl_start.is_file():
        files = [rtl_start]
        # Single file mode: allow unknown module types since we don't have all definitions
        allow_unknown = True
    else:
        rtl_root, _found = find_rtl_root_from(rtl_start)
        start_scope_dir = rtl_start if rtl_start.is_dir() else rtl_start.parent
        files = sorted(set(discover_files(rtl_root, exts)) | set(discover_files(start_scope_dir, exts)), key=lambda p: str(p))
        # Directory mode: we have all files, so enforce known types only
        allow_unknown = False
    
    modules = build_modules_db(files, allow_unknown=allow_unknown)
    if not modules:
        raise RuntimeError("No modules parsed from RTL scope")

    if not target_module:
        tops = find_top_modules(modules)
        target_module = tops[0] if tops else next(iter(modules.keys()))

    occs = find_occurrences_of_target(modules, target_module)
    
    # Build parameter environment with proper priority:
    # 1. If occurrence exists (module is instantiated), use hierarchy chain
    # 2. Otherwise, use module's default parameters
    external_params = {}
    
    if occs:
        # ✅ Module is instantiated somewhere - use occurrence-based environment
        # This handles external parameter overrides from parent modules
        env = compute_env_for_occurrence(occs[0], modules, external_params)
    else:
        # ⚠️ Module is not instantiated (or at top level)
        # Use the module's own default parameters
        if target_module in modules:
            target_mod = modules[target_module]
            if "param_defaults" in target_mod:
                external_params = dict(target_mod["param_defaults"])
        env = external_params
    ports_resolved = resolve_ports_with_params(modules, target_module, env)
    cls = classify_groups(modules[target_module]["ports"])
    ex_names = {x["name"] for x in cls.get("clocks", [])} | {x["name"] for x in cls.get("resets", [])}
    inputs_filtered = [it for it in ports_resolved["inputs"] if it["name"] not in ex_names]

    # Find the file path where this module is defined
    rtl_file_path = ""
    if target_module in modules and "file" in modules[target_module]:
        rtl_file_path = str(modules[target_module]["file"])

    # Extract parameters from param_defaults instead of port classification
    parameters = []
    if target_module in modules and "param_defaults" in modules[target_module]:
        param_defaults = modules[target_module]["param_defaults"]
        parameters = [{"name": k, "default": v} for k, v in param_defaults.items()]

    mi = ModuleInfo(
        module=target_module,
        rtl_file_path=rtl_file_path,
        inputs=inputs_filtered,
        outputs=ports_resolved["outputs"],
        inouts=ports_resolved["inouts"],
        clocks=cls.get("clocks", []),
        resets=cls.get("resets", []),
        parameters=parameters,
    )
    return modules, mi, occs


def _draw_box(win: "curses._CursesWindow", title: str) -> None:
    max_y, max_x = win.getmaxyx()
    win.box()
    label = f" {title} "
    if max_x > len(label) + 2:
        try:
            win.addstr(0, 2, label, curses.A_BOLD)
        except curses.error:
            pass


def _write_lines(win: "curses._CursesWindow", lines: List[str], start_y: int = 1, start_x: int = 1) -> None:
    max_y, max_x = win.getmaxyx()
    usable_h = max_y - start_y - 1
    usable_w = max_x - start_x - 1
    row = 0
    for line in lines:
        if row >= usable_h:
            break
        try:
            win.addnstr(start_y + row, start_x, line, usable_w)
        except curses.error:
            pass
        row += 1


def _write_lines_zebra(win: "curses._CursesWindow", lines: List[str], start_y: int = 1, start_x: int = 1, base_row_index: int = 0) -> None:
    max_y, max_x = win.getmaxyx()
    usable_h = max_y - start_y - 1
    usable_w = max_x - start_x - 1
    row = 0
    for i, line in enumerate(lines):
        if row >= usable_h:
            break
        attr = curses.A_DIM if ((base_row_index + i) % 2 == 1) else 0
        try:
            win.addnstr(start_y + row, start_x, line, usable_w, attr)
        except curses.error:
            pass
        row += 1


def _write_colored_zebra(win: "curses._CursesWindow", items: List[Tuple[str, Optional[str]]], start_y: int = 1, start_x: int = 1, base_row_index: int = 0) -> None:
    max_y, max_x = win.getmaxyx()
    usable_h = max_y - start_y - 1
    usable_w = max_x - start_x - 1
    row = 0
    for i, (line, color_name) in enumerate(items):
        if row >= usable_h:
            break
        pair = curses.color_pair(_PAIR_BY_NAME.get((color_name or "").lower(), 0))
        attr = pair | (curses.A_DIM if ((base_row_index + i) % 2 == 1) else 0)
        try:
            win.addnstr(start_y + row, start_x, line, usable_w, attr)
        except curses.error:
            pass
        row += 1


def _colorize_expression(expr: str) -> List[Tuple[str, str]]:
    """
    Parse expression and return list of (text, color) tuples for syntax highlighting.
    Colors: operators=blue(cyan), parentheses=yellow/magenta/green (nested), signals=default
    """
    if not expr:
        return [("", "")]
    
    result = []
    i = 0
    paren_depth = 0
    paren_colors = ["yellow", "magenta", "green", "cyan"]  # Cycle through colors for nested parens
    
    # Verilog operators (sorted by length, longest first for proper matching)
    # Logical: &&, ||, ==, !=, <=, >=, <, >, !
    # Bitwise: &, |, ^, ~
    # Arithmetic: +, -, *, /, %, **
    # Shift: <<, >>, <<<, >>>
    operators = [
        "<<<", ">>>",  # Arithmetic shifts (3 chars)
        "**", "&&", "||", "==", "!=", "<=", ">=", "<<", ">>",  # 2 chars
        "&", "|", "^", "~", "!", "<", ">", "+", "-", "*", "/", "%"  # 1 char
    ]
    
    while i < len(expr):
        # Check for operators (longest first)
        found_op = False
        for op in operators:
            if expr[i:i+len(op)] == op:
                result.append((op, "cyan"))
                i += len(op)
                found_op = True
                break
        
        if found_op:
            continue
        
        # Check for parentheses
        if expr[i] == '(':
            color = paren_colors[paren_depth % len(paren_colors)]
            result.append(('(', color))
            paren_depth += 1
            i += 1
        elif expr[i] == ')':
            paren_depth = max(0, paren_depth - 1)
            color = paren_colors[paren_depth % len(paren_colors)]
            result.append((')', color))
            i += 1
        else:
            # Regular character (signal name, whitespace, etc.)
            result.append((expr[i], ""))
            i += 1
    
    return result


def _write_colorized_expression(win: "curses._CursesWindow", y: int, x: int, expr: str, max_width: int, zebra_attr: int = 0) -> None:
    """Write expression with syntax highlighting"""
    tokens = _colorize_expression(expr)
    col = x
    
    for text, color_name in tokens:
        if col - x >= max_width:
            break
        
        if color_name:
            color_pair = curses.color_pair(_PAIR_BY_NAME.get(color_name, 0))
            attr = color_pair | zebra_attr
        else:
            attr = zebra_attr
        
        try:
            available = max_width - (col - x)
            display_text = text[:available]
            win.addstr(y, col, display_text, attr)
            col += len(display_text)
        except curses.error:
            break


def _format_kv_wrapped(items: List[Tuple[str, str]], total_width: int, label_width: int, add_blank_between: bool = True, value_color: Optional[str] = None) -> List[Tuple[str, Optional[str]]]:
    lines: List[Tuple[str, Optional[str]]] = []
    pad = " " * 2
    value_col = label_width + len(pad) + 2  # include ': '
    value_w = max(1, total_width - value_col)
    for label, value in items:
        lab = _truncate(label, label_width)
        # Use break_long_words=True to prevent expression truncation
        wrapped = _textwrap(value, width=value_w, break_long_words=True, break_on_hyphens=False) if value else [""]
        first = True
        for seg in wrapped:
            if first:
                line = f"{pad}{lab:{label_width}}: {seg}"
                first = False
            else:
                line = f"{pad}{'':{label_width}}  {seg}"
            lines.append((line, value_color))
        if add_blank_between:
            lines.append(("", None))
    return lines


def _paginate_list(items: List[str], page: int, page_size: int) -> Tuple[List[str], int]:
    if page_size <= 0:
        return items, 0
    total_pages = max(1, (len(items) + page_size - 1) // page_size)
    p = max(0, min(page, total_pages - 1))
    start = p * page_size
    end = min(len(items), start + page_size)
    return items[start:end], start


def _format_ports(title: str, ports: List[Dict[str, Any]], max_rows: int) -> List[str]:
    lines = [title]
    for idx, p in enumerate(ports[: max(0, max_rows - 1)]):
        width = p.get("width") or p.get("msb_lsb") or ""
        if isinstance(width, dict):
            width = f"[{width.get('msb', '')}:{width.get('lsb', '')}]"
        lines.append(f"  {p.get('name','?')} {width}")
    return lines


def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def _run_builder(state: AppState, do_fill: bool = False, do_json: bool = False, do_sv: bool = False) -> Tuple[int, str]:
    if not state.rtl_start or not state.target_module:
        return 2, "rtl/module must be set first"
    
    # Use session Excel path (not reference Excel)
    excel_path = state.session_excel_path
    if not excel_path or not Path(excel_path).exists():
        return 2, "session Excel not available - please complete onboarding"
    
    _ensure_dir(state.out_dir)
    builder = _THIS_DIR / "assertion_builder.py"
    if not builder.exists():
        return 2, "scripts/assertion_builder.py not found"
    cmd = [
        sys.executable,
        str(builder),
        "--rtl-start",
        str(state.rtl_start),
        "--target-module",
        state.target_module,
        "--excel",
        str(excel_path),
        "--out",
        str(state.out_dir),
    ]
    if do_fill:
        cmd.append("--auto-define-fill")
    if do_json:
        cmd.append("--json")
    # run_builder always attempts to generate SV if plugins succeed; allow explicit request
    # 'do_sv' flag is informational; we still run builder regardless
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True)
        out = (proc.stdout or "") + (proc.stderr or "")
        return proc.returncode, out
    except Exception as e:  # pragma: no cover
        return 2, f"execution failed: {e}"


def _list_modules_lines(modules: Dict[str, Any], max_rows: int) -> List[str]:
    names = sorted(modules.keys())
    return ["Discovered modules:"] + [f"  {n}" for n in names[: max_rows - 1]]


def run() -> None:
    try:
        curses.wrapper(_main)
    except Exception as e:
        # Graceful fallback for rare cases where stdscr is None or curses fails
        try:
            import platform
            if platform.system() == "Windows":
                print("[Error] TUI failed to initialize (curses). If not installed, run: pip install windows-curses")
        except Exception:
            pass
        raise


def _main(stdscr: "curses._CursesWindow") -> None:
    curses.curs_set(0)
    stdscr.nodelay(False)
    stdscr.keypad(True)
    curses.start_color()
    curses.use_default_colors()

    state = AppState()
    state.out_dir = Path("out/assertions")

    input_buf: List[str] = []
    cursor_pos = 0
    status_msg = "Type 'help' to get started. Set paths, then 'scan'."
    last_output: List[str] = []
    # Command history
    cmd_history: List[str] = []
    hist_idx: Optional[int] = None  # None means not browsing history
    # Path completion state
    compl_items: List[Tuple[str, bool]] = []  # (name, is_dir)
    compl_base: str = ""

    # Help overlay state
    help_pages = _load_help_pages()
    overlay_active = False
    overlay_page = 0
    overlay_scroll = 0
    # Help overlay highlight term/page
    global _OVERLAY_HL_KEY, _OVERLAY_HL_TERM, _OVERLAY_PAGE_KEY, _OVERLAY_FILTER_CMD, _ERROR_MESSAGE
    _OVERLAY_HL_KEY = None
    _OVERLAY_HL_TERM = None
    _OVERLAY_PAGE_KEY = None
    _OVERLAY_FILTER_CMD = None
    _ERROR_MESSAGE = ""
    _init_color_pairs()

    # Session load and first-run flow
    sessions = _load_sessions()
    # Always show session chooser (even if empty)
    chooser_result = _run_session_chooser(stdscr, sessions)
    if isinstance(chooser_result, dict):
        # Restore state from chosen session
        chosen = chooser_result
        try:
            state.rtl_start = Path(chosen.get("rtl_start", "")) if chosen.get("rtl_start") else None
        except Exception:
            state.rtl_start = None
        state.target_module = chosen.get("target_module") or None
        
        # Restore module hierarchy from session
        if "module_hierarchy" in chosen:
            state.module_info.module_hierarchy = chosen.get("module_hierarchy", "")
        
        # Restore clocks, resets, parameters from session.json
        if "clocks" in chosen:
            state.module_info.clocks = chosen.get("clocks", [])
        if "resets" in chosen:
            state.module_info.resets = chosen.get("resets", [])
        if "parameters" in chosen:
            state.module_info.parameters = chosen.get("parameters", [])
        
        # CRITICAL: Never restore excel_path (reference Excel)
        # Only restore session_excel_path
        state.excel_path = None  # Always None - we don't use reference paths
        
        try:
            state.session_excel_path = Path(chosen.get("session_excel_path", "")) if chosen.get("session_excel_path") else None
        except Exception:
            state.session_excel_path = None
        try:
            if chosen.get("out_dir"):
                state.out_dir = Path(chosen.get("out_dir")).resolve()
        except Exception:
            pass
        # Best-effort immediate scan if RTL is available
        if state.rtl_start:
            try:
                modules, mi, occs = build_context_from_rtl(state.rtl_start, state.target_module)
                state.modules_db = modules
                state.module_info = mi
                state.target_module = mi.module
                state.occs = occs
                
                # Try to find latest session Excel for this module if not already set
                if not state.session_excel_path:
                    latest_excel = _find_latest_session_excel(mi.module)
                    if latest_excel:
                        state.session_excel_path = latest_excel
            except Exception:
                pass
        
        # Restore condition signals from Excel after session Excel path is set
        if state.session_excel_path:
            _restore_conditions_from_excel(state)
            _restore_assertions_from_excel(state)  # Also restore assertions from Excel
    elif chooser_result == "new":
        state.onboarding_active = True
        state.onboarding_stage = 'rtl'
        state.onboarding_excel_autofound = _auto_find_excel()
    else:
        # quit or None → exit program
        return

    while True:
        max_y, max_x = stdscr.getmaxyx()
        stdscr.erase()

        if overlay_active:
            _render_help_overlay(stdscr, help_pages, overlay_page, overlay_scroll)
            stdscr.refresh()
            ch = stdscr.getch()
            if ch in (curses.KEY_RESIZE,):
                continue
            if ch in (27, ord('q'), ord('Q')):  # Esc or q
                overlay_active = False
                continue
            if ch in (curses.KEY_NPAGE,):
                overlay_scroll += 5
                continue
            if ch in (curses.KEY_PPAGE,):
                overlay_scroll = max(0, overlay_scroll - 5)
                continue
            # Page switch with n/N inside help
            if ch in (ord('n'),):
                overlay_page = (overlay_page + 1) % max(1, len(help_pages))
                overlay_scroll = 0
                continue
            if ch in (ord('N'),):
                overlay_page = (overlay_page - 1) % max(1, len(help_pages))
                overlay_scroll = 0
                continue
            if ch in (curses.KEY_UP,):
                overlay_scroll = max(0, overlay_scroll - 1)
                continue
            if ch in (curses.KEY_DOWN,):
                overlay_scroll += 1
                continue
            if ch in (9,):  # Tab next page
                overlay_page = (overlay_page + 1) % max(1, len(help_pages))
                overlay_scroll = 0
                continue
            if ch in (curses.KEY_BTAB,):  # Shift-Tab prev page
                overlay_page = (overlay_page - 1) % max(1, len(help_pages))
                overlay_scroll = 0
                continue
            # Any other key: ignore while overlay is open
            continue

        # Assertion wizard rendering
        if state.assertion_wizard_active:
            try:
                _render_assertion_wizard(stdscr, state)
            except Exception as e:
                try:
                    stdscr.addnstr(2, 2, f"ERROR in wizard rendering: {str(e)[:60]}", 80, curses.A_BOLD | curses.color_pair(_PAIR_BY_NAME.get("red", 0)))
                except:
                    pass
            max_y, max_x = stdscr.getmaxyx()
            
            # Hints
            if state.assertion_wizard_stage == 'select_type':
                hint_line = "Enter assertion type number [1-4] | 'q' to quit"
            elif state.assertion_wizard_stage == 'input_data':
                hint_line = "Enter value | [Enter] to next | 'b' for previous | 'q' to cancel"
            elif state.assertion_wizard_stage == 'confirm':
                hint_line = "[Enter] to create | 'b' to edit | 'q' to cancel"
            else:
                hint_line = ""
            try:
                stdscr.addnstr(max_y - 3, 2, _truncate(hint_line, max_x - 4), max_x - 4, curses.A_DIM)
            except curses.error:
                pass
            
            # Input prompt
            prompt = "> "
            edit_w = max_x - len(prompt) - 1
            current = "".join(input_buf)
            try:
                stdscr.addnstr(max_y - 1, 0, prompt, len(prompt))
                stdscr.addnstr(max_y - 1, len(prompt), _truncate(current, edit_w), edit_w)
                curses.curs_set(1)
                stdscr.move(max_y - 1, len(prompt) + min(cursor_pos, edit_w))
            except curses.error:
                pass
            
            stdscr.refresh()
            
            # Handle Assertion Wizard input
            ch = stdscr.getch()
            if ch in (curses.KEY_RESIZE,):
                continue
            if ch in (curses.KEY_BACKSPACE, 127, 8):
                if cursor_pos > 0:
                    del input_buf[cursor_pos - 1]
                    cursor_pos -= 1
                continue
            if ch in (curses.KEY_LEFT,):
                cursor_pos = max(0, cursor_pos - 1)
                continue
            if ch in (curses.KEY_RIGHT,):
                cursor_pos = min(len(input_buf), cursor_pos + 1)
                continue
            if ch in (10, 13):  # Enter
                cmdline = "".join(input_buf).strip()
                input_buf.clear()
                cursor_pos = 0
                msg, exit_wizard = _handle_assertion_wizard_command(state, cmdline)
                status_msg = msg
                if exit_wizard:
                    state.assertion_wizard_active = False
                    input_buf.clear()
                    cursor_pos = 0
                continue
            if ch in (ord('q'), ord('Q')):  # Quick quit
                state.assertion_wizard_active = False
                input_buf.clear()
                cursor_pos = 0
                status_msg = "Assertion wizard cancelled"
                continue
            # Regular char
            if 0 <= ch <= 255:
                try:
                    c = chr(ch)
                except Exception:
                    c = ""
                if c:
                    input_buf.insert(cursor_pos, c)
                    cursor_pos += 1
            continue
        
        elif state.gen_wizard_active:
            # File generation wizard rendering
            _render_gen_wizard(stdscr, state)
            max_y, max_x = stdscr.getmaxyx()
            
            # Hints for file generation wizard
            if state.gen_wizard_stage == 'filename':
                hint_line = "Enter output filename (no extension) | 'q' to cancel"
            elif state.gen_wizard_stage == 'file_type':
                hint_line = "Choose file type: 1=Interface(.if.sv) 2=Instance(.inst.sv) 3=Both | 'q' to cancel"
            elif state.gen_wizard_stage == 'data_source':
                hint_line = "Choose data: 1=Assertions 2=Signals 3=Both | 'b' to back | 'q' to cancel"
            elif state.gen_wizard_stage == 'preview':
                if state.gen_file_type == 3:
                    hint_line = "[Enter] to generate | n/N scroll | 'f' switch file | 'b' edit | 'q' cancel"
                else:
                    hint_line = "[Enter] to generate | n/N scroll | 'b' edit | 'q' cancel"
            else:
                hint_line = ""
            try:
                stdscr.addnstr(max_y - 3, 2, _truncate(hint_line, max_x - 4), max_x - 4, curses.A_DIM)
            except curses.error:
                pass
            
            # Input prompt
            prompt = "> "
            edit_w = max_x - len(prompt) - 1
            current = "".join(input_buf)
            try:
                stdscr.addnstr(max_y - 1, 0, prompt, len(prompt))
                stdscr.addnstr(max_y - 1, len(prompt), _truncate(current, edit_w), edit_w)
                curses.curs_set(1)
                stdscr.move(max_y - 1, len(prompt) + min(cursor_pos, edit_w))
            except curses.error:
                pass
            
            stdscr.refresh()
            
            # Handle input
            ch = stdscr.getch()
            if ch in (curses.KEY_RESIZE,):
                continue
            if ch in (curses.KEY_BACKSPACE, 127, 8):
                if cursor_pos > 0:
                    del input_buf[cursor_pos - 1]
                    cursor_pos -= 1
                continue
            if ch in (curses.KEY_LEFT,):
                cursor_pos = max(0, cursor_pos - 1)
                continue
            if ch in (curses.KEY_RIGHT,):
                cursor_pos = min(len(input_buf), cursor_pos + 1)
                continue
            if ch in (10, 13):  # Enter
                cmdline = "".join(input_buf).strip()
                input_buf.clear()
                cursor_pos = 0
                
                # Handle file generation wizard stages
                if state.gen_wizard_stage == 'filename':
                    if not cmdline:
                        status_msg = "ERROR: Filename cannot be empty"
                        continue
                    state.gen_filename = cmdline
                    state.gen_wizard_stage = 'file_type'
                    state.gen_file_type = None
                elif state.gen_wizard_stage == 'file_type':
                    if cmdline in ('1', '2', '3'):
                        state.gen_file_type = int(cmdline)
                        state.gen_wizard_stage = 'data_source'
                        state.gen_data_source = None
                    else:
                        status_msg = "ERROR: Enter 1, 2, or 3"
                elif state.gen_wizard_stage == 'data_source':
                    if cmdline in ('1', '2', '3'):
                        state.gen_data_source = cmdline
                        state.gen_wizard_stage = 'preview'
                        state.gen_preview_page = 0
                        state.gen_preview_file_idx = 0
                        # Generate full preview content
                        state.gen_preview_lines = _generate_preview_content(state)
                    elif cmdline == 'b':
                        state.gen_wizard_stage = 'file_type'
                    else:
                        status_msg = "ERROR: Enter 1, 2, 3, or 'b' to go back"
                elif state.gen_wizard_stage == 'preview':
                    if cmdline == '':  # Empty = generate
                        msg = _generate_files(state)
                        status_msg = msg
                        state.gen_wizard_active = False
                    elif cmdline == 'n':  # Next page
                        state.gen_preview_page += 1
                        status_msg = "Scrolled down"
                    elif cmdline == 'N':  # Previous page
                        state.gen_preview_page = max(0, state.gen_preview_page - 1)
                        status_msg = "Scrolled up"
                    elif cmdline == 'f':  # Switch file (for "both" mode)
                        if state.gen_file_type == 3:  # Both files
                            state.gen_preview_file_idx = 1 - state.gen_preview_file_idx
                            state.gen_preview_page = 0
                            state.gen_preview_lines = _generate_preview_content(state)
                            file_name = ["Interface (.if.sv)", "Instance (.inst.sv)"][state.gen_preview_file_idx]
                            status_msg = f"Switched to {file_name}"
                        else:
                            status_msg = "Only one file type selected"
                    elif cmdline == 'b':
                        state.gen_wizard_stage = 'data_source'
                    elif cmdline == 'q' or cmdline == 'Q':
                        state.gen_wizard_active = False
                        status_msg = "File generation cancelled"
                    else:
                        status_msg = "ERROR: Press Enter to generate, n/N to scroll, 'f' to switch file, or 'b' to edit"
                elif cmdline == 'q' or cmdline == 'Q':
                    state.gen_wizard_active = False
                    status_msg = "File generation cancelled"
                continue
            # Regular char
            if 0 <= ch <= 255:
                try:
                    c = chr(ch)
                except Exception:
                    c = ""
                if c:
                    input_buf.insert(cursor_pos, c)
                    cursor_pos += 1
            continue

        # Onboarding wizard rendering
        if state.onboarding_active:
            # Render onboarding UI
            # Avoid full clear here to prevent flicker of transient popups. Use erase only.
            try:
                stdscr.erase()
            except Exception:
                pass
            _render_onboarding(stdscr, state)
            max_y, max_x = stdscr.getmaxyx()
            # Optional status line for onboarding (errors/success) above candidates/hints
            if status_msg:
                try:
                    stdscr.addnstr(max_y - 5, 2, _truncate(status_msg, max_x - 4), max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("red",0)) | curses.A_BOLD)
                except curses.error:
                    pass
            # Decide dynamic candidates strip height before placing hint/prompt (only in RTL step)
            # Estimate grid based on current compl_items and screen width
            cand_h = 0
            prompt_row = max_y - 2
            if (state.onboarding_stage or "") == 'rtl' and compl_items:
                try:
                    items_all = [nm + ("/" if is_dir else "") for (nm, is_dir) in compl_items]
                    longest = max((len(s) for s in items_all), default=8)
                    col_w = min(max(12, longest + 2), 32)
                    cols = max(1, (max_x - 4) // col_w)
                    rows_needed = max(1, (len(items_all) + cols - 1) // cols)
                    # Header + rows, limit to 10 and to available space
                    cand_h = min(10, rows_needed + 1, max(0, max_y - 4))
                    cand_h = max(3, cand_h)
                    state.onboarding_cand_h = cand_h
                    state.onboarding_cand_visible = True
                except Exception:
                    cand_h = 3
                    state.onboarding_cand_h = cand_h
                    state.onboarding_cand_visible = True
            # Anchor candidates to the line just above the prompt; grow upwards
            if cand_h:
                cand_bottom = prompt_row - 1
                cand_top = max(1, cand_bottom - (cand_h - 1))
                hint_y = max(1, cand_top - 1)
            else:
                cand_top = 0
                hint_y = prompt_row - 1
            try:
                stdscr.move(hint_y, 0)
                stdscr.clrtoeol()
                stage = (state.onboarding_stage or "").lower()
                if stage == 'rtl':
                    hint_text = "Enter path and press Enter. Tab: complete. Esc: cancel."
                elif stage == 'module':
                    hint_text = "Type number. f <text>: filter, F: clear, n/N: page, prev/back: previous"
                elif stage == 'hierarchy':
                    hint_text = "Enter module hierarchy (e.g., top.dut.abc.u_abc). prev/back: previous"
                else:
                    hint_text = "Type path and Enter (or accept autodetected). prev/back: previous"
                stdscr.addnstr(hint_y, 2, _truncate(hint_text, max_x - 4), max_x - 4, curses.A_DIM)
            except curses.error:
                pass

            # (popup rendered after prompt, just before blocking getch)
            # Input prompt (force clear to prevent ghosting) placed near bottom
            prompt = "> "
            edit_w = max_x - len(prompt) - 1
            current = "".join(input_buf)
            try:
                stdscr.addnstr(max_y - 2, 0, " " * max_x, max_x)
                stdscr.addstr(max_y - 2, 0, prompt)
                stdscr.addnstr(max_y - 2, len(prompt), _truncate(current, edit_w), edit_w)
                try:
                    curses.curs_set(1)
                except Exception:
                    pass
                stdscr.move(max_y - 2, len(prompt) + min(cursor_pos, edit_w))
            except curses.error:
                pass
            # Draw candidates strip last on stdscr (wide, up to 10 lines) so it persists reliably and doesn't overlap (RTL step only)
            if (state.onboarding_stage or "") == 'rtl' and compl_items and cand_h:
                try:
                    # Recompute anchored top using prompt row and current cand_h so it won't drift
                    prompt_row = max_y - 2
                    cand_bottom = prompt_row - 1
                    cand_top = max(1, cand_bottom - (cand_h - 1))
                    # Clear the region
                    for r in range(cand_h):
                        stdscr.move(cand_top + r, 0); stdscr.clrtoeol()
                    items_all = [nm + ("/" if is_dir else "") for (nm, is_dir) in compl_items]
                    # Header on first line
                    stdscr.addnstr(cand_top, 2, _truncate("Candidates:", max_x - 4), max_x - 4, curses.A_BOLD)
                    # Grid layout for alignment
                    grid_rows = max(1, cand_h - 1)
                    # Determine column width from longest label, clamp for readability
                    longest = max((len(s) for s in items_all), default=8)
                    col_w = min(max(12, longest + 2), 32)
                    cols = max(1, (max_x - 4) // col_w)
                    visible_capacity = grid_rows * cols
                    # Paging
                    total_pages = max(1, (len(items_all) + visible_capacity - 1) // visible_capacity)
                    cur_page = state.onboarding_cand_page % total_pages
                    start_idx = cur_page * visible_capacity
                    items = items_all[start_idx:start_idx + visible_capacity]
                    for r in range(grid_rows):
                        for c in range(cols):
                            idx = r * cols + c
                            if idx >= len(items):
                                break
                            label = items[idx]
                            x = 2 + c * col_w
                            y = cand_top + 1 + r
                            stdscr.addnstr(y, x, _truncate(label, col_w - 1), col_w - 1, curses.color_pair(_PAIR_BY_NAME.get("cyan",0)))
                    # If truncated, show (more)
                    if total_pages > 1:
                        more_text = f"... (more {cur_page+1}/{total_pages})"
                        stdscr.addnstr(cand_top + grid_rows, max(2, max_x - len(more_text) - 2), _truncate(more_text, len(more_text)), len(more_text), curses.A_DIM)
                except curses.error:
                    pass
            stdscr.refresh()
            ch = stdscr.getch()
            # Basic editing keys
            if ch in (curses.KEY_RESIZE,):
                continue
            if ch in (27,):  # Esc cancels onboarding
                state.onboarding_active = False
                state.onboarding_stage = None
                input_buf.clear(); cursor_pos = 0
                continue
            if ch in (curses.KEY_BACKSPACE, 127, 8):
                if cursor_pos > 0:
                    del input_buf[cursor_pos - 1]
                    cursor_pos -= 1
                continue
            if ch in (curses.KEY_DC,):
                if cursor_pos < len(input_buf):
                    del input_buf[cursor_pos]
                continue
            if ch in (curses.KEY_LEFT,):
                cursor_pos = max(0, cursor_pos - 1); continue
            if ch in (curses.KEY_RIGHT,):
                cursor_pos = min(len(input_buf), cursor_pos + 1); continue
            if ch in (curses.KEY_HOME,):
                cursor_pos = 0; continue
            if ch in (curses.KEY_END,):
                cursor_pos = len(input_buf); continue
            # Tab completion for path (raw mode in onboarding): Linux-like
            if ch == 9:
                line = "".join(input_buf)
                # In onboarding, treat whole line as path prefix
                new_line, new_cursor, items, base_dir = _path_complete_raw(line, cursor_pos)
                if items:
                    # Show candidates; insert ONLY common prefix increment (no first-item commit)
                    compl_base = base_dir
                    compl_items = items
                    names = [nm for (nm, _isdir) in items]
                    common = _common_prefix(names)
                    base_str = str(base_dir).rstrip("/\\").replace("\\", "/")
                    next_prefix = (base_str + "/" + common) if base_str else common
                    # If common grows beyond current typed prefix, extend input by the new delta only
                    if next_prefix and not line.endswith(next_prefix):
                        input_buf = list(next_prefix)
                        cursor_pos = len(input_buf)
                        # typing happened implicitly → ensure candidate page stays at 0 on first Tab
                        if not state.onboarding_cand_visible:
                            state.onboarding_cand_page = 0
                        state.onboarding_cand_visible = True
                    else:
                        # No new common part → advance candidates page (wrap)
                        # Compute paging using same geometry as renderer
                        max_y, max_x = stdscr.getmaxyx()
                        # Recompute dynamic grid
                        longest = max((len(nm + ('/' if isd else '')) for (nm, isd) in items), default=8)
                        col_w = min(max(12, longest + 2), 32)
                        cols = max(1, (max_x - 4) // col_w)
                        # Use last computed cand_h from state to keep baseline anchored
                        grid_rows = max(1, (state.onboarding_cand_h - 1) if state.onboarding_cand_h else 6)
                        visible_capacity = grid_rows * cols
                        total_pages = max(1, (len(items) + visible_capacity - 1) // visible_capacity)
                        state.onboarding_cand_page = (state.onboarding_cand_page + 1) % total_pages
                else:
                    # No candidates: apply common-prefix if any; otherwise do nothing (beep)
                    if new_line != line:
                        input_buf = list(new_line)
                        cursor_pos = new_cursor
                    else:
                        try:
                            curses.beep()
                        except Exception:
                            pass
                continue
            # Enter: commit
            if ch in (10, 13, curses.KEY_ENTER):
                cmdline = "".join(input_buf).strip()
                stage = state.onboarding_stage or ""
                # Default: clear status unless error occurs
                status_msg = ""
                # RTL stage is now handled by _handle_command to ensure full instance discovery
                if stage == 'rtl':
                    # Delegate to command handler which has full instance discovery logic
                    out_msg, opened_overlay = _handle_command(state, cmdline)
                    input_buf.clear(); cursor_pos = 0
                    if opened_overlay:
                        overlay_active = True
                        overlay_page = 0
                        overlay_scroll = 0
                    else:
                        status_msg = out_msg or status_msg
                elif stage == 'hierarchy':
                    # Handle hierarchy input: empty (use default), number (select from list), or custom string
                    if not cmdline.strip():
                        # Empty input - use already selected hierarchy (first one)
                        if state.module_info.module_hierarchy:
                            state.onboarding_stage = 'excel'
                            status_msg = f"Using hierarchy: {state.module_info.module_hierarchy}"
                            input_buf.clear(); cursor_pos = 0
                            try:
                                stdscr.clear(); stdscr.refresh()
                            except Exception:
                                pass
                        else:
                            status_msg = "No hierarchy detected. Please enter manually."
                    elif cmdline.strip().isdigit():
                        # Number - select from detected hierarchies
                        idx = int(cmdline.strip()) - 1
                        if state.occs and 0 <= idx < len(state.occs):
                            selected_hierarchy = state.occs[idx].get("path", "")
                            if selected_hierarchy:
                                state.module_info.module_hierarchy = selected_hierarchy
                                state.onboarding_stage = 'excel'
                                status_msg = f"Hierarchy set: {selected_hierarchy}"
                                input_buf.clear(); cursor_pos = 0
                                try:
                                    stdscr.clear(); stdscr.refresh()
                                except Exception:
                                    pass
                            else:
                                status_msg = f"Invalid selection: no path for index {idx + 1}"
                        else:
                            status_msg = f"Invalid number. Choose 1-{len(state.occs) if state.occs else 0}"
                    else:
                        # Custom string
                        state.module_info.module_hierarchy = cmdline.strip()
                        state.onboarding_stage = 'excel'
                        status_msg = f"Hierarchy set: {cmdline.strip()}"
                        input_buf.clear(); cursor_pos = 0
                        try:
                            stdscr.clear(); stdscr.refresh()
                        except Exception:
                            pass
                elif stage == 'excel':
                    # Delegate to command handler (supports empty accept)
                    out_msg, opened_overlay = _handle_command(state, cmdline)
                    input_buf.clear(); cursor_pos = 0
                    if opened_overlay:
                        overlay_active = True
                        overlay_page = 0
                        overlay_scroll = 0
                    else:
                        status_msg = out_msg or status_msg
                else:
                    # For 'module' and any other, delegate to handler
                    out_msg, opened_overlay = _handle_command(state, cmdline)
                    input_buf.clear(); cursor_pos = 0
                    if opened_overlay:
                        overlay_active = True
                        overlay_page = 0
                        overlay_scroll = 0
                    else:
                        status_msg = out_msg or status_msg
                continue
            # Regular char (any typing clears candidates)
            if 0 <= ch <= 255:
                try:
                    c = chr(ch)
                except Exception:
                    c = ""
                if c:
                    input_buf.insert(cursor_pos, c)
                    cursor_pos += 1
                    compl_items = []
            continue

        # Normal dashboard rendering
        # Clear any leftover messages from onboarding
        if not state.onboarding_active and "Session created" in status_msg:
            last_output = []
            status_msg = "Ready"
        
        # 레이아웃 구조:
        # - Left: Module/Paths (top) + Clocks/Resets/Params (bottom) - 20% width, full height
        # - Right (80%): 상단 60% - Inputs | Outputs | Condition Signals (3등분)
        #                하단 40% - Created Assertions (전체 너비)
        
        left_w = max(24, int(max_x * 0.20))  # Fixed 20% for left panel
        right_w = max_x - left_w  # 80% for right side
        
        # Right side 상단 60%, 하단 40%
        right_top_h = max(6, int((max_y - 3) * 0.60))  # -3 for prompt area
        right_bottom_h = max(6, max_y - 3 - right_top_h)
        
        # Right top area split into 3 equal columns: Inputs | Outputs | Condition Signals
        col_w = right_w // 3
        in_w = col_w
        out_w = col_w
        cond_w = right_w - in_w - out_w  # Remaining width
        
        # Create windows
        # Left side: Module/Paths dynamically sized, Clocks/Resets/Params fills remainder
        # Calculate Module/Paths height based on content
        kv_items_preview = [
            ("Hierarchy", state.module_info.module_hierarchy or "(not set)"),
            ("Module", state.module_info.module or (state.target_module or "")),
            ("Excel", "placeholder"),
            ("Out", _sanitize_path_for_display(_safe_str(state.out_dir))),
        ]
        inner_w_preview = left_w - 2
        max_label_preview = max(len(k) for k, _ in kv_items_preview) if kv_items_preview else 8
        label_w_preview = min(max(8, max_label_preview), max(8, inner_w_preview // 3))
        kv_tuples_preview = _format_kv_wrapped(kv_items_preview, total_width=inner_w_preview, label_width=label_w_preview, add_blank_between=True, value_color=None)
        
        # Module/Paths height: title(1) + border(2) + content + margin(1)
        left_top_h = min(len(kv_tuples_preview) + 4, max_y - 10)  # Leave space for Clocks/Resets
        left_bot_h = max(6, max_y - 3 - left_top_h)  # Extend to bottom
        
        win_left_top = curses.newwin(left_top_h, left_w, 0, 0)
        win_left_bot = curses.newwin(left_bot_h, left_w, left_top_h, 0)
        
        # Right top: Inputs | Outputs | Condition Signals
        win_in = curses.newwin(right_top_h, in_w, 0, left_w)
        win_out = curses.newwin(right_top_h, out_w, 0, left_w + in_w)
        win_cond = curses.newwin(right_top_h, cond_w, 0, left_w + in_w + out_w)
        
        # Right bottom: Created Assertions (spans full right width)
        win_assertions = curses.newwin(right_bottom_h, right_w, right_top_h, left_w)

        # Draw left-top: Module/IP info (two-column KV with wrapping and zebra)
        _draw_box(win_left_top, "Module / Paths")
        
        # CRITICAL: Only show session Excel path, never reference Excel
        if state.session_excel_path and Path(state.session_excel_path).exists():
            excel_show = state.session_excel_path
            excel_error = False
        else:
            excel_show = ""
            excel_error = True
        
        kv_items = [
            ("Hierarchy", state.module_info.module_hierarchy or "(not set)"),
            ("Module", state.module_info.module or (state.target_module or "")),
            ("Excel", _sanitize_path_for_display(_safe_str(excel_show)) if excel_show else "ERROR: Session Excel not loaded"),
            ("Out", _sanitize_path_for_display(_safe_str(state.out_dir))),
        ]
        inner_w = left_w - 2
        max_label = max(len(k) for k, _ in kv_items) if kv_items else 8
        label_w = min(max(8, max_label), max(8, inner_w // 3))
        
        # Color Excel line red if missing
        value_color = None
        kv_tuples = _format_kv_wrapped(kv_items, total_width=inner_w, label_width=label_w, add_blank_between=True, value_color=value_color)
        
        # Re-color Excel line to red if error
        recolored: List[Tuple[str, Optional[str]]] = []
        for line, color in kv_tuples:
            if line.strip().startswith("Excel") and excel_error:
                recolored.append((line, "red"))
            else:
                recolored.append((line, color))
        
        # Write KV section
        row_ptr_left = 1
        _write_colored_zebra(win_left_top, recolored, row_ptr_left, 1, base_row_index=0)

        # Draw left-bottom: clocks/resets/params with numbering for clocks/resets
        _draw_box(win_left_bot, "Clocks / Resets / Params")
        left_bot_h, left_bot_w = win_left_bot.getmaxyx()
        row_ptr_left_bot = 1  # Start from row 1
        
        try:
            # Format parameters with default values using [i] numbering
            param_lines = []
            for i, p in enumerate(state.module_info.parameters):
                pname = p.get('name', '?')
                pdefault = p.get('default', '')
                if pdefault:
                    line = f"  [{i+1}] {pname} : {pdefault}"
                else:
                    line = f"  [{i+1}] {pname}"
                # Wrap if line is too long
                usable_w = max(1, left_bot_w - 4)  # Leave margin for wrapping, ensure at least 1
                if len(line) > usable_w:
                    wrapped = _wrap_text(line, usable_w)
                    param_lines.extend(wrapped)
                else:
                    param_lines.append(line)
        except Exception as e:
            param_lines = [f"  ERROR formatting params: {e}"]
        
        sections = [
            ("Clocks:", [f"  [{i+1}] {c.get('name','?')}" for i, c in enumerate(state.module_info.clocks)]),
            ("Resets:", [f"  [{i+1}] {r.get('name','?')}" for i, r in enumerate(state.module_info.resets)]),
            ("Params:", param_lines if param_lines else []),
        ]
        for title, items in sections:
            if row_ptr_left_bot >= left_bot_h - 1:
                break
            lines = [title] + items + [""]
            _write_lines_zebra(win_left_bot, lines, row_ptr_left_bot, 1, base_row_index=0)
            row_ptr_left_bot += len(lines)

        # Draw Inputs and Outputs columns with 2-column layout, numbering, and bit width
        _draw_box(win_in, "Inputs")
        _draw_box(win_out, "Outputs")
        
        # Apply port filter if set
        def _apply_filter(pl: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
            if not state.port_filter:
                return pl
            key = state.port_filter.lower()
            return [p for p in pl if key in (p.get('name','').lower())]
        
        # Include inouts into both lists if any
        in_ports = _apply_filter(state.module_info.inputs + state.module_info.inouts)
        out_ports = _apply_filter(state.module_info.outputs + state.module_info.inouts)
        
        # Pagination support (global _ports_page)
        global _ports_page
        try:
            _ports_page
        except NameError:
            _ports_page = 0
        
        in_h, in_w2 = win_in.getmaxyx()
        out_h, out_w2 = win_out.getmaxyx()
        avail_rows_in = in_h - 2
        avail_rows_out = out_h - 2
        
        # Calculate how many ports fit per page (2 columns)
        ports_per_page_in = avail_rows_in * 2
        ports_per_page_out = avail_rows_out * 2
        
        # Paginate
        in_start = _ports_page * ports_per_page_in
        out_start = _ports_page * ports_per_page_out
        in_page_ports = in_ports[in_start:in_start + ports_per_page_in]
        out_page_ports = out_ports[out_start:out_start + ports_per_page_out]
        
        # Draw in 2-column layout
        _draw_ports_two_columns(win_in, in_page_ports, start_row=1)
        _draw_ports_two_columns(win_out, out_page_ports, start_row=1)
        
        # Add pagination info and navigation hint at bottom of Input/Output boxes
        try:
            in_total_pages = max(1, (len(in_ports) + ports_per_page_in - 1) // ports_per_page_in)
            out_total_pages = max(1, (len(out_ports) + ports_per_page_out - 1) // ports_per_page_out)
            in_cur_page = (_ports_page % in_total_pages) + 1 if in_total_pages > 0 else 1
            out_cur_page = (_ports_page % out_total_pages) + 1 if out_total_pages > 0 else 1
            
            in_info = f"Page {in_cur_page}/{in_total_pages} | n/N to scroll"
            out_info = f"Page {out_cur_page}/{out_total_pages} | n/N to scroll"
            
            win_in.addnstr(in_h - 1, 2, _truncate(in_info, in_w2 - 4), in_w2 - 4, curses.A_DIM)
            win_out.addnstr(out_h - 1, 2, _truncate(out_info, out_w2 - 4), out_w2 - 4, curses.A_DIM)
        except curses.error:
            pass
        
        # Draw rightmost: Condition Signals with syntax highlighting
        _draw_box(win_cond, "Condition Signals (ms)")
        cond_h, cond_w2 = win_cond.getmaxyx()
        cond_inner_w = cond_w2 - 2
        cond_usable_h = cond_h - 2
        
        # Manually render with colorized expressions
        pad = " " * 2
        indent = " " * 6  # Indentation for wrapped expression lines
        row = 1
        for i, cond in enumerate(state.conditions):
            if row >= cond_h - 1:
                break
            
            nm = cond.get('name', '')
            bits = cond.get('width', 1)
            expr = cond.get('expr', '')
            
            # Label with bits
            label = f"{nm} ({bits}bits)" if bits and bits > 1 else nm
            
            # Calculate layout - if label is too long, put expression on next line
            max_label_w = min(25, cond_inner_w // 3)
            
            # Zebra striping
            zebra_attr = curses.A_DIM if (i % 2 == 1) else 0
            
            # Check if label is too long (> 20 chars) - if so, put expr on next line
            if len(label) > 20:
                # Long label: write label on one line, expr on next with indent
                label_line = f"{pad}{label}:"
                try:
                    win_cond.addstr(row, 1, _truncate(label_line, cond_inner_w), zebra_attr)
                except curses.error:
                    pass
                row += 1
                
                # Write expression on next line with indent
                expr_start_x = 1 + len(pad) + len(indent)
                expr_max_w = cond_inner_w - len(pad) - len(indent)
                
                expr_wrapped = _textwrap(expr, width=expr_max_w, break_long_words=True, break_on_hyphens=False) if expr else [""]
                
                for j, expr_line in enumerate(expr_wrapped):
                    if row >= cond_h - 1:
                        break
                    
                    indent_line = f"{pad}{indent}"
                    try:
                        win_cond.addstr(row, 1, indent_line, zebra_attr)
                    except curses.error:
                        pass
                    
                    _write_colorized_expression(win_cond, row, expr_start_x, expr_line, expr_max_w, zebra_attr)
                    
                    if j < len(expr_wrapped) - 1:  # More lines to come
                        row += 1
            else:
                # Short label: write label and expr on same line
                label_truncated = _truncate(label, max_label_w)
                label_line = f"{pad}{label_truncated:{max_label_w}}: "
                
                try:
                    win_cond.addstr(row, 1, label_line, zebra_attr)
                except curses.error:
                    pass
                
                # Calculate expression space correctly
                expr_start_x = len(label_line) + 1
                expr_max_w = cond_inner_w - len(label_line) + len(pad)
                
                # Wrap expression if needed
                expr_wrapped = _textwrap(expr, width=expr_max_w, break_long_words=True, break_on_hyphens=False) if expr else [""]
                
                for j, expr_line in enumerate(expr_wrapped):
                    if row >= cond_h - 1:
                        break
                    
                    if j == 0:
                        # First line: write after label
                        _write_colorized_expression(win_cond, row, expr_start_x, expr_line, expr_max_w, zebra_attr)
                    else:
                        # Continuation lines: indent
                        row += 1
                        continuation = f"{pad}{'':{max_label_w}}  "
                        try:
                            win_cond.addstr(row, 1, continuation, zebra_attr)
                        except curses.error:
                            pass
                        cont_x = len(continuation) + 1
                        cont_max_w = cond_inner_w - len(continuation) + len(pad)
                        _write_colorized_expression(win_cond, row, cont_x, expr_line, cont_max_w, zebra_attr)
            
            row += 1
            # Add blank line between items
            row += 1

        # Draw bottom: Assertion List (spans full width)
        _draw_box(win_assertions, "Created Assertions")
        assert_h, assert_w = win_assertions.getmaxyx()
        assert_inner_h = assert_h - 2
        assert_inner_w = assert_w - 2
        
        if not state.assertions:
            no_assert_msg = "No assertions created yet. Use 'new' command to create assertions."
            try:
                win_assertions.addnstr(1, 2, _truncate(no_assert_msg, assert_inner_w), assert_inner_w, curses.A_DIM)
            except curses.error:
                pass
        else:
            # Display assertions in table format: # | Type | Description
            # Column widths: Index(3) | Type(12) | Description(remaining)
            idx_w = 3
            type_w = 12
            desc_w = max(40, assert_inner_w - idx_w - type_w - 6)  # -6 for separators
            
            # Header
            header = f"{'#':<{idx_w}} | {'Type':<{type_w}} | {'Description':<{desc_w}}"
            try:
                win_assertions.addnstr(1, 2, _truncate(header, assert_inner_w), assert_inner_w, curses.A_BOLD)
            except curses.error:
                pass
            
            # Separator
            sep = "-" * min(idx_w + type_w + desc_w + 6, assert_inner_w)
            try:
                win_assertions.addnstr(2, 2, _truncate(sep, assert_inner_w), assert_inner_w)
            except curses.error:
                pass
            
            # Assertion rows
            row = 3
            for i, asrt in enumerate(state.assertions, start=1):
                if row >= assert_h - 1:
                    break
                
                atype = asrt.get('type', 'Unknown')
                adata = asrt.get('data', {})
                
                # Extract signal name and build natural language description
                # user_input_parts: list of (text, column_offset) for user-provided values
                full_desc = ""
                user_input_parts = []  # List to store user-provided values for highlighting in GREEN
                
                if atype == 'counter':
                    signal_name = adata.get('target', '?')
                    exp_cnt = adata.get('exp_cnt_val', '?')
                    # Truncate signal name to 15 chars
                    if len(str(signal_name)) > 15:
                        display_signal = str(signal_name)[:12] + "..."
                    else:
                        display_signal = str(signal_name)
                    full_desc = f"Monitor {display_signal} counter reaching {exp_cnt}"
                    # Both signal name and counter value are user inputs
                    user_input_parts.append((display_signal, len("Monitor ")))
                    user_input_parts.append((str(exp_cnt), len(f"Monitor {display_signal} counter reaching ")))
                    
                elif atype == 'handshake':
                    sender = adata.get('sender', '?')
                    receiver = adata.get('receiver', '?')
                    phase = adata.get('phase_type', '?')
                    
                    # Truncate sender/receiver to fit in 15 chars
                    sender_short = sender if len(str(sender)) <= 15 else str(sender)[:12] + "..."
                    recv_short = receiver if len(str(receiver)) <= 15 else str(receiver)[:12] + "..."
                    
                    # Build description based on phase type with clear labels
                    if phase.lower() == '2phase':
                        full_desc = f"2-Phase: req={sender_short} ack={recv_short}"
                        req_prefix_len = len("2-Phase: req=")
                        ack_prefix_len = len("2-Phase: req=") + len(sender_short) + len(" ack=")
                    elif phase.lower() == '4phase':
                        full_desc = f"4-Phase: req={sender_short} ack={recv_short}"
                        req_prefix_len = len("4-Phase: req=")
                        ack_prefix_len = len("4-Phase: req=") + len(sender_short) + len(" ack=")
                    elif phase.lower() == 'ready_valid':
                        full_desc = f"Ready-Valid: valid={sender_short} ready={recv_short}"
                        req_prefix_len = len("Ready-Valid: valid=")
                        ack_prefix_len = len("Ready-Valid: valid=") + len(sender_short) + len(" ready=")
                    else:
                        full_desc = f"{phase}: send={sender_short} recv={recv_short}"
                        req_prefix_len = len(f"{phase}: send=")
                        ack_prefix_len = len(f"{phase}: send=") + len(sender_short) + len(" recv=")
                    
                    # Add sender and receiver (both user inputs)
                    user_input_parts.append((sender_short, req_prefix_len))
                    user_input_parts.append((recv_short, ack_prefix_len))
                    
                elif atype == 'pulseWidth':
                    signal_name = adata.get('target_signal', '?')
                    min_w = adata.get('min_width', '?')
                    max_w = adata.get('max_width', '?')
                    
                    # Truncate signal name to 15 chars
                    if len(str(signal_name)) > 15:
                        display_signal = str(signal_name)[:12] + "..."
                    else:
                        display_signal = str(signal_name)
                    
                    full_desc = f"Pulse width of {display_signal} between {min_w}-{max_w} clocks"
                    # All three parts are user inputs: signal name, min_width, max_width
                    user_input_parts.append((display_signal, len("Pulse width of ")))
                    user_input_parts.append((str(min_w), len(f"Pulse width of {display_signal} between ")))
                    user_input_parts.append((str(max_w), len(f"Pulse width of {display_signal} between {min_w}-")))
                
                else:
                    full_desc = f"[{atype.upper()}] assertion configured"
                
                # Format: Index | Type | Description
                idx_str = f"{i}".ljust(idx_w)
                type_str = _truncate(atype, type_w).ljust(type_w)
                desc_str = _truncate(full_desc, desc_w)
                
                line = f"{idx_str} | {type_str} | {desc_str}"
                
                # Apply dim attribute for alternating rows
                attr = curses.A_DIM if (i % 2 == 0) else 0
                try:
                    win_assertions.addnstr(row, 2, _truncate(line, assert_inner_w), assert_inner_w, attr)
                    
                    # Highlight all user-provided values (signal names, numbers, etc.) in GREEN
                    base_col = 2 + idx_w + 3 + type_w + 3  # Position after "# | Type | "
                    for user_input, offset in user_input_parts:
                        if user_input and user_input != '?':
                            try:
                                input_col = base_col + offset
                                # Make sure we don't go beyond the line width
                                if input_col < 2 + assert_inner_w:
                                    green_pair = _PAIR_BY_NAME.get("green", 0)
                                    win_assertions.addnstr(row, input_col, _truncate(user_input, min(20, len(user_input))), 
                                                         min(20, len(user_input)), 
                                                         curses.color_pair(green_pair) | attr)
                            except curses.error:
                                pass
                except curses.error:
                    pass
                row += 1

        # Completion popup above status/hints if available
        comp_win = None
        if compl_items:
            comp_h = min(len(compl_items) + 2, 10)
            comp_y = max_y - 3 - comp_h
            if comp_y >= 0:
                comp_win = curses.newwin(comp_h, max_x, comp_y, 0)
                comp_win.box()
                title = f" {compl_base} "
                try:
                    comp_win.addnstr(0, 2, _truncate(title, max_x - 4), max_x - 4, curses.A_BOLD)
                except curses.error:
                    pass
                for i, (name, is_dir) in enumerate(compl_items[: comp_h - 2]):
                    color = _PAIR_BY_NAME.get("cyan", 0) if not is_dir else 0
                    attr = curses.color_pair(color)
                    label = name + ("/" if is_dir else "")
                    try:
                        comp_win.addnstr(1 + i, 2, _truncate(label, max_x - 4), max_x - 4, attr)
                    except curses.error:
                        pass
                comp_win.refresh()

        # Status/messages just above the command hints (or below completion)
        status_y = max_y - 3
        try:
            stdscr.addnstr(status_y, 0, _truncate(status_msg, max_x - 1), max_x - 1, curses.A_REVERSE)
        except curses.error:
            pass

        # Command hints line (second last line)
        hints = "[help] [new] [gen] [ms] [param] [f/F] [n/N] [quit|q]"
        try:
            stdscr.addnstr(max_y - 2, 0, _truncate(hints, max_x - 1), max_x - 1)
        except curses.error:
            pass

        # Input prompt (last line)
        prompt = "> "
        edit_w = max_x - len(prompt) - 1
        current = "".join(input_buf)
        vis = current
        try:
            stdscr.addnstr(max_y - 1, 0, prompt, len(prompt))
            stdscr.addnstr(max_y - 1, len(prompt), _truncate(vis, edit_w), edit_w)
            # Show cursor position if inside visible region
            curses.curs_set(1)
            stdscr.move(max_y - 1, len(prompt) + min(cursor_pos, edit_w))
        except curses.error:
            pass

        stdscr.refresh()
        win_left_top.refresh()
        win_left_bot.refresh()
        win_in.refresh()
        win_out.refresh()
        win_cond.refresh()
        win_assertions.refresh()

        ch = stdscr.getch()
        if ch in (curses.KEY_RESIZE,):
            continue
        if ch in (curses.KEY_BACKSPACE, 127, 8):
            if cursor_pos > 0:
                del input_buf[cursor_pos - 1]
                cursor_pos -= 1
            continue
        if ch in (curses.KEY_LEFT,):
            cursor_pos = max(0, cursor_pos - 1)
            continue
        if ch in (curses.KEY_RIGHT,):
            cursor_pos = min(len(input_buf), cursor_pos + 1)
            continue
        # History navigation
        if ch in (curses.KEY_UP,):
            if cmd_history:
                if hist_idx is None:
                    hist_idx = len(cmd_history) - 1
                else:
                    hist_idx = max(0, hist_idx - 1)
                recalled = cmd_history[hist_idx]
                input_buf = list(recalled)
                cursor_pos = len(input_buf)
            continue
        if ch in (curses.KEY_DOWN,):
            if cmd_history and hist_idx is not None:
                if hist_idx < len(cmd_history) - 1:
                    hist_idx += 1
                    recalled = cmd_history[hist_idx]
                    input_buf = list(recalled)
                    cursor_pos = len(input_buf)
                else:
                    hist_idx = None
                    input_buf = []
                    cursor_pos = 0
            continue
        # Removed Ctrl+N/Ctrl+P paging (now handled by 'n'/'N' commands)
        # Tab path completion
        if ch == 9:
            line = "".join(input_buf)
            new_line, new_cursor, items, base_dir = _path_complete(line, cursor_pos)
            input_buf = list(new_line)
            cursor_pos = new_cursor
            compl_items = items
            compl_base = base_dir
            continue
        if ch in (curses.KEY_HOME,):
            cursor_pos = 0
            continue
        if ch in (curses.KEY_END,):
            cursor_pos = len(input_buf)
            continue
        if ch in (curses.KEY_DC,):  # Delete key
            if cursor_pos < len(input_buf):
                del input_buf[cursor_pos]
            continue
        if ch in (10, 13):  # Enter
            cmdline = "".join(input_buf).strip()
            input_buf.clear()
            cursor_pos = 0
            compl_items = []  # clear completion popup
            if not cmdline:
                status_msg = ""
                continue
            status_msg = f"$ {cmdline}"
            # push history
            if not (cmd_history and cmd_history[-1] == cmdline):
                cmd_history.append(cmdline)
            hist_idx = None
            out_msg, opened_overlay = _handle_command(state, cmdline)
            if opened_overlay:
                overlay_active = True
                # If a target page was requested, honor it
                if _OVERLAY_PAGE_KEY:
                    idx = _find_help_page_index(help_pages, _OVERLAY_PAGE_KEY)
                    overlay_page = max(0, idx)
                else:
                    overlay_page = 0
                overlay_scroll = 0
            last_output = out_msg.splitlines() if out_msg else []
            continue
        if 0 <= ch <= 255:
            try:
                c = chr(ch)
            except Exception:
                c = ""
            if c:
                input_buf.insert(cursor_pos, c)
                cursor_pos += 1
                compl_items = []  # typing clears completion


def _handle_command(state: AppState, cmdline: str) -> Tuple[str, bool]:
    # Onboarding: allow empty input for Excel stage to accept auto-detected file
    if state.onboarding_active and (state.onboarding_stage or "") == 'excel':
        if (cmdline.strip() == "") and state.onboarding_excel_autofound and not state.excel_path:
            state.excel_path = state.onboarding_excel_autofound
            
            # CRITICAL: Create session before finishing onboarding
            ok, err = _create_session_excel_and_fill(state)
            if not ok:
                state.excel_error = err or "Session creation failed"
                _set_error_message(state.excel_error)
                return f"Session creation failed: {err}\nPlease try again or enter a different path", False
            
            # Session created successfully
            state.onboarding_active = False
            state.onboarding_stage = None
            _save_session_snapshot(state)
            # Clear screen to prevent message remnants
            status_msg = "Session created - Ready!"
            return f"✓ Session created successfully!\n✓ Excel: {state.session_excel_path}", False
    toks = cmdline.split()
    if not toks:
        return "", False
    raw_cmd = toks[0]
    cmd = raw_cmd.lower()
    args = toks[1:]

    # Onboarding-friendly shortcuts
    if state.onboarding_active:
        stage = state.onboarding_stage or ""
        # 1) RTL stage: accept raw path without 'set rtl'
        if stage == 'rtl':
            from pathlib import Path as _P
            try:
                p = _P(cmdline).expanduser().resolve()  # Add .resolve() here
                if str(p).strip() and p.exists():
                    # Validate: must be a .v or .sv file
                    if not p.is_file():
                        return "ERROR: Please provide a .v or .sv file (not a directory)", False
                    
                    if p.suffix.lower() not in ['.v', '.sv']:
                        return f"ERROR: File must be .v or .sv (got: {p.suffix})", False
                    
                    state.rtl_start = p
                    # Build instance list for next stage
                    try:
                        # DEBUG LOG 파일 설정
                        debug_log_file = _THIS_DIR.parent / "out" / "tui_step1_debug.log"
                        debug_log_file.parent.mkdir(parents=True, exist_ok=True)
                        
                        def log_debug(msg):
                            with open(debug_log_file, "a", encoding="utf-8") as f:
                                from datetime import datetime
                                f.write(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] {msg}\n")
                        
                        # 로그 파일 초기화
                        debug_log_file.write_text("=== TUI Step 1 Debug Log ===\n", encoding="utf-8")
                        log_debug(f"Processing RTL file: {p}")
                        
                        # 1. 전체 모듈 데이터베이스 구축 (같은 폴더와 하위 폴더의 모든 .v/.sv 파일)
                        rtl_root, _found = find_rtl_root_from(p)
                        start_scope_dir = p if p.is_dir() else p.parent
                        files = sorted(set(discover_files(rtl_root, [".v", ".sv"])) | set(discover_files(start_scope_dir, [".v", ".sv"])), key=lambda f: str(f))
                        log_debug(f"Files discovered: {len(files)}")
                        
                        mods_ctx = build_modules_db(files, allow_unknown=True)  # 모든 모듈 포함
                        state.modules_db = mods_ctx
                        log_debug(f"Modules parsed: {len(mods_ctx)}")
                        
                        # 2. 선택한 .v 파일에 정의된 모듈들이 사용되는 모든 인스턴스 찾기
                        log_debug("Calling find_module_instances_by_file...")
                        file_modules_hierarchy = find_module_instances_by_file(mods_ctx, p)
                        log_debug(f"Instances found: {len(file_modules_hierarchy)}")
                        
                        # 3. 파일에 정의된 모듈 확인
                        file_modules_defined = [name for name, m in mods_ctx.items() if Path(m["file"]).resolve() == p.resolve()]
                        log_debug(f"Modules in file: {file_modules_defined}")
                        
                        # 인스턴스를 반드시 찾아야 함. 없으면 에러!
                        if not file_modules_hierarchy:
                            log_debug("ERROR: No instances found!")
                            if not file_modules_defined:
                                return f"ERROR: No modules found in {p.name}", False
                            else:
                                # 모듈은 있는데 인스턴스가 없으면 에러
                                modules_str = ", ".join(file_modules_defined)
                                return (
                                    f"ERROR: Modules found in file: {modules_str}\n"
                                    f"But NO instances found where they are used!\n"
                                    f"This file may be a leaf module (not instantiated anywhere).\n"
                                    f"Or the RTL hierarchy may not be properly connected.\n"
                                    f"Debug log: {debug_log_file}"
                                ), False
                        
                        # 계층 구조를 보기 좋게 정렬하고 표시
                        instances = []
                        for item in file_modules_hierarchy:
                            hierarchy = item["hierarchy_path"]
                            file_module = item["file_module"]
                            log_debug(f"Processing instance: {hierarchy} (module: {file_module})")
                            instances.append({
                                "file_module": file_module,
                                "hierarchy": hierarchy,
                                "display": hierarchy,  # 표시는 hierarchy path만
                                "chain": item["instance_chain"],
                            })
                        
                        # 계층 경로로 정렬
                        instances.sort(key=lambda x: x["hierarchy"])
                        log_debug(f"Total instances after sorting: {len(instances)}")
                        
                        state.onboarding_instances = instances
                        state.onboarding_modules = [inst["display"] for inst in instances]
                        log_debug(f"state.onboarding_modules set to: {state.onboarding_modules}")
                        log_debug(f"SUCCESS! Moving to Step 2")
                            
                    except Exception as e:
                        import traceback
                        tb = traceback.format_exc()
                        log_debug(f"EXCEPTION: {e}\n{tb}")
                        return f"ERROR parsing RTL file: {e}\n{tb}", False
                        
                    state.onboarding_stage = 'module'
                    return f"✓ RTL file loaded: {p.name} ({len(state.onboarding_modules)} items found)", False
                else:
                    return "ERROR: File not found. Please provide a valid .v or .sv file path", False
            except Exception as e:
                return f"ERROR: Invalid path - {e}", False
        # 2) Module stage: number to pick, f filter, F clear, n/N page (wrap), prev/back to return
        if stage == 'module':
            if cmdline.isdigit():
                idx = int(cmdline) - 1
                if 0 <= idx < len(state.onboarding_modules):
                    # Check if we have instance data
                    if state.onboarding_instances and idx < len(state.onboarding_instances):
                        # User selected an instance with hierarchy path
                        selected_inst = state.onboarding_instances[idx]
                        state.target_module = selected_inst["file_module"]  # Use the actual module name
                        hierarchy_path = selected_inst["hierarchy"]
                        
                        # Store the selected instance info for hierarchy (use the hierarchy path directly)
                        state.selected_instance = selected_inst
                        state.module_info.module_hierarchy = hierarchy_path  # Pre-set hierarchy
                    else:
                        # Fallback: direct module selection (no instances, just leaf modules)
                        state.target_module = state.onboarding_modules[idx]
                        state.selected_instance = None
                        state.module_info.module_hierarchy = ""
                    
                    # If starting scope was a directory, bind rtl_start to chosen module's file
                    try:
                        if state.modules_db and state.target_module in state.modules_db:
                            fpath = Path(state.modules_db[state.target_module]["file"]).resolve()
                            if fpath.exists():
                                state.rtl_start = fpath
                        # Refresh module_info now so panels have data in Step 3
                        modules, mi, occs = build_context_from_rtl(state.rtl_start or Path("."), state.target_module)
                        state.modules_db = modules
                        state.module_info = mi
                        state.target_module = mi.module
                        state.occs = occs
                        
                        # CRITICAL: Restore the selected hierarchy from selected_instance
                        # build_context_from_rtl() creates a new module_info, so we need to re-apply the hierarchy
                        if state.selected_instance:
                            state.module_info.module_hierarchy = state.selected_instance["hierarchy"]
                        elif not state.module_info.module_hierarchy and occs and len(occs) > 0:
                            # Fallback: Use auto-detected hierarchy if available
                            auto_hierarchy = occs[0].get("path", "")
                            if auto_hierarchy:
                                state.module_info.module_hierarchy = auto_hierarchy
                    except Exception:
                        pass
                    state.onboarding_stage = 'hierarchy'
                    
                    if state.onboarding_instances and idx < len(state.onboarding_instances):
                        hierarchy = selected_inst["hierarchy"]
                        return f"✓ Picked: {hierarchy} ({state.target_module})", False
                    else:
                        return f"Picked module: {state.target_module}", False
            if cmd == 'f' and args:
                state.onboarding_filter = " ".join(args)
                state.onboarding_page = 0
                return f"Filter set: {state.onboarding_filter}", False
            if raw_cmd == 'F' or (cmd == 'f' and not args):
                state.onboarding_filter = ""
                state.onboarding_page = 0
                return "Filter cleared", False
            if cmd == 'n' and not args:
                # wrap around
                total = len(state.onboarding_modules)
                page_size = max(1, (8 if True else 8))  # placeholder, will be recomputed in render
                pages = max(1, (total + page_size - 1) // page_size)
                state.onboarding_page = (state.onboarding_page + 1) % pages
                return f"Page {state.onboarding_page}", False
            if raw_cmd == 'N' and not args:
                total = len(state.onboarding_modules)
                page_size = max(1, (8 if True else 8))
                pages = max(1, (total + page_size - 1) // page_size)
                state.onboarding_page = (state.onboarding_page - 1) % pages
                return f"Page {state.onboarding_page}", False
            if cmd in ("prev", "back"):
                state.onboarding_stage = 'rtl'
                return "Back to Step 1/4 — RTL", False
        # 2.5) Hierarchy stage: accept hierarchy string or go back
        if stage == 'hierarchy':
            if cmd in ("prev", "back"):
                state.onboarding_stage = 'module'
                return "Back to Step 2/4 — Module", False
            # Hierarchy input handled in main Enter handler
        # 3) Excel stage: accept empty to take autodetected; or raw path
        if stage == 'excel':
            from pathlib import Path as _P
            # Debug logging
            debug_log = _THIS_DIR.parent / "out" / "session_creation_debug.log"
            debug_log.parent.mkdir(parents=True, exist_ok=True)
            with open(debug_log, "a", encoding="utf-8") as f:
                from datetime import datetime
                f.write(f"\n[{datetime.now().strftime('%H:%M:%S')}] === ONBOARDING EXCEL STAGE ===\n")
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] cmdline: '{cmdline}'\n")
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] len(cmdline): {len(cmdline)}\n")
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] not cmdline: {not cmdline}\n")
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] autofound: {state.onboarding_excel_autofound}\n")
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] Condition 1 (not cmdline and autofound): {not cmdline and state.onboarding_excel_autofound}\n")
            
            # Condition 1: Empty input + autofound Excel
            if not cmdline and state.onboarding_excel_autofound:
                with open(debug_log, "a", encoding="utf-8") as f:
                    f.write(f"[{datetime.now().strftime('%H:%M:%S')}] Using autofound Excel\n")
                
                state.excel_path = state.onboarding_excel_autofound
                state.onboarding_active = False
                state.onboarding_stage = None
                
                with open(debug_log, "a", encoding="utf-8") as f:
                    f.write(f"[{datetime.now().strftime('%H:%M:%S')}] Calling _create_session_excel_and_fill...\n")
                
                # Create per-session Excel and prefill Define
                ok, err = _create_session_excel_and_fill(state)
                
                with open(debug_log, "a", encoding="utf-8") as f:
                    f.write(f"[{datetime.now().strftime('%H:%M:%S')}] _create_session_excel_and_fill returned: ok={ok}\n")
                    f.write(f"[{datetime.now().strftime('%H:%M:%S')}] state.session_excel_path after call: {state.session_excel_path}\n")
                
                if not ok:
                    state.excel_error = err or "Excel error"
                    _set_error_message(state.excel_error)
                    with open(debug_log, "a", encoding="utf-8") as f:
                        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] ERROR: {err}\n")
                    # Don't exit onboarding - let user try again
                    return f"Session creation failed: {err}\nPlease try again or enter a different Excel path", False
                else:
                    # Save after successful session creation
                    with open(debug_log, "a", encoding="utf-8") as f:
                        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] SUCCESS - saving snapshot...\n")
                    _save_session_snapshot(state)
                    with open(debug_log, "a", encoding="utf-8") as f:
                        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] Snapshot saved\n")
                    # Show success message with session folder info
                    return err or f"✓ Session created successfully!\n✓ Excel: {state.session_excel_path}", False
            
            # Condition 2: User provided a path
            if cmdline.strip() and cmd not in ("prev", "back", "help", "h"):
                try:
                    with open(debug_log, "a", encoding="utf-8") as f:
                        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] Condition 2: User provided path\n")
                    
                    p = _P(cmdline).expanduser()
                    
                    with open(debug_log, "a", encoding="utf-8") as f:
                        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] Parsed path: {p}\n")
                        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] Path exists: {p.exists() if str(p).strip() else 'N/A'}\n")
                    
                    if str(p).strip() and p.exists():
                        state.excel_path = p
                        state.onboarding_active = False
                        state.onboarding_stage = None
                        ok, err = _create_session_excel_and_fill(state)
                        if not ok:
                            state.excel_error = err or "Excel error"
                            _set_error_message(state.excel_error)
                            with open(debug_log, "a", encoding="utf-8") as f:
                                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] ERROR: {err}\n")
                            # Return to Excel stage, don't proceed
                            state.onboarding_active = True
                            state.onboarding_stage = 'excel'
                            return f"Session creation failed: {err}\nPlease try again", False
                        else:
                            # Save after successful session creation
                            _save_session_snapshot(state)
                            # Show success message with session folder info
                            return err or f"✓ Session created successfully!\n✓ Excel: {state.session_excel_path}", False
                    else:
                        return f"Excel path not found: {p}\nPlease enter a valid path", False
                except Exception as e:
                    with open(debug_log, "a", encoding="utf-8") as f:
                        import traceback
                        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] EXCEPTION in Condition 2: {e}\n")
                        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] Traceback: {traceback.format_exc()}\n")
                    return f"Error processing path: {e}\nPlease try again", False
            
            # Condition 3: Navigation commands
            if cmd in ("prev", "back"):
                state.onboarding_stage = 'hierarchy'
                return "Back to Step 3/4 — Hierarchy", False
            
            # If we reach here, no valid input was provided
            with open(debug_log, "a", encoding="utf-8") as f:
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] ERROR: NO VALID INPUT - Staying in Excel stage\n")
            
            # Stay in Excel stage, don't proceed without Excel
            return "Please enter an Excel path or press Enter to use the auto-detected one", False

    if cmd in ("help", "h"):
        return "Showing help...", True
    
    if cmd == "new":
        # Enter assertion creation wizard
        if not state.module_info.module:
            return "ERROR: Please scan RTL first (use 'scan' command)", False
        
        # Check if session Excel exists, if not create it automatically
        if not state.session_excel_path:
            if not state.excel_path or not Path(state.excel_path).exists():
                return "ERROR: Please complete onboarding first (RTL, Module, Excel setup)", False
            
            # Ensure hierarchy is set (use module name as default if not set)
            if not state.module_info.module_hierarchy:
                state.module_info.module_hierarchy = state.module_info.module
                state.log(f"Hierarchy auto-set to module name: {state.module_info.module_hierarchy}")
            
            # Automatically create session Excel
            state.log("Creating session Excel automatically...")
            ok, err = _create_session_excel_and_fill(state)
            if not ok:
                return f"ERROR: Failed to create session Excel: {err}", False
            
            _save_session_snapshot(state)
            state.log(f"✓ Session Excel created: {_sanitize_path_for_display(str(state.session_excel_path))}")
        
        # Verify session Excel path is set and exists
        if not state.session_excel_path or not Path(state.session_excel_path).exists():
            return f"ERROR: Session Excel not found or not accessible", False
        
        # DEBUG: verify state before entering wizard
        try:
            plugins = _get_assertion_plugins_info()
            if not plugins:
                return "ERROR: No assertion plugins available", False
        except Exception as e:
            return f"ERROR: Failed to load assertion plugins: {str(e)[:50]}", False
        
        # Enter assertion wizard - SET FLAGS CAREFULLY
        state.assertion_wizard_active = True
        state.assertion_wizard_stage = 'select_type'
        state.assertion_selected_type = None
        state.assertion_input_data = {}
        state.assertion_signal_ports = {}
        state.assertion_current_field_idx = 0
        return "✓ Entering assertion creator wizard...\n💡 Choose assertion type [1-4]", False
    
    if cmd in ("gen", "if", "iface", "inst", "instance"):
        # Unified file generation command
        if not state.module_info.module:
            return "ERROR: Please scan RTL first (use 'scan' command)", False
        
        if not state.session_excel_path or not Path(state.session_excel_path).exists():
            return "ERROR: Please create/load a session first (use 'new' command)", False
        
        # Initialize file generation wizard state
        state.gen_wizard_active = True
        state.gen_wizard_stage = 'filename'
        state.gen_filename = ""
        state.gen_file_type = None  # 1=interface, 2=instance, 3=both
        state.gen_data_source = None
        state.gen_preview_lines = []
        return "Entering file generation wizard...", False
    
    if cmd in ("n", "N") and not args:
        # Port paging via command
        global _ports_page
        try:
            _ports_page
        except NameError:
            _ports_page = 0
        if raw_cmd == "N":
            _ports_page = max(0, _ports_page - 1)
        else:
            _ports_page += 1
        return f"Ports page: {_ports_page}", False

    if cmd in ("quit", "q", "exit"):
        raise SystemExit(0)
    
    if cmd == "debug":
        # Hidden debug command - export comprehensive session info to file
        if not state.session_excel_path or not state.session_excel_path.exists():
            return "ERROR: No session loaded. Create session first with 'new'", False
        
        try:
            from datetime import datetime
            from openpyxl import load_workbook
            
            # Create debug export in session folder
            sess_folder = state.session_excel_path.parent
            debug_file = sess_folder / f"debug_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            
            lines = []
            lines.append("=" * 80)
            lines.append(f"DEBUG EXPORT - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            lines.append("=" * 80)
            lines.append("")
            
            # === 1. DISPLAYED VALUES (Main Page) ===
            lines.append("[ 1. MAIN PAGE DISPLAYED VALUES ]")
            lines.append("-" * 80)
            lines.append(f"Hierarchy:  {state.module_info.module_hierarchy or '(not set)'}")
            lines.append(f"Module:     {state.module_info.module or '(not set)'}")
            lines.append(f"Excel:      {_sanitize_path_for_display(str(state.session_excel_path))}")
            lines.append(f"Out Dir:    {_sanitize_path_for_display(str(state.out_dir))}")
            lines.append("")
            lines.append(f"Clocks:     {len(state.module_info.clocks)} signals")
            for i, clk in enumerate(state.module_info.clocks[:10], 1):
                lines.append(f"  [{i}] {clk.get('name', '?')} - {clk.get('width', '?')} bits")
            if len(state.module_info.clocks) > 10:
                lines.append(f"  ... ({len(state.module_info.clocks) - 10} more)")
            
            lines.append("")
            lines.append(f"Resets:     {len(state.module_info.resets)} signals")
            for i, rst in enumerate(state.module_info.resets[:10], 1):
                lines.append(f"  [{i}] {rst.get('name', '?')} - {rst.get('width', '?')} bits")
            if len(state.module_info.resets) > 10:
                lines.append(f"  ... ({len(state.module_info.resets) - 10} more)")
            
            lines.append("")
            lines.append(f"Inputs:     {len(state.module_info.inputs)} signals")
            for i, inp in enumerate(state.module_info.inputs[:20], 1):
                lines.append(f"  [{i}] {inp.get('name', '?')} - {inp.get('width', '?')} bits")
            if len(state.module_info.inputs) > 20:
                lines.append(f"  ... ({len(state.module_info.inputs) - 20} more)")
            
            lines.append("")
            lines.append(f"Outputs:    {len(state.module_info.outputs)} signals")
            for i, out in enumerate(state.module_info.outputs[:20], 1):
                lines.append(f"  [{i}] {out.get('name', '?')} - {out.get('width', '?')} bits")
            if len(state.module_info.outputs) > 20:
                lines.append(f"  ... ({len(state.module_info.outputs) - 20} more)")
            
            lines.append("")
            lines.append(f"Conditions: {len(state.conditions)} MS signals")
            for i, cond in enumerate(state.conditions, 1):
                lines.append(f"  [{i}] {cond.get('name', '?')} = {cond.get('equation', '?')} ({cond.get('bits', '?')} bits)")
            
            lines.append("")
            lines.append(f"Assertions: {len(state.assertions)} created")
            for i, asrt in enumerate(state.assertions, 1):
                lines.append(f"  [{i}] {asrt.get('type', '?')}: {asrt.get('name', '?')}")
            
            # === 2. SESSION JSON STATE ===
            lines.append("")
            lines.append("[ 2. SESSION JSON STATE ]")
            lines.append("-" * 80)
            session_json = sess_folder / "session.json"
            if session_json.exists():
                with open(session_json, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                lines.append(json.dumps(json_data, indent=2, ensure_ascii=False))
            else:
                lines.append("ERROR: session.json not found!")
            
            # === 3. EXCEL DEFINE SHEET ===
            lines.append("")
            lines.append("[ 3. EXCEL DEFINE SHEET ]")
            lines.append("-" * 80)
            
            wb = load_workbook(str(state.session_excel_path))
            if "Define" in wb.sheetnames:
                ws = wb["Define"]
                
                # Base Information
                lines.append("--- Base Section ---")
                for row_idx in range(1, min(20, ws.max_row + 1)):
                    label = ws.cell(row_idx, 1).value
                    value = ws.cell(row_idx, 2).value
                    if label:
                        lines.append(f"  {label}: {value}")
                
                # Port Information
                lines.append("")
                lines.append("--- Ports Section (first 50) ---")
                port_count = 0
                for row_idx in range(1, min(100, ws.max_row + 1)):
                    port_name = ws.cell(row_idx, 1).value
                    if port_name and isinstance(port_name, str) and not port_name.startswith(("Target", "Base", "Signal")):
                        port_type = ws.cell(row_idx, 2).value
                        direction = ws.cell(row_idx, 3).value
                        width = ws.cell(row_idx, 4).value
                        port_count += 1
                        lines.append(f"  [{port_count}] {port_name} | Type:{port_type} | Dir:{direction} | Width:{width}")
                        if port_count >= 50:
                            lines.append(f"  ... (total found: {port_count}+)")
                            break
                
                if port_count == 0:
                    lines.append("  [!] NO PORTS FOUND IN EXCEL!")
                else:
                    lines.append(f"  Total: {port_count} ports")
                
                # Signal Assignments
                lines.append("")
                lines.append("--- Signal Assignments Section ---")
                found_sa = False
                for row_idx in range(1, min(200, ws.max_row + 1)):
                    cell_val = ws.cell(row_idx, 1).value
                    if cell_val and isinstance(cell_val, str) and "signal" in cell_val.lower() and "assignment" in cell_val.lower():
                        found_sa = True
                        lines.append(f"  Found at row {row_idx}")
                        # Read MS signals
                        for offset in range(1, 50):
                            name = ws.cell(row_idx + offset, 1).value
                            equation = ws.cell(row_idx + offset, 2).value
                            bits = ws.cell(row_idx + offset, 3).value
                            if not name:
                                break
                            lines.append(f"    [{offset}] {name} = {equation} ({bits} bits)")
                        break
                
                if not found_sa:
                    lines.append("  [!] Signal Assignments section NOT FOUND")
            
            else:
                lines.append("ERROR: Define sheet not found in Excel!")
            
            wb.close()
            
            # === 4. INTERNAL STATE ===
            lines.append("")
            lines.append("[ 4. INTERNAL STATE ]")
            lines.append("-" * 80)
            lines.append(f"rtl_start:        {state.rtl_start}")
            lines.append(f"target_module:    {state.target_module}")
            lines.append(f"rtl_file_path:    {state.module_info.rtl_file_path}")
            lines.append(f"module_hierarchy: {state.module_info.module_hierarchy}")
            lines.append(f"excel_path:       {state.excel_path}")
            lines.append(f"session_excel:    {state.session_excel_path}")
            lines.append(f"session_id:       {state.session_id}")
            lines.append(f"onboarding:       {state.onboarding_active}")
            lines.append(f"wizard_active:    {state.assertion_wizard_active}")
            
            # Write to file
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            
            return f"[OK] Debug exported: {debug_file.name}\nLocation: {_sanitize_path_for_display(str(sess_folder))}", False
            
        except Exception as e:
            import traceback
            return f"Debug export failed: {e}\n{traceback.format_exc()}", False
    
    if cmd == "clear":
        state.messages.clear()
        return "Cleared messages", False

    if cmd == "set" and args:
        if args[0] == "rtl" and len(args) >= 2:
            p = Path(" ".join(args[1:])).expanduser().resolve()
            state.rtl_start = p
            return f"rtl set: {_sanitize_path_for_display(str(p))}", False
        if args[0] == "module" and len(args) >= 2:
            state.target_module = " ".join(args[1:])
            return f"module set: {state.target_module}", False
        if args[0] == "excel" and len(args) >= 2:
            p = Path(" ".join(args[1:])).expanduser().resolve()
            state.excel_path = p
            return f"excel set: {_sanitize_path_for_display(str(p))}", False
        if args[0] == "out" and len(args) >= 2:
            p = Path(" ".join(args[1:])).expanduser().resolve()
            state.out_dir = p
            return f"out set: {_sanitize_path_for_display(str(p))}", False
        return "Usage: set rtl|module|excel|out <value>", False

    if cmd == "scan":
        if not state.rtl_start:
            return "Set rtl first: set rtl <path>", False
        try:
            modules, mi, occs = build_context_from_rtl(state.rtl_start, state.target_module)
            state.modules_db = modules
            state.module_info = mi
            state.target_module = mi.module
            state.occs = occs
            tops = find_top_modules(modules)
            extra = f" (tops: {', '.join(tops[:5])}{'...' if len(tops)>5 else ''})" if tops else ""
            
            # Try to find latest session Excel for this module
            if not state.session_excel_path:
                latest_excel = _find_latest_session_excel(mi.module)
                if latest_excel:
                    state.session_excel_path = latest_excel
                    extra += f"\n✓ Found session Excel: {_sanitize_path_for_display(str(latest_excel))}"
            
            # Update Define sheet in session Excel if exists
            if state.session_excel_path and state.session_excel_path.exists():
                try:
                    _update_define_sheet(state)
                    return f"Scan complete. Target: {mi.module}{extra}\n✓ Define sheet updated", False
                except Exception as e:
                    return f"Scan complete. Target: {mi.module}{extra}\n⚠ Define sheet update failed: {e}", False
            
            return f"Scan complete. Target: {mi.module}{extra}", False
        except Exception as e:
            return f"Scan failed: {e}", False

    if cmd == "list" and args and args[0] == "modules":
        if not state.modules_db:
            return "No modules. Run scan first.", False
        names = sorted(state.modules_db.keys())
        preview = "\n".join(["Discovered modules:"] + names[:100])
        more = " (truncated)" if len(names) > 100 else ""
        return preview + more, False

    if cmd == "pick" and args:
        name = " ".join(args)
        if state.modules_db and name in state.modules_db:
            state.target_module = name
            # Recompute module_info for the new pick
            try:
                _, mi, occs = build_context_from_rtl(state.rtl_start or Path("."), state.target_module)
                state.module_info = mi
                state.occs = occs
            except Exception:
                pass
            # Persist session
            _save_session_snapshot(state)
            return f"Picked: {name}", False
        return f"Module not found: {name}", False

    if cmd in ("fill", "json", "sv"):
        do_fill = cmd == "fill"
        do_json = cmd == "json"
        do_sv = cmd == "sv"
        rc, out = _run_builder(state, do_fill=do_fill, do_json=do_json, do_sv=do_sv)
        _save_session_snapshot(state)
        return (("OK" if rc == 0 else f"RC={rc}") + "\n" + (out or "")), False

    if cmd == "ms":
        # Syntax: ms <name> = <expr>
        rest = " ".join(args).strip()
        if not rest or "=" not in rest:
            # support 'ms name expr' (no '=')
            parts = rest.split()
            if len(parts) >= 2:
                name = parts[0]
                expr = " ".join(parts[1:])
            else:
                _highlight_ms_help()
                _set_help_filter_cmd("ms")
                _set_error_message("Invalid input: usage ms <name> = <expr> or ms <name> <expr>")
                return "Invalid input. See ms help.", True
        else:
            name, expr = rest.split("=", 1)
        name = name.strip()
        expr = expr.strip()
        # Extract trailing width token if present (e.g., '... 4')
        trailing_width: Optional[int] = None
        try:
            toks = expr.split()
            if toks and toks[-1].isdigit():
                trailing_width = int(toks[-1])
                expr = " ".join(toks[:-1]).strip()
        except Exception:
            pass
        # Map aliases: i1/i2 -> inputs, o1/o2 -> outputs, p1/p2 -> params, c1/c2 -> clocks, r1/r2 -> resets
        def _alias_replace(token: str) -> str:
            import re
            
            # Input alias: i1, i2, i3...
            m = re.match(r'^i(\d+)$', token)
            if m:
                idx = int(m.group(1))
                ins = (state.module_info.inputs + state.module_info.inouts)
                if 1 <= idx <= len(ins):
                    return ins[idx-1].get('name', '')
            
            # Output alias: o1, o2, o3...
            m = re.match(r'^o(\d+)$', token)
            if m:
                idx = int(m.group(1))
                outs = (state.module_info.outputs + state.module_info.inouts)
                if 1 <= idx <= len(outs):
                    return outs[idx-1].get('name', '')
            
            # Parameter alias: p1, p2, p3...
            m = re.match(r'^p(\d+)$', token)
            if m:
                idx = int(m.group(1))
                if 1 <= idx <= len(state.module_info.parameters):
                    return state.module_info.parameters[idx-1].get('name', '')
            
            # Clock alias: c1, c2, c3...
            m = re.match(r'^c(\d+)$', token)
            if m:
                idx = int(m.group(1))
                if 1 <= idx <= len(state.module_info.clocks):
                    return state.module_info.clocks[idx-1].get('name', '')
            
            # Reset alias: r1, r2, r3...
            m = re.match(r'^r(\d+)$', token)
            if m:
                idx = int(m.group(1))
                if 1 <= idx <= len(state.module_info.resets):
                    return state.module_info.resets[idx-1].get('name', '')
            
            # Plain numbers (1, 2, 3, etc.) remain as literal numbers
            # No automatic mapping to input ports
            # Users must use i1, i2, etc. for input ports
            
            return token
        expr_tokens = _tokenize_expr(expr)
        expr_tokens = [ _alias_replace(t) for t in expr_tokens ]
        expr = _join_expr_tokens(expr_tokens)
        if not name or not expr:
            _highlight_ms_help()
            _set_help_filter_cmd("ms")
            _set_error_message("Invalid input: name/expr missing")
            return "Invalid input. See ms help.", True
        ok, err = _validate_condition_expr(expr, state)
        if not ok:
            _highlight_ms_help()
            _set_help_filter_cmd("ms")
            _set_error_message(f"Invalid signal expression: {err}")
            return f"Invalid signal expression: {err}", True
        # Cycle detection
        try:
            if _has_cycle_with_new(name, expr, state):
                _highlight_ms_help()
                _set_help_filter_cmd("ms")
                _set_error_message("Cycle detected among condition signals")
                return "Invalid: cycle detected", True
        except Exception:
            pass
        # Infer width from selects if present; fallback to trailing_width or 1
        width = trailing_width or 1
        # Simple inference: detect [msb:lsb]
        import re
        m = re.search(r"\[(\d+)\s*:\s*(\d+)\]", expr)
        if m:
            try:
                msb = int(m.group(1)); lsb = int(m.group(2))
                if msb >= lsb:
                    width = (msb - lsb + 1)
                    _set_error_message(f"Note: bit-select interpreted as {width} bits")
            except Exception:
                pass
        
        # Clean up expression: remove extra spaces between operators
        expr_cleaned = re.sub(r'([&|])\s+([&|])', r'\1\2', expr)  # &  & -> &&, |  | -> ||
        expr_cleaned = re.sub(r'\s+([&|]{2})\s+', r' \1 ', expr_cleaned)  # Normalize spaces around &&, ||
        
        # Distinguish logical vs bitwise operations
        # Logical operations (&&, ||, !, ==, !=, <, >, <=, >=) always produce 1-bit result
        # Bitwise operations (&, |, ^, ~) produce multi-bit result
        has_logical_only = bool(re.search(r'(&&|\|\||==|!=|<=|>=|<|>|!)', expr_cleaned))
        has_bitwise = bool(re.search(r'(?<![&|])([&|^~])(?![&|])', expr_cleaned))  # Single &, |, ^, ~
        
        # If user didn't specify width with trailing number
        if not trailing_width:
            if has_logical_only and not has_bitwise:
                # Pure logical expression -> 1 bit
                width = 1
            elif has_bitwise:
                # Has bitwise operations -> ask user for bit width
                # Set a flag to prompt user
                state.pending_ms_command = {
                    "name": name,
                    "expr": expr_cleaned,
                    "needs_width": True
                }
                _set_error_message(f"Bitwise operation detected. Specify bit width: ms {name} {expr} <width>")
                return f"Please specify bit width: ms {name} {expr} <width>", False
            else:
                # No operators or just comparison -> 1 bit
                width = 1
        
        # Store and refresh UI (show as name (Nbits))
        state.conditions.append({"name": name, "expr": expr_cleaned, "width": width})
        _save_session_snapshot(state)
        
        # Update Define sheet in session Excel
        try:
            if state.session_excel_path and state.session_excel_path.exists():
                _update_define_sheet(state)
                return f"✓ Condition added: {name} ({width}bits) - Define sheet updated", False
        except Exception as e:
            return f"✓ Condition added: {name} ({width}bits) - Define sheet update failed: {e}", False
        
        return f"Condition added: {name} ({width}bits)", False

    if cmd == "param":
        # Syntax: param <name>=<default> or param <name> <default>
        rest = " ".join(args).strip()
        if not rest:
            _set_error_message("Usage: param p1=10 or param p1 10")
            return "Usage: param <name>=<default_value>", False
        
        # Parse name=default or name default
        if "=" in rest:
            name, default_str = rest.split("=", 1)
            name = name.strip()
            default_str = default_str.strip()
        else:
            parts = rest.split(None, 1)
            if len(parts) != 2:
                _set_error_message("Usage: param p1=10 or param p1 10")
                return "Usage: param <name>=<default_value>", False
            name, default_str = parts
        
        # Validate name format (alphanumeric, underscore)
        if not name or not (name[0].isalpha() or name[0] == '_'):
            _set_error_message("Parameter name must start with letter or underscore")
            return "Invalid parameter name", False
        
        # Try to parse default value as number
        try:
            default_val = int(default_str)
        except ValueError:
            try:
                default_val = float(default_str)
            except ValueError:
                default_val = default_str  # Keep as string
        
        # Add to parameters
        state.module_info.parameters.append({
            'name': name,
            'default': default_val,
            'width': None
        })
        
        _save_session_snapshot(state)
        
        # Update Define sheet in session Excel
        try:
            if state.session_excel_path and state.session_excel_path.exists():
                _update_define_sheet(state)
                return f"✓ Parameter added: {name}={default_val} - Define sheet updated", False
        except Exception as e:
            return f"✓ Parameter added: {name}={default_val} - Define sheet update failed: {e}", False
        
        return f"Parameter added: {name}={default_val}", False

    if cmd == "f" or raw_cmd == "F":
        # f <substr> to set filter; F (upper) or 'f' without args to clear
        if cmd == "f" and args:
            state.port_filter = " ".join(args).strip()
            return f"Filter set: {state.port_filter}", False
        state.port_filter = None
        return "Filter cleared", False

    return f"Unknown command: {cmd} (try 'help')", False


# ---------------------------- HELP OVERLAY ---------------------------------

def _load_help_pages() -> List[Dict[str, Any]]:
    cfg_path = _THIS_DIR / "help_config.json"
    try:
        raw = json.loads(cfg_path.read_text(encoding="utf-8"))
        pages = raw.get("pages", [])
        if isinstance(pages, list):
            return pages
    except Exception:
        pass
    # Fallback minimal page
    return [
        {
            "key": "fallback",
            "title": "Help",
            "items": [
                {"text": "help | h            Open/close this overlay", "color": "white"},
                {"text": "scan                Parse RTL and populate panels", "color": "cyan"},
                {"text": "set rtl <path>      Set RTL path (Tab: path complete)", "color": "green"},
                {"text": "set module <name>   Set target module", "color": "green"},
                {"text": "set excel <path>    Set Excel path (auto-detected if possible)", "color": "green"},
                {"text": "set out <path>      Set output directory", "color": "green"},
                {"text": "fill | json | sv    Generate Define/JSON/SV", "color": "yellow"},
                {"text": "f <substr> / F      Filter ports by name / Clear filter", "color": "cyan"},
                {"text": "n / N               Ports paging next/prev", "color": "cyan"},
                {"text": "ms name = expr      Create condition signal (also: ms name expr)", "color": "magenta"},
                {"text": "Examples: ms valid = (1||2)&&3 | ms active = (I_DATA||I_DEN)&&I_HSYNC", "color": "magenta"},
                {"text": "Supports bit selects: A[0], A[1:0], [1] I_DATA[2:0]", "color": "magenta"},
                {"text": "Up/Down             Command history", "color": "white"},
                {"text": "Tab                 Path candidates popup; auto-complete common prefix", "color": "white"},
                {"text": "Esc/q               Close help", "color": "white"}
            ],
        }
    ]


_COLOR_NAME_TO_CURSES = {
    "black": curses.COLOR_BLACK,
    "red": curses.COLOR_RED,
    "green": curses.COLOR_GREEN,
    "yellow": curses.COLOR_YELLOW,
    "blue": curses.COLOR_BLUE,
    "magenta": curses.COLOR_MAGENTA,
    "cyan": curses.COLOR_CYAN,
    "white": curses.COLOR_WHITE,
}

_PAIR_BY_NAME: Dict[str, int] = {}


def _init_color_pairs() -> None:
    # Initialize pairs for supported colors with default background
    pair_id = 1
    for name, c in _COLOR_NAME_TO_CURSES.items():
        try:
            curses.init_pair(pair_id, c, -1)
            _PAIR_BY_NAME[name] = pair_id
            pair_id += 1
        except curses.error:
            # Some terminals may not support many pairs; ignore
            pass


def _highlight_ms_help() -> None:
    # Set global highlight term to 'ms' to draw attention in help overlay
    global _OVERLAY_HL_TERM
    _OVERLAY_HL_TERM = "ms"


def _set_help_filter_cmd(cmd_key: Optional[str]) -> None:
    global _OVERLAY_FILTER_CMD
    _OVERLAY_FILTER_CMD = cmd_key


def _set_error_message(msg: str) -> None:
    global _ERROR_MESSAGE
    _ERROR_MESSAGE = msg


def _find_help_page_index(pages: List[Dict[str, Any]], key: Optional[str]) -> int:
    if not key:
        return 0
    for i, p in enumerate(pages):
        if p.get("key") == key:
            return i
    return 0


def _render_help_overlay(stdscr: "curses._CursesWindow", pages: List[Dict[str, Any]], page_idx: int, scroll: int) -> None:
    max_y, max_x = stdscr.getmaxyx()
    # Frame
    stdscr.box()
    # Title
    try:
        title = pages[page_idx % len(pages)].get("title", "Help") if pages else "Help"
        label = f" {title} (Tab/Shift-Tab pages, Esc/q close) "
        if max_x > len(label) + 2:
            stdscr.addnstr(0, 2, label, max_x - 4, curses.A_BOLD)
    except curses.error:
        pass

    # Content area
    inner_h = max_y - 4
    inner_w = max_x - 4
    items = []
    current_page = pages[page_idx % len(pages)] if pages else {}
    center_mode = bool(current_page.get("center"))
    if pages:
        raw_items = current_page.get("items", [])
        # Optional filter to only show entries relevant to a specific command
        if _OVERLAY_FILTER_CMD and raw_items and isinstance(raw_items[0], dict) and ("cmd" in raw_items[0]):
            items = [it for it in raw_items if str(it.get("cmd",""))[:2].lower() == _OVERLAY_FILTER_CMD[:2].lower()]
            if not items:
                items = raw_items
        else:
            items = raw_items
    # Decide rendering mode: 3-col if items have 'cmd'
    three_col = bool(items and isinstance(items[0], dict) and ("cmd" in items[0]))
    # Zebra stripe drawing
    start = max(0, scroll)
    end = min(len(items), start + inner_h)
    if three_col:
        # Column widths: Command | Example | Description
        cmd_w = max(14, inner_w // 4)
        ex_w = max(18, inner_w // 3)
        desc_w = max(10, inner_w - cmd_w - ex_w - 2)
        # Header
        try:
            stdscr.addnstr(1, 2, _truncate("Command", cmd_w), cmd_w, curses.A_BOLD)
            stdscr.addnstr(1, 3 + cmd_w, _truncate("Example", ex_w), ex_w, curses.A_BOLD)
            stdscr.addnstr(1, 4 + cmd_w + ex_w, _truncate("Description", desc_w), desc_w, curses.A_BOLD)
        except curses.error:
            pass
        row_line = 0
        vis_rows = inner_h - 1
        for idx in range(start, end):
            if row_line >= vis_rows:
                break
            it = items[idx]
            cmd = _safe_str(it.get("cmd", ""))
            ex = _safe_str(it.get("example", ""))
            desc = _safe_str(it.get("desc", ""))
            # Colors
            cmd_pair = curses.color_pair(_PAIR_BY_NAME.get(it.get("cmd_color", "cyan"), 0)) | curses.A_BOLD
            ex_pair = curses.color_pair(_PAIR_BY_NAME.get(it.get("example_color", "yellow"), 0))
            ds_pair = curses.color_pair(_PAIR_BY_NAME.get(it.get("desc_color", "white"), 0))
            # Zebra by item index
            zebra_attr = curses.A_DIM if ((idx - start) % 2 == 1) else 0
            # Render line
            y = 2 + row_line
            try:
                stdscr.addnstr(y, 2, _truncate(cmd, cmd_w), cmd_w, cmd_pair | zebra_attr)
                stdscr.addnstr(y, 3 + cmd_w, _truncate(ex, ex_w), ex_w, ex_pair | zebra_attr)
                # Highlight term inside desc if set
                hl_term = _OVERLAY_HL_TERM or ""
                if hl_term and hl_term.lower() in desc.lower():
                    stdscr.addnstr(y, 4 + cmd_w + ex_w, _truncate(desc, desc_w), desc_w, ds_pair | zebra_attr)
                    try:
                        from re import finditer
                        for m in finditer(hl_term, desc):
                            sx = 4 + cmd_w + ex_w + m.start()
                            seg = desc[m.start():m.end()]
                            stdscr.addnstr(y, sx, _truncate(seg, desc_w - m.start()), desc_w - m.start(), curses.A_REVERSE)
                    except Exception:
                        pass
                else:
                    stdscr.addnstr(y, 4 + cmd_w + ex_w, _truncate(desc, desc_w), desc_w, ds_pair | zebra_attr)
            except curses.error:
                pass
            row_line += 1
            # separator line if room
            if row_line < vis_rows:
                try:
                    sep_y = 2 + row_line
                    stdscr.addnstr(sep_y, 2, "\u2500" * (inner_w), inner_w, curses.A_DIM)
                except curses.error:
                    pass
                row_line += 1
    else:
        for row, idx in enumerate(range(start, end)):
            item = items[idx]
            text = _safe_str(item.get("text", ""))
            color_name = _safe_str(item.get("color", "white")).lower()
            pair = curses.color_pair(_PAIR_BY_NAME.get(color_name, _PAIR_BY_NAME.get("white", 0)))
            # Alternate brightness via A_DIM every other line (approx ~10% visual change)
            attr = pair | (curses.A_DIM if (row % 2 == 1) else 0)
            # Highlight term if requested
            try:
                from re import finditer
                hl_term = _OVERLAY_HL_TERM or ""
                if hl_term and hl_term.lower() in text.lower():
                    # draw line base
                    if center_mode:
                        cx = max(2, 2 + (inner_w - len(text)) // 2)
                        stdscr.addnstr(2 + row, cx, _truncate(text, inner_w), inner_w, attr)
                    else:
                        stdscr.addnstr(2 + row, 2, _truncate(text, inner_w), inner_w, attr)
                    # overlay highlights (best-effort; may clip)
                    for m in finditer(hl_term, text, flags=0):
                        start_x = (2 + (inner_w - len(text)) // 2) + m.start() if center_mode else 2 + m.start()
                        seg = text[m.start():m.end()]
                        stdscr.addnstr(2 + row, start_x, _truncate(seg, inner_w - (start_x - 2)), inner_w - (start_x - 2), curses.A_REVERSE)
                    continue
            except Exception:
                pass
            try:
                if center_mode:
                    cx = max(2, 2 + (inner_w - len(text)) // 2)
                    stdscr.addnstr(2 + row, cx, _truncate(text, inner_w), inner_w, attr)
                else:
                    stdscr.addnstr(2 + row, 2, _truncate(text, inner_w), inner_w, attr)
            except curses.error:
                pass

    # Footer: error area (red) + instructions
    err = _ERROR_MESSAGE or ""
    footer_instr = "n/N: next/prev page   Up/Down,PgUp/PgDn: scroll   q or Esc: close"
    try:
        # Error message (red) on second-to-last row if present
        if err:
            stdscr.addnstr(max_y - 2, 2, _truncate(err, max_x - 4), max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("red",0)) | curses.A_BOLD)
        stdscr.addnstr(max_y - 1, 2, _truncate(footer_instr, max_x - 4), max_x - 4, curses.A_BOLD)
    except curses.error:
        pass


# ---------------------------- SESSIONS -------------------------------------

def _sessions_dir() -> Path:
    d = (_THIS_DIR / ".." / "out" / "sessions").resolve()
    d.mkdir(parents=True, exist_ok=True)
    return d


def _find_latest_session_excel(target_module: Optional[str] = None) -> Optional[Path]:
    """
    Find the latest session Excel file for the given module.
    If target_module is None, find the most recent session across all modules.
    Returns the path to the session Excel, or None if not found.
    """
    d = _sessions_dir()
    session_folders = [p for p in d.iterdir() if p.is_dir() and not p.name.startswith('.')]
    
    if not session_folders:
        return None
    
    # Filter by module if specified
    if target_module:
        session_folders = [p for p in session_folders if p.name.startswith(f"{target_module}-")]
    
    if not session_folders:
        return None
    
    # Sort by modification time (most recent first)
    session_folders.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    
    # Find Excel file in the most recent session
    for session_folder in session_folders:
        excel_files = list(session_folder.glob("*.xlsx"))
        if excel_files:
            return excel_files[0]
    
    return None


def _load_sessions() -> List[Dict[str, Any]]:
    """
    Load all sessions from session folders.
    Each session has a session.json file inside its folder.
    """
    d = _sessions_dir()
    sessions: List[Dict[str, Any]] = []
    
    # Iterate through all folders in sessions directory
    for folder in sorted(d.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True):
        if not folder.is_dir() or folder.name.startswith('.'):
            continue
        
        # Look for session.json in this folder
        session_json = folder / "session.json"
        if not session_json.exists():
            continue
        
        try:
            obj = json.loads(session_json.read_text(encoding="utf-8"))
            obj["_path"] = str(session_json)  # Path to JSON file
            obj["_folder"] = str(folder)      # Path to folder
            sessions.append(obj)
        except Exception:
            continue
    
    return sessions


def _save_session_snapshot(state: AppState) -> None:
    """
    Save session snapshot to session.json INSIDE the session folder.
    This keeps JSON and Excel together in the same folder.
    Now also saves conditions and assertions for full state restoration.
    """
    if not state.session_excel_path:
        # No session folder yet, cannot save
        return
    
    # Get session folder from session_excel_path
    session_folder = Path(state.session_excel_path).parent
    if not session_folder.exists():
        return
    
    # CRITICAL: Never save reference Excel path - only session Excel path
    data = {
        "rtl_start": str(state.rtl_start) if state.rtl_start else "",
        "target_module": state.target_module or "",
        "module_hierarchy": state.module_info.module_hierarchy or "",  # Save hierarchy
        "session_excel_path": str(state.session_excel_path),  # Only session Excel!
        "out_dir": str(state.out_dir),
        "conditions": state.conditions,  # Save MS signals
        "assertions": state.assertions,  # Save created assertions
        "clocks": state.module_info.clocks,  # Save clocks
        "resets": state.module_info.resets,  # Save resets
        "parameters": state.module_info.parameters,  # Save parameters
    }
    
    # Save as session.json inside the session folder
    session_json = session_folder / "session.json"
    try:
        session_json.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def _restore_conditions_from_excel(state: AppState) -> None:
    """
    Restore condition signals (MS signals) from Excel Signal Assignments section.
    Reads Define sheet and populates state.conditions list.
    If Excel reading fails, falls back to session.json.
    """
    if not state.session_excel_path or not Path(state.session_excel_path).exists():
        return
    
    # Try to restore from Excel first (authoritative source)
    excel_success = False
    if load_workbook:
        try:
            wb = load_workbook(str(state.session_excel_path))
            if "Define" in wb.sheetnames:
                ws = wb["Define"]
                
                # Find Signal Assignments header using same logic as fill_define.py
                hdr = None
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and "signal" in cell.value.strip().casefold() and "assignment" in cell.value.strip().casefold():
                            header_row = cell.row
                            # Check next row for Name/Equation/Bits
                            next_row = list(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1))
                            if next_row:
                                name_col = None
                                equation_col = None
                                bits_col = None
                                for c in next_row[0]:
                                    if isinstance(c.value, str):
                                        val = c.value.strip().casefold()
                                        if val == "name":
                                            if c.column >= cell.column - 1:
                                                name_col = c.column
                                        elif val == "equation":
                                            equation_col = c.column
                                        elif val == "bits":
                                            if equation_col and c.column == equation_col + 1:
                                                bits_col = c.column
                                if name_col and equation_col:
                                    hdr = {
                                        "header_row": header_row,
                                        "data_row": header_row + 1,
                                        "name_col": name_col,
                                        "equation_col": equation_col,
                                        "bits_col": bits_col or (equation_col + 1)
                                    }
                                    break
                        if hdr:
                            break
                    if hdr:
                        break
                
                if hdr:
                    # Read condition signals starting from data_row + 1
                    conditions = []
                    start_row = hdr["data_row"] + 1
                    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                        name_cell = row[hdr["name_col"] - 1] if len(row) >= hdr["name_col"] else None
                        equation_cell = row[hdr["equation_col"] - 1] if len(row) >= hdr["equation_col"] else None
                        bits_cell = row[hdr["bits_col"] - 1] if len(row) >= hdr["bits_col"] else None
                        
                        if not name_cell or not name_cell.value:
                            # Empty row, stop reading
                            break
                        
                        name = str(name_cell.value).strip() if name_cell.value else ""
                        expr = str(equation_cell.value).strip() if equation_cell and equation_cell.value else ""
                        bits_str = str(bits_cell.value).strip() if bits_cell and bits_cell.value else "1"
                        
                        if not name or not expr:
                            continue
                        
                        # Parse bits
                        try:
                            width = int(bits_str)
                        except (ValueError, TypeError):
                            width = 1
                        
                        conditions.append({
                            "name": name,
                            "expr": expr,
                            "width": width
                        })
                    
                    # Update state
                    state.conditions = conditions
                    excel_success = True
            
            wb.close()
            
        except Exception:
            # Excel reading failed, will try session.json fallback
            pass
    
    # Fallback: restore from session.json if Excel reading failed
    if not excel_success:
        session_folder = Path(state.session_excel_path).parent
        session_json = session_folder / "session.json"
        if session_json.exists():
            try:
                data = json.loads(session_json.read_text(encoding="utf-8"))
                if "conditions" in data and isinstance(data["conditions"], list):
                    state.conditions = data["conditions"]
                if "assertions" in data and isinstance(data["assertions"], list):
                    state.assertions = data["assertions"]
            except Exception:
                pass


def _restore_assertions_from_excel(state: AppState) -> None:
    """
    Restore assertions from Excel sheets (Counter, Handshake, PulseWidth).
    Excel is the authoritative source, so we read and populate state.assertions.
    If Excel reading fails, falls back to session.json.
    """
    if not state.session_excel_path or not Path(state.session_excel_path).exists():
        return
    
    # Try to restore from Excel first (authoritative source)
    excel_success = False
    assertions = []
    
    if load_workbook:
        try:
            wb = load_workbook(str(state.session_excel_path))
            
            # ============ Counter Sheet ============
            # Find Counter sheet (case-insensitive)
            counter_sheet = None
            for name in wb.sheetnames:
                if name.lower() == 'counter':
                    counter_sheet = name
                    break
            
            if counter_sheet:
                ws = wb[counter_sheet]
                # Counter sheet: Row 7 is header, data starts at Row 8
                # Columns: B(2)=Target, C(3)=Plus, D(4)=Reset, E(5)=Trigger, F(6)=Expect Count
                for row_idx in range(8, ws.max_row + 1):
                    target = ws.cell(row=row_idx, column=2).value  # Column B
                    if not target or str(target).strip() in ['', 'cnt', 'counter']:
                        break
                    
                    # Extract signal name (remove [...] if present)
                    import re
                    match = re.match(r'^([^\[]*)(?:\[.*\])?$', str(target))
                    target_name = match.group(1).strip() if match else str(target).strip()
                    
                    # Read all fields from correct columns
                    plus_con = ws.cell(row=row_idx, column=3).value     # Column C
                    reset_con = ws.cell(row=row_idx, column=4).value    # Column D
                    trigger_con = ws.cell(row=row_idx, column=5).value  # Column E
                    exp_cnt_val = ws.cell(row=row_idx, column=6).value  # Column F
                    
                    # Build assertion entry
                    assertion_entry = {
                        'type': 'counter',
                        'data': {
                            'target': target_name,
                            'plus_con': str(plus_con).strip() if plus_con else '',
                            'reset_con': str(reset_con).strip() if reset_con else '',
                            'trigger_con': str(trigger_con).strip() if trigger_con else '',
                            'exp_cnt_val': str(exp_cnt_val).strip() if exp_cnt_val else '',
                        },
                        'description': 'counter assertion'
                    }
                    assertions.append(assertion_entry)
            
            # ============ Handshake Sheet ============
            # Find Handshake sheet (case-insensitive)
            handshake_sheet = None
            for name in wb.sheetnames:
                if name.lower() == 'handshake':
                    handshake_sheet = name
                    break
            
            if handshake_sheet:
                ws = wb[handshake_sheet]
                # Handshake sheet: Row 6 is header, data starts at Row 7
                # Columns: C(3)=Type, D(4)=Sender, E(5)=Receiver
                for row_idx in range(7, ws.max_row + 1):
                    phase_type = ws.cell(row=row_idx, column=3).value  # Column C
                    if not phase_type or str(phase_type).strip() == '':
                        break
                    
                    # Skip sample data rows
                    phase_str = str(phase_type).strip()
                    if phase_str in ['ready_valid', '4phase', '2phase']:
                        sender_val = ws.cell(row=row_idx, column=4).value
                        if sender_val and str(sender_val).strip() in ['valid', 'req', 'ack']:
                            # This looks like original sample data, skip it
                            continue
                    
                    sender = ws.cell(row=row_idx, column=4).value    # Column D
                    receiver = ws.cell(row=row_idx, column=5).value  # Column E
                    
                    # Extract clean signal names
                    import re
                    sender_match = re.match(r'^([^\[]*)(?:\[.*\])?$', str(sender)) if sender else None
                    sender_name = sender_match.group(1).strip() if sender_match else (str(sender).strip() if sender else '')
                    
                    receiver_match = re.match(r'^([^\[]*)(?:\[.*\])?$', str(receiver)) if receiver else None
                    receiver_name = receiver_match.group(1).strip() if receiver_match else (str(receiver).strip() if receiver else '')
                    
                    # Skip if both sender and receiver are empty
                    if not sender_name and not receiver_name:
                        continue
                    
                    # Build assertion entry
                    assertion_entry = {
                        'type': 'handshake',
                        'data': {
                            'sender': sender_name,
                            'receiver': receiver_name,
                            'phase_type': phase_str,
                        },
                        'description': 'handshake assertion'
                    }
                    assertions.append(assertion_entry)
            
            # ============ PulseWidth Sheet ============
            # Find PulseWidth sheet (case-insensitive)
            pulse_sheet = None
            for name in wb.sheetnames:
                if name.lower() == 'pulsewidth':
                    pulse_sheet = name
                    break
            
            if pulse_sheet:
                ws = wb[pulse_sheet]
                # PulseWidth sheet: Row 6 is header, data starts at Row 7
                # Columns: C(3)=Type, D(4)=Count_Trigger, E(5)=Target_Pulse, F(6)=Min, G(7)=Max
                for row_idx in range(7, ws.max_row + 1):
                    pulse_type = ws.cell(row=row_idx, column=3).value  # Column C - Type
                    if not pulse_type or str(pulse_type).strip() == '':
                        break
                    
                    # Skip sample data rows (those with 'target_pulse' as signal name)
                    signal_name = ws.cell(row=row_idx, column=5).value  # Column E - Target_Pulse
                    if signal_name and str(signal_name).strip() == 'target_pulse':
                        continue
                    
                    # Extract clean signal name
                    import re
                    match = re.match(r'^([^\[]*)(?:\[.*\])?$', str(signal_name)) if signal_name else None
                    clean_signal = match.group(1).strip() if match else (str(signal_name).strip() if signal_name else '')
                    
                    if not clean_signal:
                        continue
                    
                    # Read Count_Trigger, min/max width from correct columns
                    count_trigger = ws.cell(row=row_idx, column=4).value  # Column D - Count_Trigger
                    min_width = ws.cell(row=row_idx, column=6).value  # Column F - Min
                    max_width = ws.cell(row=row_idx, column=7).value  # Column G - Max
                    
                    pulse_type_str = str(pulse_type).strip() if pulse_type else 'hpulse'
                    
                    # Build assertion entry data dict
                    pulse_data = {
                        'pulse_type': pulse_type_str,
                        'target_signal': clean_signal,
                        'min_width': str(min_width).strip() if min_width else '',
                        'max_width': str(max_width).strip() if max_width else '',
                    }
                    
                    # Add base_clock or trigger_signal depending on pulse type
                    if pulse_type_str == 'hpulse':
                        pulse_data['base_clock'] = str(count_trigger).strip() if count_trigger else '<Base Clock>'
                    elif pulse_type_str == 'vpulse':
                        pulse_data['trigger_signal'] = str(count_trigger).strip() if count_trigger else ''
                    
                    # Build assertion entry
                    assertion_entry = {
                        'type': 'pulseWidth',
                        'data': pulse_data,
                        'description': 'pulseWidth assertion'
                    }
                    assertions.append(assertion_entry)
            
            wb.close()
            
            # If any assertions were found, use them
            if assertions:
                state.assertions = assertions
                excel_success = True
        
        except Exception:
            # Excel reading failed, will try session.json fallback
            pass
    
    # Fallback: restore from session.json if Excel reading failed
    if not excel_success:
        session_folder = Path(state.session_excel_path).parent
        session_json = session_folder / "session.json"
        if session_json.exists():
            try:
                data = json.loads(session_json.read_text(encoding="utf-8"))
                if "assertions" in data and isinstance(data["assertions"], list):
                    state.assertions = data["assertions"]
            except Exception:
                pass


def _sanitize_path_for_display(p: str) -> str:
    """
    Replace non-ASCII characters (e.g., Korean) with ASCII replacements for terminal display.
    This prevents display corruption in terminals that don't handle multibyte characters well.
    Converts Korean characters to [KR] marker for better readability.
    """
    result = []
    in_korean_block = False
    
    for char in str(p):
        # Keep ASCII characters (0-127), backslash, forward slash, colon, dot, etc.
        if ord(char) < 128:
            if in_korean_block:
                in_korean_block = False
            result.append(char)
        else:
            # Mark non-ASCII sections with [*] instead of multiple underscores
            if not in_korean_block:
                result.append('[*]')
                in_korean_block = True
    
    return ''.join(result)


def _shorten_path_for_display(p: str, max_width: int, keep_segments: int = 2) -> str:
    p = str(p)
    # First sanitize the path to remove multibyte characters
    p = _sanitize_path_for_display(p)
    if len(p) <= max_width:
        return p
    parts = p.replace("\\", "/").split("/")
    tail = "/".join([seg for seg in parts if seg][:][-keep_segments:])
    base = "…/" + tail if tail else p[-max_width:]
    return _truncate(base, max_width)


def _wrap_text(s: str, width: int) -> List[str]:
    if width <= 0:
        return [""]
    s = s or ""
    lines: List[str] = []
    i = 0
    n = len(s)
    while i < n:
        lines.append(s[i:i+width])
        i += width
    if not lines:
        lines = [""]
    return lines

def _draw_ascii_box(stdscr: "curses._CursesWindow", top: int, left: int, height: int, width: int) -> None:
    if height < 2 or width < 2:
        return
    try:
        # Use curses box-drawing where possible
        horiz = "─"
        vert = "│"
        tl = "┌"
        tr = "┐"
        bl = "└"
        br = "┘"
        # Top border
        stdscr.addnstr(top, left, tl + (horiz * (width - 2)) + tr, width)
        # Middle
        for y in range(top + 1, top + height - 1):
            stdscr.addnstr(y, left, vert, 1)
            stdscr.addnstr(y, left + width - 1, vert, 1)
        # Bottom border
        stdscr.addnstr(top + height - 1, left, bl + (horiz * (width - 2)) + br, width)
    except curses.error:
        pass


def _run_session_chooser(stdscr: "curses._CursesWindow", sessions: List[Dict[str, Any]]) -> Optional[Union[Dict[str, Any], str]]:
    # Command-driven chooser: new | <number> | f <substr> | n/N | del <number> | del all | q
    filter_text = ""
    page = 0
    compl_items: List[Tuple[str, bool]] = []  # for optional candidate display (kept empty here)
    while True:
        max_y, max_x = stdscr.getmaxyx()
        # Aggressive clear to avoid ghosting from previous screens
        try:
            stdscr.clear()
        except Exception:
            stdscr.erase()
        # Banner: title + version + ASCII art + one-line guide
        banner = [
            "    _                      _   _                ____            ",
            "   / \\   ___ ___  ___ _ __| |_(_) ___  _ __    / ___| ___ _ __  ",
            "  / _ \\ / __/ __|/ _ \\ '__| __| |/ _ \\| '_ \\  | |  _ / _ \\ '_ \\",
            " / ___ \\\\__ \\__ \\  __/ |  | |_| | (_) | | | | | |_| |  __/ | | |",
            "/_/   \_\__/___/\___|_|   \__|_|\___/|_| |_|  \____|\___| |_| |_|",
        ]
        guide = "Type new to start, a number to load, f <text> to filter, n/N to page, q to quit."
        y = 0
        # Draw ASCII art centered as a block to avoid per-line rounding shifts
        ascii_width = max(len(line) for line in banner)
        base_x = 2 + max(0, (max_x - 4 - ascii_width) // 2)
        last_line_len = 0
        last_line_x = base_x
        for line in banner:
            try:
                stdscr.addnstr(y, base_x, _truncate(line, max_x - 4 - base_x + 2), max_x - 4 - base_x + 2)
            except curses.error:
                pass
            last_line_len = len(line)
            y += 1
        # Render version tag at the right side of the last ASCII line
        try:
            ver_text = f" { _APP_VERSION } "
            ver_x = min(max_x - 2 - len(ver_text), last_line_x + last_line_len + 1)
            if ver_x > 2:
                stdscr.addnstr(y - 1, ver_x, _truncate(ver_text, max_x - 4), max_x - 4, curses.A_DIM | curses.A_BOLD)
        except curses.error:
            pass
        # One blank line after banner; guide will render just above box below
        y += 1

        # Create boxed list area below banner with horizontal margins
        # Reserve 3 lines: hint, prompt (with '>'), and a trailing blank to avoid clipping
        min_reserved = 3
        list_y = y + 1
        if max_y - list_y - min_reserved < 5:
            list_y = max(1, max_y - (min_reserved + 6))
        list_margin_x = 2
        list_w = max(20, max_x - (list_margin_x * 2))
        # Clamp
        if list_margin_x + list_w > max_x:
            list_w = max(20, max_x - list_margin_x - 1)
        list_h = max(6, max_y - list_y - min_reserved)
        if list_y + list_h > max_y:
            list_h = max(6, max_y - list_y - 1)
        inner_w = list_w - 4
        # Draw guide directly above the list box
        try:
            stdscr.addnstr(max(1, list_y - 1), list_margin_x + 2, _truncate(guide, max_x - 4 - list_margin_x), max_x - 4 - list_margin_x, curses.A_DIM)
        except curses.error:
            pass
        # Show directory candidates popup immediately above the input line
        if compl_items:
            # Build popup window sized to content and ensure visibility
            max_cols = min(3, max(1, max_x // 20))
            items = [nm + ("/" if is_dir else "") for (nm, is_dir) in compl_items[:24]]
            col_w = max(8, max((len(s) for s in items), default=8)) + 2
            cols = min(max_cols, max(1, (max_x - 4) // col_w))
            rows = max(1, (len(items) + cols - 1) // cols)
            win_w = min(max_x, 2 + cols * col_w + 2)
            win_h = min(rows + 2, 10)
            # Place directly above the prompt line (which is max_y-2)
            win_y = max(1, (max_y - 2) - win_h)
            win_x = 0
            try:
                pop = curses.newwin(win_h, win_w, win_y, win_x)
                pop.box()
                try:
                    pop.addnstr(0, 2, _truncate(" Candidates ", win_w - 4), win_w - 4, curses.A_BOLD)
                except curses.error:
                    pass
                idx = 0
                for r in range(1, win_h - 1):
                    for c in range(cols):
                        if idx >= len(items):
                            break
                        label = items[idx]
                        x = 1 + c * col_w + 1
                        try:
                            pop.addnstr(r, x, _truncate(label, col_w - 1), col_w - 1, curses.color_pair(_PAIR_BY_NAME.get("cyan",0)))
                        except curses.error:
                            pass
                        idx += 1
                pop.refresh()
            except curses.error:
                pass
        # Draw ASCII box directly on stdscr
        _draw_ascii_box(stdscr, list_y, list_margin_x, list_h, list_w)
        # Column header inside list window
        no_w = 6
        module_w = max(12, inner_w // 6)
        rtl_w = max(18, inner_w // 4)
        xls_w = max(10, inner_w // 8)
        modified_w = 16  # "YYYY-MM-DD HH:MM"
        gap = 3  # 3 spaces between columns
        out_w = max(10, inner_w - no_w - module_w - rtl_w - xls_w - modified_w - (gap * 5) - 2)
        try:
            col_x = list_margin_x + 2
            stdscr.addnstr(list_y + 1, col_x, _truncate("No", no_w), no_w, curses.A_BOLD)
            col_x += no_w + gap
            stdscr.addnstr(list_y + 1, col_x, _truncate("Module", module_w), module_w, curses.A_BOLD)
            col_x += module_w + gap
            stdscr.addnstr(list_y + 1, col_x, _truncate("RTL", rtl_w), rtl_w, curses.A_BOLD)
            col_x += rtl_w + gap
            stdscr.addnstr(list_y + 1, col_x, _truncate("Excel", xls_w), xls_w, curses.A_BOLD)
            col_x += xls_w + gap
            stdscr.addnstr(list_y + 1, col_x, _truncate("Modified", modified_w), modified_w, curses.A_BOLD)
            col_x += modified_w + gap
            stdscr.addnstr(list_y + 1, col_x, _truncate("Out", out_w), out_w, curses.A_BOLD)
        except curses.error:
            pass
        # Filter and paginate
        filtered = sessions
        if filter_text:
            ft = filter_text.lower()
            def _m(s: Dict[str, Any]) -> str:
                return f"{s.get('rtl_start','')} {s.get('target_module','')} {s.get('session_excel_path','')} {s.get('out_dir','')}".lower()
            filtered = [s for s in sessions if ft in _m(s)]
        list_inner_h = list_h - 3  # rows available for list within window
        page_size = max(1, list_inner_h - 1)
        start = page * page_size
        end = min(len(filtered), start + page_size)
        # Rows with variable height (wrap RTL)
        y_ptr = list_y + 2
        y_limit = list_y + list_h - 1
        idx = start
        row_index = 0
        while idx < len(filtered) and y_ptr < y_limit:
            s = filtered[idx]
            num = f"[{idx + 1}]"
            module = s.get('target_module', '') or ''
            # RTL column: show RTL hierarchy (module_hierarchy) like main page
            # If hierarchy not set, use rtl_start as fallback
            rtl_hierarchy = s.get('module_hierarchy', '') or ''
            if not rtl_hierarchy:
                rtl_path = s.get('rtl_start', '') or ''
                rtl_hierarchy = _shorten_path_for_display(rtl_path, rtl_w) if rtl_path else module
            rtl_display = rtl_hierarchy
            # Only show session_excel_path (never reference Excel)
            xls_path = s.get('session_excel_path', '') or ''
            xls = os.path.basename(xls_path) if xls_path else 'N/A'
            outp = _shorten_path_for_display(s.get('out_dir', '') or '', out_w)
            # Get modification date from session folder
            mod_date_str = ""
            folder_path = s.get('_folder', '')
            if folder_path:
                try:
                    folder = Path(folder_path)
                    if folder.exists():
                        mtime = folder.stat().st_mtime
                        mod_dt = datetime.fromtimestamp(mtime)
                        mod_date_str = mod_dt.strftime("%Y-%m-%d %H:%M")
                except Exception:
                    mod_date_str = "N/A"
            if not mod_date_str:
                mod_date_str = "N/A"
            rtl_lines = _wrap_text(rtl_display, rtl_w)
            row_h = max(1, len(rtl_lines))
            if y_ptr + row_h > y_limit:
                break
            zebra = curses.A_DIM if (row_index % 2) else 0
            try:
                # Calculate column positions using gap variable
                col_x = list_margin_x + 2
                # number (cyan)
                stdscr.addnstr(y_ptr, col_x, _truncate(num, no_w), no_w, curses.color_pair(_PAIR_BY_NAME.get('cyan',0)) | curses.A_BOLD)
                col_x += no_w + gap
                # module (green bold)
                stdscr.addnstr(y_ptr, col_x, _truncate(module, module_w), module_w, curses.color_pair(_PAIR_BY_NAME.get('green',0)) | curses.A_BOLD | zebra)
                col_x += module_w + gap
                # rtl (wrap)
                for li, rline in enumerate(rtl_lines[:row_h]):
                    stdscr.addnstr(y_ptr + li, col_x, _truncate(rline, rtl_w), rtl_w, zebra)
                col_x += rtl_w + gap
                # excel (yellow)
                stdscr.addnstr(y_ptr, col_x, _truncate(xls, xls_w), xls_w, curses.color_pair(_PAIR_BY_NAME.get('yellow',0)) | zebra)
                col_x += xls_w + gap
                # modified
                stdscr.addnstr(y_ptr, col_x, _truncate(mod_date_str, modified_w), modified_w, zebra)
                col_x += modified_w + gap
                # out
                stdscr.addnstr(y_ptr, col_x, _truncate(outp, out_w), out_w, zebra)
            except curses.error:
                pass
            y_ptr += row_h
            idx += 1
            row_index += 1
        # Empty-state message
        if len(filtered) == 0:
            msg = "No previous sessions. Type 'new' to start."
            try:
                stdscr.addnstr(list_y + 3, list_margin_x + 2, _truncate(msg, inner_w), inner_w, curses.A_DIM)
            except curses.error:
                pass
        # Hints block (color-coded) on the line above the prompt
        hint_y = max_y - 3
        try:
            # Clear hint line fully before writing
            stdscr.addnstr(hint_y, 0, " " * max_x, max_x)
            x = 2
            def seg(text, color=None, bold=False):
                nonlocal x
                attr = curses.A_BOLD if bold else 0
                if color:
                    attr |= curses.color_pair(_PAIR_BY_NAME.get(color,0))
                stdscr.addnstr(hint_y, x, text, max_x - x - 2, attr)
                x += len(text) + 1
            seg("Type:", None, True)
            seg("new", "cyan", True); seg("start new session")
            seg("|")
            seg("<number>", "cyan", True); seg("load session")
            seg("|")
            seg("f <text>", "cyan", True); seg("filter")
            seg("|")
            seg("n/N", "cyan", True); seg("page")
            seg("|")
            seg("del <numbers | all>", "cyan", True); seg("delete")
            seg("|")
            seg("q", "white", True); seg("quit")
        except curses.error:
            pass
        # Prompt (one line below hint, not on the very last line)
        prompt = "> "
        try:
            # Clear prompt line fully to avoid residual characters
            stdscr.addnstr(max_y - 2, 0, " " * max_x, max_x)
            # Place prompt one visual line below hint
            stdscr.addstr(max_y - 2, 0, prompt)
        except curses.error:
            pass
        stdscr.refresh()
        # Read line
        curses.echo()
        try:
            cmdline = stdscr.getstr(max_y - 2, len(prompt), max(8, max_x - len(prompt) - 2)).decode(errors='ignore').strip()
        except Exception:
            cmdline = ""
        curses.noecho()
        if not cmdline:
            continue
        low = cmdline.lower()
        if low == "q":
            return None
        if low == "new":
            return "new"
        if low.startswith("f "):
            filter_text = cmdline[2:].strip()
            page = 0
            continue
        if low == "n":
            page += 1
            continue
        if cmdline == "N":
            page = max(0, page - 1)
            continue
        if cmdline.isdigit():
            idx = int(cmdline)
            if 1 <= idx <= len(filtered):
                return filtered[idx - 1]
            continue
        # del command: del <number(s)> or del all
        if low.startswith("del "):
            args = cmdline[4:].strip()
            if args == "all":
                # Delete all sessions with confirmation
                # Show all sessions highlighted and ask for confirmation
                try:
                    stdscr.clear()
                    max_y, max_x = stdscr.getmaxyx()
                    # Show banner
                    stdscr.addnstr(0, 2, "DELETE ALL SESSIONS - Confirmation Required", max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("red",0)) | curses.A_BOLD)
                    stdscr.addnstr(2, 2, f"You are about to delete {len(sessions)} session(s).", max_x - 4)
                    stdscr.addnstr(3, 2, "Type 'y' to confirm, any other key to cancel.", max_x - 4, curses.A_BOLD)
                    stdscr.refresh()
                    
                    curses.echo()
                    confirm = stdscr.getch()
                    curses.noecho()
                    
                    if confirm in (ord('y'), ord('Y')):
                        deleted_count = 0
                        failed_count = 0
                        error_msgs = []
                        
                        for s in sessions:
                            try:
                                # Delete entire session folder using force delete
                                session_folder = Path(s.get("_folder", ""))
                                if session_folder.exists() and session_folder.is_dir():
                                    success, error_msg = _force_delete_folder(session_folder)
                                    if success:
                                        deleted_count += 1
                                    else:
                                        failed_count += 1
                                        error_msgs.append(f"{session_folder.name}: {error_msg}")
                            except Exception as e:
                                failed_count += 1
                                error_msgs.append(f"Exception: {str(e)}")
                        
                        sessions.clear()
                        filtered = []
                        page = 0
                        
                        # Show result message
                        stdscr.clear()
                        y_pos = 2
                        if deleted_count > 0:
                            msg = f"[OK] {deleted_count} session(s) deleted successfully!"
                            stdscr.addnstr(y_pos, 2, msg, max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("green",0)) | curses.A_BOLD)
                            y_pos += 1
                        
                        if failed_count > 0:
                            msg = f"[X] {failed_count} session(s) failed to delete:"
                            stdscr.addnstr(y_pos, 2, msg, max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("red",0)) | curses.A_BOLD)
                            y_pos += 1
                            for err in error_msgs[:10]:  # Show first 10 errors
                                stdscr.addnstr(y_pos, 4, _truncate(err, max_x - 6), max_x - 6, curses.A_DIM)
                                y_pos += 1
                                if y_pos >= max_y - 2:
                                    break
                        
                        stdscr.addnstr(max_y - 2, 2, "Press any key to continue...", max_x - 4, curses.A_DIM)
                        stdscr.refresh()
                        stdscr.getch()
                    else:
                        # Cancelled
                        stdscr.addnstr(5, 2, "Delete cancelled.", max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("yellow",0)))
                        stdscr.refresh()
                        import time
                        time.sleep(1)
                except Exception as e:
                    # Show error message
                    try:
                        stdscr.addnstr(max_y - 4, 2, f"Error: {str(e)}", max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("red",0)) | curses.A_BOLD)
                        stdscr.refresh()
                        import time
                        time.sleep(1)
                    except curses.error:
                        pass
                continue
            else:
                # Parse multiple numbers: "del 2 3 4" or "del 2,3,4" or "del 2"
                # Support both space and comma separators
                args_normalized = args.replace(',', ' ')
                tokens = args_normalized.split()
                indices = []
                for tok in tokens:
                    if tok.isdigit():
                        idx = int(tok)
                        if 1 <= idx <= len(filtered):
                            indices.append(idx)
                
                if not indices:
                    # Invalid input
                    try:
                        stdscr.addnstr(max_y - 4, 2, "Invalid session number(s)", max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("red",0)) | curses.A_BOLD)
                        stdscr.refresh()
                        import time
                        time.sleep(1)
                    except curses.error:
                        pass
                    continue
                
                # Remove duplicates and sort
                indices = sorted(list(set(indices)))
                
                # Show confirmation with highlighted sessions
                try:
                    stdscr.clear()
                    max_y, max_x = stdscr.getmaxyx()
                    
                    # Banner
                    stdscr.addnstr(0, 2, f"DELETE {len(indices)} SESSION(S) - Confirmation Required", max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("red",0)) | curses.A_BOLD)
                    
                    # Show the sessions to be deleted (highlighted)
                    y_pos = 2
                    stdscr.addnstr(y_pos, 2, "Sessions to be deleted:", max_x - 4, curses.A_BOLD)
                    y_pos += 1
                    
                    # Limit display to prevent overflow
                    display_limit = min(len(indices), max_y - 8)
                    for i, idx in enumerate(indices[:display_limit]):
                        if y_pos >= max_y - 4:
                            break
                        s = filtered[idx - 1]
                        module = s.get('target_module', '') or 'N/A'
                        rtl = s.get('rtl_start', '') or 'N/A'
                        # Shorten paths for display
                        if len(rtl) > 40:
                            rtl = "..." + rtl[-37:]
                        line = f"  [{idx}] {module} - {rtl}"
                        stdscr.addnstr(y_pos, 2, _truncate(line, max_x - 4), max_x - 4, 
                                      curses.color_pair(_PAIR_BY_NAME.get("yellow",0)) | curses.A_BOLD)
                        y_pos += 1
                    
                    if len(indices) > display_limit:
                        stdscr.addnstr(y_pos, 2, f"  ... and {len(indices) - display_limit} more", max_x - 4, curses.A_DIM)
                        y_pos += 1
                    
                    # Confirmation prompt
                    y_pos += 1
                    stdscr.addnstr(y_pos, 2, "Type 'y' to confirm deletion, any other key to cancel.", max_x - 4, curses.A_BOLD)
                    stdscr.refresh()
                    
                    # Wait for confirmation
                    confirm = stdscr.getch()
                    
                    if confirm in (ord('y'), ord('Y')):
                        # Delete in reverse order to maintain indices
                        deleted_count = 0
                        failed_count = 0
                        error_msgs = []
                        
                        for idx in reversed(indices):
                            try:
                                s = filtered[idx - 1]
                                
                                # Delete entire session folder using force delete
                                session_folder = Path(s.get("_folder", ""))
                                if session_folder.exists() and session_folder.is_dir():
                                    success, error_msg = _force_delete_folder(session_folder)
                                    if success:
                                        deleted_count += 1
                                        # Remove from sessions list
                                        sessions.remove(s)
                                    else:
                                        failed_count += 1
                                        error_msgs.append(f"[{idx}] {session_folder.name}: {error_msg}")
                            except Exception as e:
                                failed_count += 1
                                error_msgs.append(f"[{idx}] Exception: {str(e)}")
                        
                        # Update filtered list
                        if filter_text:
                            ft = filter_text.lower()
                            def _m(s: Dict[str, Any]) -> str:
                                return f"{s.get('rtl_start','')} {s.get('target_module','')} {s.get('session_excel_path','')} {s.get('out_dir','')}".lower()
                            filtered = [s for s in sessions if ft in _m(s)]
                        else:
                            filtered = sessions
                        
                        # Show result message
                        stdscr.clear()
                        y_pos = 2
                        
                        if deleted_count > 0:
                            msg = f"[OK] {deleted_count} session(s) deleted successfully!"
                            stdscr.addnstr(y_pos, 2, msg, max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("green",0)) | curses.A_BOLD)
                            y_pos += 1
                        
                        if failed_count > 0:
                            msg = f"[X] {failed_count} session(s) failed to delete:"
                            stdscr.addnstr(y_pos, 2, msg, max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("red",0)) | curses.A_BOLD)
                            y_pos += 1
                            for err in error_msgs[:10]:  # Show first 10 errors
                                stdscr.addnstr(y_pos, 4, _truncate(err, max_x - 6), max_x - 6, curses.A_DIM)
                                y_pos += 1
                                if y_pos >= max_y - 2:
                                    break
                        
                        stdscr.addnstr(max_y - 2, 2, "Press any key to continue...", max_x - 4, curses.A_DIM)
                        stdscr.refresh()
                        stdscr.getch()
                    else:
                        # Cancelled
                        y_pos += 2
                        stdscr.addnstr(y_pos, 2, "Delete cancelled.", max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("yellow",0)))
                        stdscr.refresh()
                        import time
                        time.sleep(1)
                        
                except Exception as e:
                    # Show error message
                    try:
                        stdscr.clear()
                        stdscr.addnstr(2, 2, f"Error: {str(e)}", max_x - 4, curses.color_pair(_PAIR_BY_NAME.get("red",0)) | curses.A_BOLD)
                        stdscr.refresh()
                        import time
                        time.sleep(2)
                    except curses.error:
                        pass
                continue
        # Unknown
        continue


def _first_time_popup_and_autoscan(stdscr: "curses._CursesWindow", state: AppState) -> None:
    max_y, max_x = stdscr.getmaxyx()
    h = 9
    w = min(70, max_x - 8)
    y = (max_y - h) // 2
    x = (max_x - w) // 2
    win = curses.newwin(h, w, y, x)
    win.box()
    msg_lines = [
        "First time detected.",
        "No previous session data.",
        "We'll run an initial RTL scan to get started.",
        "Press Enter to proceed.",
    ]
    for i, t in enumerate(msg_lines):
        try:
            win.addnstr(1 + i, 2, _truncate(t, w - 4), w - 4, curses.A_BOLD if i == 0 else 0)
        except curses.error:
            pass
    win.refresh()
    while True:
        ch = win.getch()
        if ch in (10, 13, 27):
            break
    # Attempt auto-scan if rtl_start is set via env or heuristics else skip
    # Heuristic: EDA/RTL under repo root
    repo_root = _THIS_DIR.parent
    candidate = repo_root / "EDA" / "RTL"
    if candidate.exists():
        state.rtl_start = candidate
        try:
            modules, mi, occs = build_context_from_rtl(candidate, None)
            state.modules_db = modules
            state.module_info = mi
            state.target_module = mi.module
            state.occs = occs
        except Exception:
            pass
    # Auto-find excel in Data folder if present
    auto_excel = _auto_find_excel()
    if auto_excel:
        state.excel_path = auto_excel


def _auto_find_excel() -> Optional[Path]:
    """Data 폴더에서 'Assertion_TF.xlsx'를 우선 탐색, 없으면 첫 번째 xlsx 반환"""
    data_dir = _THIS_DIR.parent / "Data"
    if not data_dir.exists():
        return None
    try:
        # 1순위: Assertion_TF.xlsx (reference 파일)
        preferred = data_dir / "Assertion_TF.xlsx"
        if preferred.exists() and not preferred.name.startswith("~$"):
            return preferred.resolve()
        
        # 2순위: 정렬된 첫 번째 xlsx 파일 (세션 파일 제외)
        for x in sorted(data_dir.glob("*.xlsx")):
            # 임시 파일과 타임스탬프가 포함된 세션 파일 건너뛰기
            if x.name.startswith("~$") or "-20" in x.name:
                continue
            return x.resolve()
    except Exception:
        return None
    return None


# ---------------------- Condition Signal Parsing ---------------------------

def _tokenize_expr(expr: str) -> List[str]:
    # Tokenize a subset of SystemVerilog operators sufficient for validation
    # Multi-char ops first to avoid partial matches
    ops = [
        "<<<", ">>>", "===", "!==", "<<", ">>", "<=", ">=", "==", "!=", "&&", "||", "**",
    ]
    singles = list("&|^~!+-*/%<>()[]{}?:=,;")
    s = expr
    for op in ops:
        s = s.replace(op, f" {op} ")
    for ch in singles:
        s = s.replace(ch, f" {ch} ")
    # Collapse spaces and split
    return [t for t in s.split() if t]


def _join_expr_tokens(tokens: List[str]) -> str:
    """
    Join expression tokens back, removing spaces around operators to prevent '& &' issues.
    Multi-char operators (&&, ||, ==, etc.) should not have spaces within them.
    """
    if not tokens:
        return ""
    
    # Operators that should not have spaces before/after
    no_space_before = {")", "]", "}", ",", ";"}
    no_space_after = {"(", "[", "{", "~", "!"}
    
    # Multi-char operators that need special handling
    multi_ops = ["<<<", ">>>", "===", "!==", "<<", ">>", "<=", ">=", "==", "!=", "&&", "||", "**"]
    
    result = []
    i = 0
    while i < len(tokens):
        tok = tokens[i]
        
        # Add token
        result.append(tok)
        
        # Check if we need space after this token
        if i < len(tokens) - 1:
            next_tok = tokens[i + 1]
            
            # No space if next is closing bracket/comma/semicolon
            if next_tok in no_space_before:
                pass
            # No space if current is opening bracket or unary operator
            elif tok in no_space_after:
                pass
            # Add space otherwise
            else:
                result.append(" ")
        
        i += 1
    
    return "".join(result)


def _is_identifier(tok: str) -> bool:
    if not tok:
        return False
    if tok[0].isalpha() or tok[0] == "_":
        return all(c.isalnum() or c == "_" for c in tok)
    return tok.isdigit()  # allow numeric aliases like 1,2,3


def _resolve_signal_refs(state: AppState) -> Dict[str, Dict[str, Any]]:
    # Build a dictionary of known signals including ports and existing conditions
    refs: Dict[str, Dict[str, Any]] = {}
    all_ports = state.module_info.inputs + state.module_info.outputs + state.module_info.inouts
    for p in all_ports:
        name = p.get("name", "")
        if name:
            refs[name] = p
    # Numeric aliases: 1,2,3... map to concatenated list order
    for idx, p in enumerate(all_ports, start=1):
        refs[str(idx)] = p
    for c in state.conditions:
        nm = c.get("name", "")
        if nm:
            refs[nm] = {"name": nm, "width": 1}
    return refs


def _condition_deps(expr: str, state: AppState) -> List[str]:
    """Return list of condition-signal names referenced by this expr."""
    names = {c.get("name", "") for c in state.conditions}
    toks = _tokenize_expr(expr)
    deps: List[str] = []
    for t in toks:
        if t in names:
            deps.append(t)
    return deps


def _has_cycle_with_new(name: str, expr: str, state: AppState) -> bool:
    """Detect cycle if we add condition `name` with `expr` to existing graph."""
    # Build adjacency: existing + proposed
    adj: Dict[str, List[str]] = {}
    for c in state.conditions:
        nm = c.get("name", "")
        if not nm:
            continue
        adj[nm] = _condition_deps(c.get("expr", ""), state)
    adj[name] = _condition_deps(expr, state)

    # DFS cycle check from `name`
    seen: set[str] = set()
    stack: set[str] = set()

    def dfs(n: str) -> bool:
        if n in stack:
            return True
        if n in seen:
            return False
        seen.add(n)
        stack.add(n)
        for m in adj.get(n, []):
            if dfs(m):
                return True
        stack.remove(n)
        return False

    return dfs(name)


def _validate_condition_expr(expr: str, state: AppState) -> Tuple[bool, str]:
    tokens = _tokenize_expr(expr)
    refs = _resolve_signal_refs(state)
    stack: List[str] = []
    i = 0
    try:
        brace_stack = 0
        while i < len(tokens):
            tok = tokens[i]
            if tok in ("(",")"):
                if tok == "(":
                    stack.append(tok)
                else:
                    if not stack or stack.pop() != "(":
                        return False, "unmatched ')'"
                i += 1
                continue
            if tok == "{":
                brace_stack += 1
                i += 1
                continue
            if tok == "}":
                if brace_stack == 0:
                    return False, "unmatched '}'"
                brace_stack -= 1
                i += 1
                continue
            if tok in ("&&","||","&","|","^","~","!","+","-","*","/","%","<",
                        ">","<=",">=","==","!=","===","!==","<<",">>","<<<",
                        ">>>","?",":","**","=",")",";",","):
                i += 1
                continue
            # Bit select forms: NAME [ idx ] or NAME [ msb : lsb ] or [1] NAME[2:0]
            if tok == "[":
                # Leading bracket case: [1] NAME...
                # We accept but ignore the literal select for validation
                j = i + 1
                while j < len(tokens) and tokens[j] != "]":
                    j += 1
                if j >= len(tokens):
                    return False, "unmatched ']'"
                i = j + 1
                continue
            # Identifier possibly followed by selection
            if _is_identifier(tok):
                name = tok
                if name not in refs:
                    return False, f"unknown signal '{name}'"
                # Optional selection: [ ... ]
                j = i + 1
                if j < len(tokens) and tokens[j] == "[":
                    k = j + 1
                    while k < len(tokens) and tokens[k] != "]":
                        k += 1
                    if k >= len(tokens):
                        return False, "unmatched ']'"
                    i = k + 1
                else:
                    i = j
                continue
            return False, f"unexpected token '{tok}'"
        if stack:
            return False, "unclosed '('"
        return True, ""
    except Exception as e:
        return False, str(e)


# ---------------------- Path completion utilities --------------------------

def _path_complete(line: str, cursor_pos: int) -> Tuple[str, int, List[Tuple[str, bool]], str]:
    # Find the token at cursor that looks like a path (after 'set rtl' or 'set excel'/'set out')
    head = line[:cursor_pos]
    tail = line[cursor_pos:]
    toks = head.split()
    if len(toks) < 2:
        return line, cursor_pos, [], ""
    cmd = toks[0]
    if cmd != "set" or len(toks) < 3:
        return line, cursor_pos, [], ""
    # The third token onwards form the path prefix
    # Support spaces inside path by taking substring after 'set <key> '
    try:
        key = toks[1]
        prefix_start = head.index(key, head.index("set") + 3) + len(key) + 1
    except ValueError:
        return line, cursor_pos, [], ""
    path_prefix = head[prefix_start:].lstrip()
    if not path_prefix:
        path_prefix = ""
    # Expanduser/vars
    expanded = os.path.expandvars(os.path.expanduser(path_prefix))
    # Treat both '/' and '\\' as directory separators for completion logic
    sep_end = expanded.endswith('/') or expanded.endswith('\\')
    base_dir = Path(expanded).parent if not sep_end else Path(expanded)
    if str(base_dir) == "":
        base_dir = Path(".")
    try:
        entries: List[Tuple[str, bool]] = []
        if base_dir.exists():
            for p in sorted(base_dir.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower())):
                if p.name.startswith('.'):
                    continue
                if p.name.startswith('$') or p.name.startswith(':'):
                    continue
                if p.name.startswith("~$"):
                    continue
                # match prefix last segment
                last = Path(expanded).name
                if sep_end or p.name.lower().startswith(last.lower() if last else ""):
                    entries.append((p.name, p.is_dir()))
        # compute common prefix to auto-complete
        common = _common_prefix([e[0] for e in entries if e[0]]) if entries else ""
        if common:
            # replace the last segment in line with common
            new_prefix = os.path.join(str(base_dir), common)
            new_head = head[:prefix_start] + new_prefix
            return new_head + tail, len(new_head), entries, str(base_dir)
        return line, cursor_pos, entries, str(base_dir)
    except Exception:
        return line, cursor_pos, [], ""


def _path_complete_raw(line: str, cursor_pos: int) -> Tuple[str, int, List[Tuple[str, bool]], str]:
    # Treat the entire line up to cursor as a path prefix (used in onboarding stages)
    head = line[:cursor_pos]
    tail = line[cursor_pos:]
    try:
        prefix = head.strip()
        expanded = os.path.expandvars(os.path.expanduser(prefix))
        # If user typed nothing, start from CWD
        if expanded.strip() == "":
            expanded = os.getcwd()
        sep_end = expanded.endswith('/') or expanded.endswith('\\')
        base_dir = Path(expanded).parent if not sep_end else Path(expanded)
        if str(base_dir) == "":
            base_dir = Path(".")
        entries: List[Tuple[str, bool]] = []
        if base_dir.exists():
            for p in sorted(base_dir.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower())):
                if p.name.startswith('.'):
                    continue
                if p.name.startswith('$') or p.name.startswith(':'):
                    continue
                if p.name.startswith("~$"):
                    continue
                last = Path(expanded).name
                if sep_end or p.name.lower().startswith(last.lower() if last else ""):
                    entries.append((p.name, p.is_dir()))
        common = _common_prefix([e[0] for e in entries if e[0]]) if entries else ""
        if common:
            # Use forward slashes for consistency across platforms
            base_str = str(base_dir).rstrip("/\\").replace("\\", "/")
            new_prefix = (base_str + "/" + common) if base_str else common
            new_head = new_prefix
            return new_head + tail, len(new_head), entries, str(base_dir)
        return line, cursor_pos, entries, str(base_dir)
    except Exception:
        return line, cursor_pos, [], ""

def _common_prefix(names: List[str]) -> str:
    if not names:
        return ""
    s1 = min(names)
    s2 = max(names)
    i = 0
    while i < len(s1) and i < len(s2) and s1[i].lower() == s2[i].lower():
        i += 1
    return s1[:i]


# ---------------------- Assertion Wizard -----------------------------------

def _get_assertion_plugins_info() -> List[Dict[str, Any]]:
    """Get information about available assertion plugins."""
    from assertions import get_registered_plugins  # type: ignore
    plugins = get_registered_plugins()
    
    info_list = []
    for plugin_cls in plugins:
        plugin_info = {
            'name': plugin_cls.plugin_name,
            'sheet_name': plugin_cls.sheet_name,
            'description': _get_plugin_description(plugin_cls.plugin_name),
            'fields': _get_plugin_fields(plugin_cls.plugin_name),
        }
        info_list.append(plugin_info)
    return info_list


def _get_plugin_description(plugin_name: str) -> str:
    """Get description for each plugin type."""
    descriptions = {
        'counter': 'Generate counter-based assertions with increment/decrement/reset conditions',
        'handshake': 'Generate 2-phase or 4-phase handshake protocol assertions',
        'sequence': 'Generate temporal sequence assertions',
        'pulseWidth': 'Verify pulse width constraints within min/max clock cycle range',
    }
    return descriptions.get(plugin_name, 'Custom assertion type')


def _should_show_field(field: Dict[str, Any], current_data: Dict[str, str]) -> bool:
    """
    Check if a field should be shown based on show_if conditions.
    
    Args:
        field: Field definition with optional 'show_if' key
        current_data: Currently entered assertion data
        
    Returns:
        True if field should be shown, False otherwise
    """
    show_if = field.get('show_if')
    if not show_if:
        return True  # No condition, always show
    
    # show_if is a dict like {'pulse_type': 'hpulse'}
    for key, expected_value in show_if.items():
        actual_value = current_data.get(key)
        if actual_value != expected_value:
            return False
    
    return True


def _get_visible_fields(fields: List[Dict[str, Any]], current_data: Dict[str, str]) -> List[Dict[str, Any]]:
    """
    Filter fields based on show_if conditions.
    
    Args:
        fields: All field definitions
        current_data: Currently entered assertion data
        
    Returns:
        List of fields that should be visible
    """
    return [f for f in fields if _should_show_field(f, current_data)]


def _get_plugin_fields(plugin_name: str) -> List[Dict[str, Any]]:
    """Get required fields for each plugin type with step-by-step configuration."""
    fields = {
        'counter': [
            {
                'name': 'target',
                'type': 'string',
                'step': 1,
                'title': 'Counter Target Signal',
                'description': 'Enter the internal counter signal name (e.g., cnt, counter)',
                'example': 'cnt',
                'required': True,
            },
            {
                'name': 'plus_con',
                'type': 'signal',
                'step': 2,
                'title': 'Increment Condition',
                'description': 'Select signal/condition for when counter increments',
                'example': 'Select from available signals',
                'required': True,
            },
            {
                'name': 'reset_con',
                'type': 'signal',
                'step': 3,
                'title': 'Reset Condition',
                'description': 'Select signal/condition for when counter resets',
                'example': 'Select from available signals',
                'required': True,
            },
            {
                'name': 'trigger_con',
                'type': 'signal',
                'step': 4,
                'title': 'Check Condition (Trigger)',
                'description': 'Select when to perform the assertion check',
                'example': 'Select from available signals',
                'required': True,
            },
            {
                'name': 'exp_cnt_val',
                'type': 'signal',
                'step': 5,
                'title': 'Expected Count Value',
                'description': 'Select the expected counter value at trigger point',
                'example': 'Select from available signals',
                'required': True,
            },
        ],
        'handshake': [
            {
                'name': 'phase_type',
                'type': 'choice',
                'step': 1,
                'title': 'Handshake Protocol Type',
                'description': 'Select the handshake protocol variant',
                'options': ['2phase', '4phase', 'ready_valid'],
                'required': True,
            },
            {
                'name': 'sender',
                'type': 'signal',
                'step': 2,
                'title': 'Sender Signal',
                'description': 'Select the sender/request signal',
                'example': 'Select from available signals',
                'required': True,
            },
            {
                'name': 'receiver',
                'type': 'signal',
                'step': 3,
                'title': 'Receiver Signal',
                'description': 'Select the receiver/acknowledge signal',
                'example': 'Select from available signals',
                'required': True,
            },
        ],
        'pulseWidth': [
            {
                'name': 'pulse_type',
                'type': 'choice',
                'step': 1,
                'title': 'Pulse Type',
                'description': 'Select pulse type:\n  hpulse: Uses base clock for counting\n  vpulse: Uses trigger signal for edge detection',
                'options': ['hpulse', 'vpulse'],
                'required': True,
            },
            {
                'name': 'base_clock',
                'type': 'choice',
                'step': 2,
                'title': 'Base Clock Signal',
                'description': 'Select base clock for counting (only for hpulse)',
                'options': [],  # Will be populated from state.clocks
                'required': True,
                'show_if': {'pulse_type': 'hpulse'},
            },
            {
                'name': 'trigger_signal',
                'type': 'signal',
                'step': 2,
                'title': 'Trigger Signal',
                'description': 'Select trigger signal for edge detection (only for vpulse)',
                'example': 'Select from available signals',
                'required': True,
                'show_if': {'pulse_type': 'vpulse'},
            },
            {
                'name': 'target_signal',
                'type': 'signal',
                'step': 3,
                'title': 'Target Pulse Signal',
                'description': 'Select which signal to monitor for pulse width check',
                'example': 'Select from available signals',
                'required': True,
            },
            {
                'name': 'min_width',
                'type': 'string',
                'step': 4,
                'title': 'Minimum Pulse Width',
                'description': 'Enter minimum width (number or parameter like p1, p2)\nExample: 10 or PARAM_WIDTH',
                'example': '10 or PARAM_WIDTH',
                'required': True,
            },
            {
                'name': 'max_width',
                'type': 'string',
                'step': 5,
                'title': 'Maximum Pulse Width',
                'description': 'Enter maximum width (number or parameter like p1, p2)\nExample: 20 or MAX_COUNT',
                'example': '20 or MAX_COUNT',
                'required': True,
            },
        ],
    }
    return fields.get(plugin_name, [])


def _render_gen_wizard(stdscr: "curses._CursesWindow", state: AppState) -> None:
    """Render file generation wizard for creating interface/instance files."""
    max_y, max_x = stdscr.getmaxyx()
    
    # Title
    title = "File Generation Wizard"
    try:
        stdscr.box()
        stdscr.addstr(0, 2, f" {title} ", curses.A_BOLD)
    except curses.error:
        pass
    
    margin_x = 3
    y_start = 2
    inner_w = max_x - 6
    
    try:
        if state.gen_wizard_stage == 'filename':
            stdscr.addnstr(y_start, margin_x, "Step 1/4: Output Filename", inner_w, curses.A_BOLD)
            stdscr.addnstr(y_start + 2, margin_x, "Enter filename (without extension):", inner_w)
            stdscr.addnstr(y_start + 3, margin_x, "Example: 'my_assertions' or 'my_checker'", inner_w, curses.A_DIM)
            stdscr.addnstr(y_start + 5, margin_x, f"Current: {state.gen_filename if state.gen_filename else '(empty)'}", inner_w)
        
        elif state.gen_wizard_stage == 'file_type':
            stdscr.addnstr(y_start, margin_x, "Step 2/4: File Type", inner_w, curses.A_BOLD)
            stdscr.addnstr(y_start + 2, margin_x, "Select which files to generate:", inner_w)
            stdscr.addnstr(y_start + 3, margin_x, "  1) Interface (.if.sv)       - Assertion interface definition", inner_w)
            stdscr.addnstr(y_start + 4, margin_x, "  2) Instance (.inst.sv)       - Assertion instance/binding", inner_w)
            stdscr.addnstr(y_start + 5, margin_x, "  3) Both files                - Generate both files", inner_w)
            if state.gen_file_type:
                type_str = ["Interface", "Instance", "Both"][state.gen_file_type - 1]
                stdscr.addnstr(y_start + 7, margin_x, f"Selected: {type_str}", inner_w, curses.color_pair(_PAIR_BY_NAME.get("green", 0)))
        
        elif state.gen_wizard_stage == 'data_source':
            stdscr.addnstr(y_start, margin_x, "Step 3/4: Data Source", inner_w, curses.A_BOLD)
            stdscr.addnstr(y_start + 2, margin_x, "What data to include in generated files:", inner_w)
            stdscr.addnstr(y_start + 3, margin_x, "  1) Assertions                - Include only assertions from Define sheet", inner_w)
            stdscr.addnstr(y_start + 4, margin_x, "  2) Signals                   - Include only signals from Define sheet", inner_w)
            stdscr.addnstr(y_start + 5, margin_x, "  3) Both                      - Include assertions and signals", inner_w)
            if state.gen_data_source:
                src_str = ["Assertions", "Signals", "Both"][int(state.gen_data_source[0]) - 1] if state.gen_data_source else "?"
                stdscr.addnstr(y_start + 7, margin_x, f"Selected: {src_str}", inner_w, curses.color_pair(_PAIR_BY_NAME.get("green", 0)))
        
        elif state.gen_wizard_stage == 'preview':
            stdscr.addnstr(y_start, margin_x, "Step 4/4: Generate Files - Preview", inner_w, curses.A_BOLD)
            type_str = ["Interface", "Instance", "Both"][state.gen_file_type - 1] if state.gen_file_type else "?"
            src_str = ["Assertions", "Signals", "Both"][int(state.gen_data_source[0]) - 1] if state.gen_data_source else "?"
            
            # File info header
            stdscr.addnstr(y_start + 2, margin_x, f"Filename: {state.gen_filename}", inner_w)
            stdscr.addnstr(y_start + 3, margin_x, f"File Type: {type_str}", inner_w)
            stdscr.addnstr(y_start + 4, margin_x, f"Data Source: {src_str}", inner_w)
            
            # For "both" mode, show which file is currently displayed
            if state.gen_file_type == 3:
                current_file = ["Interface (.if.sv)", "Instance (.inst.sv)"][state.gen_preview_file_idx]
                stdscr.addnstr(y_start + 5, margin_x, f"Viewing: {current_file} (press 'f' to switch)", inner_w, 
                              curses.color_pair(_PAIR_BY_NAME.get("yellow", 0)))
                preview_start_y = y_start + 7
            else:
                preview_start_y = y_start + 6
            
            # Calculate pagination
            available_lines = max_y - preview_start_y - 4
            total_lines = len(state.gen_preview_lines) if state.gen_preview_lines else 0
            lines_per_page = max(10, available_lines)
            total_pages = max(1, (total_lines + lines_per_page - 1) // lines_per_page) if total_lines > 0 else 1
            
            # Clamp current page
            state.gen_preview_page = max(0, min(state.gen_preview_page, total_pages - 1))
            
            # Get lines for current page
            start_line = state.gen_preview_page * lines_per_page
            end_line = min(start_line + lines_per_page, total_lines)
            
            # Show preview header
            preview_header = f"Preview (lines {start_line + 1}-{end_line} of {total_lines}):"
            stdscr.addnstr(preview_start_y, margin_x, preview_header, inner_w, 
                          curses.A_BOLD | curses.color_pair(_PAIR_BY_NAME.get("cyan", 0)))
            
            # Show preview lines
            if state.gen_preview_lines:
                preview_lines = state.gen_preview_lines[start_line:end_line]
                for i, line in enumerate(preview_lines):
                    line_y = preview_start_y + 1 + i
                    if line_y < max_y - 4:
                        # Add line number
                        line_num = f"{start_line + i + 1:4d} "
                        stdscr.addnstr(line_y, margin_x, line_num, len(line_num), curses.A_DIM)
                        # Add line content
                        stdscr.addnstr(line_y, margin_x + len(line_num), _truncate(line, inner_w - len(line_num)), 
                                      inner_w - len(line_num))
            else:
                stdscr.addnstr(preview_start_y + 1, margin_x + 2, "(No preview available)", inner_w - 2, curses.A_DIM)
            
            # Show pagination info
            if total_pages > 1:
                page_info = f"Page {state.gen_preview_page + 1}/{total_pages} - Use n/N to navigate"
                stdscr.addnstr(max_y - 4, margin_x, page_info, inner_w, 
                              curses.A_DIM | curses.color_pair(_PAIR_BY_NAME.get("yellow", 0)))
    
    except curses.error:
        pass


def _render_assertion_wizard(stdscr: "curses._CursesWindow", state: AppState) -> None:
    """Render assertion creation wizard with step-by-step flow."""
    max_y, max_x = stdscr.getmaxyx()
    
    # Title
    title = "Assertion Creator - Step by Step"
    try:
        stdscr.addnstr(0, 2, title, max_x - 4, curses.A_BOLD | curses.color_pair(_PAIR_BY_NAME.get("cyan", 0)))
    except curses.error:
        pass
    
    # Box for content
    margin_x = 2
    top = 2
    reserved_bottom = 5
    box_h = max(15, max_y - top - reserved_bottom)
    box_w = max(80, max_x - (margin_x * 2))
    
    _draw_ascii_box(stdscr, top, margin_x, box_h, box_w)
    
    if state.assertion_wizard_stage == 'select_type':
        _render_type_selection_step(stdscr, state, top, margin_x, box_h, box_w)
    elif state.assertion_wizard_stage == 'input_data':
        _render_field_input_step(stdscr, state, top, margin_x, box_h, box_w)
    elif state.assertion_wizard_stage == 'confirm':
        _render_confirmation_step(stdscr, state, top, margin_x, box_h, box_w)


def _render_type_selection_step(stdscr: "curses._CursesWindow", state: AppState, top: int, margin_x: int, box_h: int, box_w: int) -> None:
    """Render type selection screen."""
    plugins = _get_assertion_plugins_info()
    
    y = top + 2
    try:
        stdscr.addnstr(y, margin_x + 2, "Step 1: Select Assertion Type", box_w - 4, curses.A_BOLD)
        y += 2
    except curses.error:
        pass
    
    for i, plugin in enumerate(plugins, start=1):
        if y >= top + box_h - 4:
            break
        
        try:
            option_line = f"[{i}] {plugin['name'].upper()}"
            stdscr.addnstr(y, margin_x + 4, _truncate(option_line, box_w - 6), box_w - 6, curses.A_BOLD)
            y += 1
            
            desc_line = f"    {plugin['description']}"
            stdscr.addnstr(y, margin_x + 4, _truncate(desc_line, box_w - 6), box_w - 6, curses.A_DIM)
            y += 1
        except curses.error:
            pass
    
    try:
        y = top + box_h - 3
        inst = "Enter [1-9] to select type, or q to quit"
        stdscr.addnstr(y, margin_x + 2, _truncate(inst, box_w - 4), box_w - 4, curses.A_DIM)
    except curses.error:
        pass


def _render_field_input_step(stdscr: "curses._CursesWindow", state: AppState, top: int, margin_x: int, box_h: int, box_w: int) -> None:
    """Render step-by-step field input with left panel (form) and right panel (preview)."""
    plugin_name = state.assertion_selected_type
    if not plugin_name:
        return
    
    plugins = _get_assertion_plugins_info()
    plugin = next((p for p in plugins if p['name'] == plugin_name), None)
    if not plugin:
        return
    
    all_fields = plugin['fields']
    
    # Populate base_clock options dynamically from state.module_info.clocks
    for field in all_fields:
        if field.get('name') == 'base_clock' and field.get('type') == 'choice':
            if state.module_info and state.module_info.clocks:
                field['options'] = [clk.get('name', '') for clk in state.module_info.clocks if clk.get('name')]
            else:
                field['options'] = ['I_CLK']  # Default if no clocks defined
    
    # Filter fields based on show_if conditions
    visible_fields = _get_visible_fields(all_fields, state.assertion_input_data)
    
    if state.assertion_current_field_idx >= len(visible_fields):
        return
    
    current_field = visible_fields[state.assertion_current_field_idx]
    current_step = current_field.get('step', 1)
    total_steps = len(visible_fields)
    
    # Split screen
    split_x = margin_x + (box_w - 4) // 2 + 2
    left_w = (box_w - 4) // 2 - 2
    right_w = (box_w - 4) // 2 - 2
    
    # ===== LEFT PANEL: Step Input =====
    y = top + 2
    
    # Special rendering when waiting for custom number input
    if state.assertion_waiting_custom_number:
        try:
            title = f"Step {current_step}/{total_steps}: Custom Number Input"
            stdscr.addnstr(y, margin_x + 2, _truncate(title, left_w), left_w, 
                          curses.A_BOLD | curses.color_pair(_PAIR_BY_NAME.get("yellow", 0)))
            y += 2
            
            desc = "Enter a custom number value for the expected count"
            stdscr.addnstr(y, margin_x + 2, _truncate(desc, left_w), left_w, curses.A_DIM)
            y += 2
            
            prompt = "Type number and press Enter:"
            stdscr.addnstr(y, margin_x + 2, _truncate(prompt, left_w), left_w, 
                          curses.color_pair(_PAIR_BY_NAME.get("green", 0)))
        except curses.error:
            pass
    else:
        # Title
        try:
            title = f"Step {current_step}/{total_steps}: {current_field.get('title', '')}"
            stdscr.addnstr(y, margin_x + 2, _truncate(title, left_w), left_w, 
                          curses.A_BOLD | curses.color_pair(_PAIR_BY_NAME.get("cyan", 0)))
            y += 2
        except curses.error:
            pass
    
    # Skip normal field rendering if waiting for custom number input
    if not state.assertion_waiting_custom_number:
        # Description
        try:
            desc = current_field.get('description', '')
            desc_y = y
            for line in desc.split('\n'):
                if desc_y >= top + box_h - 8:
                    break
                stdscr.addnstr(desc_y, margin_x + 2, _truncate(line, left_w), left_w, curses.A_DIM)
                desc_y += 1
            y = desc_y + 1
        except curses.error:
            pass
        
        # Current value display
        field_name = current_field['name']
        current_val = state.assertion_input_data.get(field_name, '')
        
        try:
            if current_field['type'] == 'choice':
                # Show options for choice fields
                options = current_field.get('options', [])
                if y < top + box_h - 8:
                    stdscr.addnstr(y, margin_x + 2, "Options:", left_w, curses.A_BOLD)
                    y += 1
                    for i, opt in enumerate(options, 1):
                        if y >= top + box_h - 5:
                            break
                        marker = " ✓" if opt == current_val else ""
                        opt_line = f"  [{i}] {opt}{marker}"
                        color = _PAIR_BY_NAME.get("green", 0) if opt == current_val else 0
                        stdscr.addnstr(y, margin_x + 2, _truncate(opt_line, left_w), left_w, 
                                      curses.color_pair(color))
                        y += 1
            
            elif current_field['type'] == 'signal':
                # Show available signals in single-column layout for readability
                if y < top + box_h - 5:
                    # Build complete signal list (all signals, no limit)
                    all_signals = []  # (idx, name, type, port_dict)
                    idx = 0
                    
                    # Special option for reset_con field: "Only Base Reset"
                    if field_name == 'reset_con':
                        all_signals.append((idx, '<Only Base Reset>', 'special', {}))
                        idx += 1
                    
                    # Special option for exp_cnt_val field: "Custom Number Input"
                    if field_name == 'exp_cnt_val':
                        all_signals.append((idx, '<Custom Number Input>', 'special', {}))
                        idx += 1
                    
                    # Input signals - ALL of them
                    if state.module_info and state.module_info.inputs:
                        for inp in state.module_info.inputs:
                            inp_name = inp.get('name', '')
                            all_signals.append((idx, inp_name, 'input', inp))
                            idx += 1
                    
                    # Output signals - ALL of them
                    if state.module_info and state.module_info.outputs:
                        for out in state.module_info.outputs:
                            out_name = out.get('name', '')
                            all_signals.append((idx, out_name, 'output', out))
                            idx += 1
                    
                    # MS Signals (user-defined) - ALL of them
                    if state.conditions:
                        for cond in state.conditions:
                            cond_name = cond.get('name', '')
                            all_signals.append((idx, cond_name, 'ms_signal', cond))
                            idx += 1
                    
                    # Store full list in state for pagination
                    state.assertion_signal_list = all_signals
                    
                    # Calculate pagination
                    signals_per_page = max(10, top + box_h - y - 5)  # Available lines for signals
                    total_pages = max(1, (len(all_signals) + signals_per_page - 1) // signals_per_page)
                    current_page = state.assertion_signal_page % total_pages if total_pages > 0 else 0
                    
                    # Get signals for current page
                    start_idx = current_page * signals_per_page
                    end_idx = start_idx + signals_per_page
                    page_signals = all_signals[start_idx:end_idx]
                    
                    # Build complete signal map for all signals
                    signal_map = {}
                    for idx_num, name, sig_type, port_dict in all_signals:
                        signal_map[idx_num] = (name, port_dict)
                    state.assertion_signal_map = signal_map
                    
                    # Draw signals with proper spacing
                    for idx_num, name, sig_type, port_dict in page_signals:
                        if y >= top + box_h - 4:
                            break
                        
                        try:
                            marker = "✓" if name == current_val else " "
                            
                            # Color by signal type
                            if sig_type == 'input':
                                color = _PAIR_BY_NAME.get("cyan", 0)
                                prefix = "[I]"
                            elif sig_type == 'output':
                                color = _PAIR_BY_NAME.get("yellow", 0)
                                prefix = "[O]"
                            elif sig_type == 'special':
                                color = _PAIR_BY_NAME.get("green", 0)
                                prefix = "[*]"
                            else:  # ms_signal
                                color = _PAIR_BY_NAME.get("magenta", 0)
                                prefix = "[M]"
                            
                            # Highlight current selection
                            if name == current_val:
                                color = _PAIR_BY_NAME.get("green", 0)
                            
                            line = f"  {marker} [{idx_num}] {prefix} {name}"
                            line = _truncate(line, left_w)
                            stdscr.addnstr(y, margin_x + 2, line, left_w, curses.color_pair(color))
                            y += 1
                        except curses.error:
                            pass
                    
                    # Show pagination info if multiple pages exist
                    if total_pages > 1:
                        try:
                            page_info = f"  ... (page {current_page + 1}/{total_pages}) [n/N to navigate]"
                            stdscr.addnstr(y, margin_x + 2, _truncate(page_info, left_w), left_w, curses.A_DIM)
                            y += 1
                        except curses.error:
                            pass
            
            elif current_field['type'] == 'string':
                if current_val:
                    val_line = f"Current: {current_val}"
                    stdscr.addnstr(y, margin_x + 2, _truncate(val_line, left_w), left_w, 
                                  curses.color_pair(_PAIR_BY_NAME.get("green", 0)))
                else:
                    val_line = "[Enter value]"
                    stdscr.addnstr(y, margin_x + 2, _truncate(val_line, left_w), left_w, 
                                  curses.color_pair(_PAIR_BY_NAME.get("yellow", 0)))
        except curses.error:
            pass
    
    # ===== RIGHT PANEL: Preview and Timing Diagram =====
    py = top + 2
    
    # Show preview
    preview_lines = _generate_assertion_preview(plugin_name, state.assertion_input_data, state)
    for line in preview_lines:
        if py >= top + box_h - 3:
            break
        try:
            stdscr.addnstr(py, split_x, _truncate(line, right_w), right_w)
            py += 1
        except curses.error:
            pass
    
    # ===== BOTTOM Instructions =====
    try:
        y = top + box_h - 3
        
        if state.assertion_waiting_custom_number:
            inst = "Enter number value | 'q' to cancel"
        elif current_field['type'] == 'choice':
            inst = "Enter [1-9] to select | 'prev'/'p' for previous | 'q' to cancel"
        elif current_field['type'] == 'signal':
            # Special instruction for exp_cnt_val field
            if field_name == 'exp_cnt_val':
                inst = "Enter [0-N], number, or expression (e.g. 'i1 - 1') | n/N page | 'prev'/'p' | 'q'"
            # Special instruction for reset_con field
            elif field_name == 'reset_con':
                inst = "Enter [0-N] (0=Only Base Reset) | n/N page | 'prev'/'p' for previous | 'q' to cancel"
            else:
                inst = "Enter signal [1-N], number, or expression (e.g. 'i1 - 1') | n/N page | 'prev'/'p' | 'q'"
        else:  # string
            inst = "Enter value | 'prev'/'p' for previous | 'q' to cancel"
        
        stdscr.addnstr(y, margin_x + 2, _truncate(inst, box_w - 4), box_w - 4, curses.A_DIM)
    except curses.error:
        pass


def _render_confirmation_step(stdscr: "curses._CursesWindow", state: AppState, top: int, margin_x: int, box_h: int, box_w: int) -> None:
    """Render final confirmation screen with preview on right panel."""
    y = top + 2
    
    try:
        stdscr.addnstr(y, margin_x + 2, "Step: Review & Create", box_w - 4, 
                      curses.A_BOLD | curses.color_pair(_PAIR_BY_NAME.get("green", 0)))
        y += 2
    except curses.error:
        pass
    
    # ===== SPLIT LAYOUT: Left panel for fields, Right panel for preview =====
    split_x = margin_x + box_w // 2 + 1
    left_w = box_w // 2 - 2
    right_w = box_w // 2 - 2
    
    # ===== LEFT PANEL: Show all configured fields =====
    plugins = _get_assertion_plugins_info()
    plugin = next((p for p in plugins if p['name'] == state.assertion_selected_type), None)
    
    ly = top + 2
    if plugin:
        all_fields = plugin['fields']
        visible_fields = _get_visible_fields(all_fields, state.assertion_input_data)
        
        try:
            for field in visible_fields:
                if ly >= top + box_h - 4:
                    break
                
                field_name = field['name']
                title = field.get('title', field_name)
                val = state.assertion_input_data.get(field_name, '')
                
                line = f"  {title}: "
                if val:
                    line += val
                    color = _PAIR_BY_NAME.get("green", 0)
                else:
                    line += "[NOT SET]"
                    color = _PAIR_BY_NAME.get("red", 0)
                
                stdscr.addnstr(ly, margin_x + 2, _truncate(line, left_w), left_w, 
                              curses.color_pair(color))
                ly += 1
        except curses.error:
            pass
    
    # ===== RIGHT PANEL: Preview and Timing Diagram =====
    py = top + 2
    
    # Show preview
    preview_lines = _generate_assertion_preview(state.assertion_selected_type, state.assertion_input_data, state)
    for line in preview_lines:
        if py >= top + box_h - 4:
            break
        try:
            stdscr.addnstr(py, split_x, _truncate(line, right_w), right_w)
            py += 1
        except curses.error:
            pass
    
    # ===== BOTTOM Instructions =====
    try:
        y = top + box_h - 3
        inst = "[Enter] to create | 'b' to edit | 'q' to cancel"
        stdscr.addnstr(y, margin_x + 2, _truncate(inst, box_w - 4), box_w - 4, 
                      curses.A_BOLD | curses.color_pair(_PAIR_BY_NAME.get("yellow", 0)))
    except curses.error:
        pass


def _render_type_selection(stdscr: "curses._CursesWindow", state: AppState, top: int, margin_x: int, box_h: int, box_w: int) -> None:
    """Render assertion type selection screen with full type list and status indicators."""
    plugins = _get_assertion_plugins_info()
    
    y = top + 2
    try:
        stdscr.addnstr(y, margin_x + 2, "Select Assertion Type:", box_w - 4, curses.A_BOLD)
        y += 2
    except curses.error:
        pass
    
    for i, plugin in enumerate(plugins, start=1):
        if y >= top + box_h - 3:
            # Show continuation indicator
            try:
                stdscr.addnstr(y, margin_x + 2, "... (more types)", box_w - 4, curses.A_DIM)
            except curses.error:
                pass
            break
        
        # Check if plugin file exists
        plugin_file_path = Path(__file__).parent / "assertions" / f"{plugin['name']}.py"
        plugin_exists = plugin_file_path.exists()
        
        # Check if sheet exists in Excel (use session excel if available)
        sheet_exists = False
        excel_to_check = state.session_excel_path or state.excel_path
        if excel_to_check:
            try:
                from openpyxl import load_workbook  # type: ignore
                from assertions.base import BaseAssertionPlugin  # type: ignore
                wb = load_workbook(str(excel_to_check), read_only=True)
                actual_sheet = BaseAssertionPlugin.find_sheet_case_insensitive(wb.sheetnames, plugin['sheet_name'])
                sheet_exists = bool(actual_sheet)
                wb.close()
            except Exception:
                pass
        
        # Build status line: show both plugin and sheet status
        status_parts = []
        if not plugin_exists:
            status_parts.append("Plugin_missing")
        if not sheet_exists:
            status_parts.append("Excel_missing")
        
        status_indicator = ""
        if status_parts:
            status_indicator = f" [{', '.join(status_parts)}]"
        
        # Display option
        try:
            option_line = f"[{i}] {plugin['name'].upper()}{status_indicator}"
            # Use red if either plugin or sheet is missing
            color = "red" if status_parts else "green"
            color_code = _PAIR_BY_NAME.get(color, 0)
            stdscr.addnstr(y, margin_x + 4, _truncate(option_line, box_w - 6), box_w - 6, 
                          curses.A_BOLD | curses.color_pair(color_code))
            y += 1
            
            # Description
            desc_line = f"    {plugin['description']}"
            stdscr.addnstr(y, margin_x + 4, _truncate(desc_line, box_w - 6), box_w - 6, curses.A_DIM)
            y += 1
            y += 1  # Blank line between options
        except curses.error:
            pass
    
    # Instructions
    try:
        y = top + box_h - 3
        inst_line = "Enter number to select, or 'q' to quit"
        stdscr.addnstr(y, margin_x + 2, inst_line, box_w - 4, curses.A_DIM)
    except curses.error:
        pass


def _render_data_input(stdscr: "curses._CursesWindow", state: AppState, top: int, margin_x: int, box_h: int, box_w: int) -> None:
    """Render data input screen for selected assertion type."""
    plugin_name = state.assertion_selected_type
    if not plugin_name:
        return
    
    plugins = _get_assertion_plugins_info()
    plugin = next((p for p in plugins if p['name'] == plugin_name), None)
    if not plugin:
        return
    
    fields = plugin['fields']
    
    # Title
    y = top + 2
    try:
        title = f"Setup {plugin_name.upper()} Assertion"
        stdscr.addnstr(y, margin_x + 2, title, box_w - 4, curses.A_BOLD | curses.color_pair(_PAIR_BY_NAME.get("cyan", 0)))
        y += 2
    except curses.error:
        pass
    
    # Show all fields with their status
    for i, field in enumerate(fields):
        if y >= top + box_h - 4:
            break
        
        field_name = field['name']
        prompt = field['prompt']
        current_val = state.assertion_input_data.get(field_name, '')
        is_current = (i == state.assertion_current_field_idx)
        is_required = field.get('required', True)
        
        try:
            # Field line with number and prompt
            field_line = f"[{i+1}] {prompt}"
            
            # Highlight if current field
            attr = curses.A_BOLD
            if is_current:
                attr |= curses.color_pair(_PAIR_BY_NAME.get("yellow", 0))
            
            stdscr.addnstr(y, margin_x + 2, _truncate(field_line, box_w - 4), box_w - 4, attr)
            y += 1
            
            # Show current value or status
            if current_val:
                val_line = f"    -> {current_val}"
                stdscr.addnstr(y, margin_x + 2, _truncate(val_line, box_w - 4), box_w - 4, 
                              curses.color_pair(_PAIR_BY_NAME.get("green", 0)))
            else:
                req_marker = "*" if is_required else ""
                val_line = f"    {req_marker} (not set)"
                color = _PAIR_BY_NAME.get("red", 0) if is_required else _PAIR_BY_NAME.get("dim", 0)
                stdscr.addnstr(y, margin_x + 2, _truncate(val_line, box_w - 4), box_w - 4, 
                              curses.color_pair(color))
            y += 1
        except curses.error:
            pass
    
    # Bottom instructions
    try:
        y = top + box_h - 4
        inst1 = "Enter [#] to set field | set [#] value | b to go back | done to finish"
        stdscr.addnstr(y, margin_x + 2, _truncate(inst1, box_w - 4), box_w - 4, curses.A_DIM)
        y += 1
        inst2 = "q to cancel | * = required field"
        stdscr.addnstr(y, margin_x + 2, _truncate(inst2, box_w - 4), box_w - 4, curses.A_DIM)
    except curses.error:
        pass


def _generate_assertion_preview(plugin_name: str, data: Dict[str, Any], state: "AppState") -> List[str]:
    """Generate assertion preview with timing diagram and Pass/Fail conditions."""
    lines = []
    
    def format_signal_name(name: str, role: str, width: int = 20) -> str:
        """Format signal name right-aligned with role in parentheses."""
        formatted = f"{name} ({role})"
        return formatted.rjust(width)
    
    def format_waveform_line(waveform: str, width: int = 20) -> str:
        """Format waveform data right-aligned."""
        return waveform.rjust(width)
    
    if plugin_name == 'counter':
        target = data.get('target', '?')
        plus_con = data.get('plus_con', '?')
        reset_con = data.get('reset_con', '?')
        trigger_con = data.get('trigger_con', '?')
        exp_cnt_val = data.get('exp_cnt_val', '?')
        
        # Get base clock and reset from module_info
        base_clk = '?'
        base_rst = '?'
        if state.module_info.clocks:
            base_clk = state.module_info.clocks[0].get('name', '?')
        if state.module_info.resets:
            base_rst = state.module_info.resets[0].get('name', '?')
        
        # Truncate signal names to 10 characters
        display_target = target if len(str(target)) <= 10 else str(target)[:7] + "..."
        display_plus = plus_con if len(str(plus_con)) <= 10 else str(plus_con)[:7] + "..."
        display_reset = reset_con if len(str(reset_con)) <= 10 else str(reset_con)[:7] + "..."
        display_trigger = trigger_con if len(str(trigger_con)) <= 10 else str(trigger_con)[:7] + "..."
        
        lines.append("=" * 60)
        lines.append("COUNTER ASSERTION")
        lines.append("=" * 60)
        lines.append("")
        lines.append(f"Counter Signal: {display_target}")
        lines.append(f"Increments when: {display_plus}")
        lines.append(f"Resets when: {display_reset}")
        lines.append(f"Checked at: {display_trigger}")
        lines.append(f"Expected value: {exp_cnt_val}")
        lines.append(f"Base Clock: {base_clk}")
        lines.append(f"Base Reset: {base_rst}")
        lines.append("")
        
        lines.append("Timing Diagram:")
        lines.append("-" * 60)
        lines.append("")
        lines.append("Clock cycles: 0   1   2   3   4   5   6   7")
        lines.append(format_waveform_line("clk") + " |___|‾‾‾|___|‾‾‾|___|‾‾‾|___|‾‾‾|")
        lines.append("")
        lines.append(format_signal_name(display_target, "counter") + " 0   0   1   1   1   0   0   0")
        lines.append(format_signal_name(display_plus, "increment") + " └─────┘   └─────┘   └─────┘")
        lines.append(format_signal_name(display_reset, "reset") + " └───────────────┘       └───────┘")
        lines.append(format_signal_name(display_trigger, "trigger") + " └─────┘       └─────┘   └─────┘")
        lines.append("")
        
        lines.append("Pass:")
        lines.append(f"  {display_trigger}=1 -> {display_target}={exp_cnt_val}")
        lines.append("")
        
        lines.append("Fail:")
        lines.append(f"  {display_trigger}=1 -> {display_target}!={exp_cnt_val}")
        lines.append("")
        
    elif plugin_name == 'handshake':
        phase = data.get('phase_type', '?')
        sender = data.get('sender', '?')
        receiver = data.get('receiver', '?')
        
        # Get base clock and reset from module_info
        base_clk = '?'
        base_rst = '?'
        if state.module_info.clocks:
            base_clk = state.module_info.clocks[0].get('name', '?')
        if state.module_info.resets:
            base_rst = state.module_info.resets[0].get('name', '?')
        
        # Truncate signal names to 10 characters
        display_sender = sender if len(str(sender)) <= 10 else str(sender)[:7] + "..."
        display_receiver = receiver if len(str(receiver)) <= 10 else str(receiver)[:7] + "..."
        
        lines.append("=" * 60)
        lines.append(f"{phase.upper()} HANDSHAKE ASSERTION")
        lines.append("=" * 60)
        lines.append("")
        lines.append(f"Protocol Type: {phase}")
        lines.append(f"Sender Signal: {display_sender}")
        lines.append(f"Receiver Signal: {display_receiver}")
        lines.append(f"Base Clock: {base_clk}")
        lines.append(f"Base Reset: {base_rst}")
        lines.append("")
        
        if phase == '2phase':
            lines.append("Timing Diagram (2-Phase Handshake):")
            lines.append("-" * 60)
            lines.append("")
            lines.append("Clock cycles: 0   1   2   3   4   5   6   7   8")
            lines.append(format_waveform_line("clk") + " |___|‾‾‾|___|‾‾‾|___|‾‾‾|___|‾‾‾|___|")
            lines.append("")
            lines.append(format_signal_name(display_sender, "sender") + " └─────────────┘   └─────────────┘")
            lines.append(format_signal_name(display_receiver, "receiver") + "     └─────────────┘   └─────────────┘")
            lines.append("")
            
            lines.append("Pass:")
            lines.append(f"  1. {display_sender}=1 (multi-cycle hold)")
            lines.append(f"  2. {display_receiver}=1 (after {display_sender})")
            lines.append(f"  3. {display_sender}=1 & {display_receiver}=1 (overlap)")
            lines.append(f"  4. {display_sender}=0 & {display_receiver}=0")
            lines.append("")
            
            lines.append("Fail:")
            lines.append(f"  1. {display_sender}=1 but {display_receiver}=0")
            lines.append(f"  2. wait-timeout (no-ack)")
            lines.append(f"  3. glitch (noise)")
            lines.append("")
            
        elif phase == '4phase':
            lines.append("Timing Diagram (4-Phase Handshake):")
            lines.append("-" * 60)
            lines.append("")
            lines.append("Clock cycles: 0   1   2   3   4   5   6   7   8")
            lines.append(format_waveform_line("clk") + " |___|‾‾‾|___|‾‾‾|___|‾‾‾|___|‾‾‾|___|")
            lines.append("")
            lines.append(format_signal_name(display_sender, "sender") + " └───────┘   └───────┘   └───────┘")
            lines.append(format_signal_name(display_receiver, "receiver") + "     └───────┘   └───────┘   └───────┘")
            lines.append("")
            
            lines.append("Pass:")
            lines.append(f"  1. {display_sender}: 1->0->1")
            lines.append(f"  2. {display_receiver}: 1->0->1 (after {display_sender})")
            lines.append(f"  3. {display_sender}=0 before {display_receiver}=1")
            lines.append(f"  4. both=0 (sync-complete)")
            lines.append("")
            
            lines.append("Fail:")
            lines.append(f"  1. {display_sender}!=pulse (stays-high/low)")
            lines.append(f"  2. {display_receiver}!=pulse")
            lines.append(f"  3. {display_sender} & {display_receiver} same-time (race)")
            lines.append(f"  4. wait-timeout")
            lines.append("")
            
        elif phase == 'ready_valid':
            lines.append("Timing Diagram (Ready-Valid):")
            lines.append("-" * 60)
            lines.append("")
            lines.append("Clock cycles: 0   1   2   3   4   5   6   7   8")
            lines.append(format_waveform_line("clk") + " |___|‾‾‾|___|‾‾‾|___|‾‾‾|___|‾‾‾|___|")
            lines.append("")
            lines.append(format_signal_name(display_sender, "sender") + " └─────────────────────┘   └─────────┘")
            lines.append(format_signal_name(display_receiver, "receiver") + "     └───────┘   └───────┘   └───────┘")
            lines.append("")
            
            lines.append("Pass:")
            lines.append(f"  1. {display_sender}=1 & {display_receiver}=1 -> transfer-OK")
            lines.append(f"  2. {display_sender}=1 (hold multi-cycle)")
            lines.append(f"  3. {display_receiver}=0 -> {display_sender}=0 (pause)")
            lines.append(f"  4. no-hang (hang=both-wait)")
            lines.append("")
            
            lines.append("Fail:")
            lines.append(f"  1. {display_receiver}=0 but data-send (mismatch)")
            lines.append(f"  2. wrong-data (corruption=bad-bits)")
            lines.append(f"  3. hang (both-stuck=no-progress)")
            lines.append(f"  4. bad-sequence (wrong-order)")
            lines.append("")
    
    elif plugin_name == 'pulseWidth':
        pulse_type = data.get('pulse_type', '?')
        target_signal = data.get('target_signal', '?')
        min_width = data.get('min_width', '?')
        max_width = data.get('max_width', '?')
        base_clock = data.get('base_clock', '?')
        trigger_signal = data.get('trigger_signal', '?')
        
        # Get base reset from module_info
        base_rst = '?'
        if state.module_info.resets:
            base_rst = state.module_info.resets[0].get('name', '?')
        
        # Truncate signal names to 10 characters
        display_signal = target_signal if len(str(target_signal)) <= 10 else str(target_signal)[:7] + "..."
        display_base_clk = base_clock if len(str(base_clock)) <= 10 else str(base_clock)[:7] + "..."
        display_trigger = trigger_signal if len(str(trigger_signal)) <= 10 else str(trigger_signal)[:7] + "..."
        
        lines.append("=" * 60)
        lines.append("PULSE WIDTH ASSERTION")
        lines.append("=" * 60)
        lines.append("")
        lines.append(f"Pulse Type: {pulse_type}")
        lines.append(f"Target Signal: {display_signal}")
        
        if pulse_type == 'hpulse':
            lines.append(f"Base Clock: {display_base_clk} (for counting)")
        elif pulse_type == 'vpulse':
            lines.append(f"Trigger Signal: {display_trigger} (for edge detection)")
        
        lines.append(f"Min Width: {min_width}")
        lines.append(f"Max Width: {max_width}")
        lines.append(f"Base Reset: {base_rst}")
        lines.append("")
        
        # Show available parameters
        if state.module_info and state.module_info.parameters:
            lines.append("Available Parameters:")
            for param in state.module_info.parameters[:5]:  # Show first 5
                param_name = param.get('name', '')
                param_default = param.get('default', '')
                lines.append(f"  {param_name} = {param_default}")
            if len(state.module_info.parameters) > 5:
                lines.append(f"  ... and {len(state.module_info.parameters) - 5} more")
            lines.append("")
        
        lines.append("Pulse Width Range:")
        lines.append("-" * 60)
        lines.append("")
        lines.append(f"  {display_signal} pulse width must be between {min_width} and {max_width}")
        lines.append("")
        lines.append(f"  _____|{'‾' * 20}|_____")
        lines.append(f"       <--  {min_width}-{max_width}  -->")
        lines.append("")
        
        lines.append("Result:")
        lines.append(f"  OK:   pulse_width is between {min_width} and {max_width}")
        lines.append(f"  FAIL: pulse_width is outside {min_width}-{max_width} range")
        lines.append("")
    
    else:
        lines.append(f"Assertion Type: {plugin_name}")
        lines.append("Configuration:")
        for key, val in sorted(data.items()):
            if val:
                lines.append(f"  {key}: {val}")
    
    return lines


def _render_confirmation(stdscr: "curses._CursesWindow", state: AppState, top: int, margin_x: int, box_h: int, box_w: int) -> None:
    """Render confirmation screen before creating assertion."""
    y = top + 2
    
    try:
        title = "Confirm Assertion Configuration"
        stdscr.addnstr(y, margin_x + 2, title, box_w - 4, curses.A_BOLD | curses.color_pair(_PAIR_BY_NAME.get("green", 0)))
        y += 2
    except curses.error:
        pass
    
    # Show all configured fields
    plugins = _get_assertion_plugins_info()
    plugin = next((p for p in plugins if p['name'] == state.assertion_selected_type), None)
    if plugin:
        try:
            for field in plugin['fields']:
                if y >= top + box_h - 5:
                    break
                
                field_name = field['name']
                prompt = field['prompt']
                val = state.assertion_input_data.get(field_name, '')
                
                line = f"  {prompt}: "
                if val:
                    line += val
                    color = _PAIR_BY_NAME.get("green", 0)
                else:
                    line += "[not set]"
                    color = _PAIR_BY_NAME.get("dim", 0)
                
                stdscr.addnstr(y, margin_x + 2, _truncate(line, box_w - 4), box_w - 4, curses.color_pair(color))
                y += 1
        except curses.error:
            pass
    
    # Action instruction
    try:
        y = top + box_h - 4
        inst = "Press Enter to create or type 'q' to cancel"
        stdscr.addnstr(y, margin_x + 2, _truncate(inst, box_w - 4), box_w - 4, curses.A_BOLD | curses.color_pair(_PAIR_BY_NAME.get("yellow", 0)))
    except curses.error:
        pass


def _generate_preview_content(state: AppState) -> List[str]:
    """Generate full preview content for file generation wizard."""
    if not state.gen_filename or not state.gen_file_type or not state.gen_data_source:
        return ["(Configuration incomplete)"]
    
    # Determine what to include
    include_asserts = state.gen_data_source in ('1', '3')
    include_signals = state.gen_data_source in ('2', '3')
    
    # For "both" mode, show the file selected by gen_preview_file_idx
    if state.gen_file_type == 3:  # Both
        if state.gen_preview_file_idx == 0:
            # Show interface
            content = _generate_interface_content(state, include_asserts, include_signals)
        else:
            # Show instance
            content = _generate_instance_content(state, include_asserts, include_signals)
    elif state.gen_file_type == 1:  # Interface only
        content = _generate_interface_content(state, include_asserts, include_signals)
    else:  # Instance only
        content = _generate_instance_content(state, include_asserts, include_signals)
    
    return content.split('\n')


def _generate_files(state: AppState) -> str:
    """Generate interface and/or instance files based on wizard configuration."""
    if not state.gen_filename:
        return "ERROR: Filename not set"
    if not state.gen_file_type:
        return "ERROR: File type not selected"
    if not state.gen_data_source:
        return "ERROR: Data source not selected"
    
    try:
        import json
        from pathlib import Path
        
        out_dir = state.out_dir or Path.cwd()
        out_dir = Path(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        
        # Determine what to generate
        gen_iface = state.gen_file_type in (1, 3)
        gen_inst = state.gen_file_type in (2, 3)
        include_asserts = state.gen_data_source in ('1', '3')
        include_signals = state.gen_data_source in ('2', '3')
        
        generated_files = []
        
        # Generate interface file
        if gen_iface:
            iface_path = out_dir / f"{state.gen_filename}.if.sv"
            iface_content = _generate_interface_content(state, include_asserts, include_signals)
            iface_path.write_text(iface_content, encoding='utf-8')
            generated_files.append(str(iface_path))
        
        # Generate instance file
        if gen_inst:
            inst_path = out_dir / f"{state.gen_filename}.inst.sv"
            inst_content = _generate_instance_content(state, include_asserts, include_signals)
            inst_path.write_text(inst_content, encoding='utf-8')
            generated_files.append(str(inst_path))
        
        msg = f"Generated {len(generated_files)} file(s):\n" + "\n".join(f"  {Path(f).name}" for f in generated_files)
        state.gen_wizard_active = False
        return msg
        
    except Exception as e:
        return f"ERROR generating files: {str(e)[:100]}"


def _generate_interface_content(state: AppState, include_asserts: bool, include_signals: bool) -> str:
    """Generate interface file content."""
    lines = []
    lines.append("// Auto-generated interface file")
    lines.append(f"// Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    
    if include_asserts and state.assertions:
        lines.append("// ===== ASSERTIONS =====")
        for i, asrt in enumerate(state.assertions, 1):
            lines.append(f"// [{i}] {asrt.get('type', 'unknown').upper()} assertion")
            lines.append(f"//     Name: {asrt.get('name', 'unnamed')}")
        lines.append("")
    
    if include_signals and (state.module_info.inputs or state.module_info.outputs):
        lines.append("// ===== SIGNALS =====")
        lines.append("// Input signals:")
        for inp in state.module_info.inputs[:20]:
            lines.append(f"//   - {inp.get('name', '?')} ({inp.get('width', '?')} bits)")
        if len(state.module_info.inputs) > 20:
            lines.append(f"//   ... and {len(state.module_info.inputs) - 20} more")
        
        lines.append("// Output signals:")
        for out in state.module_info.outputs[:20]:
            lines.append(f"//   - {out.get('name', '?')} ({out.get('width', '?')} bits)")
        if len(state.module_info.outputs) > 20:
            lines.append(f"//   ... and {len(state.module_info.outputs) - 20} more")
        lines.append("")
    
    lines.append("// Please add your interface properties here")
    return "\n".join(lines)


def _generate_instance_content(state: AppState, include_asserts: bool, include_signals: bool) -> str:
    """Generate instance file content."""
    lines = []
    lines.append("// Auto-generated instance file")
    lines.append(f"// Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    
    if include_asserts and state.assertions:
        lines.append("// ===== ASSERTIONS =====")
        for i, asrt in enumerate(state.assertions, 1):
            lines.append(f"// [{i}] {asrt.get('type', 'unknown').upper()} assertion")
            lines.append(f"//     Name: {asrt.get('name', 'unnamed')}")
        lines.append("")
    
    if include_signals and (state.module_info.inputs or state.module_info.outputs):
        lines.append("// ===== SIGNALS =====")
        lines.append("// Input signals:")
        for inp in state.module_info.inputs[:20]:
            lines.append(f"//   - {inp.get('name', '?')} ({inp.get('width', '?')} bits)")
        if len(state.module_info.inputs) > 20:
            lines.append(f"//   ... and {len(state.module_info.inputs) - 20} more")
        
        lines.append("// Output signals:")
        for out in state.module_info.outputs[:20]:
            lines.append(f"//   - {out.get('name', '?')} ({out.get('width', '?')} bits)")
        if len(state.module_info.outputs) > 20:
            lines.append(f"//   ... and {len(state.module_info.outputs) - 20} more")
        lines.append("")
    
    lines.append("// Please add your binding or instance here")
    return "\n".join(lines)


def _handle_assertion_wizard_command(state: AppState, cmdline: str) -> Tuple[str, bool]:
    """
    Handle commands for step-by-step wizard.
    Returns (message, should_exit_wizard).
    
    Auto-advance to next step after each input (no need to type 'next').
    """
    cmd = cmdline.strip().lower()
    
    # Allow empty command to proceed to Confirm stage handling
    # (empty Enter means 'confirm' in confirm stage)
    if not cmd and state.assertion_wizard_stage != 'confirm':
        return "", False
    
    # Universal: quit
    if cmd in ('q', 'quit'):
        state.assertion_wizard_active = False
        state.assertion_wizard_stage = ""
        state.assertion_selected_type = None
        state.assertion_input_data.clear()
        state.assertion_signal_ports.clear()
        state.assertion_current_field_idx = 0
        state.assertion_waiting_custom_number = False
        return "Cancelled", True
    
    # Stage 1: Select Type
    if state.assertion_wizard_stage == 'select_type':
        if cmd.isdigit():
            idx = int(cmd) - 1
            plugins = _get_assertion_plugins_info()
            if 0 <= idx < len(plugins):
                selected = plugins[idx]
                state.assertion_selected_type = selected['name']
                state.assertion_wizard_stage = 'input_data'
                state.assertion_current_field_idx = 0
                state.assertion_input_data.clear()
                state.assertion_signal_ports.clear()
                
                fields = selected.get('fields', [])
                if not fields:
                    return f"No fields defined for {selected['name']}", True
                current_field = fields[0]
                msg = f"Selected {selected['name'].upper()}\n"
                msg += f"\nStep 1/{len(fields)}: {current_field.get('title', '')}\n"
                msg += current_field.get('description', '')
                return msg, False
            else:
                return f"Invalid. Choose 1-{len(plugins)}", False
        return "Enter number to select type", False
    
    # Stage 2: Input Data (step-by-step, input saves immediately, Enter to advance)
    elif state.assertion_wizard_stage == 'input_data':
        # Special handling: If waiting for custom number input (exp_cnt_val [0] selected)
        if state.assertion_waiting_custom_number:
            # Validate that input is a number
            if not cmd.isdigit() and cmd != '':
                return "Please enter a valid number", False
            
            if cmd == '':
                return "Please enter a number value", False
            
            # Save the custom number as the expected count value
            state.assertion_input_data['exp_cnt_val'] = cmd
            state.assertion_signal_ports['exp_cnt_val'] = {}  # No port for custom number
            state.assertion_waiting_custom_number = False
            
            # Auto-advance to next field or confirmation
            plugins = _get_assertion_plugins_info()
            plugin = next((p for p in plugins if p['name'] == state.assertion_selected_type), None)
            if plugin:
                all_fields = plugin.get('fields', [])
                fields = _get_visible_fields(all_fields, state.assertion_input_data)
                
                if state.assertion_current_field_idx < len(fields) - 1:
                    state.assertion_current_field_idx += 1
                    next_field = fields[state.assertion_current_field_idx]
                    step = state.assertion_current_field_idx + 1
                    msg = f"\nStep {step}/{len(fields)}: {next_field.get('title', '')}\n"
                    msg += next_field.get('description', '')
                    return msg, False
                else:
                    # All fields done, move to confirm
                    state.assertion_wizard_stage = 'confirm'
                    return "\nAll steps complete. Review and press Enter to create.", False
            
            return "Error processing custom number", False
        
        plugins = _get_assertion_plugins_info()
        plugin = next((p for p in plugins if p['name'] == state.assertion_selected_type), None)
        if not plugin:
            return "Plugin not found", False
        
        all_fields = plugin.get('fields', [])
        if not all_fields:
            return "No fields", False
        
        # Populate base_clock options dynamically from state.module_info.clocks
        for field in all_fields:
            if field.get('name') == 'base_clock' and field.get('type') == 'choice':
                if state.module_info and state.module_info.clocks:
                    field['options'] = [clk.get('name', '') for clk in state.module_info.clocks if clk.get('name')]
                else:
                    field['options'] = ['I_CLK']  # Default if no clocks defined
        
        # Filter fields based on show_if conditions
        fields = _get_visible_fields(all_fields, state.assertion_input_data)
        
        if state.assertion_current_field_idx >= len(fields):
            return "No fields or invalid field index", False
        
        current_field = fields[state.assertion_current_field_idx]
        field_name = current_field['name']
        field_type = current_field['type']
        
        # Command: previous step
        if cmd in ('prev', 'p', 'back', 'b'):
            if state.assertion_current_field_idx > 0:
                state.assertion_current_field_idx -= 1
                prev_field = fields[state.assertion_current_field_idx]
                step = state.assertion_current_field_idx + 1
                msg = f"\nStep {step}/{len(fields)}: {prev_field.get('title', '')}\n"
                msg += prev_field.get('description', '')
                return msg, False
            else:
                return "Already at first step", False
        
        # Empty command: move to next field if current is filled
        if cmd == '':
            if field_name not in state.assertion_input_data:
                return "Please enter a value", False
            
            # Advance to next field
            if state.assertion_current_field_idx < len(fields) - 1:
                state.assertion_current_field_idx += 1
                next_field = fields[state.assertion_current_field_idx]
                step = state.assertion_current_field_idx + 1
                msg = f"\nStep {step}/{len(fields)}: {next_field.get('title', '')}\n"
                msg += next_field.get('description', '')
                return msg, False
            else:
                # All fields done, move to confirm
                state.assertion_wizard_stage = 'confirm'
                return "\nAll steps complete. Review and press Enter to create.", False
        
        # Field-specific input handling
        if field_type == 'choice':
            options = current_field.get('options', [])
            selected_option = None
            
            if cmd.isdigit():
                idx = int(cmd) - 1
                if 0 <= idx < len(options):
                    selected_option = options[idx]
            else:
                # Try to match option by name
                for opt in options:
                    if opt.lower() == cmd:
                        selected_option = opt
                        break
            
            if not selected_option:
                return f"Invalid. Choose [1-{len(options)}] or type option name", False
            
            # Save the selection
            state.assertion_input_data[field_name] = selected_option
            
            # Auto-advance to next field
            if state.assertion_current_field_idx < len(fields) - 1:
                state.assertion_current_field_idx += 1
                next_field = fields[state.assertion_current_field_idx]
                step = state.assertion_current_field_idx + 1
                msg = f"\nStep {step}/{len(fields)}: {next_field.get('title', '')}\n"
                msg += next_field.get('description', '')
                return msg, False
            else:
                # All fields done, move to confirm
                state.assertion_wizard_stage = 'confirm'
                return "\nAll steps complete. Review and press Enter to create.", False
        
        elif field_type == 'signal':
            # Command: navigate signal pages (n = next, N = previous)
            if cmd == 'n':
                if state.assertion_signal_list:
                    signals_per_page = 10
                    total_pages = max(1, (len(state.assertion_signal_list) + signals_per_page - 1) // signals_per_page)
                    state.assertion_signal_page = (state.assertion_signal_page + 1) % total_pages
                    return "Signal list advanced", False
                else:
                    return "No signals to navigate", False
            
            if cmd == 'N':
                if state.assertion_signal_list:
                    signals_per_page = 10
                    total_pages = max(1, (len(state.assertion_signal_list) + signals_per_page - 1) // signals_per_page)
                    state.assertion_signal_page = (state.assertion_signal_page - 1) % total_pages
                    return "Signal list rewound", False
                else:
                    return "No signals to navigate", False
            
            # Accept signal by number index, name, plain number, or expression
            selected_signal = None
            selected_port = None
            
            # Check if cmd is a plain number (no prefix like 'i' or 'o')
            # or an expression (contains operators like +, -, *, /)
            is_expression = any(op in cmd for op in ['+', '-', '*', '/', '(', ')'])
            
            if cmd.isdigit():
                # This could be: [1] signal index, or plain number like "5"
                idx = int(cmd)
                
                # Special handling for exp_cnt_val field: [0] = Custom Number Input
                if field_name == 'exp_cnt_val' and idx == 0:
                    state.assertion_waiting_custom_number = True
                    return "Enter custom number value for expected count:", False
                
                # Check if this index exists in signal map (small number = index)
                # If index >= 100, likely a plain number value
                if idx in state.assertion_signal_map and idx < 100:
                    # Treat as signal index
                    selected_signal, selected_port = state.assertion_signal_map[idx]
                else:
                    # Treat as plain number value (no signal, just number)
                    selected_signal = cmd  # Store the number as-is
                    selected_port = {}  # No port info for plain numbers
            
            elif is_expression:
                # Expression like "i1 - 1" or "o5 + 2"
                # Parse and validate the expression contains valid signal references
                # For now, accept the expression as-is and store it
                # The code generation will handle the expression properly
                selected_signal = cmd
                selected_port = {}  # No single port for expressions
            
            else:
                # Try to match by name (signal name without brackets)
                found = False
                for signal_name, port_dict in state.assertion_signal_map.values():
                    if signal_name.lower() == cmd.lower():
                        selected_signal = signal_name
                        selected_port = port_dict
                        found = True
                        break
                
                if not found:
                    # Could be a plain number value or expression that wasn't caught above
                    # Accept it as literal value
                    selected_signal = cmd
                    selected_port = {}
            
            # Save signal name to assertion_input_data and port_dict to assertion_signal_ports
            state.assertion_input_data[field_name] = selected_signal
            state.assertion_signal_ports[field_name] = selected_port
            
            # Auto-advance to next field
            if state.assertion_current_field_idx < len(fields) - 1:
                state.assertion_current_field_idx += 1
                next_field = fields[state.assertion_current_field_idx]
                step = state.assertion_current_field_idx + 1
                msg = f"\nStep {step}/{len(fields)}: {next_field.get('title', '')}\n"
                msg += next_field.get('description', '')
                return msg, False
            else:
                # All fields done, move to confirm
                state.assertion_wizard_stage = 'confirm'
                return "\nAll steps complete. Review and press Enter to create.", False
        
        elif field_type == 'string':
            # Accept any string input
            
            # Special validation for PulseWidth max_width field
            if field_name == 'max_width':
                # Check if max_width < min_width
                try:
                    min_val = int(state.assertion_input_data.get('min_width', '0'))
                    max_val = int(cmd)
                    if max_val < min_val:
                        return f"Error: Maximum width ({max_val}) must be >= Minimum width ({min_val}). Please re-enter.", False
                except ValueError:
                    return "Error: Please enter a valid number for maximum width.", False
            
            state.assertion_input_data[field_name] = cmd
            
            # Auto-advance to next field
            if state.assertion_current_field_idx < len(fields) - 1:
                state.assertion_current_field_idx += 1
                next_field = fields[state.assertion_current_field_idx]
                step = state.assertion_current_field_idx + 1
                msg = f"\nStep {step}/{len(fields)}: {next_field.get('title', '')}\n"
                msg += next_field.get('description', '')
                return msg, False
            else:
                # All fields done, move to confirm
                state.assertion_wizard_stage = 'confirm'
                return "\nAll steps complete. Review and press Enter to create.", False
        
        return "Invalid input for this field", False
    
    # Stage 3: Confirm
    elif state.assertion_wizard_stage == 'confirm':
        if cmd == '' or cmd in ('yes', 'y', 'confirm', 'ok', 'create'):
            result = _create_assertion_from_wizard(state)
            state.assertion_wizard_active = False
            state.assertion_wizard_stage = ""
            state.assertion_selected_type = None
            state.assertion_input_data.clear()
            state.assertion_signal_ports.clear()
            state.assertion_current_field_idx = 0
            state.assertion_waiting_custom_number = False
            return result, True
        
        elif cmd in ('prev', 'p', 'back', 'b'):
            # Go back to last field
            plugins = _get_assertion_plugins_info()
            plugin = next((p for p in plugins if p['name'] == state.assertion_selected_type), None)
            if plugin:
                all_fields = plugin['fields']
                visible_fields = _get_visible_fields(all_fields, state.assertion_input_data)
                
                state.assertion_wizard_stage = 'input_data'
                state.assertion_current_field_idx = len(visible_fields) - 1
                last_field = visible_fields[state.assertion_current_field_idx]
                step = state.assertion_current_field_idx + 1
                msg = f"\nStep {step}/{len(visible_fields)}: {last_field.get('title', '')}\n"
                msg += last_field.get('description', '')
                return msg, False
        
        return "Press Enter to create or type 'b' to edit", False
    
    return "", False


def _show_field_selector(state: AppState, fields: List[Dict[str, Any]]) -> str:
    """
    Show selector for the current field.
    Returns a message describing what to select.
    """
    if state.assertion_current_field_idx >= len(fields):
        return "Error: Field index out of range"
    
    field = fields[state.assertion_current_field_idx]
    field_name = field['name']
    prompt = field['prompt']
    field_type = field['type']
    field_num = state.assertion_current_field_idx + 1
    
    separator = "-" * 50
    lines = [f"\n{separator}"]
    lines.append(f"[{field_num}] {prompt}")
    lines.append(separator)
    
    # Build signal/option list based on field type
    if field_type == 'select':
        # For select fields, list options
        options = field.get('options', [])
        lines.append("Options:")
        for i, opt in enumerate(options, 1):
            lines.append(f"  [{i}] {opt}")
        lines.append(f"\nExample: set {field_num} 2phase")
    
    elif field_type == 'port':
        # For port fields, list available signals
        lines.append("Available signals:")
        lines.append("")
        
        # Add module inputs/outputs
        idx = 1
        if hasattr(state, 'module_info') and state.module_info:
            if state.module_info.get('inputs'):
                lines.append("[Module Inputs]")
                for port in state.module_info.get('inputs', [])[:10]:
                    port_name = port.get('name', f'in{idx}')
                    lines.append(f"  {port_name}")
                    idx += 1
                lines.append("")
            
            if state.module_info.get('outputs'):
                lines.append("[Module Outputs]")
                for port in state.module_info.get('outputs', [])[:10]:
                    port_name = port.get('name', f'out{idx}')
                    lines.append(f"  {port_name}")
                    idx += 1
                lines.append("")
        
        # Add user-defined signals (MS Signals)
        if state.conditions:
            lines.append("[User-Defined Signals]")
            for i, cond in enumerate(state.conditions, 1):
                cond_name = cond.get('name', f'signal{i}')
                lines.append(f"  {cond_name}")
        
        lines.append(f"\nExample: set {field_num} clk")
    
    elif field_type == 'string':
        lines.append("Text input field")
        current_val = state.assertion_input_data.get(field_name, '')
        if current_val:
            lines.append(f"Current: {current_val}")
        lines.append(f"\nExample: set {field_num} my_signal_name")
    
    lines.append(f"\nCommands: [#] next field | b back | done finish | q cancel")
    
    return "\n".join(lines)


def _create_assertion_from_wizard(state: AppState) -> str:
    """Create assertion entry and write to Excel using the actual plugin."""
    plugin_name = state.assertion_selected_type
    data = state.assertion_input_data
    
    try:
        # Get the plugin instance
        from assertions import get_registered_plugins  # type: ignore
        plugin_cls = None
        for cls in get_registered_plugins():
            if cls.plugin_name == plugin_name:
                plugin_cls = cls
                break
        
        if not plugin_cls:
            return f"Error: Plugin '{plugin_name}' not found"
        
        # Prepare data in the format the plugin expects
        plugin_data = dict(data)
        
        # Write to Excel using the Excel path
        excel_path = state.session_excel_path or state.excel_path
        if not excel_path or not Path(excel_path).exists():
            return f"Error: Excel file not found at {excel_path}"
        
        # The plugin's parse method writes to Excel
        # For now, we add to in-memory state
        assertion_entry = {
            'type': plugin_name,
            'data': plugin_data,
            'description': f"{plugin_name} assertion",
        }
        state.assertions.append(assertion_entry)
        
        # Try to write to Excel (pass state for signal width info)
        _write_assertion_to_excel(str(excel_path), plugin_name, plugin_data, state)
        
        return f"OK. {plugin_name.upper()} assertion created and saved to Excel."
    
    except Exception as e:
        return f"OK. Assertion created in memory. Excel write failed: {str(e)}"


def _write_assertion_to_excel(excel_path: str, plugin_name: str, data: Dict[str, Any], state: Optional['AppState'] = None) -> None:
    """Write assertion data to the corresponding Excel sheet using plugin structure."""
    try:
        from openpyxl import load_workbook  # type: ignore
        
        wb = load_workbook(excel_path)
        
        # Helper: Find sheet case-insensitively
        def find_sheet_ci(target_name: str) -> Optional[str]:
            """Find sheet by name (case-insensitive). Returns actual sheet name or None."""
            target_lower = target_name.lower()
            for name in wb.sheetnames:
                if name.lower() == target_lower:
                    return name
            return None
        
        # Helper: Format bit width as "[msb:lsb]" from calculated_bit_width
        def format_bit_width(bit_width: int) -> str:
            """
            Format calculated bit width as "[msb:lsb]".
            Input: 8 returns "[7:0]"
            Input: 0 returns ""
            """
            if bit_width > 0:
                return f"[{bit_width-1}:0]"
            return ""
        
        # Helper: Get width string for a signal field (with parameter detection)
        def get_signal_width(field_name: str) -> Tuple[str, bool]:
            """
            Get bit width string for a signal field.
            Returns: (width_str, has_unresolved_params)
            Examples:
              - ("[7:0]", False) - resolved width
              - ("[DATA_WIDTH-1:0]", True) - unresolved parameter expression
              - ("", False) - no width info
            """
            port_dict = port_map.get(field_name)
            if port_dict:
                # Try calculated_bit_width first (from rtl_parser)
                calculated_width = port_dict.get('calculated_bit_width', 0)
                if calculated_width > 0:
                    return (format_bit_width(calculated_width), False)
                
                # Check if there's a width expression with unresolved parameters
                signal_name = data.get(field_name, '')
                import re
                match = re.match(r'^([^\[]*)\[([^\]]*)\]$', signal_name)
                if match:
                    width_expr = match.group(2).strip()
                    # Check if expression contains identifiers (parameter references)
                    has_identifiers = bool(re.search(r'[A-Za-z_]\w*', width_expr))
                    if has_identifiers:
                        return (f"[{width_expr}]", True)  # Unresolved parameter
                    else:
                        return (f"[{width_expr}]", False)
            
            return ("", False)
        
        # Map from assertion_signal_ports passed via state
        port_map = state.assertion_signal_ports if hasattr(state, 'assertion_signal_ports') else {}
        
        # Determine sheet name based on plugin (case-insensitive lookup)
        if plugin_name == 'counter':
            sheet_name = find_sheet_ci('Counter')
            if not sheet_name:
                sheet_name = 'Counter'
                wb.create_sheet(sheet_name)
            
            ws = wb[sheet_name]
            
            # Find next empty row starting from row 8 (after header at row 7)
            # Counter sheet has header at row 7, data starts at row 8
            # Column 2 is "Target" column (B)
            target_col = 2
            next_row = 8
            
            # Check if this is the first real data or if we have sample data
            # Sample data has specific values like 'cnt', 'plus_condition', etc.
            first_target = ws.cell(row=8, column=target_col).value
            is_sample_data = (first_target and str(first_target).strip() in ['cnt', 'counter', 'sample'])
            
            if is_sample_data:
                # Clear sample data only (rows 8+) - keep headers
                # Clear from row 8 onwards (skip merged cells)
                from openpyxl.cell import MergedCell
                for row in range(8, ws.max_row + 1):
                    for col in range(1, 15):
                        cell = ws.cell(row=row, column=col)
                        # Skip merged cells (they're read-only)
                        if not isinstance(cell, MergedCell):
                            cell.value = None
                next_row = 8
            else:
                # Find the next empty row after existing data
                while ws.cell(row=next_row, column=target_col).value:
                    next_row += 1
            
            # Get signal name and width
            target_signal = data.get('target', '')
            target_width, has_unresolved = get_signal_width('target')
            
            # Extract clean signal name (remove [...] if present)
            import re
            match = re.match(r'^([^\[]*)(?:\[.*\])?$', target_signal)
            target_name = match.group(1).strip() if match else target_signal.strip()
            
            # Write data in plugin's expected format
            # Counter sheet columns (from row 7): col2=Target, col3=Plus, col4=Reset, col5=Trigger, col6=Expect Count
            ws.cell(row=next_row, column=2, value=target_name)
            ws.cell(row=next_row, column=3, value=data.get('plus_con', ''))
            ws.cell(row=next_row, column=4, value=data.get('reset_con', ''))
            ws.cell(row=next_row, column=5, value=data.get('trigger_con', ''))
            ws.cell(row=next_row, column=6, value=data.get('exp_cnt_val', ''))
        
        elif plugin_name == 'handshake':
            sheet_name = find_sheet_ci('Handshake')
            if not sheet_name:
                sheet_name = 'Handshake'
                wb.create_sheet(sheet_name)
            
            ws = wb[sheet_name]
            
            # Find next empty row starting from row 7 (after header at row 6)
            # handshake sheet has header at row 6, data starts at row 7
            # Column 3 is "Type" column (C)
            type_col = 3
            next_row = 7
            
            # Check if this is the first real data or if we have sample data
            first_type = ws.cell(row=7, column=type_col).value
            first_sender = ws.cell(row=7, column=4).value
            
            # Check if this looks like sample data (has specific sender values or empty)
            is_sample_data = False
            if first_type and first_sender:
                type_str = str(first_type).strip()
                sender_str = str(first_sender).strip()
                # If sender is one of the generic sample values, clear it
                if type_str in ['ready_valid', '4phase', '2phase'] and sender_str in ['valid', 'req']:
                    is_sample_data = True
            
            if is_sample_data:
                # This appears to be the original sample - clear it
                from openpyxl.cell import MergedCell
                for row in range(7, ws.max_row + 1):
                    for col in range(1, 10):
                        cell = ws.cell(row=row, column=col)
                        # Skip merged cells (they're read-only)
                        if not isinstance(cell, MergedCell):
                            cell.value = None
                next_row = 7
            else:
                # Find the next empty row after existing data
                while ws.cell(row=next_row, column=type_col).value:
                    next_row += 1
            
            # Get signal names and widths
            sender_signal = data.get('sender', '')
            receiver_signal = data.get('receiver', '')
            sender_width, sender_unresolved = get_signal_width('sender')
            receiver_width, receiver_unresolved = get_signal_width('receiver')
            
            # Extract clean signal names
            import re
            sender_match = re.match(r'^([^\[]*)(?:\[.*\])?$', sender_signal)
            sender_name = sender_match.group(1).strip() if sender_match else sender_signal.strip()
            receiver_match = re.match(r'^([^\[]*)(?:\[.*\])?$', receiver_signal)
            receiver_name = receiver_match.group(1).strip() if receiver_match else receiver_signal.strip()
            
            # Write data in plugin's expected format
            # handshake sheet columns (from row 6): col3=Type, col4=Sender, col5=Receiver
            ws.cell(row=next_row, column=3, value=data.get('phase_type', ''))
            ws.cell(row=next_row, column=4, value=sender_name)
            ws.cell(row=next_row, column=5, value=receiver_name)
        
        elif plugin_name == 'pulseWidth':
            sheet_name = find_sheet_ci('PulseWidth')
            if not sheet_name:
                sheet_name = 'PulseWidth'
                wb.create_sheet(sheet_name)
            
            ws = wb[sheet_name]
            
            # Find next empty row starting from row 7 (after header at row 6)
            # pulseWidth sheet has header at row 6, data starts at row 7
            # Column 3 is "Type" column (C)
            type_col = 3
            next_row = 7
            
            # Check if this is the first real data or if we have sample data
            first_type = ws.cell(row=7, column=type_col).value
            is_sample_data = (first_type and str(first_type).strip() in ['hpulse', 'vpulse', 'sample'])
            
            # Only clear if it's still sample data with default values
            first_target = ws.cell(row=7, column=5).value  # Target_Pulse column
            if is_sample_data and first_target and str(first_target).strip() == 'target_pulse':
                # This appears to be the original sample - clear it
                from openpyxl.cell import MergedCell
                for row in range(7, ws.max_row + 1):
                    for col in range(1, 10):
                        cell = ws.cell(row=row, column=col)
                        # Skip merged cells (they're read-only)
                        if not isinstance(cell, MergedCell):
                            cell.value = None
                next_row = 7
            else:
                # Find the next empty row after existing data
                while ws.cell(row=next_row, column=type_col).value:
                    next_row += 1
            
            # Get signal name and width
            target_signal = data.get('target_signal', '')
            signal_width, signal_unresolved = get_signal_width('target_signal')
            
            # Extract clean signal name
            import re
            match = re.match(r'^([^\[]*)(?:\[.*\])?$', target_signal)
            signal_name = match.group(1).strip() if match else target_signal.strip()
            
            # Write data
            # pulseWidth sheet columns (from row 6): col3=Type, col4=Count_Trigger, col5=Target_Pulse, col6=Min, col7=Max
            pulse_type = data.get('pulse_type', 'hpulse')  # Get actual pulse type
            
            # Determine Count_Trigger value based on pulse type
            if pulse_type == 'hpulse':
                # Use actual base clock name from data
                count_trigger = data.get('base_clock', '<Base Clock>')
            elif pulse_type == 'vpulse':
                # Use actual trigger signal name from data
                trigger_sig = data.get('trigger_signal', '')
                # Extract clean trigger signal name
                trigger_match = re.match(r'^([^\[]*)(?:\[.*\])?$', trigger_sig)
                count_trigger = trigger_match.group(1).strip() if trigger_match else trigger_sig.strip()
            else:
                count_trigger = '<Base Clock>'  # Fallback
            
            ws.cell(row=next_row, column=3, value=pulse_type)
            ws.cell(row=next_row, column=4, value=count_trigger)
            ws.cell(row=next_row, column=5, value=signal_name)
            ws.cell(row=next_row, column=6, value=data.get('min_width', ''))
            ws.cell(row=next_row, column=7, value=data.get('max_width', ''))
        
        wb.save(excel_path)
        wb.close()
    
    except Exception as e:
        raise RuntimeError(f"Failed to write to Excel: {str(e)}")


# ---------------------- Onboarding Wizard ----------------------------------

def _render_onboarding(stdscr: "curses._CursesWindow", state: AppState) -> None:
    max_y, max_x = stdscr.getmaxyx()
    title = "First Run - Guided Setup"
    try:
        stdscr.addnstr(0, 2, title, max_x - 4, curses.A_BOLD)
    except curses.error:
        pass
    # Box below title, with margins and safe height
    margin_x = 2
    top = 2
    reserved_bottom = 14  # space for candidates strip (up to 10) + hint + prompt
    box_h = max(6, max_y - top - reserved_bottom)
    box_w = max(20, max_x - (margin_x * 2))
    # Compact the box to avoid large empty area and overlap
    stage = (state.onboarding_stage or "").lower()
    if stage == 'rtl':
        box_h = min(box_h, 8)
    elif stage == 'module':
        box_h = min(box_h, 12)
    elif stage == 'hierarchy':
        box_h = min(box_h, 20)  # Increased for hierarchy list
    elif stage == 'excel':
        box_h = min(box_h, 13)  # Increased by 3 lines for better readability
    if top + box_h > max_y:
        box_h = max(6, max_y - top - 1)
    _draw_ascii_box(stdscr, top, margin_x, box_h, box_w)
    # No extra subtitle for a more professional look; rely on clear step guides below
    # Stage-specific content
    if state.onboarding_stage == 'rtl':
        lines = [
            "Step 1/4 — RTL File Path",
            "- Enter file (.v/.sv) to pick modules in that file only.",
            "- Enter folder to scan all modules under it (recursive).",
            "- Tab: show candidates + extend common prefix. Use '/'.",
            "- Esc: cancel. prev/back: previous (exit).",
            "",
            "Note: After selecting module, you'll enter the hierarchy path.",
        ]
        for i, t in enumerate(lines):
            try:
                stdscr.addnstr(top + 2 + i, margin_x + 2, _truncate(t, box_w - 4), box_w - 4)
            except curses.error:
                pass
        # Candidates are rendered in a reserved strip near the bottom; do not draw here
    elif state.onboarding_stage == 'module':
        # Expand box to use available height except candidates/hint/prompt area
        box_h = max(8, max_y - top - 10)
        _draw_ascii_box(stdscr, top, margin_x, box_h, box_w)
        content_bottom = top + box_h - 2  # last drawable row inside the box
        try:
            stdscr.addnstr(min(top + 2, content_bottom), margin_x + 2, "Step 2/4 — Module: number+Enter. f <text> filter, F clear, n/N page (wrap), prev/back.", box_w - 4, curses.A_BOLD)
        except curses.error:
            pass
        # Render filtered, paginated module list in 3 columns
        mods = state.onboarding_modules
        if state.onboarding_filter:
            mods = [m for m in mods if state.onboarding_filter.lower() in m.lower()]
        page = max(0, state.onboarding_page)
        inner_h = max(1, content_bottom - (top + 4) + 1)
        inner_w = box_w - 4
        cols = 3
        col_w = max(10, inner_w // cols)
        rows = max(1, inner_h)
        items_per_page = rows * cols
        total = len(mods)
        start = (page * items_per_page) % max(1, ((total + items_per_page - 1) // items_per_page) * items_per_page)
        slice_mods = mods[start:start + items_per_page]
        # Grid render with clamping
        for i, m in enumerate(slice_mods):
            r = i % rows
            c = i // rows
            y = top + 4 + r
            if y > content_bottom:
                break
            x = margin_x + 2 + c * col_w
            label = f"[{start + i + 1}] {m}"
            try:
                stdscr.addnstr(y, x, _truncate(label, col_w - 2), col_w - 2)
            except curses.error:
                pass
        # Footer with page info (wrap-aware)
        try:
            pages = max(1, (total + items_per_page - 1) // items_per_page)
            stdscr.addnstr(content_bottom, margin_x + 2, _truncate(f"Page {page % pages + 1}/{pages}  Total: {total}", inner_w), inner_w, curses.A_DIM)
        except curses.error:
            pass
    elif state.onboarding_stage == 'hierarchy':
        try:
            stdscr.addnstr(top + 2, margin_x + 2, _truncate("Step 3/4 — Module Hierarchy", box_w - 4), box_w - 4, curses.A_BOLD)
            y = top + 4
            content_bottom = top + box_h - 2
            
            # Show the selected instance's hierarchy as the DUT path (in green - this is what user should use)
            if state.selected_instance and y <= content_bottom:
                hierarchy = state.selected_instance.get("hierarchy", "")
                
                # Check if hierarchy is complete (contains a dot) or is incomplete
                is_complete = "." in hierarchy
                
                if is_complete:
                    stdscr.addnstr(y, margin_x + 2, _truncate("[DUT Hierarchy Path]", box_w - 4), box_w - 4, curses.A_BOLD)
                    y += 1
                    
                    # Display hierarchy path in GREEN - user should copy this
                    if hierarchy and y <= content_bottom:
                        hierarchy_display = f"  {hierarchy}"
                        stdscr.addnstr(y, margin_x + 2, _truncate(hierarchy_display, box_w - 4), box_w - 4, 
                                     curses.color_pair(_PAIR_BY_NAME.get("green", 0)) | curses.A_BOLD)
                        y += 2
                else:
                    # Incomplete hierarchy - show warning in RED
                    stdscr.addnstr(y, margin_x + 2, _truncate("[DUT Hierarchy Path] - INCOMPLETE", box_w - 4), box_w - 4, 
                                 curses.color_pair(_PAIR_BY_NAME.get("red", 0)) | curses.A_BOLD)
                    y += 1
                    
                    if hierarchy and y <= content_bottom:
                        hierarchy_display = f"  {hierarchy}"
                        stdscr.addnstr(y, margin_x + 2, _truncate(hierarchy_display, box_w - 4), box_w - 4, 
                                     curses.color_pair(_PAIR_BY_NAME.get("red", 0)) | curses.A_BOLD)
                        y += 1
                    
                    if y <= content_bottom:
                        stdscr.addnstr(y, margin_x + 2, _truncate("WARNING: Could not find full hierarchy path.", box_w - 4), box_w - 4, 
                                     curses.color_pair(_PAIR_BY_NAME.get("red", 0)))
                        y += 1
                    
                    if y <= content_bottom:
                        stdscr.addnstr(y, margin_x + 2, _truncate("Please type the full path manually (e.g., top.dut.u1_sync_signal).", box_w - 4), box_w - 4, 
                                     curses.color_pair(_PAIR_BY_NAME.get("red", 0)))
                        y += 2
            
            # Show auto-detected alternative hierarchies
            if state.occs and len(state.occs) > 0 and y <= content_bottom:
                stdscr.addnstr(y, margin_x + 2, _truncate("[Alternative Hierarchies]", box_w - 4), box_w - 4, curses.A_BOLD)
                y += 1
                
                # Show up to 5 hierarchies
                for idx, occ in enumerate(state.occs[:5], 1):
                    if y > content_bottom:
                        break
                    hierarchy = occ.get("path", "")
                    if hierarchy:
                        line = f"  [{idx}] {hierarchy}"
                        stdscr.addnstr(y, margin_x + 2, _truncate(line, box_w - 4), box_w - 4, 
                                     curses.color_pair(_PAIR_BY_NAME.get("cyan", 0)))
                        y += 1
                
                if len(state.occs) > 5:
                    stdscr.addnstr(y, margin_x + 2, _truncate(f"  ... and {len(state.occs) - 5} more", box_w - 4), box_w - 4, curses.A_DIM)
                    y += 1
                
                y += 1
            
            # Instructions
            if y <= content_bottom:
                if is_complete:
                    stdscr.addnstr(y, margin_x + 2, _truncate("Press Enter to accept the DUT hierarchy above,", box_w - 4), box_w - 4)
                else:
                    stdscr.addnstr(y, margin_x + 2, _truncate("Press Enter to accept the path above (even if incomplete),", box_w - 4), box_w - 4)
                y += 1
            
            if y <= content_bottom:
                stdscr.addnstr(y, margin_x + 2, _truncate("or type number to select an alternative,", box_w - 4), box_w - 4)
                y += 1
            
            if y <= content_bottom:
                stdscr.addnstr(y, margin_x + 2, _truncate("or type custom path (e.g., top.dut.abc.u_abc).", box_w - 4), box_w - 4)
                y += 1
            
            if y <= content_bottom:
                stdscr.addnstr(y, margin_x + 2, _truncate("Type 'prev' or 'back' to return.", box_w - 4), box_w - 4, curses.A_DIM)
        except curses.error:
            pass
    elif state.onboarding_stage == 'excel':
        title = "Step 4/4 — Excel"
        found = state.onboarding_excel_autofound or state.excel_path
        try:
            stdscr.addnstr(top + 2, margin_x + 2, _truncate(title, box_w - 4), box_w - 4, curses.A_BOLD)
            content_bottom = top + box_h - 2
            y = top + 4
            
            if found:
                # Instruction lines
                if y <= content_bottom:
                    stdscr.addnstr(y, margin_x + 2, _truncate("Specify the path to the reference Excel file.", box_w - 4), box_w - 4)
                    y += 1
                
                if y <= content_bottom:
                    stdscr.addnstr(y, margin_x + 2, _truncate("Use '/' separator. Type 'prev' or 'back' to return to previous step.", box_w - 4), box_w - 4, curses.A_DIM)
                    y += 2  # Extra spacing before auto-detected section
                
                # Auto-detected file display with green highlight
                if y <= content_bottom:
                    stdscr.addnstr(y, margin_x + 2, _truncate("✓ Auto-detected:", box_w - 4), box_w - 4, curses.color_pair(_PAIR_BY_NAME.get("green", 0)) | curses.A_BOLD)
                    y += 1
                
                # Show the file path in green
                if y <= content_bottom:
                    file_display = f"  {_sanitize_path_for_display(str(found))}"
                    stdscr.addnstr(y, margin_x + 2, _truncate(file_display, box_w - 4), box_w - 4, curses.color_pair(_PAIR_BY_NAME.get("green", 0)))
                    y += 2  # Extra spacing
                
                # Confirmation prompt with color emphasis
                if y <= content_bottom:
                    prompt_msg = "Press Enter to accept"
                    stdscr.addnstr(y, margin_x + 2, _truncate(prompt_msg, box_w - 4), box_w - 4, curses.A_BOLD)
                    y += 1
                
                if y <= content_bottom:
                    or_msg = "or type a custom path below."
                    stdscr.addnstr(y, margin_x + 2, _truncate(or_msg, box_w - 4), box_w - 4, curses.A_DIM)
            else:
                # No auto-detected file
                if y <= content_bottom:
                    stdscr.addnstr(y, margin_x + 2, _truncate("Type the path to your Excel file and press Enter.", box_w - 4), box_w - 4)
                    y += 2
                
                if y <= content_bottom:
                    stdscr.addnstr(y, margin_x + 2, _truncate("No Excel found in Data/ folder.", box_w - 4), box_w - 4, curses.color_pair(_PAIR_BY_NAME.get("red", 0)) | curses.A_BOLD)
        except curses.error:
            pass


def _handle_onboarding_input(stdscr: "curses._CursesWindow", state: AppState, ch: int) -> bool:
    # Returns True if handled
    if state.onboarding_stage == 'rtl':
        if ch == 9:
            # Use existing completion on input line; nothing else to do
            return True
        if ch in (10, 13):
            # Accept current input line as rtl path
            # Read from prompt line content via main buffer; here we just move to next stage if set by main handler
            if state.rtl_start and Path(state.rtl_start).exists():
                # REMOVED: Build modules for next stage
                # This was overwriting the hierarchy paths set in _handle_rtl_selection_for_onboarding
                # The correct instance hierarchy is already set there (line ~1840)
                # OLD CODE (DISABLED):
                # try:
                #     mods_ctx, _mi, _occs = build_context_from_rtl(Path(state.rtl_start), None)
                #     mods = sorted(list(mods_ctx.keys()))
                #     state.onboarding_modules = mods
                # except Exception:
                #     state.onboarding_modules = []
                
                # Check if onboarding_modules was already set by _handle_rtl_selection_for_onboarding
                if not state.onboarding_modules:
                    # Fallback removed per user request - show error instead
                    state.onboarding_modules = ["ERROR: No instances found for selected RTL file"]
                
                state.onboarding_stage = 'module'
            return True
        return False
    if state.onboarding_stage == 'module':
        if ch in (ord('n'), ord('N')):
            # wrap page advance/back based on last rendered geometry
            total = len(state.onboarding_modules)
            # Recompute items per page similar to renderer (uses 3 cols)
            # We don't have box_h here; approximate from screen height
            max_y, max_x = stdscr.getmaxyx()
            inner_h = max(8, max_y - 10) - 6
            cols = 3
            items_per_page = max(1, inner_h) * cols
            pages = max(1, (total + items_per_page - 1) // items_per_page)
            cur = state.onboarding_page % pages
            if ch == ord('n'):
                cur = (cur + 1) % pages
            else:
                cur = (cur - 1) % pages
            state.onboarding_page = cur
            return True
        # Filtering with 'f <substr>' handled in command parser; here we just refresh
        if ch in (10, 13):
            # Expect the user to type: pick <module> or set module <name>
            if state.target_module:
                state.onboarding_stage = 'hierarchy'
            return True
        return False
    if state.onboarding_stage == 'hierarchy':
        # Simply wait for Enter - handled in main loop
        return False
    if state.onboarding_stage == 'excel':
        if ch in (10, 13):
            # Accept excel path; if empty and autofound exists, take autofound
            if not state.excel_path and state.onboarding_excel_autofound:
                state.excel_path = state.onboarding_excel_autofound
            
            if state.excel_path and Path(state.excel_path).exists():
                # CRITICAL: Create session before finishing onboarding
                ok, err = _create_session_excel_and_fill(state)
                if ok:
                    # Session created successfully
                    state.onboarding_active = False
                    state.onboarding_stage = None
                    # Save session snapshot
                    _save_session_snapshot(state)
                else:
                    # Session creation failed - show error and stay in Excel stage
                    state.excel_error = err or "Session creation failed"
                    _set_error_message(state.excel_error)
            return True
        return False
    return False


if __name__ == "__main__":
    run()



