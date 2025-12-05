#!/usr/bin/env python3
"""
Unified, modular assertion builder.

Features:
- RTL scan using scripts/rtl_parser.py internal APIs to extract ports for a target module.
- Define Excel population (reuses scripts/fill_define.py logic via JSON handoff).
- Extensible assertion plugins (see scripts/assertions/*) for each Excel sheet type.
- Configurable inputs via CLI or optional config file.

Quick start:
  python scripts/assertion_builder.py \
    --rtl-start EDA/RTL \
    --target-module blur_scaler \
    --excel Data/Assertion_TF.xlsx \
    --auto-define-fill \
    --out out/assertions

No-args interactive mode:
- Run without options to launch an interactive wizard that lets you pick
  RTL start path, target module, Excel file from Data/, plugins, and modes
  (Define fill, JSON output, SV generation).

Adding new assertion sheet:
- Create a plugin under scripts/assertions (see base.py and counter.py).
- Implement parse/generate_sv, register in registry.
- Run with --enable <plugin_name> or leave enabled by default.

Assistant prompt hint to extend:
- "Add a plugin 'sequence' for sheet 'sequence_gen' with headers ... that
    emits sequence ..." The AI should add scripts/assertions/sequence.py
    inheriting BaseAssertionPlugin and update registry.
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import re
from pathlib import Path
import platform
import shutil
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple, Type
import os
import importlib

# Make local 'scripts' directory importable no matter the CWD
_THIS_DIR = Path(__file__).resolve().parent
if str(_THIS_DIR) not in sys.path:
    sys.path.insert(0, str(_THIS_DIR))

# Import RTL parsing helpers from scripts/rtl_parser.py (as a library)
from rtl_parser import (  # type: ignore
    discover_files,
    find_rtl_root_from,
    build_modules_db,
    find_top_modules,
    find_occurrences_of_target,
    compute_env_for_occurrence,
    resolve_ports_with_params,
    classify_groups,
)

from assertions import get_registered_plugins, BaseAssertionPlugin  # type: ignore


def _run_pip_install(packages: List[str]) -> Tuple[bool, str]:
    """Attempt to install given pip packages non-interactively. Returns (ok, output)."""
    if not packages:
        return True, ""
    cmd = [
        sys.executable,
        "-m",
        "pip",
        "install",
        "--disable-pip-version-check",
        "--no-input",
    ] + packages
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True)
        out = (proc.stdout or "") + (proc.stderr or "")
        return proc.returncode == 0, out
    except Exception as e:
        return False, f"pip failed: {e}"


def _ensure_runtime_deps(require_tui: bool) -> None:
    """
    Ensure required Python dependencies are available. If missing, try to install.
    - Always checks: pandas, openpyxl (used by plugins/fill scripts)
    - If require_tui: also ensure curses (windows-curses on Windows)
    Fails with SystemExit if a required package cannot be installed.
    """
    missing: List[str] = []

    # Core Excel stack
    try:
        import pandas  # type: ignore  # noqa: F401
    except Exception:
        missing.append("pandas")
    try:
        import openpyxl  # type: ignore  # noqa: F401
    except Exception:
        missing.append("openpyxl")

    # TUI dependency
    win_needs_curses = False
    if require_tui:
        try:
            import curses  # type: ignore  # noqa: F401
        except Exception:
            if platform.system() == "Windows":
                # On Windows, the shim package provides _curses
                missing.append("windows-curses")
                win_needs_curses = True
            else:
                # On non-Windows, curses should be in stdlib; if not, abort
                raise SystemExit("curses module not available; cannot launch TUI")

    if missing:
        ok, out = _run_pip_install(missing)
        if not ok:
            print(out)
            raise SystemExit(f"Failed to install required packages: {', '.join(missing)}")

    # Verify imports after install
    try:
        import pandas  # type: ignore  # noqa: F401
        import openpyxl  # type: ignore  # noqa: F401
    except Exception as e:
        raise SystemExit(f"Dependencies not satisfied after install attempt: {e}")

    if require_tui:
        try:
            import curses  # type: ignore  # noqa: F401
        except Exception as e:
            if win_needs_curses:
                raise SystemExit("windows-curses installed but curses still unavailable; cannot launch TUI")
            raise SystemExit(f"curses unavailable: {e}")


def build_module_context(rtl_start: Path, target_module: Optional[str]) -> Dict[str, Any]:
    exts = [".v", ".sv"]
    rtl_root, found = find_rtl_root_from(rtl_start)
    start_scope_dir = rtl_start if rtl_start.is_dir() else rtl_start.parent
    files = sorted(set(discover_files(rtl_root, exts)) | set(discover_files(start_scope_dir, exts)), key=lambda p: str(p))
    # tb_top 등 Non-ANSI 전용 파일 제외
    files = _filter_rtl_ansi(files)
    modules = build_modules_db(files, allow_unknown=False)
    if not modules:
        raise SystemExit("No modules parsed from RTL scope")

    if not target_module:
        tops = find_top_modules(modules)
        # Fallback: pick the first top as target
        target_module = tops[0] if tops else next(iter(modules.keys()))

    # Resolve first occurrence for environment
    occs = find_occurrences_of_target(modules, target_module)
    env = compute_env_for_occurrence(occs[0], modules, {}) if occs else {}
    ports_resolved = resolve_ports_with_params(modules, target_module, env)
    cls = classify_groups(modules[target_module]["ports"])
    # Exclude clock/reset names from inputs
    ex_names = {x["name"] for x in cls.get("clocks", [])} | {x["name"] for x in cls.get("resets", [])}
    inputs_filtered = [it for it in ports_resolved["inputs"] if it["name"] not in ex_names]

    module_info = {
        "module": target_module,
        "inputs": inputs_filtered,
        "outputs": ports_resolved["outputs"],
        "inouts": ports_resolved["inouts"],
        "clocks": cls.get("clocks", []),
        "resets": cls.get("resets", []),
        "parameters": cls.get("parameters", []),
    }
    return {"modules": modules, "module_info": module_info, "occs": occs}


def fill_define_excel_if_needed(excel_path: Path, module_info: Dict[str, Any], out_json_dir: Path):
    out_json_dir.mkdir(parents=True, exist_ok=True)
    define_json_path = out_json_dir / "module_define.json"
    define_json = {
        "top_path": "",
        "module": module_info["module"],
        "rtl_file_path": module_info.get("rtl_file_path", ""),  # Add RTL file path
        "paths": [],
        "instances": [],
        "clocks": module_info["clocks"],
        "resets": module_info["resets"],
        "inputs": module_info["inputs"],
        "outputs": module_info["outputs"],
        "inouts": module_info["inouts"],
        "parameters": module_info["parameters"],
        "conditions": module_info.get("conditions", []),  # Add condition signals
    }
    define_json_path.write_text(json.dumps(define_json, ensure_ascii=False, indent=2), encoding="utf-8")

    # Reuse the fill_define.py script via import and call its main logic? It is CLI-oriented.
    # For simplicity and to avoid tight coupling, invoke it as a subprocess-like call would be ideal,
    # but here we keep it as an external step. The caller can run:
    #   python scripts/fill_define.py <excel_path> <define_json_path>
    return define_json_path


def _create_session_excel_copy(reference_excel: Path, target_module: str, out_dir: Path) -> Tuple[Path, Path]:
    """
    Reference 엑셀을 세션 폴더로 복사하고 구조화된 디렉터리를 생성합니다.

    Args:
        reference_excel: 원본 엑셀 파일 경로
        target_module: 대상 모듈명
        out_dir: 출력 디렉터리 (일반적으로 out/assertions)

    Returns:
        (session_excel_path, session_dir): 복사된 엑셀 경로와 세션 디렉터리 경로
    """
    # 세션 디렉터리: out/sessions/<module>-<timestamp>/
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    session_name = f"{target_module}-{ts}"
    session_dir = out_dir.parent / "sessions" / session_name
    session_dir.mkdir(parents=True, exist_ok=True)

    # 엑셀 파일명: <module>.xlsx
    dest_excel = session_dir / f"{target_module}.xlsx"

    try:
        shutil.copy2(reference_excel, dest_excel)
        print(f"✓ Session Excel created: {dest_excel}")
        print(f"✓ Session directory: {session_dir}")
        return dest_excel, session_dir
    except Exception as e:
        print(f"[Warn] Failed to copy Excel: {e}")
        print(f"[Info] Using reference Excel directly: {reference_excel}")
        return reference_excel, out_dir


# ANSI/Non-ANSI 모듈 헤더 정규식 및 필터 유틸 추가
_ANSI_MOD_RE = re.compile(r"(?ims)^\s*module\s+[A-Za-z_]\w*\s*(?:#\s*\(.*?\)\s*)?\(")
_NONANSI_MOD_RE = re.compile(r"(?im)^\s*module\s+[A-Za-z_]\w*\s*(?:#\s*\(.*?\)\s*)?;")

def _read_text_safe(p: Path) -> str:
    try:
        return p.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""

def _filter_rtl_ansi(files: List[Path]) -> List[Path]:
    """포트 리스트가 있는 ANSI 헤더 포함 파일만 통과. Non-ANSI만 있는 파일은 제외."""
    out: List[Path] = []
    skipped = 0
    for p in files:
        if p.suffix.lower() not in (".v", ".sv"):
            continue
        txt = _read_text_safe(p)
        if _ANSI_MOD_RE.search(txt):
            out.append(p)
            continue
        if _NONANSI_MOD_RE.search(txt):
            skipped += 1
            continue
    if skipped:
        print(f"[Info] Skipped {skipped} Non-ANSI module header file(s) during RTL scan")
    return out


def _verify_excel_write(excel_path: Path, plugin_name: str, data: Dict[str, Any]) -> Tuple[bool, List[str]]:
    """
    Verify that assertion data was correctly written to the Excel sheet.
    
    Returns True if verification passed, False otherwise.
    Prints detailed verification results.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
        
        wb = load_workbook(str(excel_path))
        
        # Find the sheet for this plugin (case-insensitive)
        sheet_name = None
        target_sheet_lower = None
        
        # Map plugin_name to expected sheet_name
        plugin_to_sheet = {
            'hact': 'HACT',
            'hsw': 'HSW',
            'hbp': 'HBP',
            'counter': 'Counter',
            'handshake': 'Handshake',
            'pulseWidth': 'pulseWidth',
        }
        target_sheet = plugin_to_sheet.get(plugin_name, plugin_name.title())
        target_sheet_lower = target_sheet.lower()
        
        for name in wb.sheetnames:
            if name.lower() == target_sheet_lower:
                sheet_name = name
                break
        
        if not sheet_name:
            print(f"  ✗ Sheet '{target_sheet}' not found in workbook")
            wb.close()
            return False
        
        ws = wb[sheet_name]
        
        # Helper function to find cell by label
        def _find_cell_by_label(ws, label: str) -> Tuple[Optional[int], Optional[int]]:
            """Find a cell by its value (label). Returns (row, col) or (None, None)."""
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if cell.value.strip() == label:
                            return cell.row, cell.column
            return None, None
        
        # Plugin-specific verification logic
        verification_passed = True
        found_items = []
        
        if plugin_name in ['hact', 'hsw']:
            # HACT and HSW have the same structure: blocks with specific fields
            blocks = data.get("blocks", [])
            if not blocks:
                print(f"  ✗ No blocks found in parsed data")
                verification_passed = False
            else:
                block = blocks[0]
                labels_to_check = [
                    "Base Clock",
                    "Base Reset",
                    "Hsync Signal",
                    "Data Enable Signal",
                    "Expected Min Value",
                    "Expected Max Value",
                ]
                
                for label in labels_to_check:
                    row, col = _find_cell_by_label(ws, label)
                    if row is None:
                        print(f"    ✗ Label '{label}' not found")
                        verification_passed = False
                        continue
                    
                    # Read value from row+1 (data row)
                    data_cell = ws.cell(row=row + 1, column=col)
                    data_value = data_cell.value
                    expected_value = block.get(label)
                    
                    if data_value == expected_value:
                        found_items.append((label, data_value, row + 1, col))
                        print(f"    ✓ {label}: '{data_value}' (row {row + 1}, col {col})")
                    else:
                        print(f"    ✗ {label}: Expected '{expected_value}', got '{data_value}' (row {row + 1}, col {col})")
                        verification_passed = False
        
        elif plugin_name == 'hbp':
            # HBP is similar to HACT/HSW but with Vsync Signal
            blocks = data.get("blocks", [])
            if not blocks:
                print(f"  ✗ No blocks found in parsed data")
                verification_passed = False
            else:
                block = blocks[0]
                labels_to_check = [
                    "Base Clock",
                    "Base Reset",
                    "Hsync Signal",
                    "Vsync Signal",
                    "Data Enable Signal",
                    "Expected Min Value",
                    "Expected Max Value",
                ]
                
                for label in labels_to_check:
                    row, col = _find_cell_by_label(ws, label)
                    if row is None:
                        print(f"    ✗ Label '{label}' not found")
                        verification_passed = False
                        continue
                    
                    # Read value from row+1 (data row)
                    data_cell = ws.cell(row=row + 1, column=col)
                    data_value = data_cell.value
                    expected_value = block.get(label)
                    
                    if data_value == expected_value:
                        found_items.append((label, data_value, row + 1, col))
                        print(f"    ✓ {label}: '{data_value}' (row {row + 1}, col {col})")
                    else:
                        print(f"    ✗ {label}: Expected '{expected_value}', got '{data_value}' (row {row + 1}, col {col})")
                        verification_passed = False
        
        elif plugin_name == 'counter':
            # Counter: columns are Target, Plus, Reset, Trigger, Expect Count
            # Row 8 is first data row
            fields = ['target', 'plus_con', 'reset_con', 'trigger_con', 'exp_cnt_val']
            col_map = {
                'target': 2,
                'plus_con': 3,
                'reset_con': 4,
                'trigger_con': 5,
                'exp_cnt_val': 6,
            }
            
            # Find next empty row starting from row 8
            next_row = 8
            while ws.cell(row=next_row, column=2).value:
                next_row += 1
            
            for field, col in col_map.items():
                cell_value = ws.cell(row=next_row, column=col).value
                expected_value = data.get(field)
                
                if cell_value == expected_value:
                    found_items.append((field, cell_value, next_row, col))
                    print(f"    ✓ {field}: '{cell_value}' (row {next_row}, col {col})")
                else:
                    print(f"    ✗ {field}: Expected '{expected_value}', got '{cell_value}' (row {next_row}, col {col})")
                    verification_passed = False
        
        elif plugin_name == 'handshake':
            # Handshake: Type, Sender, Receiver at row 7+
            next_row = 7
            while ws.cell(row=next_row, column=3).value:
                next_row += 1
            
            fields = [('phase_type', 3), ('sender', 4), ('receiver', 5)]
            for field_name, col in fields:
                cell_value = ws.cell(row=next_row, column=col).value
                expected_value = data.get(field_name)
                
                if cell_value == expected_value:
                    found_items.append((field_name, cell_value, next_row, col))
                    print(f"    ✓ {field_name}: '{cell_value}' (row {next_row}, col {col})")
                else:
                    print(f"    ✗ {field_name}: Expected '{expected_value}', got '{cell_value}' (row {next_row}, col {col})")
                    verification_passed = False
        
        elif plugin_name == 'pulseWidth':
            # PulseWidth: Signal, Target, Min, Max at row 8+
            next_row = 8
            while ws.cell(row=next_row, column=2).value:
                next_row += 1
            
            fields = [('signal', 2), ('target', 3), ('min_val', 4), ('max_val', 5)]
            for field_name, col in fields:
                cell_value = ws.cell(row=next_row, column=col).value
                expected_value = data.get(field_name)
                
                if cell_value == expected_value:
                    found_items.append((field_name, cell_value, next_row, col))
                    print(f"    ✓ {field_name}: '{cell_value}' (row {next_row}, col {col})")
                else:
                    print(f"    ✗ {field_name}: Expected '{expected_value}', got '{cell_value}' (row {next_row}, col {col})")
                    verification_passed = False
        
        wb.close()
        
        if verification_passed and found_items:
            print(f"  ✓ All {len(found_items)} values verified successfully")
        
        return verification_passed
    
    except Exception as e:
        print(f"  ✗ Verification error: {e}")
        import traceback
        traceback.print_exc()
        return False


def run_builder(
    rtl_start: Path,
    target_module: Optional[str],
    excel_path: Path,
    out_dir: Path,
    auto_define_fill: bool,
    enabled_plugins: Optional[List[str]],
    emit_json: bool,
    handshake_cfg: Optional[Dict[str, str]] = None,
    base_clk: str = "",
    base_rst: str = "",
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    ctx = build_module_context(rtl_start, target_module)
    module_info = ctx["module_info"]
    actual_module = module_info["module"]

    if handshake_cfg and handshake_cfg.get("force_type"):
        os.environ["ASSERTION_FORCE_TYPE"] = handshake_cfg["force_type"]

    session_excel, session_dir = _create_session_excel_copy(excel_path, actual_module, out_dir)

    # Write Base Clock and Base Reset to Define sheet before fill_define.py runs
    if base_clk or base_rst:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(session_excel)
            # Find "Define" sheet (case-insensitive)
            ws = None
            for sheet_name in wb.sheetnames:
                if sheet_name.lower() == "define":
                    ws = wb[sheet_name]
                    break
            if not ws:
                ws = wb.active
            
            found_clk = False
            found_rst = False
            
            # Find Base Clock and Base Reset labels and write values
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_val = cell.value.strip()
                        if cell_val == "Base Clock" and base_clk:
                            right_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            right_cell.value = base_clk
                            found_clk = True
                            print(f"✓ Base Clock set to: {base_clk} (row {cell.row}, col {cell.column})")
                        elif cell_val == "Base Reset" and base_rst:
                            right_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            right_cell.value = base_rst
                            found_rst = True
                            print(f"✓ Base Reset set to: {base_rst} (row {cell.row}, col {cell.column})")
            
            wb.save(session_excel)
            if not found_clk and base_clk:
                print(f"[Warn] 'Base Clock' label not found in Define sheet, but value provided: {base_clk}")
            if not found_rst and base_rst:
                print(f"[Warn] 'Base Reset' label not found in Define sheet, but value provided: {base_rst}")
        except Exception as e:
            print(f"[Warn] Failed to write Base Clock/Reset to Define sheet: {e}")
            import traceback
            traceback.print_exc()

    if auto_define_fill:
        define_json_path = fill_define_excel_if_needed(session_excel, module_info, session_dir)
        print(f"✓ Define JSON written: {define_json_path}")
        fill_script = _THIS_DIR / "fill_define.py"
        if fill_script.exists():
            try:
                result = subprocess.run(
                    [sys.executable, "-X", "utf8", str(fill_script), str(session_excel), str(define_json_path)],
                    check=False, capture_output=True, text=True, encoding="utf-8", errors="replace",
                    env=dict(os.environ, PYTHONIOENCODING="utf-8", PYTHONUTF8="1"),
                )
                if result.returncode == 0:
                    print("✓ Define sheet populated successfully")
                else:
                    print("[Warn] fill_define.py execution failed")
            except Exception as e:
                print(f"[Warn] fill_define.py execution failed: {e}")
        else:
            print("[Info] scripts/fill_define.py not found; please run it manually.")

    plugin_types: List[Type[BaseAssertionPlugin]] = get_registered_plugins()
    if enabled_plugins:
        enabled_names = set(enabled_plugins)
        plugin_types = [p for p in plugin_types if p.plugin_name in enabled_names]

    parsed_by_plugin: Dict[str, Dict[str, Any]] = {}
    for pcls in plugin_types:
        try:
            parsed_by_plugin[pcls.plugin_name] = pcls().parse(session_excel)
        except Exception as e:
            print(f"[Warn] Plugin {pcls.plugin_name} parse failed: {e}")

    if emit_json:
        json_blob = {
            "module": module_info["module"],
            "clocks": module_info["clocks"],
            "resets": module_info["resets"],
            "inputs": module_info["inputs"],
            "outputs": module_info["outputs"],
            "inouts": module_info["inouts"],
            "parameters": module_info["parameters"],
            "sheets": parsed_by_plugin,
        }
        json_path = session_dir / "assertion_inputs.json"
        json_path.write_text(json.dumps(json_blob, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"✓ Inputs JSON written: {json_path}")

    common_context = {
        "module_info": module_info,
        "define_excel_path": str(session_excel),
        "output_dir": str(session_dir),
        "session_dir": str(session_dir),
        "config": {
            "auto_define_fill": auto_define_fill,
            "enabled_plugins": enabled_plugins,
            "emit_json": emit_json,
        },
    }

    # ---------------- New multi-generation aggregation logic ----------------
    agg_headers: List[str] = []
    agg_inputs: Dict[str, str] = {}      # name -> width token
    agg_bodies: List[str] = []           # plain SV (no module/interface wrappers)
    agg_ports_order: List[str] = []      # preserve insertion order

    # Regex helpers
    # 헤더 제거: ; 유무 모두 허용, 공백 허용
    re_header = re.compile(r'^\s*(?:`include\s+"[^"]+"\s*;?\s*|import\s+uvm_pkg::\*\s*;?\s*)$', re.MULTILINE)
    # module/interface 래퍼 제거: 파라미터(#(...))와 개행 허용
    re_module = re.compile(r'(?is)^\s*module\b\s+[A-Za-z_]\w*\s*(?:#\s*\(.*?\)\s*)?\(\s*(?P<ports>.*?)\s*\)\s*;\s*(?P<body>.*?)\s*endmodule\b')
    re_interface = re.compile(r'(?is)^\s*interface\b\s+[A-Za-z_]\w*\s*(?:#\s*\(.*?\)\s*)?\s*;?\s*(?P<body>.*?)\s*endinterface\b')
    re_input_decl = re.compile(r'^\s*input\s+logic\s+(?:(\[[^\]]+\])\s+)?([A-Za-z_]\w*)', re.MULTILINE)

    # 표준 헤더 상수
    STD_IMPORT = "import uvm_pkg::*;"
    STD_INCLUDE = '`include "uvm_macros.svh"'
    def _is_std_header_line(line: str) -> bool:
        s = line.strip().rstrip(";")
        return s == STD_IMPORT.rstrip(";") or s == STD_INCLUDE.rstrip(";")

    def _collect_return(ret) -> Tuple[str, str]:
        sv_txt, inst_txt = "", ""
        if isinstance(ret, dict):
            sv_txt = str(ret.get("sv", "") or "")
            inst_txt = str(ret.get("sv_inst", "") or "")
        elif isinstance(ret, (list, tuple)):
            if len(ret) >= 1 and ret[0]:
                sv_txt = str(ret[0])
            if len(ret) >= 2 and ret[1]:
                inst_txt = str(ret[1])
        elif isinstance(ret, str):
            sv_txt = ret
        return sv_txt, inst_txt

    def _add_input(name: str, width: str):
        if not name:
            return
        if name not in agg_inputs:
            agg_inputs[name] = width or "[0:0]"
            agg_ports_order.append(name)

    def _extract_inputs_from_port_block(port_block: str):
        for line in port_block.splitlines():
            m = re_input_decl.search(line)
            if m:
                w = (m.group(1) or "[0:0]").strip()
                n = m.group(2)
                _add_input(n, w)

    def _strip_and_collect(sv_txt: str, plugin_name: str):
        if not sv_txt.strip():
            return
        # headers: 표준 2개는 무시, 그 외만 수집. 본문에서는 모두 제거
        for m in re_header.finditer(sv_txt):
            hdr = m.group(0).strip()
            if not _is_std_header_line(hdr) and hdr not in agg_headers:
                agg_headers.append(hdr)
        core = re_header.sub("", sv_txt)

        # interface unwrap
        m_iface = re_interface.search(core)
        if m_iface:
            core = m_iface.group("body").strip()

        # module unwrap
        m_mod = re_module.search(core)
        if m_mod:
            ports_block = m_mod.group("ports")
            _extract_inputs_from_port_block(ports_block)
            mod_body = m_mod.group("body").strip()
            core = mod_body

        # standalone input decl lines in body
        for m in re_input_decl.finditer(core):
            w = (m.group(1) or "[0:0]").strip()
            n = m.group(2)
            _add_input(n, w)
        # remove those lines
        core = "\n".join(
            line for line in core.splitlines()
            if not re_input_decl.search(line)
        ).strip()

        if core:
            agg_bodies.append(f"// {plugin_name}\n{core}")

    def _deduplicate_logic_declarations(sv_text: str) -> str:
        """
        interface 블록 내 logic 선언의 중복을 제거
        같은 신호 이름으로 여러 번 선언된 logic은 첫 번째만 유지
        """
        lines = sv_text.splitlines(keepends=True)
        result_lines = []
        seen_logics = set()
        
        # logic 패턴: logic [width] name;
        logic_pattern = re.compile(r'^\s*logic\s+(?:\[[^\]]+\])?\s+([A-Za-z_]\w*)\s*;', re.MULTILINE)
        
        for line in lines:
            m = logic_pattern.match(line)
            if m:
                sig_name = m.group(1)
                if sig_name in seen_logics:
                    # 이미 봤으면 건너뛰기 (중복 제거)
                    continue
                seen_logics.add(sig_name)
            result_lines.append(line)
        
        return "".join(result_lines)

    def _merge_interface_code(existing_sv: str, new_sv: str, plugin_name: str) -> str:
        """
        기존 interface 코드와 새 코드를 병합
        - 헤더 (`include, import) 중복 제거
        - interface 선언 통합
        - logic 선언 중복 제거
        - 본문(property, sequence 등) 추가
        """
        # 1. 헤더 추출 (중복 제거)
        header_pattern = re.compile(r'^\s*(`include\s+"[^"]+"|import\s+\w+::\*;)\s*$', re.MULTILINE)
        
        existing_headers = set(header_pattern.findall(existing_sv))
        new_headers = set(header_pattern.findall(new_sv))
        all_headers = sorted(existing_headers | new_headers)
        
        # 2. 기존 코드에서 interface 내부 추출
        existing_no_header = header_pattern.sub("", existing_sv).strip()
        new_no_header = header_pattern.sub("", new_sv).strip()
        
        # 3. interface 내부 본문 추출
        interface_pattern = re.compile(
            r'interface\s+assertion_intf\s*\(\s*\)\s*;(.*?)endinterface',
            re.DOTALL | re.IGNORECASE
        )
        
        existing_match = interface_pattern.search(existing_no_header)
        new_match = interface_pattern.search(new_no_header)
        
        if not existing_match or not new_match:
            # interface 패턴이 없으면 그냥 추가
            return existing_sv.rstrip() + f"\n\n// ===== {plugin_name} (Additional) =====\n\n" + new_sv
        
        existing_body = existing_match.group(1).strip()
        new_body = new_match.group(1).strip()
        
        # 4. logic 선언 추출 및 병합 (더 나은 정규식으로 모든 spacing 처리)
        # logic [width] name; 형식을 모두 찾음
        logic_pattern = re.compile(r'logic\s+(?:\[[^\]]+\])?\s+([A-Za-z_]\w*)\s*;', re.MULTILINE)
        
        # 기존 코드에서 logic 신호 이름 추출
        existing_logic_names = set()
        existing_logics = []  # 순서 유지
        for line in existing_body.splitlines():
            m = logic_pattern.search(line)
            if m:
                sig_name = m.group(1)
                existing_logic_names.add(sig_name)
                existing_logics.append(line.strip())
        
        # 새 코드에서 logic 신호 이름 추출 (중복 제거)
        new_logics = []
        seen_in_new = set()
        for line in new_body.splitlines():
            m = logic_pattern.search(line)
            if m:
                sig_name = m.group(1)
                # 기존 logic에 없고, 현재 새 로직에서도 아직 안 봤으면 추가
                if sig_name not in existing_logic_names and sig_name not in seen_in_new:
                    new_logics.append(line.strip())
                    seen_in_new.add(sig_name)
        
        # 5. logic 선언 제거한 본문 추출
        existing_body_no_logic = logic_pattern.sub("", existing_body).strip()
        new_body_no_logic = logic_pattern.sub("", new_body).strip()
        
        # 6. parameter 추출 및 병합
        param_pattern = re.compile(r'^\s*parameter\s+\w+\s*=\s*[^;]+;', re.MULTILINE)
        
        existing_params = set(param_pattern.findall(existing_body_no_logic))
        new_params = set(param_pattern.findall(new_body_no_logic))
        all_params = existing_params | new_params
        
        existing_body_no_param = param_pattern.sub("", existing_body_no_logic).strip()
        new_body_no_param = param_pattern.sub("", new_body_no_logic).strip()
        
        # 7. 최종 병합
        merged_headers = "\n".join(all_headers)
        
        # logic 선언 재구성
        merged_logic_lines = existing_logics + new_logics
        merged_logics = "\n".join(merged_logic_lines) if merged_logic_lines else ""
        
        # parameter 재구성
        merged_params = "\n".join(sorted(all_params)) if all_params else ""
        
        # 본문 병합
        merged_body_parts = []
        if existing_body_no_param.strip():
            merged_body_parts.append(existing_body_no_param.strip())
        if new_body_no_param.strip():
            merged_body_parts.append(f"// ===== {plugin_name} (Additional) =====\n" + new_body_no_param.strip())
        merged_body = "\n\n".join(merged_body_parts)
        
        # 최종 조립
        result_parts = [merged_headers, "", "interface assertion_intf();", ""]
        
        if merged_logics:
            result_parts.append(merged_logics)
            result_parts.append("")
        
        if merged_params:
            result_parts.append(merged_params)
            result_parts.append("")
        
        if merged_body:
            result_parts.append(merged_body)
            result_parts.append("")
        
        result_parts.append("endinterface")
        result_parts.append("")
        
        return "\n".join(result_parts)

    def _merge_inst_code(existing_inst: str, new_inst: str, plugin_name: str) -> str:
        """
        inst 파일 병합
        - 헤더 중복 제거
        - u_assertion_intf 인스턴스는 한 번만
        - assign 문 중복 제거
        """
        # 1. 헤더 추출
        header_pattern = re.compile(r'^\s*(`include\s+"[^"]+"|import\s+\w+::\*;)\s*$', re.MULTILINE)
        
        existing_headers = set(header_pattern.findall(existing_inst))
        new_headers = set(header_pattern.findall(new_inst))
        all_headers = sorted(existing_headers | new_headers)
        
        # 2. 헤더 제거
        existing_no_header = header_pattern.sub("", existing_inst).strip()
        new_no_header = header_pattern.sub("", new_inst).strip()
        
        # 3. 인스턴스 선언 추출 (한 번만)
        inst_pattern = re.compile(r'^\s*(assertion_intf\s+u_assertion_intf\s*\(\s*\)\s*;)', re.MULTILINE)
        inst_match = inst_pattern.search(existing_no_header)
        inst_decl = inst_match.group(1) if inst_match else "assertion_intf u_assertion_intf();"
        
        # 4. assign 문 추출 및 병합
        assign_pattern = re.compile(r'^\s*assign\s+u_assertion_intf\.(\w+)\s*=\s*([^;]+);', re.MULTILINE)
        
        existing_assigns = {}  # {signal_name: full_assign_statement}
        for m in assign_pattern.finditer(existing_no_header):
            sig = m.group(1)
            existing_assigns[sig] = m.group(0).strip()
        
        new_assigns = []
        for m in assign_pattern.finditer(new_no_header):
            sig = m.group(1)
            stmt = m.group(0).strip()
            if sig not in existing_assigns:
                new_assigns.append(stmt)
                existing_assigns[sig] = stmt
        
        # 5. 최종 조립
        result_parts = []
        if all_headers:
            result_parts.append("\n".join(all_headers))
            result_parts.append("")
        
        result_parts.append(inst_decl)
        result_parts.append("")
        
        all_assign_lines = [existing_assigns[sig] for sig in sorted(existing_assigns.keys())]
        if all_assign_lines:
            result_parts.append("\n".join(all_assign_lines))
        
        return "\n".join(result_parts) + "\n"

    def _run_generation(p_types: List[Type[BaseAssertionPlugin]], append_mode: bool = False):
        for pcls in p_types:
            parsed = parsed_by_plugin.get(pcls.plugin_name)
            if not parsed:
                continue
            try:
                ret = pcls().generate_sv(parsed, common_context)
                sv_txt, inst_txt = _collect_return(ret)
                
                # 모든 플러그인이 완전한 interface를 생성하므로 직접 파일 작성
                session_dir_path = Path(common_context["session_dir"])
                intf_sv_path = session_dir_path / "assertion_intf.sv"
                intf_inst_path = session_dir_path / "assertion_intf_inst.sv"
                
                if append_mode and intf_sv_path.exists():
                    # 추가 모드: 중복 제거하며 병합
                    existing_sv = intf_sv_path.read_text(encoding="utf-8")
                    existing_inst = intf_inst_path.read_text(encoding="utf-8") if intf_inst_path.exists() else ""
                    
                    # 지능형 병합
                    combined_sv = _merge_interface_code(existing_sv, sv_txt, pcls.plugin_name)
                    # 최종 중복 제거
                    combined_sv = _deduplicate_logic_declarations(combined_sv)
                    combined_inst = _merge_inst_code(existing_inst, inst_txt, pcls.plugin_name)
                    
                    intf_sv_path.write_text(combined_sv, encoding="utf-8")
                    intf_inst_path.write_text(combined_inst, encoding="utf-8")
                    print(f"✓ {pcls.plugin_name.upper()}: Merged to {intf_sv_path.name} and {intf_inst_path.name}")
                else:
                    # 첫 생성 또는 덮어쓰기 모드: 중복 제거 적용
                    sv_txt = _deduplicate_logic_declarations(sv_txt)
                    intf_sv_path.write_text(sv_txt, encoding="utf-8")
                    intf_inst_path.write_text(inst_txt, encoding="utf-8")
                    print(f"✓ {pcls.plugin_name.upper()}: Wrote {intf_sv_path.name} and {intf_inst_path.name}")
                
                # ===== Excel 검증 =====
                print(f"\n--- Verifying {pcls.plugin_name.upper()} Excel write ---")
                session_excel_path = Path(common_context["define_excel_path"])
                if _verify_excel_write(session_excel_path, pcls.plugin_name, parsed):
                    print(f"✓ {pcls.plugin_name.upper()} Excel verification PASSED\n")
                else:
                    print(f"✗ {pcls.plugin_name.upper()} Excel verification FAILED\n")
            
            except Exception as e:
                print(f"[Warn] Plugin {pcls.plugin_name} generate failed: {e}")
                import traceback
                traceback.print_exc()

    # First pass
    _run_generation(plugin_types)

    # Loop for additional generations
    while True:
        ans = input("\nDo you want to continue generating Assertion Code?\n[1] Yes\n[2] No\n> ").strip()
        if ans == "2":
            break
        if ans != "1":
            print("Invalid selection. Try again.")
            continue
        # Re-select plugins
        _import_all_plugins()
        all_types = get_registered_plugins()
        names = [p.plugin_name for p in all_types]
        picks = _pick_multi("Select plugins", names)
        next_types = [p for p in all_types if p.plugin_name in set(picks)]
        # Handshake 타입 선택 지원
        handshake_cls = next((p for p in all_types if p.plugin_name == "handshake"), None)
        chosen_has_handshake = any(p.plugin_name == "handshake" for p in next_types)
        # 비-handshake 플러그인 먼저 처리
        non_hs_types = [p for p in next_types if p.plugin_name != "handshake"]
        for pcls in non_hs_types:
            try:
                parsed_by_plugin[pcls.plugin_name] = pcls().parse(session_excel)
            except Exception as e:
                print(f"[Warn] Plugin {pcls.plugin_name} parse failed: {e}")
        _run_generation(non_hs_types, append_mode=True)  # 추가 모드로 실행
        # handshake가 선택된 경우, 타입(복수 선택 가능)별로 개별 실행
        if chosen_has_handshake and handshake_cls is not None:
            hs_options = ["2phase", "4phase", "ready_valid"]
            hs_picks = _pick_multi("Select handshake type(s) (or 'all')", hs_options)
            prev = os.environ.get("ASSERTION_FORCE_TYPE")
            try:
                for hs in hs_picks:
                    os.environ["ASSERTION_FORCE_TYPE"] = hs
                    try:
                        parsed_by_plugin["handshake"] = handshake_cls().parse(session_excel)
                    except Exception as e:
                        print(f"[Warn] Plugin handshake parse failed ({hs}): {e}")
                        continue
                    try:
                        ret = handshake_cls().generate_sv(parsed_by_plugin["handshake"], common_context)
                        sv_txt, _ = _collect_return(ret)
                        _strip_and_collect(sv_txt, "handshake")
                    except Exception as e:
                        print(f"[Warn] Plugin handshake generate failed ({hs}): {e}")
            finally:
                if prev is None:
                    os.environ.pop("ASSERTION_FORCE_TYPE", None)
                else:
                    os.environ["ASSERTION_FORCE_TYPE"] = prev

    # Build final assertion_intf.sv
    session_dir_path = Path(common_context["session_dir"])
    gen_sv_path = session_dir_path / "assertion_intf.sv"
    gen_inst_path = session_dir_path / "assertion_intf_inst.sv"

    # 모든 플러그인이 완전한 interface를 생성하므로 파일이 존재하는지만 확인
    if gen_sv_path.exists():
        print(f"✓ Generated {gen_sv_path.name} and {gen_inst_path.name}")
    else:
        print(f"[Warn] No assertion files were generated")
    
    print(f"\n===== Outputs saved to: {session_dir} =====")



def _prompt(msg: str, default: Optional[str] = None) -> str:
    hint = f" [{default}]" if default else ""
    s = input(f"{msg}{hint}: ").strip()
    return s or (default or "")


def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    print(title)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Custom Input")
    while True:
        s = input("Select > ").strip()
        if allow_custom and s == "0":
            return _prompt("Enter custom value")
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i-1][1]
        print("Invalid selection. Try again.")


def _pick_multi(title: str, options: List[str]) -> List[str]:
    print(title)
    for i, label in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    print("Enter numbers separated by space/comma (or 'all').")
    while True:
        s = input("Select > ").strip().lower()
        if s in ("all", "a"):
            return list(options)
        toks = [t for t in s.replace(",", " ").split() if t.isdigit()]
        picks: List[str] = []
        for t in toks:
            i = int(t)
            if 1 <= i <= len(options):
                picks.append(options[i-1])
        if picks:
            return picks
        print("Invalid selection. Try again.")


def interactive_wizard():
    print("==== Assertion Builder Interactive Mode ====")
    repo_root = _THIS_DIR.parent

    # RTL start candidates
    candidates: List[Tuple[str, str]] = []
    eda_rtl = repo_root / "EDA" / "RTL"
    if eda_rtl.exists():
        candidates.append((f"{eda_rtl} (EDA/RTL)", str(eda_rtl.resolve())))
    # Scan for other 'RTL' dirs (limit to 10)
    found = []
    try:
        for p in repo_root.rglob("RTL"):
            if p.is_dir() and p != eda_rtl:
                found.append(p)
                if len(found) >= 10:
                    break
    except Exception:
        pass
    for p in found:
        candidates.append((str(p), str(p.resolve())))
    if not candidates:
        candidates.append((str(repo_root), str(repo_root.resolve())))
    rtl_start_str = _pick_one("Select RTL start directory", candidates, allow_custom=True)
    rtl_start = Path(rtl_start_str).resolve()

    # Build modules DB to offer module selection
    try:
        ctx = build_module_context(rtl_start, None)
        modules = ctx["modules"]
    except SystemExit as e:
        print(str(e))
        sys.exit(1)

    mod_names = sorted(list(modules.keys()))
    # Prefer tops at front
    tops = set(find_top_modules(modules))
    mod_names_sorted = sorted(mod_names, key=lambda m: (0 if m in tops else 1, m))
    target_module = _pick_one("Select target module", [(m, m) for m in mod_names_sorted])

    # Excel selection from Data/
    data_dir = repo_root / "Data"
    excel_opts: List[Tuple[str, str]] = []
    if data_dir.exists():
        for x in sorted(data_dir.glob("*.xlsx")):
            if x.name.startswith("~$"):
                continue
            excel_opts.append((x.name, str(x.resolve())))
    if not excel_opts:
        excel_path = Path(_prompt("Enter Excel file path"))
    else:
        excel_path_str = _pick_one("Select Excel file (Data/)", excel_opts, allow_custom=True)
        excel_path = Path(excel_path_str).resolve()

    # Get properly classified ports for the selected target module
    try:
        ctx = build_module_context(rtl_start, target_module)
        module_info = ctx["module_info"]
        clocks = module_info.get("clocks", [])
        resets = module_info.get("resets", [])
        inputs = module_info.get("inputs", [])
    except Exception as e:
        print(f"[Warn] Failed to classify module ports: {e}")
        clocks = []
        resets = []
        inputs = []
    
    # Build clock options (all ports: clocks, resets, inputs, outputs)
    clk_opts: List[Tuple[str, str]] = []
    all_ports = (
        module_info.get("clocks", []) +
        module_info.get("resets", []) +
        module_info.get("inputs", []) +
        module_info.get("outputs", [])
    )
    seen_clk = set()
    for p in all_ports:
        port_name = p.get("name")
        if port_name and port_name not in seen_clk:
            clk_opts.append((port_name, port_name))
            seen_clk.add(port_name)
    base_clk = _pick_one("\nSelect Base Clock signal", clk_opts, allow_custom=True)
    
    # Build reset options (all ports: clocks, resets, inputs, outputs)
    rst_opts: List[Tuple[str, str]] = []
    seen_rst = set()
    for p in all_ports:
        port_name = p.get("name")
        if port_name and port_name not in seen_rst:
            rst_opts.append((port_name, port_name))
            seen_rst.add(port_name)
    base_rst = _pick_one("Select Base Reset signal", rst_opts, allow_custom=True)

    # Remove Output directory prompt; set default silently
    repo_root = Path(__file__).resolve().parents[1]
    out_dir = (repo_root / "out" / "assertions").resolve()

    # 모드 프롬프트 제거: 항상 1/2/3 모두 수행
    auto_define_fill = True
    emit_json = True

    # Plugin selection
    _import_all_plugins()
    plugin_types = get_registered_plugins()
    plugin_names = [p.plugin_name for p in plugin_types]
    enabled = _pick_multi("Select plugins", plugin_names)

    # handshake 플러그인이 포함되면 ready_valid까지 포함한 타입을 미리 선택(플러그인 기본값으로 전달)
    handshake_cfg = {}
    if "handshake" in enabled:
        hs_type = _pick_one("Select handshake type (2phase/4phase/ready_valid)",
                            [("2phase", "2phase"), ("4phase", "4phase"), ("ready_valid", "ready_valid")])
        handshake_cfg = {"force_type": hs_type}

    return {
        "rtl_start": rtl_start,
        "target_module": target_module,
        "excel_path": excel_path,
        "out_dir": out_dir,
        "auto_define_fill": auto_define_fill,
        "enabled_plugins": enabled,
        "emit_json": emit_json,
        "handshake_cfg": handshake_cfg,
        "base_clk": base_clk,
        "base_rst": base_rst,
    }


def _import_all_plugins() -> None:
    """
    scripts/assertions 폴더 내의 모든 플러그인 모듈을 동적 import하여
    @register 데코레이터가 실행되도록 보장합니다.
    """
    pkg_dir = _THIS_DIR / "assertions"
    pkg_name = "assertions"
    if not pkg_dir.exists():
        return
    for py in pkg_dir.glob("*.py"):
        name = py.stem
        if name in ("__init__", "base", "registry"):
            continue
        mod_name = f"{pkg_name}.{name}"
        try:
            importlib.import_module(mod_name)
        except Exception as e:
            print(f"[Warn] Failed to import plugin {name}: {e}")


def main():
    # Default: launch TUI when no args
    if len(sys.argv) == 1:
        try:
            _ensure_runtime_deps(require_tui=True)
            from cli_tui import run as run_tui  # type: ignore
        except Exception as e:
            print(f"[Info] TUI failed to start ({e}); falling back to wizard.")
            cfg = interactive_wizard()
            return run_builder(**cfg)
        return run_tui()

    parser = argparse.ArgumentParser(description="Modular Assertion Builder")
    parser.add_argument("--rtl-start", help="Start path (file or dir) to scan RTL")
    parser.add_argument("--target-module", help="Target module name (optional)")
    parser.add_argument("--excel", help="Reference Excel path containing sheets")
    parser.add_argument("--out", default="out/assertions", help="Output directory for generated files")
    parser.add_argument("--auto-define-fill", action="store_true", help="Produce Define JSON for fill_define.py and log path")
    parser.add_argument("--use-default-excel", action="store_true", help="Ignore --excel and use default Data/Assertion_TF.xlsx")
    parser.add_argument("--enable", action="append", help="Enable only listed plugins (name). Can repeat.")
    parser.add_argument("--json", action="store_true", help="Emit consolidated JSON of inputs/outputs/parameters")
    parser.add_argument("--tui", action="store_true", help="Launch full-screen TUI")
    parser.add_argument("--wizard", action="store_true", help="Launch interactive wizard instead of TUI")
    args = parser.parse_args()

    # Explicit TUI launch
    if getattr(args, "tui", False) and not getattr(args, "wizard", False):
        _ensure_runtime_deps(require_tui=True)
        from cli_tui import run as run_tui  # type: ignore
        return run_tui()

    # Explicit wizard
    if getattr(args, "wizard", False):
        cfg = interactive_wizard()
        return run_builder(**cfg)

    # Validate required when not interactive
    _ensure_runtime_deps(require_tui=False)
    if not args.rtl_start:
        raise SystemExit("--rtl-start is required when not using interactive mode")
    # Excel path selection with optional default
    if args.use_default_excel:
        excel_path = Path("Data/Assertion_TF.xlsx").resolve()
    else:
        if not args.excel:
            raise SystemExit("--excel is required when not using interactive mode (or use --use-default-excel)")
        excel_path = Path(args.excel).resolve()

    run_builder(
        rtl_start=Path(args.rtl_start).resolve(),
        target_module=args.target_module,
        excel_path=excel_path,
        out_dir=Path(args.out).resolve(),
        auto_define_fill=bool(args.auto_define_fill),
        enabled_plugins=list(args.enable) if args.enable else None,
        emit_json=bool(args.json),
        handshake_cfg=None,
    )


if __name__ == "__main__":
    main()


