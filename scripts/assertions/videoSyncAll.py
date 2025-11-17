from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json

from openpyxl import load_workbook

from .registry import register
from .base import BaseAssertionPlugin

# ===== 공통 유틸 =====
def _find_cell(ws, value: str) -> Tuple[Optional[int], Optional[int]]:
    tgt = (value or "").strip().lower()
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if str(v).strip().lower() == tgt:
                return c.row, c.column
    return None, None

def _get_sheet_ci(wb, want_name: str):
    target = (want_name or "").strip().lower()
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == target:
            return wb[nm]
    raise KeyError(f"Worksheet {want_name} does not exist.")

def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    print(title, flush=True)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Enter custom")
    while True:
        try:
            s = input("Select > ").strip()
        except EOFError:
            return options[0][1] if options else ""
        if allow_custom and s == "0":
            try:
                return input("Enter value > ").strip()
            except EOFError:
                return ""
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.", flush=True)

def _read_json(p: Path) -> Optional[Dict[str, Any]]:
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None

def _load_module_define(xls_path: Path) -> Dict[str, Any]:
    session_dir = xls_path.parent
    md = _read_json(session_dir / "module_define.json")
    if md:
        return md
    ai = _read_json(session_dir / "assertion_inputs.json")
    if ai:
        return {
            "module": ai.get("module") or "",
            "clocks": ai.get("clocks") or [],
            "resets": ai.get("resets") or [],
            "inputs": ai.get("inputs") or [],
            "outputs": ai.get("outputs") or [],
            "inouts": ai.get("inouts") or [],
            "parameters": ai.get("parameters") or [],
        }
    return {}

def _normalize_range_token(token: Any) -> str:
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    try:
        n = int(t, 10)
        return f"[{n-1}:0]" if n >= 1 else "[0:0]"
    except Exception:
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    candidates = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in candidates:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            for key in ("packed_range", "range", "packed", "decl"):
                pr = it.get(key)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            for key in ("width", "bit_width", "width_bits"):
                w = it.get(key)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            for lk, rk in (("msb", "lsb"), ("left", "right")):
                msb = it.get(lk); lsb = it.get(rk)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            return "[0:0]"
    return "[0:0]"

def _read_define_clk_rst(wb) -> Tuple[str, str]:
    try:
        ws = _get_sheet_ci(wb, "Define")
    except KeyError:
        return "", ""
    clk_r, clk_c = _find_cell(ws, "Base Clock")
    rst_r, rst_c = _find_cell(ws, "Base Reset")
    clk = ws.cell(row=clk_r, column=clk_c + 1).value if clk_r else None
    rst = ws.cell(row=rst_r, column=rst_c + 1).value if rst_r else None
    return (str(clk).strip() if clk else ""), (str(rst).strip() if rst else "")

def _pick_from(names: List[str], title: str, allow_custom: bool = False) -> str:
    opts = [(n, n) for n in names] if names else []
    return _pick_one(title, opts, allow_custom=allow_custom)

def _pick_int(title: str) -> str:
    while True:
        try:
            s = input(f"{title} (integer) > ").strip()
        except EOFError:
            return "0"
        if s.lstrip("-").isdigit():
            return s
        print("Please enter an integer.", flush=True)


@register
class VideoSyncAllPlugin(BaseAssertionPlugin):
    """
    VideoSyncAll: HSW, HBP, HACT, HFP, VSW, VBP, VACT, VFP 8개 타이밍을
    순차적으로 입력받고 각 시트에 기록한 뒤, 통합된 assertion_intf.sv 생성
    """
    plugin_name = "videosyncall"
    
    # 8개 timing type 순서 정의
    TIMING_TYPES = ["HSW", "HBP", "HACT", "HFP", "VSW", "VBP", "VACT", "VFP"]
    
    def parse(self, excel_path: Path) -> Dict[str, Any]:
        """8개 타이밍에 대해 순차적으로 사용자 입력 받기"""
        wb = load_workbook(excel_path, data_only=False)
        mod = _load_module_define(excel_path)
        base_clk, base_rst = _read_define_clk_rst(wb)
        
        if not base_clk:
            base_clk = "clk"
        if not base_rst:
            base_rst = "rst_n"
        
        # 공통 신호 수집
        in_names = [it["name"] for it in mod.get("inputs", [])]
        out_names = [it["name"] for it in mod.get("outputs", [])]
        all_ports = in_names + out_names
        
        print("\n===== VideoSyncAll: Video Timing Assertions =====")
        
        # Hsync, Vsync, Data Enable 신호 선택 (공통)
        print("\n[Common Signals Selection]")
        hsync_signal = _pick_from(all_ports, "Select Hsync signal", allow_custom=True) or "i_hsync"
        vsync_signal = _pick_from(all_ports, "Select Vsync signal", allow_custom=True) or "i_vsync"
        de_signal = _pick_from(all_ports, "Select Data Enable signal", allow_custom=True) or "i_de"
        
        # Width 정보
        w_clk = _port_width_token(mod, base_clk) or "[0:0]"
        w_rst = _port_width_token(mod, base_rst) or "[0:0]"
        w_hs = _port_width_token(mod, hsync_signal) or "[0:0]"
        w_vs = _port_width_token(mod, vsync_signal) or "[0:0]"
        w_de = _port_width_token(mod, de_signal) or "[0:0]"
        
        # 8개 타이밍에 대해 min/max 입력 받기
        timing_data = {}
        for timing_type in self.TIMING_TYPES:
            print(f"\n--- {timing_type} Timing ---")
            exp_min = _pick_int(f"  {timing_type} Expected Min Value")
            exp_max = _pick_int(f"  {timing_type} Expected Max Value")
            timing_data[timing_type.lower()] = {
                "min": exp_min,
                "max": exp_max
            }
        
        # 각 시트에 데이터 쓰기
        for timing_type in self.TIMING_TYPES:
            try:
                ws = _get_sheet_ci(wb, timing_type)
                self._write_to_sheet(ws, base_clk, base_rst, hsync_signal, vsync_signal, de_signal,
                                   timing_data[timing_type.lower()]["min"],
                                   timing_data[timing_type.lower()]["max"])
                print(f"✓ Written to {timing_type} sheet")
            except KeyError:
                print(f"[Warn] Sheet '{timing_type}' not found, skipping write")
        
        wb.save(excel_path)
        wb.close()
        
        return {
            "blocks": [{
                "Base Clock": base_clk,
                "Base Reset": base_rst,
                "Base Clock Width": w_clk,
                "Base Reset Width": w_rst,
                "Hsync Signal": hsync_signal,
                "Vsync Signal": vsync_signal,
                "Data Enable Signal": de_signal,
                "Hsync Signal Width": w_hs,
                "Vsync Signal Width": w_vs,
                "Data Enable Signal Width": w_de,
                "timing_data": timing_data
            }]
        }
    
    def _write_to_sheet(self, ws, base_clk: str, base_rst: str, hsync: str, vsync: str, de: str, exp_min: str, exp_max: str):
        """개별 시트에 데이터 기록"""
        # Base Clock
        clk_row, clk_col = _find_cell(ws, "Base Clock")
        if clk_row:
            ws.cell(row=clk_row, column=clk_col + 1, value=base_clk)
        
        # Base Reset
        rst_row, rst_col = _find_cell(ws, "Base Reset")
        if rst_row:
            ws.cell(row=rst_row, column=rst_col + 1, value=base_rst)
        
        # Hsync Signal
        hs_row, hs_col = _find_cell(ws, "Hsync Signal")
        if hs_row:
            ws.cell(row=hs_row + 1, column=hs_col, value=hsync)
        
        # Vsync Signal
        vs_row, vs_col = _find_cell(ws, "Vsync Signal")
        if vs_row:
            ws.cell(row=vs_row + 1, column=vs_col, value=vsync)
        
        # Data Enable Signal
        de_row, de_col = _find_cell(ws, "Data Enable Signal")
        if de_row:
            ws.cell(row=de_row + 1, column=de_col, value=de)
        
        # Expected Min Value
        min_row, min_col = _find_cell(ws, "Expected Min Value")
        if min_row:
            ws.cell(row=min_row + 1, column=min_col, value=exp_min)
        
        # Expected Max Value
        max_row, max_col = _find_cell(ws, "Expected Max Value")
        if max_row:
            ws.cell(row=max_row + 1, column=max_col, value=exp_max)
    
    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
        """8개 타이밍 assertion을 통합하여 하나의 interface 생성"""
        blocks = parsed.get("blocks") or []
        if not blocks:
            return ["// No VideoSyncAll assertions generated.\n", ""]
        
        b = blocks[0]
        base_clk = b.get("Base Clock", "") or "clk"
        base_rst = b.get("Base Reset", "") or "rst_n"
        hsync_signal = b.get("Hsync Signal", "") or "i_hsync"
        vsync_signal = b.get("Vsync Signal", "") or "i_vsync"
        data_enable_signal = b.get("Data Enable Signal", "") or "i_de"
        w_clk = b.get("Base Clock Width", "[0:0]")
        w_rst = b.get("Base Reset Width", "[0:0]")
        w_hs = b.get("Hsync Signal Width", "[0:0]")
        w_vs = b.get("Vsync Signal Width", "[0:0]")
        w_de = b.get("Data Enable Signal Width", "[0:0]")
        timing_data = b.get("timing_data") or {}

        lines: List[str] = []
        
        # 1. Include/Import (최상단)
        lines.append("`include \"uvm_macros.svh\"")
        lines.append("import uvm_pkg::*;")
        lines.append("")
        
        # 2. Interface 선언
        lines.append("interface assertion_intf();")
        lines.append("")
        
        # 3. Logic 선언 (공통 신호)
        lines.append(f"logic {w_clk} {base_clk};")
        lines.append(f"logic {w_rst} {base_rst};")
        lines.append(f"logic {w_hs} {hsync_signal};")
        lines.append(f"logic {w_vs} {vsync_signal};")
        lines.append(f"logic {w_de} {data_enable_signal};")
        lines.append("")
        
        # 4. Int/Logic 변수 선언
        self._add_variables(lines, timing_data)
        lines.append("")
        
        # 5. Always 블록 (module 역할)
        self._add_always_blocks(lines, timing_data, base_clk, base_rst, hsync_signal, vsync_signal, data_enable_signal)
        
        # 6. Sequence 선언
        self._add_sequences(lines, timing_data, base_clk, base_rst, data_enable_signal)
        
        # 7. Property 선언
        self._add_properties(lines, timing_data, base_clk, base_rst, hsync_signal, vsync_signal, data_enable_signal)
        
        # 8. Assert 선언
        self._add_asserts(lines, timing_data)
        
        # 9. Endinterface (최하단)
        lines.append("endinterface")
        lines.append("")
        
        sv_text = "\n".join(lines)
        
        # Instantiation 파일
        inst_lines: List[str] = []
        inst_lines.append("`include \"uvm_macros.svh\"")
        inst_lines.append("import uvm_pkg::*;")
        inst_lines.append("")
        inst_lines.append("assertion_intf")
        inst_lines.append("u_assertion_intf();")
        inst_lines.append("")
        inst_lines.append(f"assign u_assertion_intf.{base_clk} = top.dut.{base_clk};")
        inst_lines.append(f"assign u_assertion_intf.{base_rst} = top.dut.{base_rst};")
        inst_lines.append(f"assign u_assertion_intf.{hsync_signal} = top.dut.{hsync_signal};")
        inst_lines.append(f"assign u_assertion_intf.{vsync_signal} = top.dut.{vsync_signal};")
        inst_lines.append(f"assign u_assertion_intf.{data_enable_signal} = top.dut.{data_enable_signal};")
        inst_text = "\n".join(inst_lines) + "\n"
        
        return [sv_text, inst_text]
    
    def _add_variables(self, lines: List[str], timing_data: Dict[str, Any]):
        """변수 선언 추가 (int형 먼저, logic형 나중에)"""
        # int형 변수들 먼저 선언
        if "hsw" in timing_data:
            lines.append("int hsw_value_count = 0;")
        if "hbp" in timing_data:
            lines.append("int hbp_value_count = 0;")
        if "hact" in timing_data:
            lines.append("int hact_de_cnt = 0;")
        if "hfp" in timing_data:
            lines.append("int hfp_hsync_cnt = 0;")
        if "vsw" in timing_data:
            lines.append("int vsw_value_count = 0;")
        if "vbp" in timing_data:
            lines.append("int vbp_hsync_cnt = 0;")
        if "vact" in timing_data:
            lines.append("int vact_de_cnt = 0;")
        if "vfp" in timing_data:
            lines.append("int vfp_hsync_cnt = 0;")
        
        # logic형 변수들 나중에 선언
        if "hfp" in timing_data:
            lines.append("logic l_hfp = 0;")
        if "vbp" in timing_data:
            lines.append("logic l_vbp = 0;")
        if "vact" in timing_data:
            lines.append("logic l_vbp_vact_vfp = 0;")
        if "vfp" in timing_data:
            lines.append("logic l_vact_vfp = 0;")
    
    def _add_always_blocks(self, lines: List[str], timing_data: Dict[str, Any], 
                          clk: str, rst: str, hsync: str, vsync: str, de: str):
        """Always 블록 추가"""
        # HACT
        if "hact" in timing_data:
            lines.append(f"always @(posedge {clk} or negedge {rst}) begin")
            lines.append(f"    if (!{rst}) begin")
            lines.append("        hact_de_cnt <= 0;")
            lines.append("    end else begin")
            lines.append(f"        if ({de}) begin")
            lines.append("            hact_de_cnt <= hact_de_cnt + 1;")
            lines.append("        end else begin")
            lines.append("            hact_de_cnt <= 0;")
            lines.append("        end")
            lines.append("    end")
            lines.append("end")
            lines.append("")
        
        # HFP
        if "hfp" in timing_data:
            lines.append(f"always @(posedge {clk} or negedge {rst}) begin")
            lines.append(f"    if (!{rst}) begin")
            lines.append("        l_hfp <= 0;")
            lines.append("        hfp_hsync_cnt <= 0;")
            lines.append("    end else begin")
            lines.append(f"        if ({de} && !l_hfp) begin")
            lines.append("            l_hfp <= 1;")
            lines.append("            hfp_hsync_cnt <= 0;")
            lines.append(f"        end else if ({hsync} && l_hfp) begin")
            lines.append("            hfp_hsync_cnt <= hfp_hsync_cnt + 1;")
            lines.append("        end else begin")
            lines.append("            l_hfp <= 0;")
            lines.append("        end")
            lines.append("    end")
            lines.append("end")
            lines.append("")
        
        # VBP
        if "vbp" in timing_data:
            lines.append(f"always @(posedge {clk} or negedge {rst}) begin")
            lines.append(f"    if (!{rst}) begin")
            lines.append("        l_vbp <= 0;")
            lines.append("        vbp_hsync_cnt <= 0;")
            lines.append("    end else begin")
            lines.append(f"        if (!{vsync} && !l_vbp) begin")
            lines.append("            l_vbp <= 1;")
            lines.append("            vbp_hsync_cnt <= 0;")
            lines.append(f"        end else if ({hsync} && l_vbp) begin")
            lines.append("            vbp_hsync_cnt <= vbp_hsync_cnt + 1;")
            lines.append("        end else begin")
            lines.append("            l_vbp <= 0;")
            lines.append("        end")
            lines.append("    end")
            lines.append("end")
            lines.append("")
        
        # VACT
        if "vact" in timing_data:
            lines.append(f"always @(posedge {clk} or negedge {rst}) begin")
            lines.append(f"    if (!{rst}) begin")
            lines.append("        l_vbp_vact_vfp <= 0;")
            lines.append("        vact_de_cnt <= 0;")
            lines.append("    end else begin")
            lines.append(f"        if (!{vsync} && !l_vbp_vact_vfp) begin")
            lines.append("            l_vbp_vact_vfp <= 1;")
            lines.append("            vact_de_cnt <= 0;")
            lines.append(f"        end else if ({de} && l_vbp_vact_vfp) begin")
            lines.append("            vact_de_cnt <= vact_de_cnt + 1;")
            lines.append("        end else begin")
            lines.append("            l_vbp_vact_vfp <= 0;")
            lines.append("        end")
            lines.append("    end")
            lines.append("end")
            lines.append("")
        
        # VFP
        if "vfp" in timing_data:
            lines.append(f"always @(posedge {clk} or negedge {rst}) begin")
            lines.append(f"    if (!{rst}) begin")
            lines.append("        l_vact_vfp <= 0;")
            lines.append("        vfp_hsync_cnt <= 0;")
            lines.append("    end else begin")
            lines.append(f"        if (!{de} && !l_vact_vfp) begin")
            lines.append("            l_vact_vfp <= 1;")
            lines.append("            vfp_hsync_cnt <= 0;")
            lines.append(f"        end else if ({hsync} && l_vact_vfp) begin")
            lines.append("            vfp_hsync_cnt <= vfp_hsync_cnt + 1;")
            lines.append("        end else begin")
            lines.append("            l_vact_vfp <= 0;")
            lines.append("        end")
            lines.append("    end")
            lines.append("end")
            lines.append("")
    
    def _add_sequences(self, lines: List[str], timing_data: Dict[str, Any], clk: str, rst: str, de: str):
        """Sequence 선언 추가"""
        if "hact" in timing_data:
            lines.append("sequence s_hact;")
            lines.append(f"    @(posedge {clk}) disable iff (!{rst})")
            lines.append(f"        $rose({de}) ##1 ({de})[*] ##1 $fell({de});")
            lines.append("endsequence")
            lines.append("")
    
    def _add_properties(self, lines: List[str], timing_data: Dict[str, Any],
                       clk: str, rst: str, hsync: str, vsync: str, de: str):
        """Property 선언 추가"""
        # HSW
        if "hsw" in timing_data:
            hsw_min = timing_data["hsw"]["min"]
            hsw_max = timing_data["hsw"]["max"]
            lines.append("property p_hsw;")
            lines.append(f"    @(posedge {clk}) disable iff (!{rst})")
            lines.append(f"        $rose({hsync}) |-> (1, hsw_value_count = 0) ##1")
            lines.append(f"            ({hsync}, hsw_value_count++)[*0:$] ##1")
            lines.append(f"            (!{hsync}, hsw_value_count++) ##0")
            lines.append(f"            ({hsw_min} <= hsw_value_count && hsw_value_count <= {hsw_max});")
            lines.append("endproperty")
            lines.append("")
        
        # HBP
        if "hbp" in timing_data:
            hbp_min = timing_data["hbp"]["min"]
            hbp_max = timing_data["hbp"]["max"]
            lines.append("property p_hbp;")
            lines.append(f"    @(posedge {clk}) disable iff (!{rst})")
            lines.append(f"        $fell({hsync}) |-> (1, hbp_value_count = 0) ##1")
            lines.append(f"            (!{de}, hbp_value_count++)[*0:$] ##1")
            lines.append(f"            ({de}, hbp_value_count++) ##0")
            lines.append(f"            ({hbp_min} <= hbp_value_count && hbp_value_count <= {hbp_max});")
            lines.append("endproperty")
            lines.append("")
        
        # HACT
        if "hact" in timing_data:
            hact_min = timing_data["hact"]["min"]
            hact_max = timing_data["hact"]["max"]
            lines.append("property p_hact;")
            lines.append(f"    @(posedge {clk}) disable iff (!{rst})")
            lines.append(f"        s_hact |-> ({hact_min} <= hact_de_cnt && hact_de_cnt <= {hact_max});")
            lines.append("endproperty")
            lines.append("")
        
        # HFP
        if "hfp" in timing_data:
            hfp_min = timing_data["hfp"]["min"]
            hfp_max = timing_data["hfp"]["max"]
            lines.append("property p_hfp;")
            lines.append(f"    @(posedge {clk}) disable iff (!{rst})")
            lines.append(f"        $fell({de}) |-> ##1 ({hfp_min} <= hfp_hsync_cnt && hfp_hsync_cnt <= {hfp_max});")
            lines.append("endproperty")
            lines.append("")
        
        # VSW
        if "vsw" in timing_data:
            vsw_min = timing_data["vsw"]["min"]
            vsw_max = timing_data["vsw"]["max"]
            lines.append("property p_vsw;")
            lines.append(f"    @(posedge {clk}) disable iff (!{rst})")
            lines.append(f"        $rose({vsync}) |-> (1, vsw_value_count = 0) ##1")
            lines.append(f"            ({vsync}, vsw_value_count++)[*0:$] ##1")
            lines.append(f"            (!{vsync}, vsw_value_count++) ##0")
            lines.append(f"            ({vsw_min} <= vsw_value_count && vsw_value_count <= {vsw_max});")
            lines.append("endproperty")
            lines.append("")
        
        # VBP
        if "vbp" in timing_data:
            vbp_min = timing_data["vbp"]["min"]
            vbp_max = timing_data["vbp"]["max"]
            lines.append("property p_vbp;")
            lines.append(f"    @(posedge {clk}) disable iff (!{rst})")
            lines.append(f"        $fell({vsync}) |-> ##1 ({vbp_min} <= vbp_hsync_cnt && vbp_hsync_cnt <= {vbp_max});")
            lines.append("endproperty")
            lines.append("")
        
        # VACT
        if "vact" in timing_data:
            vact_min = timing_data["vact"]["min"]
            vact_max = timing_data["vact"]["max"]
            lines.append("property p_vact;")
            lines.append(f"    @(posedge {clk}) disable iff (!{rst})")
            lines.append(f"        $fell({vsync}) |-> ##1 ({vact_min} <= vact_de_cnt && vact_de_cnt <= {vact_max});")
            lines.append("endproperty")
            lines.append("")
        
        # VFP
        if "vfp" in timing_data:
            vfp_min = timing_data["vfp"]["min"]
            vfp_max = timing_data["vfp"]["max"]
            lines.append("property p_vfp;")
            lines.append(f"    @(posedge {clk}) disable iff (!{rst})")
            lines.append(f"        $fell({de}) |-> ##1 ({vfp_min} <= vfp_hsync_cnt && vfp_hsync_cnt <= {vfp_max});")
            lines.append("endproperty")
            lines.append("")
    
    def _add_asserts(self, lines: List[str], timing_data: Dict[str, Any]):
        """Assert 선언 추가"""
        if "hsw" in timing_data:
            lines.append("assert property (p_hsw) $display(\"[%0t] HSW Assertion PASS\", $time);")
            lines.append("else `uvm_error(\"ASSERTION\", \"[%0t] HSW Assertion FAIL\")")
            lines.append("")
        
        if "hbp" in timing_data:
            lines.append("assert property (p_hbp) $display(\"[%0t] HBP Assertion PASS\", $time);")
            lines.append("else `uvm_error(\"ASSERTION\", \"[%0t] HBP Assertion FAIL\")")
            lines.append("")
        
        if "hact" in timing_data:
            lines.append("assert property (p_hact) $display(\"[%0t] HACT Assertion PASS\", $time);")
            lines.append("else `uvm_error(\"ASSERTION\", \"[%0t] HACT Assertion FAIL\")")
            lines.append("")
        
        if "hfp" in timing_data:
            lines.append("assert property (p_hfp) $display(\"[%0t] HFP Assertion PASS\", $time);")
            lines.append("else `uvm_error(\"ASSERTION\", \"[%0t] HFP Assertion FAIL\")")
            lines.append("")
        
        if "vsw" in timing_data:
            lines.append("assert property (p_vsw) $display(\"[%0t] VSW Assertion PASS\", $time);")
            lines.append("else `uvm_error(\"ASSERTION\", \"[%0t] VSW Assertion FAIL\")")
            lines.append("")
        
        if "vbp" in timing_data:
            lines.append("assert property (p_vbp) $display(\"[%0t] VBP Assertion PASS\", $time);")
            lines.append("else `uvm_error(\"ASSERTION\", \"[%0t] VBP Assertion FAIL\")")
            lines.append("")
        
        if "vact" in timing_data:
            lines.append("assert property (p_vact) $display(\"[%0t] VACT Assertion PASS\", $time);")
            lines.append("else `uvm_error(\"ASSERTION\", \"[%0t] VACT Assertion FAIL\")")
            lines.append("")
        
        if "vfp" in timing_data:
            lines.append("assert property (p_vfp) $display(\"[%0t] VFP Assertion PASS\", $time);")
            lines.append("else `uvm_error(\"ASSERTION\", \"[%0t] VFP Assertion FAIL\")")
            lines.append("")
