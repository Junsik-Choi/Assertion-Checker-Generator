from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json

from openpyxl import load_workbook

from .registry import register
from .base import BaseAssertionPlugin

# ===== 공통 유틸리티 함수 =====
def _find_cell(ws, value: str) -> Tuple[Optional[int], Optional[int]]:
    """시트에서 특정 값을 가진 셀의 위치(row, column)를 찾기"""
    tgt = (value or "").strip().lower()
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if str(v).strip().lower() == tgt:
                return c.row, c.column
    return None, None

def _get_sheet_ci(wb, want_name: str):
    """대소문자 구분 없이 워크시트 찾기"""
    target = (want_name or "").strip().lower()
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == target:
            return wb[nm]
    raise KeyError(f"Worksheet '{want_name}' does not exist.")

def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    """프롬프트로 옵션 선택 또는 커스텀 입력"""
    print(title, flush=True)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Enter custom")
    while True:
        try:
            s = input("Select > ").strip()
        except EOFError:
            return options[0][1] if options else ""
        if allow_custom and s == "0":
            try:
                return input("Enter value > ").strip()
            except EOFError:
                return ""
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.", flush=True)

def _read_json(p: Path) -> Optional[Dict[str, Any]]:
    """JSON 파일 읽기"""
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None

def _load_module_define(xls_path: Path) -> Dict[str, Any]:
    """module_define.json 또는 assertion_inputs.json에서 RTL 정보 로드"""
    session_dir = xls_path.parent
    md = _read_json(session_dir / "module_define.json")
    if md:
        return md
    ai = _read_json(session_dir / "assertion_inputs.json")
    if ai:
        return {
            "module": ai.get("module") or "",
            "clocks": ai.get("clocks") or [],
            "resets": ai.get("resets") or [],
            "inputs": ai.get("inputs") or [],
            "outputs": ai.get("outputs") or [],
            "inouts": ai.get("inouts") or [],
            "parameters": ai.get("parameters") or [],
        }
    return {}

def _normalize_range_token(token: Any) -> str:
    """포트 width를 [msb:lsb] 형식으로 정규화"""
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    try:
        n = int(t, 10)
        return f"[{n-1}:0]" if n >= 1 else "[0:0]"
    except Exception:
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    """포트 이름으로 width 토큰 찾기"""
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    candidates = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in candidates:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            # packed_range, range 등에서 width 정보 찾기
            for key in ("packed_range", "range", "packed", "decl"):
                pr = it.get(key)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            for key in ("width", "bit_width", "width_bits"):
                w = it.get(key)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            for lk, rk in (("msb", "lsb"), ("left", "right")):
                msb = it.get(lk); lsb = it.get(rk)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            return "[0:0]"
    return "[0:0]"

def _pick_from(names: List[str], title: str, allow_custom: bool = False) -> str:
    """리스트에서 선택 또는 커스텀 입력"""
    opts = [(n, n) for n in names] if names else []
    return _pick_one(title, opts, allow_custom=allow_custom)

# ===== QCH 플러그인 =====
@register
class QCHPlugin(BaseAssertionPlugin):
    plugin_name = "QCH"
    sheet_name = "QCH"

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        """
        Excel 파일 파싱 및 사용자 입력 처리
        1. QCH 시트 확인
        2. Base Clock/Reset 입력받기
        3. QREQn, QACCEPTn, QACTIVE, QDENY 신호 입력받기
        """
        mod = _load_module_define(Path(xls_path))
        wb = load_workbook(xls_path)

        # 1. QCH 시트 확인
        try:
            ws = _get_sheet_ci(wb, self.sheet_name)
        except KeyError:
            print(f"ERROR: '{self.sheet_name}' sheet does not exist in the Excel file.", flush=True)
            raise

        # 2. 모든 포트 수집 (입력/출력/inout)
        all_ports: List[str] = []
        for it in (mod.get("inputs") or []):
            n = it.get("name")
            if n:
                all_ports.append(n)
        for it in (mod.get("outputs") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)
        for it in (mod.get("inouts") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)

        # 3. Base Clock 확인 및 입력
        clk_r, clk_c = _find_cell(ws, "Base Clock")
        if not clk_r:
            print("ERROR: 'Base Clock' cell not found in QCH sheet.", flush=True)
            raise ValueError("'Base Clock' cell not found")

        # 4. Base Reset 확인 및 입력
        rst_r, rst_c = _find_cell(ws, "Base Reset")
        if not rst_r:
            print("ERROR: 'Base Reset' cell not found in QCH sheet.", flush=True)
            raise ValueError("'Base Reset' cell not found")

        # 5. QREQn, QACCEPTn, QACTIVE, QDENY 셀 위치 확인
        qreqn_r, qreqn_c = _find_cell(ws, "QREQn")
        if not qreqn_r:
            print("ERROR: 'QREQn' cell not found in QCH sheet.", flush=True)
            raise ValueError("'QREQn' cell not found")

        qacceptn_r, qacceptn_c = _find_cell(ws, "QACCEPTn")
        if not qacceptn_r:
            print("ERROR: 'QACCEPTn' cell not found in QCH sheet.", flush=True)
            raise ValueError("'QACCEPTn' cell not found")

        qactive_r, qactive_c = _find_cell(ws, "QACTIVE")
        if not qactive_r:
            print("ERROR: 'QACTIVE' cell not found in QCH sheet.", flush=True)
            raise ValueError("'QACTIVE' cell not found")

        qdeny_r, qdeny_c = _find_cell(ws, "QDENY")
        if not qdeny_r:
            print("ERROR: 'QDENY' cell not found in QCH sheet.", flush=True)
            raise ValueError("'QDENY' cell not found")

        # 6. 기존 셀 값 읽기
        # Base Clock/Reset은 Define 시트에서 직접 읽기 (수식 참조 문제 방지)
        wb_data = load_workbook(xls_path, data_only=True)
        try:
            ws_define = _get_sheet_ci(wb_data, "Define")
            define_clk_r, define_clk_c = _find_cell(ws_define, "Base Clock")
            define_rst_r, define_rst_c = _find_cell(ws_define, "Base Reset")
            clk_cell = ws_define.cell(row=define_clk_r, column=define_clk_c + 1).value if define_clk_r else None
            rst_cell = ws_define.cell(row=define_rst_r, column=define_rst_c + 1).value if define_rst_r else None
        except Exception:
            clk_cell = None
            rst_cell = None
        
        qreqn_cell = ws.cell(row=qreqn_r + 1, column=qreqn_c).value
        qacceptn_cell = ws.cell(row=qacceptn_r + 1, column=qacceptn_c).value
        qactive_cell = ws.cell(row=qactive_r + 1, column=qactive_c).value
        qdeny_cell = ws.cell(row=qdeny_r + 1, column=qdeny_c).value

        base_clk = str(clk_cell).strip() if clk_cell and str(clk_cell).strip() else "<Base Clock>"
        base_rst = str(rst_cell).strip() if rst_cell and str(rst_cell).strip() else "<Base Reset>"
        qreqn = str(qreqn_cell).strip() if qreqn_cell and str(qreqn_cell).strip() else "<QREQn>"
        qacceptn = str(qacceptn_cell).strip() if qacceptn_cell and str(qacceptn_cell).strip() else "<QACCEPTn>"
        qactive = str(qactive_cell).strip() if qactive_cell and str(qactive_cell).strip() else "<QACTIVE>"
        qdeny = str(qdeny_cell).strip() if qdeny_cell and str(qdeny_cell).strip() else "<QDENY>"

        # 7. 인터랙티브 입력
        while True:
            print("\n==================== QCH Settings ====================")
            print(f"[1] Base Clock  : {base_clk}")
            print(f"[2] Base Reset  : {base_rst}")
            print(f"[3] QREQn       : {qreqn}")
            print(f"[4] QACCEPTn    : {qacceptn}")
            print(f"[5] QACTIVE     : {qactive}")
            print(f"[6] QDENY       : {qdeny}")
            print("======================================================")
            print("Select item number to edit")
            print("Press Enter twice to confirm all")
            choice = input("> ").strip()
            if choice == "":
                missing = []
                if base_clk in ("", "<Base Clock>"):   missing.append("[1]")
                if base_rst in ("", "<Base Reset>"):   missing.append("[2]")
                if qreqn in ("", "<QREQn>"):           missing.append("[3]")
                if qacceptn in ("", "<QACCEPTn>"):     missing.append("[4]")
                if qactive in ("", "<QACTIVE>"):       missing.append("[5]")
                if qdeny in ("", "<QDENY>"):           missing.append("[6]")
                if missing:
                    print(f"{','.join(missing)} has NOT been entered yet.")
                print("Press Enter again to confirm, or select item number to edit")
                choice = input("> ").strip()
                if choice == "":
                    break
            if choice == "1":
                print("\n=== Base Clock ===")
                base_clk = _pick_from(all_ports, "Select Base Clock:", allow_custom=True)
            elif choice == "2":
                print("\n=== Base Reset ===")
                base_rst = _pick_from(all_ports, "Select Base Reset:", allow_custom=True)
            elif choice == "3":
                print("\n=== QREQn ===")
                qreqn = _pick_from(all_ports, "Select QREQn signal:", allow_custom=True)
            elif choice == "4":
                print("\n=== QACCEPTn ===")
                qacceptn = _pick_from(all_ports, "Select QACCEPTn signal:", allow_custom=True)
            elif choice == "5":
                print("\n=== QACTIVE ===")
                qactive = _pick_from(all_ports, "Select QACTIVE signal:", allow_custom=True)
            elif choice == "6":
                print("\n=== QDENY ===")
                qdeny = _pick_from(all_ports, "Select QDENY signal:", allow_custom=True)
            else:
                print("Invalid selection. Try again.")
                continue

        # 8. 최종값을 시트에 기록
        ws.cell(row=clk_r, column=clk_c + 1, value=base_clk)
        ws.cell(row=rst_r, column=rst_c + 1, value=base_rst)
        ws.cell(row=qreqn_r + 1, column=qreqn_c, value=qreqn)
        ws.cell(row=qacceptn_r + 1, column=qacceptn_c, value=qacceptn)
        ws.cell(row=qactive_r + 1, column=qactive_c, value=qactive)
        ws.cell(row=qdeny_r + 1, column=qdeny_c, value=qdeny)

        # 9. Excel 저장
        wb.save(xls_path)

        # 10. Width 정보 수집
        clk_width = _port_width_token(mod, base_clk)
        rst_width = _port_width_token(mod, base_rst)
        qreqn_width = _port_width_token(mod, qreqn)
        qacceptn_width = _port_width_token(mod, qacceptn)
        qactive_width = _port_width_token(mod, qactive)
        qdeny_width = _port_width_token(mod, qdeny)

        # 11. 결과 반환
        blocks = [{
            "Base Clock": base_clk,
            "Base Reset": base_rst,
            "QREQn": qreqn,
            "QACCEPTn": qacceptn,
            "QACTIVE": qactive,
            "QDENY": qdeny,
            "Base Clock Width": clk_width,
            "Base Reset Width": rst_width,
            "QREQn Width": qreqn_width,
            "QACCEPTn Width": qacceptn_width,
            "QACTIVE Width": qactive_width,
            "QDENY Width": qdeny_width,
        }]
        return {"blocks": blocks}

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
        """
        SystemVerilog assertion 코드 생성
        - assertion_intf.sv: interface 정의
        - assertion_intf_inst.sv: 인스턴스 및 연결
        """
        blocks = parsed.get("blocks") or []
        if not blocks:
            return ["// No QCH assertions generated.\n", ""]

        b = blocks[0]
        base_clk = b.get("Base Clock", "") or "clk"
        base_rst = b.get("Base Reset", "") or "rst_n"
        qreqn = b.get("QREQn", "") or "QREQn"
        qacceptn = b.get("QACCEPTn", "") or "QACCEPTn"
        qactive = b.get("QACTIVE", "") or "QACTIVE"
        qdeny = b.get("QDENY", "") or "QDENY"

        clk_w = b.get("Base Clock Width", "[0:0]")
        rst_w = b.get("Base Reset Width", "[0:0]")
        qreqn_w = b.get("QREQn Width", "[0:0]")
        qacceptn_w = b.get("QACCEPTn Width", "[0:0]")
        qactive_w = b.get("QACTIVE Width", "[0:0]")
        qdeny_w = b.get("QDENY Width", "[0:0]")

        # ===== assertion_intf.sv 생성 =====
        lines: List[str] = []
        lines.append("`include \"uvm_macros.svh\"")
        lines.append("import uvm_pkg::*;")
        lines.append("")
        lines.append("interface assertion_intf();")
        lines.append("")
        lines.append(f"logic {clk_w} {base_clk};")
        lines.append(f"logic {rst_w} {base_rst};")
        lines.append(f"logic {qreqn_w} {qreqn};")
        lines.append(f"logic {qacceptn_w} {qacceptn};")
        lines.append(f"logic {qactive_w} {qactive};")
        lines.append(f"logic {qdeny_w} {qdeny};")
        lines.append("")
        lines.append(f"always @(posedge {qreqn}) begin")
        lines.append(f"    if ({base_rst}) begin")
        lines.append(f"        assert (({qacceptn} == 1'b0) || ({qdeny} == 1'b1)) else $error(\"Full handshake violation on rising edge of QREQn\");")
        lines.append("    end")
        lines.append("end")
        lines.append("")
        lines.append(f"always @(negedge {qreqn}) begin")
        lines.append(f"    if ({base_rst}) begin")
        lines.append(f"        assert ({qacceptn} == 1'b1) else $error(\"Full handshake violation on falling edge of QREQn\");")
        lines.append("    end")
        lines.append("end")
        lines.append("")
        lines.append(f"always @(posedge {qacceptn}) begin")
        lines.append(f"    if ({base_rst}) begin")
        lines.append(f"        assert ({qreqn} == 1'b1) else $error(\"Full handshake violation on rising edge of QACCEPTn\");")
        lines.append("    end")
        lines.append("end")
        lines.append("")
        lines.append(f"always @(negedge {qacceptn}) begin")
        lines.append(f"    if ({base_rst}) begin")
        lines.append(f"        assert ({qreqn} == 1'b0) else $error(\"Full handshake violation on falling edge of QACCEPTn\");")
        lines.append("    end")
        lines.append("end")
        lines.append("")
        lines.append(f"always @(posedge {qdeny}) begin")
        lines.append(f"    if ({base_rst}) begin")
        lines.append(f"        assert ({qreqn} == 1'b0) else $error(\"Full handshake violation on rising edge of QDENY\");")
        lines.append("    end")
        lines.append("end")
        lines.append("")
        lines.append(f"always @(negedge {qdeny}) begin")
        lines.append(f"    if ({base_rst}) begin")
        lines.append(f"        assert ({qreqn} == 1'b1) else $error(\"Full handshake violation on falling edge of QDENY\");")
        lines.append("    end")
        lines.append("end")
        lines.append("")
        lines.append(f"always @(posedge {base_clk}) begin")
        lines.append(f"    if ({base_rst}) begin")
        lines.append(f"        assert (!(({qacceptn} == 1'b0) && ({qdeny} == 1'b1))) else $error(\"Conflicting QACCEPTn and QDENY\");")
        lines.append("    end")
        lines.append("end")
        lines.append("")
        lines.append("property P_QRUN;")
        lines.append(f"    @(posedge {base_clk}) disable iff(!{base_rst})")
        lines.append(f"    $rose({qactive}) |-> ##1 ({qreqn} && {qacceptn})")
        lines.append(f"                         ##1 {qactive};")
        lines.append("endproperty")
        lines.append("")
        lines.append("assert property (P_QRUN) $display(\"[%0t] P_QRUN Assertion PASS\", $time);")
        lines.append("else $display (\"[%0t] P_QRUN Assertion FAIL\", $time);")
        lines.append("")
        lines.append("endinterface")
        lines.append("")
        sv_text = "\n".join(lines)

        # ===== assertion_intf_inst.sv 생성 =====
        inst_lines: List[str] = []
        inst_lines.append("`include \"uvm_macros.svh\"")
        inst_lines.append("import uvm_pkg::*;")
        inst_lines.append("")
        inst_lines.append("assertion_intf u_assertion_intf();")
        inst_lines.append("")
        inst_lines.append(f"assign u_assertion_intf.{base_clk} = top.dut.{base_clk};")
        inst_lines.append(f"assign u_assertion_intf.{base_rst} = top.dut.{base_rst};")
        inst_lines.append(f"assign u_assertion_intf.{qreqn} = top.dut.{qreqn};")
        inst_lines.append(f"assign u_assertion_intf.{qacceptn} = top.dut.{qacceptn};")
        inst_lines.append(f"assign u_assertion_intf.{qactive} = top.dut.{qactive};")
        inst_lines.append(f"assign u_assertion_intf.{qdeny} = top.dut.{qdeny};")
        inst_text = "\n".join(inst_lines) + "\n"

        return [sv_text, inst_text]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        """JSON 출력 (필요시)"""
        return parsed
