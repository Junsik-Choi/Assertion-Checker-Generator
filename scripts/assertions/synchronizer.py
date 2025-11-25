from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json

from openpyxl import load_workbook

from .registry import register
from .base import BaseAssertionPlugin

# ===== 공통 유틸리티 함수 =====
def _find_cell(ws, value: str) -> Tuple[Optional[int], Optional[int]]:
    """시트에서 특정 값을 가진 셀의 위치(row, column)를 찾기"""
    tgt = (value or "").strip().lower()
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if str(v).strip().lower() == tgt:
                return c.row, c.column
    return None, None

def _get_sheet_ci(wb, want_name: str):
    """대소문자 구분 없이 워크시트 찾기"""
    target = (want_name or "").strip().lower()
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == target:
            return wb[nm]
    raise KeyError(f"Worksheet '{want_name}' does not exist.")

def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    """프롬프트로 옵션 선택 또는 커스텀 입력"""
    print(title, flush=True)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Enter custom")
    while True:
        try:
            s = input("Select > ").strip()
        except EOFError:
            return options[0][1] if options else ""
        if allow_custom and s == "0":
            try:
                return input("Enter value > ").strip()
            except EOFError:
                return ""
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.", flush=True)

def _read_json(p: Path) -> Optional[Dict[str, Any]]:
    """JSON 파일 읽기"""
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None

def _load_module_define(xls_path: Path) -> Dict[str, Any]:
    """module_define.json 또는 assertion_inputs.json에서 RTL 정보 로드"""
    session_dir = xls_path.parent
    md = _read_json(session_dir / "module_define.json")
    if md:
        return md
    ai = _read_json(session_dir / "assertion_inputs.json")
    if ai:
        return {
            "module": ai.get("module") or "",
            "clocks": ai.get("clocks") or [],
            "resets": ai.get("resets") or [],
            "inputs": ai.get("inputs") or [],
            "outputs": ai.get("outputs") or [],
            "inouts": ai.get("inouts") or [],
            "parameters": ai.get("parameters") or [],
        }
    return {}

def _normalize_range_token(token: Any) -> str:
    """포트 width를 [msb:lsb] 형식으로 정규화"""
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    try:
        n = int(t, 10)
        return f"[{n-1}:0]" if n >= 1 else "[0:0]"
    except Exception:
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    """포트 이름으로 width 토큰 찾기"""
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    candidates = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in candidates:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            # packed_range, range 등에서 width 정보 찾기
            for key in ("packed_range", "range", "packed", "decl"):
                pr = it.get(key)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            for key in ("width", "bit_width", "width_bits"):
                w = it.get(key)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            for lk, rk in (("msb", "lsb"), ("left", "right")):
                msb = it.get(lk); lsb = it.get(rk)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            return "[0:0]"
    return "[0:0]"

def _read_define_clk_rst(wb) -> Tuple[str, str]:
    """Define 시트에서 Base Clock/Reset 읽기"""
    try:
        ws = _get_sheet_ci(wb, "Define")
    except KeyError:
        return "", ""
    clk_r, clk_c = _find_cell(ws, "Base Clock")
    rst_r, rst_c = _find_cell(ws, "Base Reset")
    clk = ws.cell(row=clk_r, column=clk_c + 1).value if clk_r else None
    rst = ws.cell(row=rst_r, column=rst_c + 1).value if rst_r else None
    return (str(clk).strip() if clk else ""), (str(rst).strip() if rst else "")

def _pick_from(names: List[str], title: str, allow_custom: bool = False) -> str:
    """리스트에서 선택 또는 커스텀 입력"""
    opts = [(n, n) for n in names] if names else []
    return _pick_one(title, opts, allow_custom=allow_custom)

def _pick_int(title: str, default: str = "2") -> str:
    """정수 입력받기"""
    while True:
        try:
            s = input(f"{title} (integer) [default: {default}] > ").strip()
        except EOFError:
            return default
        if not s:
            return default
        if s.lstrip("-").isdigit():
            return s
        print("Please enter an integer.", flush=True)

# ===== Synchronizer 플러그인 =====
@register
class SynchronizerPlugin(BaseAssertionPlugin):
    plugin_name = "synchronizer"
    sheet_name = "Synchronizer"

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        """
        Excel 파일 파싱 및 사용자 입력 처리
        1. Synchronizer 시트 확인
        2. Base Clock/Reset 읽기 (Define 시트에서)
        3. Depth Sync Value, Enable Signal, Input Signal, Output Signal 입력받기
        """
        mod = _load_module_define(Path(xls_path))
        wb = load_workbook(xls_path)

        # 1. Synchronizer 시트 확인
        try:
            ws = _get_sheet_ci(wb, self.sheet_name)
        except KeyError:
            print(f"ERROR: '{self.sheet_name}' sheet does not exist in the Excel file.", flush=True)
            raise

        # 2. Base Clock/Reset은 Define 시트에서 읽기
        base_clk, base_rst = _read_define_clk_rst(wb)
        if not base_clk:
            print("ERROR: Base Clock value is empty in Define sheet.", flush=True)
            raise ValueError("Base Clock value is empty")
        if not base_rst:
            print("ERROR: Base Reset value is empty in Define sheet.", flush=True)
            raise ValueError("Base Reset value is empty")

        # 3. 모든 포트 수집 (입력/출력/inout)
        all_ports: List[str] = []
        for it in (mod.get("inputs") or []):
            n = it.get("name")
            if n:
                all_ports.append(n)
        for it in (mod.get("outputs") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)
        for it in (mod.get("inouts") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)

        # 4. Depth Sync / Enable / Input / Output 시그널 위치 찾기
        depth_r, depth_c = _find_cell(ws, "Depth Sync Value")
        enable_r, enable_c = _find_cell(ws, "Enable Signal")
        input_r,  input_c  = _find_cell(ws, "Input Signal")
        output_r, output_c = _find_cell(ws, "Output Signal")

        if not depth_r or not enable_r or not input_r or not output_r:
            print("ERROR: One or more required cells "
                  "('Depth Sync Value' / 'Enable Signal' / 'Input Signal' / 'Output Signal') "
                  "not found in Synchronizer sheet.", flush=True)
            raise ValueError("Missing required Synchronizer labels")

        # 기존 값 읽어서 없으면 플레이스홀더로 초기화
        depth_cell  = ws.cell(row=depth_r  + 1, column=depth_c).value
        enable_cell = ws.cell(row=enable_r + 1, column=enable_c).value
        input_cell  = ws.cell(row=input_r  + 1, column=input_c).value
        output_cell = ws.cell(row=output_r + 1, column=output_c).value

        depth_val  = str(depth_cell).strip()  if depth_cell  and str(depth_cell).strip()  else "<Depth Sync Value>"
        enable_sig = str(enable_cell).strip() if enable_cell and str(enable_cell).strip() else "<Enable Signal>"
        input_sig  = str(input_cell).strip()  if input_cell  and str(input_cell).strip()  else "<Input Signal>"
        output_sig = str(output_cell).strip() if output_cell and str(output_cell).strip() else "<Output Signal>"

        while True:
            print("\n==================== Synchronizer Settings ====================")
            print(f"[1] Depth Sync Value : {depth_val}")
            print(f"[2] Enable Signal    : {enable_sig}")
            print(f"[3] Input Signal     : {input_sig}")
            print(f"[4] Output Signal    : {output_sig}")
            print("================================================================")
            print("Select item number to edit")
            print("Press Enter twice to confirm all")
            choice = input("> ").strip()
            if choice == "":
                missing = []
                if depth_val  in ("", "<Depth Sync Value>"): missing.append("[1]")
                if enable_sig in ("", "<Enable Signal>"):     missing.append("[2]")
                if input_sig  in ("", "<Input Signal>"):      missing.append("[3]")
                if output_sig in ("", "<Output Signal>"):     missing.append("[4]")
                if missing: print(f"{','.join(missing)} has NOT been entered yet.")
                print("Press Enter again to confirm, or select item number to edit")
                choice = input("> ").strip()
                if choice == "":
                    break
            if choice == "1":
                print("\n=== Depth Sync Value ===")
                print("Enter the synchronizer depth (e.g., 2 for 2-stage, 3 for 3-stage)")
                default_depth = depth_val if depth_val not in ("", "<Depth Sync Value>") else "2"
                depth_val = _pick_int("Enter Depth Sync Value", default=default_depth)
            elif choice == "2":
                print("\n=== Enable Signal ===")
                enable_sig = _pick_from(all_ports, "Select Enable Signal:", allow_custom=True)
            elif choice == "3":
                print("\n=== Input Signal ===")
                input_sig = _pick_from(all_ports, "Select Input Signal:", allow_custom=True)
            elif choice == "4":
                print("\n=== Output Signal ===")
                output_sig = _pick_from(all_ports, "Select Output Signal:", allow_custom=True)
            else:
                print("Invalid selection. Try again.")
                continue

        # 최종 값 시트에 기록
        ws.cell(row=depth_r  + 1, column=depth_c, value=depth_val)
        ws.cell(row=enable_r + 1, column=enable_c, value=enable_sig)
        ws.cell(row=input_r  + 1, column=input_c, value=input_sig)
        ws.cell(row=output_r + 1, column=output_c, value=output_sig)

        # 8. Synchronizer 시트의 Base Clock/Reset 셀에도 값 기록
        clk_row, clk_col = _find_cell(ws, "Base Clock")
        if clk_row:
            ws.cell(row=clk_row, column=clk_col + 1, value=base_clk)

        rst_row, rst_col = _find_cell(ws, "Base Reset")
        if rst_row:
            ws.cell(row=rst_row, column=rst_col + 1, value=base_rst)

        # 9. Excel 저장
        wb.save(xls_path)

        # 10. Width 정보 수집
        enable_width = _port_width_token(mod, enable_sig)
        input_width = _port_width_token(mod, input_sig)
        output_width = _port_width_token(mod, output_sig)
        clk_width = _port_width_token(mod, base_clk)
        rst_width = _port_width_token(mod, base_rst)

        # 11. 결과 반환
        blocks = [{
            "Base Clock": base_clk,
            "Base Reset": base_rst,
            "Depth Sync Value": depth_val,
            "Enable Signal": enable_sig,
            "Input Signal": input_sig,
            "Output Signal": output_sig,
            "Base Clock Width": clk_width,
            "Base Reset Width": rst_width,
            "Enable Signal Width": enable_width,
            "Input Signal Width": input_width,
            "Output Signal Width": output_width,
        }]
        return {"blocks": blocks}

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
        """
        SystemVerilog assertion 코드 생성
        - assertion_intf.sv: interface 정의
        - assertion_intf_inst.sv: 인스턴스 및 연결
        """
        blocks = parsed.get("blocks") or []
        if not blocks:
            return ["// No Synchronizer assertions generated.\n", ""]

        b = blocks[0]
        clk = b.get("Base Clock", "") or "clk"
        rst = b.get("Base Reset", "") or "rst_n"
        depth = b.get("Depth Sync Value", "") or "2"
        enable = b.get("Enable Signal", "") or "enable"
        input_sig = b.get("Input Signal", "") or "data_in"
        output_sig = b.get("Output Signal", "") or "data_out"

        clk_w = b.get("Base Clock Width", "[0:0]")
        rst_w = b.get("Base Reset Width", "[0:0]")
        enable_w = b.get("Enable Signal Width", "[0:0]")
        input_w = b.get("Input Signal Width", "[0:0]")
        output_w = b.get("Output Signal Width", "[0:0]")

        # ===== assertion_intf.sv 생성 =====
        lines: List[str] = []
        lines.append("`include \"uvm_macros.svh\"")
        lines.append("import uvm_pkg::*;")
        lines.append("")
        lines.append("interface assertion_intf();")
        lines.append("")
        lines.append(f"logic {clk_w} {clk};")
        lines.append(f"logic {rst_w} {rst};")
        lines.append(f"logic {enable_w} {enable};")
        lines.append(f"logic {input_w} {input_sig};")
        lines.append(f"logic {output_w} {output_sig};")
        lines.append("")
        lines.append(f"parameter DEPTH_SYNC = {depth};")
        lines.append("")
        lines.append("property p_synchronizer;")
        lines.append(f"    @(posedge {clk}) disable iff(!{rst})")
        lines.append(f"    {enable} |-> ##(DEPTH_SYNC) ({output_sig} == $past({input_sig}, DEPTH_SYNC));")
        lines.append("endproperty")
        lines.append("")
        lines.append("assert property (p_synchronizer) else `uvm_error(\"ASSERTION\", \"Synchronizer check failed\")")
        lines.append("")
        lines.append("endinterface")
        lines.append("")
        sv_text = "\n".join(lines)

        # ===== assertion_intf_inst.sv 생성 =====
        inst_lines: List[str] = []
        inst_lines.append("`include \"uvm_macros.svh\"")
        inst_lines.append("import uvm_pkg::*;")
        inst_lines.append("")
        inst_lines.append("assertion_intf u_assertion_intf();")
        inst_lines.append("")
        inst_lines.append(f"assign u_assertion_intf.{clk} = top.dut.{clk};")
        inst_lines.append(f"assign u_assertion_intf.{rst} = top.dut.{rst};")
        inst_lines.append(f"assign u_assertion_intf.{enable} = top.dut.{enable};")
        inst_lines.append(f"assign u_assertion_intf.{input_sig} = top.dut.{input_sig};")
        inst_lines.append(f"assign u_assertion_intf.{output_sig} = top.dut.{output_sig};")
        inst_text = "\n".join(inst_lines) + "\n"

        return [sv_text, inst_text]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        """JSON 출력 (필요시)"""
        return parsed
