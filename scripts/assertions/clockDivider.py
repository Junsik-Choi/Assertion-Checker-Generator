from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json

from openpyxl import load_workbook

from .registry import register
from .base import BaseAssertionPlugin

# ===== 공통 유틸리티 함수 =====
def _find_cell(ws, value: str) -> Tuple[Optional[int], Optional[int]]:
    """시트에서 특정 값을 가진 셀의 위치(row, column)를 찾기"""
    tgt = (value or "").strip().lower()
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if str(v).strip().lower() == tgt:
                return c.row, c.column
    return None, None

def _get_sheet_ci(wb, want_name: str):
    """대소문자 구분 없이 워크시트 찾기"""
    target = (want_name or "").strip().lower()
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == target:
            return wb[nm]
    raise KeyError(f"Worksheet '{want_name}' does not exist.")

def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    """프롬프트로 옵션 선택 또는 커스텀 입력"""
    print(title, flush=True)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Enter custom")
    while True:
        try:
            s = input("Select > ").strip()
        except EOFError:
            return options[0][1] if options else ""
        if allow_custom and s == "0":
            try:
                return input("Enter value > ").strip()
            except EOFError:
                return ""
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.", flush=True)

def _read_json(p: Path) -> Optional[Dict[str, Any]]:
    """JSON 파일 읽기"""
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None

def _load_module_define(xls_path: Path) -> Dict[str, Any]:
    """module_define.json 또는 assertion_inputs.json에서 RTL 정보 로드"""
    session_dir = xls_path.parent
    md = _read_json(session_dir / "module_define.json")
    if md:
        return md
    ai = _read_json(session_dir / "assertion_inputs.json")
    if ai:
        return {
            "module": ai.get("module") or "",
            "clocks": ai.get("clocks") or [],
            "resets": ai.get("resets") or [],
            "inputs": ai.get("inputs") or [],
            "outputs": ai.get("outputs") or [],
            "inouts": ai.get("inouts") or [],
            "parameters": ai.get("parameters") or [],
        }
    return {}

def _normalize_range_token(token: Any) -> str:
    """포트 width를 [msb:lsb] 형식으로 정규화"""
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    try:
        n = int(t, 10)
        return f"[{n-1}:0]" if n >= 1 else "[0:0]"
    except Exception:
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    """포트 이름으로 width 토큰 찾기"""
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    candidates = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in candidates:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            # packed_range, range 등에서 width 정보 찾기
            for key in ("packed_range", "range", "packed", "decl"):
                pr = it.get(key)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            for key in ("width", "bit_width", "width_bits"):
                w = it.get(key)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            for lk, rk in (("msb", "lsb"), ("left", "right")):
                msb = it.get(lk); lsb = it.get(rk)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            return "[0:0]"
    return "[0:0]"

def _read_define_clk_rst(wb) -> Tuple[str, str]:
    """Define 시트에서 Base Clock/Reset 읽기"""
    try:
        ws = _get_sheet_ci(wb, "Define")
    except KeyError:
        return "", ""
    clk_r, clk_c = _find_cell(ws, "Base Clock")
    rst_r, rst_c = _find_cell(ws, "Base Reset")
    clk = ws.cell(row=clk_r, column=clk_c + 1).value if clk_r else None
    rst = ws.cell(row=rst_r, column=rst_c + 1).value if rst_r else None
    return (str(clk).strip() if clk else ""), (str(rst).strip() if rst else "")

def _pick_from(names: List[str], title: str, allow_custom: bool = False) -> str:
    """리스트에서 선택 또는 커스텀 입력"""
    opts = [(n, n) for n in names] if names else []
    return _pick_one(title, opts, allow_custom=allow_custom)

def _pick_int(title: str, default: str = "") -> str:
    """정수 입력받기"""
    while True:
        try:
            prompt = f"{title} (integer)"
            if default:
                prompt += f" [default: {default}]"
            prompt += " > "
            s = input(prompt).strip()
        except EOFError:
            return default if default else "0"
        if not s:
            return default if default else "0"
        if s.lstrip("-").isdigit():
            return s
        print("Please enter an integer.", flush=True)

# ===== Clock Divider 플러그인 =====
@register
class ClockDividerPlugin(BaseAssertionPlugin):
    plugin_name = "clockDivider"
    sheet_name = "clockDivider"

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        """
        Excel 파일 파싱 및 사용자 입력 처리
        1. clockDivider 시트 확인
        2. Base Clock/Reset 읽기 (Define 시트에서)
        3. Reference Clock, MAX Value, DIVRATIO, CLKOUT, START FLAG, DISABLE 입력받기
        """
        mod = _load_module_define(Path(xls_path))
        wb = load_workbook(xls_path)

        # 1. clockDivider 시트 확인
        try:
            ws = _get_sheet_ci(wb, self.sheet_name)
        except KeyError:
            print(f"ERROR: '{self.sheet_name}' sheet does not exist in the Excel file.", flush=True)
            raise

        # 2. Base Clock/Reset은 Define 시트에서 읽기
        base_clk, base_rst = _read_define_clk_rst(wb)
        if not base_clk:
            print("ERROR: Base Clock value is empty in Define sheet.", flush=True)
            raise ValueError("Base Clock value is empty")
        if not base_rst:
            print("ERROR: Base Reset value is empty in Define sheet.", flush=True)
            raise ValueError("Base Reset value is empty")

        # 3. 모든 포트 수집 (입력/출력/inout)
        all_ports: List[str] = []
        for it in (mod.get("inputs") or []):
            n = it.get("name")
            if n:
                all_ports.append(n)
        for it in (mod.get("outputs") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)
        for it in (mod.get("inouts") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)

        # 4. Reference Clock 확인 및 입력
        ref_clk_r, ref_clk_c = _find_cell(ws, "Reference Clock")
        if not ref_clk_r:
            print("ERROR: 'Reference Clock' cell not found in clockDivider sheet.", flush=True)
            raise ValueError("'Reference Clock' cell not found")
        
        ref_clk = ws.cell(row=ref_clk_r + 1, column=ref_clk_c).value
        if not ref_clk or str(ref_clk).strip() == "":
            print(f"\n=== Reference Clock ===")
            ref_clk = _pick_from(all_ports, "Select Reference Clock:", allow_custom=True)
            ws.cell(row=ref_clk_r + 1, column=ref_clk_c, value=ref_clk)
        else:
            ref_clk = str(ref_clk).strip()

        # 5. MAX Value 확인 및 입력
        max_r, max_c = _find_cell(ws, "MAX Value")
        if not max_r:
            print("ERROR: 'MAX Value' cell not found in clockDivider sheet.", flush=True)
            raise ValueError("'MAX Value' cell not found")
        
        max_val = ws.cell(row=max_r + 1, column=max_c).value
        if not max_val or str(max_val).strip() == "":
            print(f"\n=== MAX Value ===")
            print("Enter the maximum divider ratio value")
            max_val = _pick_int("Enter MAX Value", default="")
            ws.cell(row=max_r + 1, column=max_c, value=max_val)
        else:
            max_val = str(max_val).strip()

        # 6. DIVRATIO 확인 및 입력
        divratio_r, divratio_c = _find_cell(ws, "DIVRATIO")
        if not divratio_r:
            print("ERROR: 'DIVRATIO' cell not found in clockDivider sheet.", flush=True)
            raise ValueError("'DIVRATIO' cell not found")
        
        divratio = ws.cell(row=divratio_r + 1, column=divratio_c).value
        if not divratio or str(divratio).strip() == "":
            print(f"\n=== DIVRATIO ===")
            divratio = _pick_from(all_ports, "Select DIVRATIO signal:", allow_custom=True)
            ws.cell(row=divratio_r + 1, column=divratio_c, value=divratio)
        else:
            divratio = str(divratio).strip()

        # 7. CLKOUT 확인 및 입력
        clkout_r, clkout_c = _find_cell(ws, "CLKOUT")
        if not clkout_r:
            print("ERROR: 'CLKOUT' cell not found in clockDivider sheet.", flush=True)
            raise ValueError("'CLKOUT' cell not found")
        
        clkout = ws.cell(row=clkout_r + 1, column=clkout_c).value
        if not clkout or str(clkout).strip() == "":
            print(f"\n=== CLKOUT ===")
            clkout = _pick_from(all_ports, "Select CLKOUT signal:", allow_custom=True)
            ws.cell(row=clkout_r + 1, column=clkout_c, value=clkout)
        else:
            clkout = str(clkout).strip()

        # 8. START FLAG 확인 및 입력
        start_flag_r, start_flag_c = _find_cell(ws, "START FLAG")
        if not start_flag_r:
            print("ERROR: 'START FLAG' cell not found in clockDivider sheet.", flush=True)
            raise ValueError("'START FLAG' cell not found")
        
        start_flag = ws.cell(row=start_flag_r + 1, column=start_flag_c).value
        if not start_flag or str(start_flag).strip() == "":
            print(f"\n=== START FLAG ===")
            start_flag = _pick_from(all_ports, "Select START FLAG signal:", allow_custom=True)
            ws.cell(row=start_flag_r + 1, column=start_flag_c, value=start_flag)
        else:
            start_flag = str(start_flag).strip()

        # 9. DISABLE 확인 및 입력
        disable_r, disable_c = _find_cell(ws, "DISABLE")
        if not disable_r:
            print("ERROR: 'DISABLE' cell not found in clockDivider sheet.", flush=True)
            raise ValueError("'DISABLE' cell not found")
        
        disable = ws.cell(row=disable_r + 1, column=disable_c).value
        if not disable or str(disable).strip() == "":
            print(f"\n=== DISABLE ===")
            disable = _pick_from(all_ports, "Select DISABLE signal:", allow_custom=True)
            ws.cell(row=disable_r + 1, column=disable_c, value=disable)
        else:
            disable = str(disable).strip()

        # 10. clockDivider 시트의 Base Clock/Reset 셀에도 값 기록
        clk_row, clk_col = _find_cell(ws, "Base Clock")
        if clk_row:
            ws.cell(row=clk_row, column=clk_col + 1, value=base_clk)
        
        rst_row, rst_col = _find_cell(ws, "Base Reset")
        if rst_row:
            ws.cell(row=rst_row, column=rst_col + 1, value=base_rst)

        # 11. Excel 저장
        wb.save(xls_path)

        # 12. Width 정보 수집
        ref_clk_width = _port_width_token(mod, ref_clk)
        base_rst_width = _port_width_token(mod, base_rst)
        divratio_width = _port_width_token(mod, divratio)
        clkout_width = _port_width_token(mod, clkout)
        start_flag_width = _port_width_token(mod, start_flag)
        disable_width = _port_width_token(mod, disable)

        # 13. 결과 반환
        blocks = [{
            "Base Clock": base_clk,
            "Base Reset": base_rst,
            "Reference Clock": ref_clk,
            "MAX Value": max_val,
            "DIVRATIO": divratio,
            "CLKOUT": clkout,
            "START FLAG": start_flag,
            "DISABLE": disable,
            "Reference Clock Width": ref_clk_width,
            "Base Reset Width": base_rst_width,
            "DIVRATIO Width": divratio_width,
            "CLKOUT Width": clkout_width,
            "START FLAG Width": start_flag_width,
            "DISABLE Width": disable_width,
        }]
        return {"blocks": blocks}

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
        """
        SystemVerilog assertion 코드 생성
        - assertion_intf.sv: interface 정의
        - assertion_intf_inst.sv: 인스턴스 및 연결
        """
        blocks = parsed.get("blocks") or []
        if not blocks:
            return ["// No Clock Divider assertions generated.\n", ""]
        
        b = blocks[0]
        base_clk = b.get("Base Clock", "") or "clk"
        base_rst = b.get("Base Reset", "") or "rst_n"
        ref_clk = b.get("Reference Clock", "") or "ref_clk"
        max_val = b.get("MAX Value", "") or "100"
        divratio = b.get("DIVRATIO", "") or "divratio"
        clkout = b.get("CLKOUT", "") or "clkout"
        start_flag = b.get("START FLAG", "") or "start_flag"
        disable = b.get("DISABLE", "") or "disable"

        ref_clk_w = b.get("Reference Clock Width", "[0:0]")
        base_rst_w = b.get("Base Reset Width", "[0:0]")
        divratio_w = b.get("DIVRATIO Width", "[0:0]")
        clkout_w = b.get("CLKOUT Width", "[0:0]")
        start_flag_w = b.get("START FLAG Width", "[0:0]")
        disable_w = b.get("DISABLE Width", "[0:0]")

        # ===== assertion_intf.sv 생성 =====
        lines: List[str] = []
        lines.append("`include \"uvm_macros.svh\"")
        lines.append("import uvm_pkg::*;")
        lines.append("")
        lines.append("interface assertion_intf();")
        lines.append("")
        lines.append(f"    logic {ref_clk_w}  {ref_clk};")
        lines.append(f"    logic {base_rst_w}  {base_rst};")
        lines.append(f"    logic {divratio_w}  {divratio};")
        lines.append(f"    logic {clkout_w}  {clkout};")
        lines.append(f"    logic {start_flag_w}  {start_flag};")
        lines.append(f"    logic {disable_w}  {disable};")
        lines.append("")
        lines.append(f"    parameter MAX = {max_val};")
        lines.append("")
        lines.append("")
        lines.append("    sequence s_clkdiv_counter(int high_cnt, int low_cnt);")
        lines.append(f"        @(posedge {ref_clk} or negedge {ref_clk})")
        lines.append("         (1, high_cnt = 0, low_cnt = 0)")
        lines.append(f"         ##[0:MAX*2]   $rose({clkout})")
        lines.append(f"         ##0           ({clkout}, high_cnt = high_cnt + 1)[*1:$]")
        lines.append("         ##1           $fell(" + clkout + ")")
        lines.append(f"         ##0           (!{clkout}, low_cnt = low_cnt + 1)[*1:$]")
        lines.append(f"         ##1           $rose({clkout})")
        lines.append("    endsequence")
        lines.append("")
        lines.append("")
        lines.append("    property p_clkdiv_check0;")
        lines.append("        int h_cnt, l_cnt;")
        lines.append(f"        @(posedge {ref_clk}) disable iff (!{base_rst} || {disable})")
        lines.append(f"        $rose({start_flag})")
        lines.append("        |-> first_match(s_clkdiv_counter(h_cnt, l_cnt))")
        lines.append(f"        |-> ((h_cnt == ({divratio} + 1)) && (l_cnt == ({divratio} + 1)))")
        lines.append(f"        ##0 (1,$display(\"[%0t] [CLKDIV_CHK] (div=%0d) -> HIGH=%0d (%.1f {ref_clk}), LOW=%0d (%.1f {ref_clk})\",")
        lines.append(f"                        $realtime,{divratio},h_cnt, h_cnt/2.0, l_cnt, l_cnt/2.0));")
        lines.append("    endproperty")
        lines.append("")
        lines.append("    assert property (p_clkdiv_check0) else `uvm_error(\"ASSERTION\", \"Clock divider check failed\")")
        lines.append("")
        lines.append("endinterface")
        lines.append("")
        sv_text = "\n".join(lines)

        # ===== assertion_intf_inst.sv 생성 =====
        inst_lines: List[str] = []
        inst_lines.append("`include \"uvm_macros.svh\"")
        inst_lines.append("import uvm_pkg::*;")
        inst_lines.append("")
        inst_lines.append("assertion_intf u_assertion_intf();")
        inst_lines.append("")
        inst_lines.append(f"assign u_assertion_intf.{ref_clk} = top.dut.{ref_clk};")
        inst_lines.append(f"assign u_assertion_intf.{base_rst} = top.dut.{base_rst};")
        inst_lines.append(f"assign u_assertion_intf.{divratio} = top.dut.{divratio};")
        inst_lines.append(f"assign u_assertion_intf.{clkout} = top.dut.{clkout};")
        inst_lines.append(f"assign u_assertion_intf.{start_flag} = top.dut.{start_flag};")
        inst_lines.append(f"assign u_assertion_intf.{disable} = top.dut.{disable};")
        inst_text = "\n".join(inst_lines) + "\n"

        return [sv_text, inst_text]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        """JSON 출력 (필요시)"""
        return parsed
