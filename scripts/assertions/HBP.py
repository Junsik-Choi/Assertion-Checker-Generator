from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json

from openpyxl import load_workbook

from .registry import register
from .base import BaseAssertionPlugin

# ===== 공통 유틸 =====
def _find_cell(ws, value: str) -> Tuple[Optional[int], Optional[int]]:
    tgt = (value or "").strip().lower()
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if str(v).strip().lower() == tgt:
                return c.row, c.column
    return None, None

def _get_sheet_ci(wb, want_name: str):
    target = (want_name or "").strip().lower()
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == target:
            return wb[nm]
    raise KeyError(f"Worksheet {want_name} does not exist.")

def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    print(title, flush=True)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Enter custom")
    while True:
        try:
            s = input("Select > ").strip()
        except EOFError:
            return options[0][1] if options else ""
        if allow_custom and s == "0":
            try:
                return input("Enter value > ").strip()
            except EOFError:
                return ""
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.", flush=True)

def _read_json(p: Path) -> Optional[Dict[str, Any]]:
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None

def _load_module_define(xls_path: Path) -> Dict[str, Any]:
    session_dir = xls_path.parent
    md = _read_json(session_dir / "module_define.json")
    if md:
        return md
    ai = _read_json(session_dir / "assertion_inputs.json")
    if ai:
        return {
            "module": ai.get("module") or "",
            "clocks": ai.get("clocks") or [],
            "resets": ai.get("resets") or [],
            "inputs": ai.get("inputs") or [],
            "outputs": ai.get("outputs") or [],
            "inouts": ai.get("inouts") or [],
            "parameters": ai.get("parameters") or [],
        }
    return {}

def _normalize_range_token(token: Any) -> str:
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    try:
        n = int(t, 10)
        return f"[{n-1}:0]" if n >= 1 else "[0:0]"
    except Exception:
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    candidates = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in candidates:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            for key in ("packed_range", "range", "packed", "decl"):
                pr = it.get(key)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            for key in ("width", "bit_width", "width_bits"):
                w = it.get(key)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            for lk, rk in (("msb", "lsb"), ("left", "right")):
                msb = it.get(lk); lsb = it.get(rk)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            return "[0:0]"
    return "[0:0]"

def _fmt_input_decl(sig: str, width_tok: str) -> str:
    tok = (width_tok or "").strip() or "[0:0]"
    return f"input logic {tok} {sig}"

# ===== HBP 시트 처리 =====
def _ensure_hbp_layout(ws) -> Tuple[int, int]:
    h_row, h_col = _find_cell(ws, "HBP")
    if h_row is None:
        h_row, h_col = 1, 1
        ws.cell(row=h_row, column=h_col, value="HBP")
    labels = ["Count Trigger", "Target Pulse", "Data Enable Signal", "Expected Min Value", "Expected Max Value"]
    for i, lab in enumerate(labels, start=1):
        if ws.cell(row=h_row + i, column=h_col).value is None:
            ws.cell(row=h_row + i, column=h_col, value=lab)
    return h_row, h_col

def _read_define_clk_rst(wb) -> Tuple[str, str]:
    try:
        ws = _get_sheet_ci(wb, "Define")
    except KeyError:
        return "", ""
    clk_r, clk_c = _find_cell(ws, "Base Clock")
    rst_r, rst_c = _find_cell(ws, "Base Reset")
    clk = ws.cell(row=clk_r, column=clk_c + 1).value if clk_r else None
    rst = ws.cell(row=rst_r, column=rst_c + 1).value if rst_r else None
    return (str(clk).strip() if clk else ""), (str(rst).strip() if rst else "")

def _pick_from(names: List[str], title: str, allow_custom: bool = False) -> str:
    opts = [(n, n) for n in names] if names else []
    return _pick_one(title, opts, allow_custom=allow_custom)

def _pick_int(title: str) -> str:
    while True:
        try:
            s = input(f"{title} (integer) > ").strip()
        except EOFError:
            return "0"
        if s.lstrip("-").isdigit():
            return s
        print("Please enter an integer.", flush=True)

# ===== 플러그인 =====
@register
class HBPPlugin(BaseAssertionPlugin):
    plugin_name = "hbp"
    sheet_name = "HBP"

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        mod = _load_module_define(Path(xls_path))
        wb_w = load_workbook(xls_path)

        # HBP 시트 확보 및 레이아웃
        try:
            ws_w = _get_sheet_ci(wb_w, self.sheet_name)
        except KeyError:
            ws_w = wb_w.create_sheet(title=self.sheet_name)
        h_row, h_col = _ensure_hbp_layout(ws_w)

        # Define 시트로부터 Base Clock/Reset
        base_clk, base_rst = _read_define_clk_rst(wb_w)
        count_trig = base_clk

        # Target Pulse 후보: i_hsync 포함(대소문자 무시)
        all_ports: List[str] = []
        for it in (mod.get("inputs") or []):
            n = it.get("name")
            if n:
                all_ports.append(n)
        for it in (mod.get("outputs") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)
        hs_candidates = [n for n in all_ports if "i_hsync" in n.lower()]
        if not hs_candidates:
            target_pulse = _pick_from(all_ports, "Select Target Pulse (no i_hsync found)", allow_custom=True)
        elif len(hs_candidates) == 1:
            target_pulse = hs_candidates[0]
        else:
            target_pulse = _pick_from(hs_candidates, "Select Target Pulse (matched i_hsync)", allow_custom=False)

        # Data Enable Signal 후보: i_de 포함(대소문자 무시)
        de_candidates = [n for n in all_ports if "i_de" in n.lower()]
        if not de_candidates:
            data_enable_signal = _pick_from(all_ports, "Select Data Enable Signal (no i_de found)", allow_custom=True)
        elif len(de_candidates) == 1:
            data_enable_signal = de_candidates[0]
        else:
            data_enable_signal = _pick_from(de_candidates, "Select Data Enable Signal (matched i_de)", allow_custom=False)

        # Expected Min/Max 입력
        exp_min = _pick_int("Enter Expected Min Value")
        exp_max = _pick_int("Enter Expected Max Value")

        # 시트 기록
        ws_w.cell(row=h_row + 1, column=h_col + 1, value=count_trig)
        ws_w.cell(row=h_row + 2, column=h_col + 1, value=target_pulse)
        ws_w.cell(row=h_row + 3, column=h_col + 1, value=data_enable_signal)
        ws_w.cell(row=h_row + 4, column=h_col + 1, value=exp_min)
        ws_w.cell(row=h_row + 5, column=h_col + 1, value=exp_max)
        wb_w.save(xls_path)

        # 파싱 결과
        w_clk = _port_width_token(mod, base_clk)
        w_rst = _port_width_token(mod, base_rst)
        w_tp  = _port_width_token(mod, target_pulse)
        w_de  = _port_width_token(mod, data_enable_signal)
        blocks = [{
            "Base Clock": base_clk,
            "Base Reset": base_rst,
            "Count Trigger": count_trig,
            "Target Pulse": target_pulse,
            "Data Enable Signal": data_enable_signal,
            "Expected Min Value": exp_min,
            "Expected Max Value": exp_max,
            "Base Clock Width": w_clk,
            "Base Reset Width": w_rst,
            "Target Pulse Width": w_tp,
            "Data Enable Signal Width": w_de,
        }]
        return {"blocks": blocks}

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
        blocks = parsed.get("blocks") or []
        if not blocks:
            return ["// No HBP assertions generated.\n", ""]
        b = blocks[0]
        base_clk = b.get("Base Clock", "") or "clk"
        base_rst = b.get("Base Reset", "") or "rst_n"
        target_pulse = b.get("Target Pulse", "") or "i_hsync"
        data_enable_signal = b.get("Data Enable Signal", "") or "i_de"
        exp_min = b.get("Expected Min Value", "") or "0"
        exp_max = b.get("Expected Max Value", "") or "0"
        w_clk = b.get("Base Clock Width", "[0:0]")
        w_rst = b.get("Base Reset Width", "[0:0]")
        w_tp  = b.get("Target Pulse Width", "[0:0]")
        w_de  = b.get("Data Enable Signal Width", "[0:0]")

        # 모듈 래퍼(빌더가 포트/본문 분리 집계)
        lines: List[str] = []
        lines.append("module assertion_hbp")
        lines.append("(")
        lines.append(f"    {_fmt_input_decl(base_clk, w_clk)},")
        lines.append(f"    {_fmt_input_decl(base_rst, w_rst)},")
        lines.append(f"    {_fmt_input_decl(target_pulse, w_tp)},")
        lines.append(f"    {_fmt_input_decl(data_enable_signal, w_de)}")
        lines.append(");")
        lines.append("")
        lines.append("property p_hbp;")
        lines.append("    int value_count;")
        lines.append(f"    @(posedge {base_clk}) disable iff(!{base_rst})")
        lines.append(f"    $fell({target_pulse}) |-> (1, value_count  = 0)")
        lines.append(f"    ##1 ({data_enable_signal}, value_count = value_count + 1)[*0:$]")
        lines.append(f"    ##1 (!{data_enable_signal}, value_count = value_count + 1)")
        lines.append(f"    ##0 ({exp_min} <= value_count && value_count <= {exp_max});")
        lines.append("endproperty")
        lines.append("")
        lines.append("assert property (p_hbp)  else $error(\"failed at %t\", $time);")
        lines.append("")
        lines.append("endmodule")
        lines.append("")
        sv_text = "\n".join(lines)

        # 인스턴스: 빌더가 헤더/선언 생성 → assign만 반환
        inst_lines: List[str] = []
        inst_lines.append(f"assign u_assertion_gen.{base_clk} = top.dut.{base_clk};")
        inst_lines.append(f"assign u_assertion_gen.{base_rst} = top.dut.{base_rst};")
        inst_lines.append(f"assign u_assertion_gen.{target_pulse} = top.dut.{target_pulse};")
        inst_lines.append(f"assign u_assertion_gen.{data_enable_signal} = top.dut.{data_enable_signal};")
        inst_text = "\n".join(inst_lines) + "\n"
        return [sv_text, inst_text]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        return parsed