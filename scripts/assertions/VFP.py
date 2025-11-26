from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json

from openpyxl import load_workbook

from .registry import register
from .base import BaseAssertionPlugin

# ===== 공통 유틸 =====
def _find_cell(ws, value: str) -> Tuple[Optional[int], Optional[int]]:
    tgt = (value or "").strip().lower()
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if str(v).strip().lower() == tgt:
                return c.row, c.column
    return None, None

def _get_sheet_ci(wb, want_name: str):
    target = (want_name or "").strip().lower()
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == target:
            return wb[nm]
    raise KeyError(f"Worksheet {want_name} does not exist.")

def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    print(title, flush=True)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Enter custom")
    while True:
        try:
            s = input("Select > ").strip()
        except EOFError:
            return options[0][1] if options else ""
        if allow_custom and s == "0":
            try:
                return input("Enter value > ").strip()
            except EOFError:
                return ""
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.", flush=True)

def _read_json(p: Path) -> Optional[Dict[str, Any]]:
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None

def _load_module_define(xls_path: Path) -> Dict[str, Any]:
    session_dir = xls_path.parent
    md = _read_json(session_dir / "module_define.json")
    if md:
        return md
    ai = _read_json(session_dir / "assertion_inputs.json")
    if ai:
        return {
            "module": ai.get("module") or "",
            "clocks": ai.get("clocks") or [],
            "resets": ai.get("resets") or [],
            "inputs": ai.get("inputs") or [],
            "outputs": ai.get("outputs") or [],
            "inouts": ai.get("inouts") or [],
            "parameters": ai.get("parameters") or [],
        }
    return {}

def _normalize_range_token(token: Any) -> str:
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    try:
        n = int(t, 10)
        return f"[{n-1}:0]" if n >= 1 else "[0:0]"
    except Exception:
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    candidates = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in candidates:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            for key in ("packed_range", "range", "packed", "decl"):
                pr = it.get(key)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            for key in ("width", "bit_width", "width_bits"):
                w = it.get(key)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            for lk, rk in (("msb", "lsb"), ("left", "right")):
                msb = it.get(lk); lsb = it.get(rk)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            return "[0:0]"
    return "[0:0]"

def _fmt_input_decl(sig: str, width_tok: str) -> str:
    tok = (width_tok or "").strip() or "[0:0]"
    return f"input {tok} {sig}"

def _read_define_clk_rst(wb) -> Tuple[str, str]:
    try:
        ws = _get_sheet_ci(wb, "Define")
    except KeyError:
        return "", ""
    clk_r, clk_c = _find_cell(ws, "Base Clock")
    rst_r, rst_c = _find_cell(ws, "Base Reset")
    clk = ws.cell(row=clk_r, column=clk_c + 1).value if clk_r else None
    rst = ws.cell(row=rst_r, column=rst_c + 1).value if rst_r else None
    return (str(clk).strip() if clk else ""), (str(rst).strip() if rst else "")

def _pick_from(names: List[str], title: str, allow_custom: bool = False) -> str:
    opts = [(n, n) for n in names] if names else []
    return _pick_one(title, opts, allow_custom=allow_custom)

def _pick_int(title: str) -> str:
    while True:
        try:
            s = input(f"{title} (integer) > ").strip()
        except EOFError:
            return "0"
        if s.lstrip("-").isdigit():
            return s
        print("Please enter an integer.", flush=True)

def _pick_int_with_validation(title: str, min_val: Optional[int] = None) -> str:
    """정수를 입력받되, min_val이 주어지면 그보다 크거나 같은지 검증"""
    while True:
        try:
            s = input(f"{title} (integer) > ").strip()
        except EOFError:
            return "0"
        if s.lstrip("-").isdigit():
            val = int(s)
            if min_val is not None and val < min_val:
                print(f"Value must be >= {min_val}. Try again.", flush=True)
                continue
            return s
        print("Please enter an integer.", flush=True)

# ===== 플러그인 =====
@register
class VFPPlugin(BaseAssertionPlugin):
    plugin_name = "vfp"
    sheet_name = "VFP"

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        mod = _load_module_define(Path(xls_path))
        wb_w = load_workbook(xls_path)

        # VFP 시트 확인
        try:
            ws_vfp = _get_sheet_ci(wb_w, self.sheet_name)
        except KeyError:
            print(f"ERROR: '{self.sheet_name}' sheet does not exist in the Excel file.", flush=True)
            raise

        # Base Clock/Reset은 Define 시트에서 직접 읽기 (수식 참조 문제 방지)
        base_clk, base_rst = _read_define_clk_rst(wb_w)
        if not base_clk:
            print("ERROR: Base Clock value is empty in Define sheet.", flush=True)
            raise ValueError("Base Clock value is empty")
        if not base_rst:
            print("ERROR: Base Reset value is empty in Define sheet.", flush=True)
            raise ValueError("Base Reset value is empty")

        # 모든 포트 수집
        all_ports: List[str] = []
        for it in (mod.get("inputs") or []):
            n = it.get("name")
            if n:
                all_ports.append(n)
        for it in (mod.get("outputs") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)

        # 필수 라벨 위치 찾기
        hs_row, hs_col = _find_cell(ws_vfp, "Hsync Signal")
        vs_row, vs_col = _find_cell(ws_vfp, "Vsync Signal")
        de_row, de_col = _find_cell(ws_vfp, "Data Enable Signal")
        min_row, min_col = _find_cell(ws_vfp, "Expected Min Value")
        max_row, max_col = _find_cell(ws_vfp, "Expected Max Value")

        if (
            hs_row is None
            or vs_row is None
            or de_row is None
            or min_row is None
            or max_row is None
        ):
            print(
                "ERROR: One or more required cells "
                "(Hsync / Vsync / Data Enable / Expected Min / Expected Max) "
                "not found in VFP sheet.",
                flush=True,
            )
            raise ValueError("Missing required VFP labels")

        # 기존 시트 값 읽기 (라벨 바로 아래 칸)
        hs_cell  = ws_vfp.cell(row=hs_row  + 1, column=hs_col).value
        vs_cell  = ws_vfp.cell(row=vs_row  + 1, column=vs_col).value
        de_cell  = ws_vfp.cell(row=de_row  + 1, column=de_col).value
        min_cell = ws_vfp.cell(row=min_row + 1, column=min_col).value
        max_cell = ws_vfp.cell(row=max_row + 1, column=max_col).value

        # 후보 포트 목록
        hs_candidates = [n for n in all_ports if "i_hsync" in n.lower()]
        vs_candidates = [n for n in all_ports if "i_vsync" in n.lower()]
        de_candidates = [n for n in all_ports if "i_de" in n.lower()]

        # Hsync Signal 초기값
        if hs_cell and str(hs_cell).strip():
            hsync_signal = str(hs_cell).strip()
        elif not hs_candidates:
            print("ERROR: No port containing 'i_hsync' found in RTL.", flush=True)
            hsync_signal = "<Hsync Signal>"
            raise ValueError("i_hsync port not found")
        elif len(hs_candidates) == 1:
            hsync_signal = hs_candidates[0]
        else:
            hsync_signal = "<Hsync Signal>"

        # Vsync Signal 초기값
        if vs_cell and str(vs_cell).strip():
            vsync_signal = str(vs_cell).strip()
        elif not vs_candidates:
            print("ERROR: No port containing 'i_vsync' found in RTL.", flush=True)
            vsync_signal = "<Vsync Signal>"
            raise ValueError("i_vsync port not found")
        elif len(vs_candidates) == 1:
            vsync_signal = vs_candidates[0]
        else:
            vsync_signal = "<Vsync Signal>"

        # Data Enable Signal 초기값
        if de_cell and str(de_cell).strip():
            data_enable_signal = str(de_cell).strip()
        elif not de_candidates:
            print("ERROR: No port containing 'i_de' found in RTL.", flush=True)
            data_enable_signal = "<Data Enable Signal>"
            raise ValueError("i_de port not found")
        elif len(de_candidates) == 1:
            data_enable_signal = de_candidates[0]
        else:
            data_enable_signal = "<Data Enable Signal>"

        # Min/Max 초기값 (없으면 placeholder)
        exp_min = (
            str(min_cell).strip()
            if min_cell is not None and str(min_cell).strip()
            else "<Expected Min Value>"
        )
        exp_max = (
            str(max_cell).strip()
            if max_cell is not None and str(max_cell).strip()
            else "<Expected Max Value>"
        )

        # 인터랙티브 편집 루프 (HACT/HFP/HSW/VACT/VBP와 동일 UX)
        while True:
            print("\n==================== VFP Settings ====================")
            print(f"[1] Hsync Signal        : {hsync_signal}")
            print(f"[2] Vsync Signal        : {vsync_signal}")
            print(f"[3] Data Enable Signal  : {data_enable_signal}")
            print(f"[4] Expected Min Value  : {exp_min}")
            print(f"[5] Expected Max Value  : {exp_max}")
            print("=======================================================")
            print("Select item number to edit")
            print("Press Enter twice to confirm all")
            choice = input("> ").strip()

            if choice == "":
                missing = []
                if hsync_signal       in ("", "<Hsync Signal>"):
                    missing.append("[1]")
                if vsync_signal       in ("", "<Vsync Signal>"):
                    missing.append("[2]")
                if data_enable_signal in ("", "<Data Enable Signal>"):
                    missing.append("[3]")
                if exp_min            in ("", "<Expected Min Value>"):
                    missing.append("[4]")
                if exp_max            in ("", "<Expected Max Value>"):
                    missing.append("[5]")

                if missing:
                    print(f"{','.join(missing)} has NOT been entered yet.")
                print("Press Enter again to confirm, or select item number to edit")
                choice = input("> ").strip()
                if choice == "":
                    break

            if choice == "1":
                hsync_signal = _pick_from(
                    hs_candidates,
                    "Select Hsync Signal (matched i_hsync)",
                    allow_custom=False,
                )
            elif choice == "2":
                vsync_signal = _pick_from(
                    vs_candidates,
                    "Select Vsync Signal (matched i_vsync)",
                    allow_custom=False,
                )
            elif choice == "3":
                data_enable_signal = _pick_from(
                    de_candidates,
                    "Select Data Enable Signal (matched i_de)",
                    allow_custom=False,
                )
            elif choice == "4":
                exp_min = _pick_int("Enter Expected Min Value")
            elif choice == "5":
                try:
                    min_val_int = int(exp_min)
                except Exception:
                    min_val_int = 0
                exp_max = _pick_int_with_validation(
                    "Enter Expected Max Value",
                    min_val=min_val_int,
                )
            else:
                print("Invalid selection. Try again.")
                continue

        # 최종 선택 값을 시트에 기록
        ws_vfp.cell(row=hs_row  + 1, column=hs_col,   value=hsync_signal)
        ws_vfp.cell(row=vs_row  + 1, column=vs_col,   value=vsync_signal)
        ws_vfp.cell(row=de_row  + 1, column=de_col,   value=data_enable_signal)
        ws_vfp.cell(row=min_row + 1, column=min_col,  value=exp_min)
        ws_vfp.cell(row=max_row + 1, column=max_col,  value=exp_max)

        # VFP 시트의 Base Clock/Reset 셀에도 Define에서 읽은 값 기록 (수식이 아닌 실제 값)
        clk_row, clk_col = _find_cell(ws_vfp, "Base Clock")
        if clk_row:
            ws_vfp.cell(row=clk_row, column=clk_col + 1, value=base_clk)

        rst_row, rst_col = _find_cell(ws_vfp, "Base Reset")
        if rst_row:
            ws_vfp.cell(row=rst_row, column=rst_col + 1, value=base_rst)

        wb_w.save(xls_path)

        # 포트 너비 추출
        w_clk = _port_width_token(mod, base_clk)
        w_rst = _port_width_token(mod, base_rst)
        w_hs  = _port_width_token(mod, hsync_signal)
        w_vs  = _port_width_token(mod, vsync_signal)
        w_de  = _port_width_token(mod, data_enable_signal)

        blocks = [{
            "Base Clock": base_clk,
            "Base Reset": base_rst,
            "Hsync Signal": hsync_signal,
            "Vsync Signal": vsync_signal,
            "Data Enable Signal": data_enable_signal,
            "Expected Min Value": exp_min,
            "Expected Max Value": exp_max,
            "Base Clock Width": w_clk,
            "Base Reset Width": w_rst,
            "Hsync Signal Width": w_hs,
            "Vsync Signal Width": w_vs,
            "Data Enable Signal Width": w_de,
        }]
        return {"blocks": blocks}

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
        blocks = parsed.get("blocks") or []
        if not blocks:
            return ["// No VFP assertions generated.\n", ""]

        b = blocks[0]
        base_clk = b.get("Base Clock", "") or "clk"
        base_rst = b.get("Base Reset", "") or "rst_n"
        hsync_signal = b.get("Hsync Signal", "") or "i_hsync"
        vsync_signal = b.get("Vsync Signal", "") or "i_vsync"
        data_enable_signal = b.get("Data Enable Signal", "") or "i_de"
        exp_min = b.get("Expected Min Value", "") or "0"
        exp_max = b.get("Expected Max Value", "") or "0"
        w_clk = b.get("Base Clock Width", "[0:0]")
        w_rst = b.get("Base Reset Width", "[0:0]")
        w_hs = b.get("Hsync Signal Width", "[0:0]")
        w_vs = b.get("Vsync Signal Width", "[0:0]")
        w_de = b.get("Data Enable Signal Width", "[0:0]")

        # interface 코드 생성 (포트 없는 interface, 내부에 logic 선언)
        lines: List[str] = []
        lines.append("`include \"uvm_macros.svh\"")
        lines.append("import uvm_pkg::*;")
        lines.append("")
        lines.append("interface assertion_intf();")
        lines.append("")
        lines.append(f"logic {w_clk} {base_clk};")
        lines.append(f"logic {w_rst} {base_rst};")
        lines.append(f"logic {w_hs} {hsync_signal};")
        lines.append(f"logic {w_vs} {vsync_signal};")
        lines.append(f"logic {w_de} {data_enable_signal};")
        lines.append("")
        lines.append("int vfp_hsync_cnt;")
        lines.append("int vfp_de_cnt;")
        lines.append("logic l_vact_vfp;")
        lines.append("")
        lines.append(f"always @(posedge {base_clk} or negedge {base_rst}) begin")
        lines.append(f"    if(!{base_rst}) begin")
        lines.append("        l_vact_vfp <= 0;")
        lines.append("    end")
        lines.append(f"    else if($rose({data_enable_signal})) begin")
        lines.append("        l_vact_vfp <= 1;")
        lines.append("    end")
        lines.append(f"    else if({vsync_signal}) begin")
        lines.append("        l_vact_vfp <= 0;")
        lines.append("    end")
        lines.append("    else begin")
        lines.append("        l_vact_vfp <= l_vact_vfp;")
        lines.append("    end")
        lines.append("end")
        lines.append("")
        lines.append(f"always @(posedge {base_clk} or negedge {base_rst}) begin")
        lines.append(f"    if(!{base_rst}) begin")
        lines.append("        vfp_hsync_cnt <= 0;")
        lines.append("    end")
        lines.append("    else begin")
        lines.append(f"        if(l_vact_vfp && $rose({hsync_signal})) begin")
        lines.append("            vfp_hsync_cnt <= vfp_hsync_cnt + 1;")
        lines.append("        end")
        lines.append(f"        else if($fell({vsync_signal})) begin")
        lines.append("            vfp_hsync_cnt <= 0;")
        lines.append("        end")
        lines.append("        else begin")
        lines.append("            vfp_hsync_cnt <= vfp_hsync_cnt;")
        lines.append("        end")
        lines.append("    end")
        lines.append("end")
        lines.append("")
        lines.append("property p_vfp;")
        lines.append(f"    @(posedge {base_clk}) disable iff (!{base_rst})")
        lines.append(f"    $fell(l_vact_vfp) |=> ({exp_min} <= (vfp_hsync_cnt - vfp_de_cnt) ) && ((vfp_hsync_cnt - vfp_de_cnt) <= {exp_max});")
        lines.append("endproperty")
        lines.append("")
        lines.append("assert property (p_vfp) $display(\"[%0t] Assertion PASS\", $time);")
        lines.append("else $display (\"[%0t] Assertion FAIL\", $time);")
        lines.append("")
        lines.append("endinterface")
        lines.append("")
        sv_text = "\n".join(lines)

        # 인스턴스 파일 생성
        inst_lines: List[str] = []
        inst_lines.append("`include \"uvm_macros.svh\"")
        inst_lines.append("import uvm_pkg::*;")
        inst_lines.append("")
        inst_lines.append("assertion_intf")
        inst_lines.append("u_assertion_intf();")
        inst_lines.append("")
        inst_lines.append(f"assign u_assertion_intf.{base_clk} = top.dut.{base_clk};")
        inst_lines.append(f"assign u_assertion_intf.{base_rst} = top.dut.{base_rst};")
        inst_lines.append(f"assign u_assertion_intf.{hsync_signal} = top.dut.{hsync_signal};")
        inst_lines.append(f"assign u_assertion_intf.{vsync_signal} = top.dut.{vsync_signal};")
        inst_lines.append(f"assign u_assertion_intf.{data_enable_signal} = top.dut.{data_enable_signal};")
        inst_text = "\n".join(inst_lines) + "\n"

        return [sv_text, inst_text]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        return parsed