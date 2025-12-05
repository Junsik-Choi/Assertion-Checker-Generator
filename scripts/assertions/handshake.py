from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Type
import json
import os
import importlib
import re

from openpyxl import load_workbook
from .base import BaseAssertionPlugin
from .registry import register

ALLOWED_TYPES = ("2phase", "4phase", "ready_valid")

def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    print(title, flush=True)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Enter custom")
    while True:
        try:
            s = input("Select > ").strip()
        except EOFError:
            return options[0][1] if options else ""
        if allow_custom and s == "0":
            try:
                return input("Enter value > ").strip()
            except EOFError:
                return ""
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.", flush=True)

def find_cell(ws, value: str) -> Tuple[Optional[int], Optional[int]]:
    tgt = value.strip().lower()
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if str(c.value).strip().lower() == tgt:
                return c.row, c.column
    return None, None

def _get_sheet_ci(wb, want_name: str = "Handshake", create: bool = False):
    target = (want_name or "").strip().lower()
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == target:
            return wb[nm]
    if create:
        return wb.create_sheet(title=want_name)
    raise KeyError(f"Worksheet {want_name} does not exist.")

def _ensure_handshake_layout(ws) -> Tuple[int, int, int]:
    h_row, h_col = find_cell(ws, "Handshake")
    if h_row is None:
        h_row, h_col = 1, 1
        ws.cell(row=h_row, column=h_col, value="Handshake")
        ws.cell(row=h_row + 1, column=h_col, value="Type")
        ws.cell(row=h_row + 1, column=h_col + 1, value="Sender")
        ws.cell(row=h_row + 1, column=h_col + 2, value="Receiver")
        ws.cell(row=h_row + 3, column=h_col, value="Base Clock")
        ws.cell(row=h_row + 4, column=h_col, value="Base Reset")
    min_col = h_col
    for mr in ws.merged_cells.ranges:
        if ws.cell(row=h_row, column=h_col).coordinate in mr:
            min_col = mr.min_col
            break
    type_row = h_row + 1
    type_col = min_col
    if not ws.cell(row=type_row, column=type_col).value:
        ws.cell(row=type_row, column=type_col, value="Type")
    if not ws.cell(row=type_row, column=type_col + 1).value:
        ws.cell(row=type_row, column=type_col + 1, value="Sender")
    if not ws.cell(row=type_row, column=type_col + 2).value:
        ws.cell(row=type_row, column=type_col + 2, value="Receiver")
    data_row = type_row + 1
    return h_row, type_col, data_row

def parse_handshake_block_for_row(ws, row: int, type_col: int) -> Dict[str, Any]:
    clk_row, clk_col = find_cell(ws, "Base Clock")
    rst_row, rst_col = find_cell(ws, "Base Reset")
    base_clk = ws.cell(row=clk_row, column=clk_col + 1).value if clk_row else None
    base_rst = ws.cell(row=rst_row, column=rst_col + 1).value if rst_row else None
    phase_val = ws.cell(row=row, column=type_col).value
    phase_type = str(phase_val).strip().lower() if phase_val is not None else ""
    sender = ws.cell(row=row, column=type_col + 1).value
    receiver = ws.cell(row=row, column=type_col + 2).value
    return {
        "phase_type": phase_type or "2phase",
        "Base Clock": str(base_clk).strip() if base_clk else "",
        "Reset": str(base_rst).strip() if base_rst else "",
        "Sender": str(sender).strip() if sender else "",
        "Receiver": str(receiver).strip() if receiver else "",
    }

def _auto_pick_clk_rst(mod: Dict[str, Any]) -> Tuple[str, str]:
    clk = ""
    rst = ""
    clocks = mod.get("clocks") or []
    resets = mod.get("resets") or []
    if clocks:
        clk = clocks[0].get("name") or ""
    if resets:
        rst = resets[0].get("name") or ""
    if not clk:
        for it in mod.get("inputs", []):
            n = (it.get("name") or "").lower()
            if "clk" in n or n.endswith("clock"):
                clk = it.get("name") or ""
                break
    if not rst:
        for it in mod.get("inputs", []):
            n = (it.get("name") or "").lower()
            if "rst" in n or "reset" in n:
                rst = it.get("name") or ""
                break
    return clk, rst

# -----------------------------
# Width helpers
# -----------------------------
def _normalize_range_token(token: str) -> str:
    """Return normalized [msb:lsb] token; default to [0:0] for 1-bit/unknown."""
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    # numeric like '32' or '1'
    try:
        n = int(t, 10)
        if n >= 1:
            return f"[{n-1}:0]"
        return "[0:0]"
    except Exception:
        # unknown format -> treat as 1-bit
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    """Find port by name and return width token '[msb:lsb]' (defaults to [0:0])."""
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    # 1) search a broader 'ports' list first if present
    candidates = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in candidates:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            # Typical fields
            for key in ("packed_range", "range", "packed", "decl"):
                pr = it.get(key)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            # Width as integer/string
            for key in ("width", "bit_width", "width_bits"):
                w = it.get(key)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            # Separate ends
            for left_key, right_key in (("msb", "lsb"), ("left", "right")):
                msb = it.get(left_key)
                lsb = it.get(right_key)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            # Fallback 1-bit
            return "[0:0]"
    return "[0:0]"

def _fmt_input_decl(sig: str, width_tok: str) -> str:
    """Format input logic with width; defaults to [0:0] if empty."""
    width_tok = (width_tok or "").strip() or "[0:0]"
    return f"input logic {width_tok} {sig}"

def _update_handshake_sheet(ws, hs_cfg: Dict[str, str], module_info: Dict[str, Any]) -> int:
    """
    Handshake 시트에 값을 기록하고, 기록한 '행 번호'를 반환한다.
    기존 데이터 아래의 첫 빈 행(Type 셀이 비어있는 곳)에 기록.
    """
    h_row, type_col, data_row = _ensure_handshake_layout(ws)
    # 아래로 내려가며 Type 셀이 비어있는 첫 행을 찾음
    write_row = data_row
    while True:
        val = ws.cell(row=write_row, column=type_col).value
        if val is None or str(val).strip() == "":
            break
        write_row += 1
    # 선택 정보 기록
    ws.cell(row=write_row, column=type_col,     value=hs_cfg.get("phase_type", "2phase"))
    ws.cell(row=write_row, column=type_col + 1, value=(hs_cfg.get("sender", "") or "").strip())
    ws.cell(row=write_row, column=type_col + 2, value=(hs_cfg.get("receiver", "") or "").strip())

    # Base Clock / Base Reset 라벨 및 값 정리
    clk_label = find_cell(ws, "Base Clock")
    rst_label = find_cell(ws, "Base Reset")
    if clk_label[0] is None:
        ws.cell(row=h_row + 3, column=type_col, value="Base Clock"); clk_label = (h_row + 3, type_col)
    if rst_label[0] is None:
        ws.cell(row=h_row + 4, column=type_col, value="Base Reset"); rst_label = (h_row + 4, type_col)
    clk_name, rst_name = _auto_pick_clk_rst(module_info)
    if clk_name:
        ws.cell(row=clk_label[0], column=clk_label[1] + 1, value=clk_name)
    if rst_name:
        ws.cell(row=rst_label[0], column=rst_label[1] + 1, value=rst_name)
    return write_row

def generate_verilog(info: Dict[str, Any]) -> str:
    clk = info["Base Clock"]; rst = info["Reset"]
    s = info["Sender"]; r = info["Receiver"]
    # width tokens (always emitted; 1-bit => [0:0])
    w_clk = info.get("Base Clock Width", "") or "[0:0]"
    w_rst = info.get("Reset Width", "") or "[0:0]"
    w_s   = info.get("Sender Width", "") or "[0:0]"
    w_r   = info.get("Receiver Width", "") or "[0:0]"
    header = '`include "uvm_macros.svh"\nimport uvm_pkg::*;\n\n'
    # ready_valid
    if info["phase_type"] == "ready_valid":
        return header + f"""interface assertion_intf();

logic {w_clk} {clk};
logic {w_rst} {rst};
logic {w_s} {s};
logic {w_r} {r};

property p_ready_valid_check(ready, valid);
    @(posedge {clk}) disable iff(!{rst})
    valid && !ready |-> ##[1:$] (ready || (valid && !ready));
endproperty

assert property (p_ready_valid_check({s}, {r})) else `uvm_error("ASSERTION", "Ready-Valid check failed")

endinterface
"""
    if info["phase_type"] == "4phase":
        return header + f"""interface assertion_intf();

logic {w_clk} {clk};
logic {w_rst} {rst};
logic {w_s} {s};
logic {w_r} {r};

property p_4ph_check_0(req,ack);
    @(posedge {clk}) disable iff(!{rst})
    (~req & ~ack) |-> ##1 ((req & ~ack) or (req & ack) or (~req & ~ack));
endproperty

property p_4ph_check_1(req,ack);
    @(posedge {clk}) disable iff(!{rst})
    (req & ~ack) |-> ##1 ((req & ack) or (req & ~ack));
endproperty

property p_4ph_check_2(req,ack);
    @(posedge {clk}) disable iff(!{rst})
    (req & ~ack) |-> ##1 ((req & ~ack) or (~req & ~ack) or (req & ack));
endproperty

assert property (p_4ph_check_0({s}, {r})) else `uvm_error("ASSERTION", "4-phase check 0 failed")
assert property (p_4ph_check_1({s}, {r})) else `uvm_error("ASSERTION", "4-phase check 1 failed")
assert property (p_4ph_check_2({s}, {r})) else `uvm_error("ASSERTION", "4-phase check 2 failed")

endinterface
"""
    else:
        return header + f"""interface assertion_intf();

logic {w_clk} {clk};
logic {w_rst} {rst};
logic {w_s} {s};
logic {w_r} {r};

property p_2phase_check_0(req, ack);
  @(posedge {clk}) disable iff (!{rst})
  (~req & ~ack) |-> ##1 ((req & ~ack) or (req & ack) or (~req & ~ack));
endproperty

property p_2phase_check_1(req, ack);
  @(posedge {clk}) disable iff (!{rst})
  (~req & ack) |-> ##1 ((~req & ~ack) or (~req & ack));
endproperty

property p_2phase_check_2(req, ack);
  @(posedge {clk}) disable iff (!{rst})
  (req & ~ack) |-> ##1 ((req & ack) or (req & ~ack));
endproperty

property p_2phase_check_3(req, ack);
  @(posedge {clk}) disable iff (!{rst})
  (req & ack) |-> ##1 ((~req & ack) or (~req & ~ack) or (req & ack));
endproperty

assert property (p_2phase_check_0({s}, {r})) else `uvm_error("ASSERTION", "2-phase check 0 failed")
assert property (p_2phase_check_1({s}, {r})) else `uvm_error("ASSERTION", "2-phase check 1 failed")
assert property (p_2phase_check_2({s}, {r})) else `uvm_error("ASSERTION", "2-phase check 2 failed")
assert property (p_2phase_check_3({s}, {r})) else `uvm_error("ASSERTION", "2-phase check 3 failed")

endinterface
"""

def generate_inst_verilog(info: Dict[str, Any]) -> str:
    clk = info["Base Clock"]; rst = info["Reset"]
    s = info["Sender"]; r = info["Receiver"]
    header = '`include "uvm_macros.svh"\nimport uvm_pkg::*;\n\n'
    # inst 파일은 모듈 래퍼 없이 인스턴스와 assign만 생성
    return header + (
        f"assertion_intf u_assertion_intf();\n\n"
        f"assign u_assertion_intf.{clk} = top.dut.{clk};\n"
        f"assign u_assertion_intf.{rst} = top.dut.{rst};\n"
        f"assign u_assertion_intf.{s} = top.dut.{s};\n"
        f"assign u_assertion_intf.{r} = top.dut.{r};\n"
    )

def _get_forced_type() -> Optional[str]:
    t = (os.environ.get("ASSERTION_FORCE_TYPE") or "").strip().lower()
    return t if t in ALLOWED_TYPES else None

@register
class HandshakePlugin(BaseAssertionPlugin):
    plugin_name = "handshake"
    sheet_name = "Handshake"

    def _load_module_define(self, xls_path: Path) -> Dict[str, Any]:
        try:
            md = xls_path.parent / "module_define.json"
            if md.exists():
                return json.loads(md.read_text(encoding="utf-8"))
            # fallback: assertion_inputs.json (same schema subset)
            ai = xls_path.parent / "assertion_inputs.json"
            if ai.exists():
                return json.loads(ai.read_text(encoding="utf-8"))
        except Exception:
            pass
        return {}

    def _interactive_collect(self, mod: Dict[str, Any]) -> Dict[str, str]:
        # 빌더에서 이미 타입을 선택했으므로 재질문하지 않음
        phase = _get_forced_type() or "2phase"
        in_names = [(f"in : {p.get('name')}", p.get("name") or "") for p in (mod.get("inputs") or []) if p.get("name")]
        out_names = [(f"out: {p.get('name')}", p.get("name") or "") for p in (mod.get("outputs") or []) if p.get("name")]
        sig_opts = in_names + out_names or [("manual input", "")]
        sender = _pick_one("Select Sender signal", sig_opts, allow_custom=True)
        receiver = _pick_one("Select Receiver signal", sig_opts, allow_custom=True)
        return {"phase_type": phase or "2phase", "sender": sender, "receiver": receiver}

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        # 1) 플러그인 선택 직후: 타입/신호 선택 프롬프트
        mod = self._load_module_define(Path(xls_path))
        hs_cfg = self._interactive_collect(mod)
        # 2) Excel에 기록
        wb_w = load_workbook(xls_path)
        try:
            ws_w = _get_sheet_ci(wb_w, self.sheet_name, create=False)
        except KeyError:
            ws_w = _get_sheet_ci(wb_w, self.sheet_name, create=True)
        write_row = _update_handshake_sheet(ws_w, hs_cfg, mod)
        wb_w.save(xls_path)
        # 3) data_only로 재오픈하여 '방금 기록한 행'만 파싱
        wb = load_workbook(xls_path, data_only=True)
        try:
            ws = _get_sheet_ci(wb, self.sheet_name, create=False)
        except KeyError:
            return {"blocks": []}
        _, type_col, _ = _ensure_handshake_layout(ws)
        info = parse_handshake_block_for_row(ws, write_row, type_col)
        # attach bit widths from module define
        info["Base Clock Width"] = _port_width_token(mod, info.get("Base Clock", ""))
        info["Reset Width"] = _port_width_token(mod, info.get("Reset", ""))
        info["Sender Width"] = _port_width_token(mod, info.get("Sender", ""))
        info["Receiver Width"] = _port_width_token(mod, info.get("Receiver", ""))
        blocks: List[Dict[str, Any]] = []
        pt = (info.get("phase_type", "") or "").lower()
        if pt in ALLOWED_TYPES and info.get("Sender") and info.get("Receiver"):
            blocks.append(info)

        # 선택된 타입만 남기기
        forced = _get_forced_type()
        if forced:
            blocks = [b for b in blocks if (b.get("phase_type", "").lower() == forced)]
        return {"blocks": blocks}

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
        # 파일 저장은 assertion_builder.py에서 처리. 여기서는 문자열만 반환.
        sv_parts: List[str] = []
        inst_parts: List[str] = []
        forced = _get_forced_type()
        for info in parsed.get("blocks", []):
            if forced and (info.get("phase_type", "").lower() != forced):
                continue
            sv_parts.append(generate_verilog(info))
            inst_parts.append(generate_inst_verilog(info))
        combined_sv = "\n\n".join([s.strip() for s in sv_parts if str(s).strip()]) + ("\n" if sv_parts else "")
        combined_inst = "\n\n".join([s.strip() for s in inst_parts if str(s).strip()]) + ("\n" if inst_parts else "")
        return [combined_sv, combined_inst]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        return parsed
    
    @classmethod
    def write_to_excel(cls, excel_path: Path, data: Dict[str, Any], state: Optional[Any] = None) -> None:
        """Write handshake assertion data to Excel sheet."""
        from openpyxl import load_workbook  # type: ignore
        
        wb = load_workbook(str(excel_path))
        
        # Find Handshake sheet
        sheet_name = cls.find_sheet_case_insensitive(wb.sheetnames, 'Handshake')
        if not sheet_name:
            sheet_name = 'Handshake'
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
        
        ws = wb[sheet_name]
        
        # Find next empty row (Handshake sheet: data starts at row 7)
        next_row = 7
        while ws.cell(row=next_row, column=3).value:
            next_row += 1
        
        # Handshake sheet columns (from row 6): col3=Type, col4=Sender, col5=Receiver
        ws.cell(row=next_row, column=3, value=data.get('phase_type', ''))
        ws.cell(row=next_row, column=4, value=data.get('sender', ''))
        ws.cell(row=next_row, column=5, value=data.get('receiver', ''))
        
        wb.save(str(excel_path))
        wb.close()

# Optional: standalone script entry point for legacy usage
def main(xlsx_path: str) -> None:
    # 간단 실행 테스트: 시트 준비만 수행
    wb = load_workbook(xlsx_path)
    try:
        ws = _get_sheet_ci(wb, "Handshake", create=False)
    except KeyError:
        ws = _get_sheet_ci(wb, "Handshake", create=True)
    _ensure_handshake_layout(ws)
    wb.save(xlsx_path)

