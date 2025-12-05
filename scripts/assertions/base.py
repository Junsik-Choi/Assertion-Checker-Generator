from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional


class BaseAssertionPlugin:
    """
    Base interface for assertion generation plugins.

    Required class attributes:
    - plugin_name: short unique name (e.g., "counter").
    - sheet_name: Excel sheet this plugin expects (e.g., "counter_gen").

    Required methods:
    - parse(xls_path: Path) -> Dict[str, Any]
      Read and validate plugin-specific sheet, return structured data.
    - generate_sv(parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]
      Produce a list of SystemVerilog sections (strings) for this plugin.

    Context keys (common across plugins):
    - module_info: Dict with module name, clocks/resets, ports, parameters
    - define_excel_path: target Excel path for define sheet filling
    - output_dir: directory to write results
    - config: global CLI/config settings
    """

    plugin_name: str = "base"
    sheet_name: str = ""

    @staticmethod
    def find_sheet_case_insensitive(sheet_names: List[str], target_name: str) -> Optional[str]:
        """
        Find a sheet name case-insensitively.
        Returns the actual sheet name if found, None otherwise.
        """
        target_lower = target_name.lower()
        for name in sheet_names:
            if name.lower() == target_lower:
                return name
        return None

    def parse(self, xls_path: Path) -> Dict[str, Any]:  # pragma: no cover
        raise NotImplementedError

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:  # pragma: no cover
        raise NotImplementedError

    # Optional: per-plugin JSON to emit alongside SV
    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        return None
    
    @classmethod
    def write_to_excel(cls, excel_path: Path, data: Dict[str, Any], state: Optional[Any] = None) -> None:
        """
        Write assertion data to the corresponding Excel sheet.
        Default implementation: Generic write to first available columns.
        Subclasses should override for plugin-specific column mapping.
        
        Args:
            excel_path: Path to Excel file
            data: Dict of field_name -> value pairs
            state: Optional AppState for additional context
        """
        try:
            from openpyxl import load_workbook  # type: ignore
            from openpyxl.cell import MergedCell
            
            wb = load_workbook(str(excel_path))
            
            # Find sheet case-insensitively
            sheet_name = cls.find_sheet_case_insensitive(wb.sheetnames, cls.sheet_name)
            if not sheet_name:
                sheet_name = cls.sheet_name
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
            
            ws = wb[sheet_name]
            
            # Find next empty row starting from row 8
            next_row = 8
            while ws.cell(row=next_row, column=2).value:
                next_row += 1
            
            # Write all data fields starting from column 2
            col = 2
            for key, value in data.items():
                if value:  # Only write non-empty values
                    ws.cell(row=next_row, column=col, value=value)
                    col += 1
            
            wb.save(str(excel_path))
            wb.close()
        except Exception as e:
            raise RuntimeError(f"Failed to write {cls.plugin_name} to Excel: {str(e)}")


