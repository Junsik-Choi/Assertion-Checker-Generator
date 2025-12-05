from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json

from openpyxl import load_workbook

from .registry import register
from .base import BaseAssertionPlugin

# ===== 공통 유틸리티 함수 =====
def _find_cell(ws, value: str) -> Tuple[Optional[int], Optional[int]]:
    """시트에서 특정 값을 가진 셀의 위치(row, column)를 찾기"""
    tgt = (value or "").strip().lower()
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if str(v).strip().lower() == tgt:
                return c.row, c.column
    return None, None

def _get_sheet_ci(wb, want_name: str):
    """대소문자 구분 없이 워크시트 찾기"""
    target = (want_name or "").strip().lower()
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == target:
            return wb[nm]
    raise KeyError(f"Worksheet '{want_name}' does not exist.")

def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    """프롬프트로 옵션 선택 또는 커스텀 입력"""
    print(title, flush=True)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Enter custom")
    while True:
        try:
            s = input("Select > ").strip()
        except EOFError:
            return options[0][1] if options else ""
        if allow_custom and s == "0":
            try:
                return input("Enter value > ").strip()
            except EOFError:
                return ""
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.", flush=True)

def _read_json(p: Path) -> Optional[Dict[str, Any]]:
    """JSON 파일 읽기"""
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None

def _load_module_define(xls_path: Path) -> Dict[str, Any]:
    """module_define.json 또는 assertion_inputs.json에서 RTL 정보 로드"""
    session_dir = xls_path.parent
    md = _read_json(session_dir / "module_define.json")
    if md:
        return md
    ai = _read_json(session_dir / "assertion_inputs.json")
    if ai:
        return {
            "module": ai.get("module") or "",
            "clocks": ai.get("clocks") or [],
            "resets": ai.get("resets") or [],
            "inputs": ai.get("inputs") or [],
            "outputs": ai.get("outputs") or [],
            "inouts": ai.get("inouts") or [],
            "parameters": ai.get("parameters") or [],
        }
    return {}

def _normalize_range_token(token: Any) -> str:
    """포트 width를 [msb:lsb] 형식으로 정규화"""
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    try:
        n = int(t, 10)
        return f"[{n-1}:0]" if n >= 1 else "[0:0]"
    except Exception:
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    """포트 이름으로 width 토큰 찾기"""
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    candidates = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in candidates:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            # packed_range, range 등에서 width 정보 찾기
            for key in ("packed_range", "range", "packed", "decl"):
                pr = it.get(key)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            for key in ("width", "bit_width", "width_bits"):
                w = it.get(key)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            for lk, rk in (("msb", "lsb"), ("left", "right")):
                msb = it.get(lk); lsb = it.get(rk)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            return "[0:0]"
    return "[0:0]"

def _read_define_clk_rst(wb) -> Tuple[str, str]:
    """Define 시트에서 Base Clock/Reset 읽기"""
    try:
        ws = _get_sheet_ci(wb, "Define")
    except KeyError:
        return "", ""
    clk_r, clk_c = _find_cell(ws, "Base Clock")
    rst_r, rst_c = _find_cell(ws, "Base Reset")
    clk = ws.cell(row=clk_r, column=clk_c + 1).value if clk_r else None
    rst = ws.cell(row=rst_r, column=rst_c + 1).value if rst_r else None
    return (str(clk).strip() if clk else ""), (str(rst).strip() if rst else "")

def _pick_from(names: List[str], title: str, allow_custom: bool = False) -> str:
    """리스트에서 선택 또는 커스텀 입력"""
    opts = [(n, n) for n in names] if names else []
    return _pick_one(title, opts, allow_custom=allow_custom)

# ===== Pulse Width 플러그인 =====
@register
class PulseWidthPlugin(BaseAssertionPlugin):
    plugin_name = "pulseWidth"
    sheet_name = "pulseWidth"

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        """
        Excel 파일 파싱 및 사용자 입력 처리
        1. pulseWidth 시트 확인
        2. Base Clock/Reset 읽기 (Define 시트에서)
        3. Pulse Signal, Width Parameter, Disable Condition 입력받기
        """
        mod = _load_module_define(Path(xls_path))
        wb = load_workbook(xls_path)

        # 1. pulseWidth 시트 확인
        try:
            ws = _get_sheet_ci(wb, self.sheet_name)
        except KeyError:
            print(f"ERROR: '{self.sheet_name}' sheet does not exist in the Excel file.", flush=True)
            raise

        # 2. Base Clock/Reset은 Define 시트에서 읽기
        base_clk, base_rst = _read_define_clk_rst(wb)
        if not base_clk:
            print("ERROR: Base Clock value is empty in Define sheet.", flush=True)
            raise ValueError("Base Clock value is empty")
        if not base_rst:
            print("ERROR: Base Reset value is empty in Define sheet.", flush=True)
            raise ValueError("Base Reset value is empty")

        # 3. 모든 포트 수집 (입력/출력/inout)
        all_ports: List[str] = []
        for it in (mod.get("inputs") or []):
            n = it.get("name")
            if n:
                all_ports.append(n)
        for it in (mod.get("outputs") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)
        for it in (mod.get("inouts") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)

        # 4. Pulse / Width / Disable 셀 위치 찾기
        pulse_r, pulse_c   = _find_cell(ws, "Pulse Signal")
        width_r, width_c   = _find_cell(ws, "Width Parameter")
        disable_r, disable_c = _find_cell(ws, "Disable Condition")

        if not pulse_r or not width_r or not disable_r:
            print("ERROR: One or more required cells (Pulse Signal / Width Parameter / Disable Condition) "
                  "not found in pulseWidth sheet.", flush=True)
            raise ValueError("Missing required pulseWidth labels")

        # 5. 셀 값 읽어서 플레이스홀더 세팅
        pulse_cell   = ws.cell(row=pulse_r + 1, column=pulse_c).value
        width_cell   = ws.cell(row=width_r + 1, column=width_c).value
        disable_cell = ws.cell(row=disable_r + 1, column=disable_c).value

        pulse_sig    = str(pulse_cell).strip()   if pulse_cell   and str(pulse_cell).strip()   else "<Pulse Signal>"
        width_param  = str(width_cell).strip()   if width_cell   and str(width_cell).strip()   else "<Width Parameter>"
        disable_cond = str(disable_cell).strip() if disable_cell and str(disable_cell).strip() else "<Disable Condition>"

        # 6. 한 화면에서 수정 + Enter 두 번으로 확정
        while True:
            print("\n==================== Pulse Width Settings ====================")
            print(f"[1] Pulse Signal      : {pulse_sig}")
            print(f"[2] Width Parameter   : {width_param}")
            print(f"[3] Disable Condition : {disable_cond}")
            print("============================================================")
            print("Select item number to edit")
            print("Press Enter twice to confirm all")
            choice = input("> ").strip()
            if choice == "":
                missing = []
                if pulse_sig    in ("", "<Pulse Signal>"):       missing.append("[1]")
                if width_param  in ("", "<Width Parameter>"):    missing.append("[2]")
                if disable_cond in ("", "<Disable Condition>"):  missing.append("[3]")
                if missing:
                    print(f"{','.join(missing)} has NOT been entered yet.")
                print("Press Enter again to confirm, or select item number to edit")
                choice = input("> ").strip()
                if choice == "":
                    break
            if choice == "1":
                pulse_sig = _pick_from(all_ports, "Select Pulse Signal:", allow_custom=True)
            elif choice == "2":
                width_param = _pick_from(all_ports, "Select Width Parameter:", allow_custom=True)
            elif choice == "3":
                disable_cond = _pick_from(all_ports, "Select Disable Condition:", allow_custom=True)
            else:
                print("Invalid selection. Try again.")
                continue

        # 7. 최종값을 시트에 기록
        ws.cell(row=pulse_r   + 1, column=pulse_c,   value=pulse_sig)
        ws.cell(row=width_r   + 1, column=width_c,   value=width_param)
        ws.cell(row=disable_r + 1, column=disable_c, value=disable_cond)

        # 8. pulseWidth 시트의 Base Clock/Reset 셀에도 값 기록
        clk_row, clk_col = _find_cell(ws, "Base Clock")
        if clk_row:
            ws.cell(row=clk_row, column=clk_col + 1, value=base_clk)

        rst_row, rst_col = _find_cell(ws, "Base Reset")
        if rst_row:
            ws.cell(row=rst_row, column=rst_col + 1, value=base_rst)

        # 9. Excel 저장
        wb.save(xls_path)

        # 10. Width 정보 수집
        pulse_width  = _port_width_token(mod, pulse_sig)
        disable_width = _port_width_token(mod, disable_cond)
        clk_width    = _port_width_token(mod, base_clk)
        rst_width    = _port_width_token(mod, base_rst)

        # 11. 결과 반환
        blocks = [{
            "Base Clock": base_clk,
            "Base Reset": base_rst,
            "Pulse Signal": pulse_sig,
            "Width Parameter": width_param,
            "Disable Condition": disable_cond,
            "Base Clock Width": clk_width,
            "Base Reset Width": rst_width,
            "Pulse Signal Width": pulse_width,
            "Disable Condition Width": disable_width,
        }]
        return {"blocks": blocks}


    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
        """
        SystemVerilog assertion 코드 생성
        - assertion_intf.sv: interface 정의
        - assertion_intf_inst.sv: 인스턴스 및 연결
        """
        blocks = parsed.get("blocks") or []
        if not blocks:
            return ["// No Pulse Width assertions generated.\n", ""]

        b = blocks[0]
        base_clk = b.get("Base Clock", "") or "clk"
        base_rst = b.get("Base Reset", "") or "rst_n"
        pulse_sig = b.get("Pulse Signal", "") or "pulse"
        width_param = b.get("Width Parameter", "") or "8"
        disable_cond = b.get("Disable Condition", "") or "1'b0"

        clk_w = b.get("Base Clock Width", "[0:0]")
        rst_w = b.get("Base Reset Width", "[0:0]")
        pulse_w = b.get("Pulse Signal Width", "[0:0]")
        disable_w = b.get("Disable Condition Width", "[0:0]")

        # ===== assertion_intf.sv 생성 =====
        lines: List[str] = []
        lines.append("`include \"uvm_macros.svh\"")
        lines.append("import uvm_pkg::*;")
        lines.append("")
        lines.append("interface assertion_intf();")
        lines.append("")
        lines.append(f"logic {clk_w} {base_clk};")
        lines.append(f"logic {rst_w} {base_rst};")
        lines.append(f"logic {pulse_w} {pulse_sig};")
        lines.append(f"logic {disable_w} {disable_cond};")
        lines.append("")
        lines.append(f"parameter WIDTH_PARAM = {width_param};")
        lines.append("")
        lines.append("// It primarily checks High Width. For Low Width checking, substitute $fell to $rose, and $rose to $fell")
        lines.append("property p_pulseWidth;")
        lines.append(f"    @(posedge {base_clk}) disable iff(!{base_rst} || {disable_cond})")
        lines.append(f"    $rose({pulse_sig})")
        lines.append(f"    |-> {pulse_sig}[*(WIDTH_PARAM)]")
        lines.append(f"    ##(WIDTH_PARAM) $fell({pulse_sig});")
        lines.append("endproperty")
        lines.append("")
        lines.append("assert property (p_pulseWidth)")
        lines.append("    $display(\"[%0t] Pulse Width Assertion PASS\", $time);")
        lines.append("else")
        lines.append("    $display(\"[%0t] Pulse Width Assertion FAIL\", $time);")
        lines.append("")
        lines.append("endinterface")
        lines.append("")
        sv_text = "\n".join(lines)

        # ===== assertion_intf_inst.sv 생성 =====
        inst_lines: List[str] = []
        inst_lines.append("`include \"uvm_macros.svh\"")
        inst_lines.append("import uvm_pkg::*;")
        inst_lines.append("")
        inst_lines.append("assertion_intf u_assertion_intf();")
        inst_lines.append("")
        inst_lines.append(f"assign u_assertion_intf.{base_clk} = top.dut.{base_clk};")
        inst_lines.append(f"assign u_assertion_intf.{base_rst} = top.dut.{base_rst};")
        inst_lines.append(f"assign u_assertion_intf.{pulse_sig} = top.dut.{pulse_sig};")
        inst_lines.append(f"assign u_assertion_intf.{disable_cond} = top.dut.{disable_cond};")
        inst_text = "\n".join(inst_lines) + "\n"

        return [sv_text, inst_text]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        """JSON 출력 (필요시)"""
        return parsed
    
    @classmethod
    def write_to_excel(cls, excel_path: Path, data: Dict[str, Any], state: Optional[Any] = None) -> None:
        """Write pulseWidth assertion data to Excel sheet."""
        from openpyxl import load_workbook  # type: ignore
        
        wb = load_workbook(str(excel_path))
        
        # Find pulseWidth sheet
        sheet_name = cls.find_sheet_case_insensitive(wb.sheetnames, 'pulseWidth')
        if not sheet_name:
            sheet_name = 'pulseWidth'
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
        
        ws = wb[sheet_name]
        
        # Find next empty row (pulseWidth sheet: data starts at row 7)
        next_row = 7
        while ws.cell(row=next_row, column=3).value:
            next_row += 1
        
        # Determine pulse type and count trigger
        pulse_type = data.get('pulse_type', 'hpulse')
        if pulse_type == 'hpulse':
            count_trigger = data.get('base_clock', '')
        else:
            count_trigger = data.get('trigger_signal', '')
        
        # pulseWidth sheet columns (from row 6): col3=Type, col4=Count_Trigger, col5=Target_Pulse, col6=Min, col7=Max
        ws.cell(row=next_row, column=3, value=pulse_type)
        ws.cell(row=next_row, column=4, value=count_trigger)
        ws.cell(row=next_row, column=5, value=data.get('target_signal', ''))
        ws.cell(row=next_row, column=6, value=data.get('min_width', ''))
        ws.cell(row=next_row, column=7, value=data.get('max_width', ''))
        
        wb.save(str(excel_path))
        wb.close()
