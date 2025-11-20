from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json

from openpyxl import load_workbook

from .registry import register
from .base import BaseAssertionPlugin

# ===== 공통 유틸리티 함수 =====
def _find_cell(ws, value: str) -> Tuple[Optional[int], Optional[int]]:
    """시트에서 특정 값을 가진 셀의 위치(row, column)를 찾기"""
    tgt = (value or "").strip().lower()
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if str(v).strip().lower() == tgt:
                return c.row, c.column
    return None, None

def _get_sheet_ci(wb, want_name: str):
    """대소문자 구분 없이 워크시트 찾기"""
    target = (want_name or "").strip().lower()
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == target:
            return wb[nm]
    raise KeyError(f"Worksheet '{want_name}' does not exist.")

def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    """프롬프트로 옵션 선택 또는 커스텀 입력"""
    print(title, flush=True)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Enter custom")
    while True:
        try:
            s = input("Select > ").strip()
        except EOFError:
            return options[0][1] if options else ""
        if allow_custom and s == "0":
            try:
                return input("Enter value > ").strip()
            except EOFError:
                return ""
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.", flush=True)

def _read_json(p: Path) -> Optional[Dict[str, Any]]:
    """JSON 파일 읽기"""
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None

def _load_module_define(xls_path: Path) -> Dict[str, Any]:
    """module_define.json 또는 assertion_inputs.json에서 RTL 정보 로드"""
    session_dir = xls_path.parent
    md = _read_json(session_dir / "module_define.json")
    if md:
        return md
    ai = _read_json(session_dir / "assertion_inputs.json")
    if ai:
        return {
            "module": ai.get("module") or "",
            "clocks": ai.get("clocks") or [],
            "resets": ai.get("resets") or [],
            "inputs": ai.get("inputs") or [],
            "outputs": ai.get("outputs") or [],
            "inouts": ai.get("inouts") or [],
            "parameters": ai.get("parameters") or [],
        }
    return {}

def _normalize_range_token(token: Any) -> str:
    """포트 width를 [msb:lsb] 형식으로 정규화"""
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    try:
        n = int(t, 10)
        return f"[{n-1}:0]" if n >= 1 else "[0:0]"
    except Exception:
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    """포트 이름으로 width 토큰 찾기"""
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    candidates = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in candidates:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            # packed_range, range 등에서 width 정보 찾기
            for key in ("packed_range", "range", "packed", "decl"):
                pr = it.get(key)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            for key in ("width", "bit_width", "width_bits"):
                w = it.get(key)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            for lk, rk in (("msb", "lsb"), ("left", "right")):
                msb = it.get(lk); lsb = it.get(rk)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            return "[0:0]"
    return "[0:0]"

def _pick_from(names: List[str], title: str, allow_custom: bool = False) -> str:
    """리스트에서 선택 또는 커스텀 입력"""
    opts = [(n, n) for n in names] if names else []
    return _pick_one(title, opts, allow_custom=allow_custom)

# ===== AHB Slave 플러그인 =====
@register
class AHB_SPlugin(BaseAssertionPlugin):
    plugin_name = "AHB_S"
    sheet_name = "AHB_S"

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        """
        Excel 파일 파싱 및 사용자 입력 처리
        1. AHB_S 시트 확인
        2. Base Clock/Reset 읽기
        3. AHB 신호들 입력받기
        """
        mod = _load_module_define(Path(xls_path))
        wb = load_workbook(xls_path)

        # 1. AHB_S 시트 확인
        try:
            ws = _get_sheet_ci(wb, self.sheet_name)
        except KeyError:
            print(f"ERROR: '{self.sheet_name}' sheet does not exist in the Excel file.", flush=True)
            raise

        # 2. 모든 포트 수집 (입력/출력/inout)
        all_ports: List[str] = []
        for it in (mod.get("inputs") or []):
            n = it.get("name")
            if n:
                all_ports.append(n)
        for it in (mod.get("outputs") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)
        for it in (mod.get("inouts") or []):
            n = it.get("name")
            if n and n not in all_ports:
                all_ports.append(n)

        # 3. Base Clock 확인 및 입력
        clk_r, clk_c = _find_cell(ws, "Base Clock")
        if not clk_r:
            print("ERROR: 'Base Clock' cell not found in AHB_S sheet.", flush=True)
            raise ValueError("'Base Clock' cell not found")
        
        base_clk = ws.cell(row=clk_r, column=clk_c + 1).value
        if not base_clk or str(base_clk).strip() == "":
            print(f"ERROR: Base Clock value is empty in AHB_S sheet.", flush=True)
            raise ValueError("Base Clock value is empty")
        base_clk = str(base_clk).strip()

        # 4. Base Reset 확인 및 입력
        rst_r, rst_c = _find_cell(ws, "Base Reset")
        if not rst_r:
            print("ERROR: 'Base Reset' cell not found in AHB_S sheet.", flush=True)
            raise ValueError("'Base Reset' cell not found")
        
        base_rst = ws.cell(row=rst_r, column=rst_c + 1).value
        if not base_rst or str(base_rst).strip() == "":
            print(f"ERROR: Base Reset value is empty in AHB_S sheet.", flush=True)
            raise ValueError("Base Reset value is empty")
        base_rst = str(base_rst).strip()

        # AHB 신호들 정의 (Slave용)
        ahb_signals = [
            "HADDR", "HBURST", "HSIZE", "HTRANS", 
            "HWRITE", "HWDATA", "HPROT", "HRDATA", 
            "HRESP", "HREADY"
        ]
        
        signal_values = {}
        
        # 5. 각 AHB 신호 확인 및 입력
        for sig_name in ahb_signals:
            sig_r, sig_c = _find_cell(ws, sig_name)
            if not sig_r:
                print(f"ERROR: '{sig_name}' cell not found in AHB_S sheet.", flush=True)
                raise ValueError(f"'{sig_name}' cell not found")
            
            sig_val = ws.cell(row=sig_r + 1, column=sig_c).value
            if not sig_val or str(sig_val).strip() == "":
                print(f"\n=== {sig_name} ===")
                sig_val = _pick_from(all_ports, f"Select {sig_name} signal:", allow_custom=True)
                ws.cell(row=sig_r + 1, column=sig_c, value=sig_val)
            else:
                sig_val = str(sig_val).strip()
            
            signal_values[sig_name] = sig_val

        # 6. Excel 저장
        wb.save(xls_path)

        # 7. Width 정보 수집
        signal_widths = {}
        for sig_name, sig_val in signal_values.items():
            signal_widths[f"{sig_name} Width"] = _port_width_token(mod, sig_val)

        # 8. 결과 반환
        blocks = [{
            "Base Clock": base_clk,
            "Base Reset": base_rst,
            "Base Clock Width": _port_width_token(mod, base_clk),
            "Base Reset Width": _port_width_token(mod, base_rst),
            **signal_values,
            **signal_widths,
        }]
        return {"blocks": blocks}

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
        """
        SystemVerilog assertion 코드 생성
        - assertion_intf.sv: interface 정의
        - assertion_intf_inst.sv: 인스턴스 및 연결
        """
        blocks = parsed.get("blocks") or []
        if not blocks:
            return ["// No AHB Slave assertions generated.\n", ""]
        
        b = blocks[0]
        base_clk = b.get("Base Clock", "") or "clk"
        base_rst = b.get("Base Reset", "") or "nreset"
        
        # AHB 신호들
        haddr = b.get("HADDR", "") or "HADDR"
        hburst = b.get("HBURST", "") or "HBURST"
        hsize = b.get("HSIZE", "") or "HSIZE"
        htrans = b.get("HTRANS", "") or "HTRANS"
        hwrite = b.get("HWRITE", "") or "HWRITE"
        hwdata = b.get("HWDATA", "") or "HWDATA"
        hprot = b.get("HPROT", "") or "HPROT"
        hrdata = b.get("HRDATA", "") or "HRDATA"
        hresp = b.get("HRESP", "") or "HRESP"
        hready = b.get("HREADY", "") or "HREADY"

        # Width 정보
        clk_w = b.get("Base Clock Width", "[0:0]")
        rst_w = b.get("Base Reset Width", "[0:0]")
        haddr_w = b.get("HADDR Width", "[31:0]")
        hburst_w = b.get("HBURST Width", "[2:0]")
        hsize_w = b.get("HSIZE Width", "[2:0]")
        htrans_w = b.get("HTRANS Width", "[1:0]")
        hwrite_w = b.get("HWRITE Width", "[0:0]")
        hwdata_w = b.get("HWDATA Width", "[31:0]")
        hprot_w = b.get("HPROT Width", "[3:0]")
        hrdata_w = b.get("HRDATA Width", "[31:0]")
        hresp_w = b.get("HRESP Width", "[0:0]")
        hready_w = b.get("HREADY Width", "[0:0]")

        # ===== assertion_intf.sv 생성 =====
        sv_text = f"""`include "uvm_macros.svh"
import uvm_pkg::*;

interface assertion_intf();

logic {clk_w} {base_clk};
logic {rst_w} {base_rst};
logic {haddr_w} HADDR;
logic {hburst_w} HBURST;
logic {hsize_w} HSIZE;
logic {htrans_w} HTRANS;
logic {hwrite_w} HWRITE;
logic {hwdata_w} HWDATA;
logic {hprot_w} HPROT;
logic {hrdata_w} HRDATA;
logic {hresp_w} HRESP;
logic {hready_w} HREADY;
logic HSEL;
logic HREADYOUT;

logic [1:0] slave_state;

//=================================================
//	VARIABLES
//=================================================

parameter AHB_DATA_WIDTH = 32;

bit [31:0] cur_addr;

bit [AHB_DATA_WIDTH-1:0] hrdata_reserved_byte;
bit [AHB_DATA_WIDTH-1:0] hrdata_reserved_halfword;
bit [AHB_DATA_WIDTH-1:0] hrdata_reserved_word;
bit [AHB_DATA_WIDTH-1:0] hrdata_reserved_2word;
bit [AHB_DATA_WIDTH-1:0] hrdata_reserved_4word;
bit [AHB_DATA_WIDTH-1:0] hrdata_reserved_8word;
bit [AHB_DATA_WIDTH-1:0] hrdata_reserved_16word;

//=================================================
//	ASSIGN, ALWAYS
//=================================================

always_ff @(posedge {base_clk} or negedge {base_rst}) begin : update_cur_addr
  if(!{base_rst}) begin
    cur_addr <= 0;
  end else if (HREADYOUT && HTRANS inside {{2,3}}) begin
    cur_addr <= HADDR;
  end else begin
    cur_addr <= cur_addr;
  end
end : update_cur_addr

always_comb begin : hrdata_reserved_byte_update
  if      (AHB_DATA_WIDTH == 32)   hrdata_reserved_byte = HRDATA & ~(   32'hff << 8 * cur_addr[1:0]);
  else if (AHB_DATA_WIDTH == 64)   hrdata_reserved_byte = HRDATA & ~(   64'hff << 8 * cur_addr[2:0]);
  else if (AHB_DATA_WIDTH == 128)  hrdata_reserved_byte = HRDATA & ~(  128'hff << 8 * cur_addr[3:0]);
  else if (AHB_DATA_WIDTH == 256)  hrdata_reserved_byte = HRDATA & ~(  256'hff << 8 * cur_addr[4:0]);
  else if (AHB_DATA_WIDTH == 512)  hrdata_reserved_byte = HRDATA & ~(  512'hff << 8 * cur_addr[5:0]);
  else if (AHB_DATA_WIDTH == 1024) hrdata_reserved_byte = HRDATA & ~( 1024'hff << 8 * cur_addr[6:0]);
end : hrdata_reserved_byte_update

always_comb begin : hrdata_reserved_halfword_update
  if      (AHB_DATA_WIDTH == 32)   hrdata_reserved_halfword = HRDATA & ~(   32'hffff << 8 * cur_addr[1  ]);
  else if (AHB_DATA_WIDTH == 64)   hrdata_reserved_halfword = HRDATA & ~(   64'hffff << 8 * cur_addr[2:1]);
  else if (AHB_DATA_WIDTH == 128)  hrdata_reserved_halfword = HRDATA & ~(  128'hffff << 8 * cur_addr[3:1]);
  else if (AHB_DATA_WIDTH == 256)  hrdata_reserved_halfword = HRDATA & ~(  256'hffff << 8 * cur_addr[4:1]);
  else if (AHB_DATA_WIDTH == 512)  hrdata_reserved_halfword = HRDATA & ~(  512'hffff << 8 * cur_addr[5:1]);
  else if (AHB_DATA_WIDTH == 1024) hrdata_reserved_halfword = HRDATA & ~( 1024'hffff << 8 * cur_addr[6:1]);
end : hrdata_reserved_halfword_update

always_comb begin : hrdata_reserved_word_update
  if      (AHB_DATA_WIDTH == 64)   hrdata_reserved_word = HRDATA & ~(   64'hffffffff << 8 * cur_addr[2  ]);
  else if (AHB_DATA_WIDTH == 128)  hrdata_reserved_word = HRDATA & ~(  128'hffffffff << 8 * cur_addr[3:2]);
  else if (AHB_DATA_WIDTH == 256)  hrdata_reserved_word = HRDATA & ~(  256'hffffffff << 8 * cur_addr[4:2]);
  else if (AHB_DATA_WIDTH == 512)  hrdata_reserved_word = HRDATA & ~(  512'hffffffff << 8 * cur_addr[5:2]);
  else if (AHB_DATA_WIDTH == 1024) hrdata_reserved_word = HRDATA & ~( 1024'hffffffff << 8 * cur_addr[6:2]);
end : hrdata_reserved_word_update

always_comb begin : hrdata_reserved_2word_update
  if      (AHB_DATA_WIDTH == 128)  hrdata_reserved_2word = HRDATA & ~(  128'hffffffffffffffff << 8 * cur_addr[3  ]);
  else if (AHB_DATA_WIDTH == 256)  hrdata_reserved_2word = HRDATA & ~(  256'hffffffffffffffff << 8 * cur_addr[4:3]);
  else if (AHB_DATA_WIDTH == 512)  hrdata_reserved_2word = HRDATA & ~(  512'hffffffffffffffff << 8 * cur_addr[5:3]);
  else if (AHB_DATA_WIDTH == 1024) hrdata_reserved_2word = HRDATA & ~( 1024'hffffffffffffffff << 8 * cur_addr[6:3]);
end : hrdata_reserved_2word_update

always_comb begin : hrdata_reserved_4word_update
  if      (AHB_DATA_WIDTH == 256)  hrdata_reserved_4word = HRDATA & ~(  256'hffffffffffffffffffffffffffffffff << 8 * cur_addr[4  ]);
  else if (AHB_DATA_WIDTH == 512)  hrdata_reserved_4word = HRDATA & ~(  512'hffffffffffffffffffffffffffffffff << 8 * cur_addr[5:4]);
  else if (AHB_DATA_WIDTH == 1024) hrdata_reserved_4word = HRDATA & ~( 1024'hffffffffffffffffffffffffffffffff << 8 * cur_addr[6:4]);
end : hrdata_reserved_4word_update

always_comb begin : hrdata_reserved_8word_update
  if      (AHB_DATA_WIDTH == 512)  hrdata_reserved_8word = HRDATA & ~(  512'hffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffff << 8 * cur_addr[5  ]);
  if      (AHB_DATA_WIDTH == 1024) hrdata_reserved_8word = HRDATA & ~( 1024'hffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffff << 8 * cur_addr[6:5]);
end : hrdata_reserved_8word_update

always_comb begin : hrdata_reserved_16word_update
  if      (AHB_DATA_WIDTH == 1024) hrdata_reserved_16word = HRDATA & ~( 1024'hffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffff << 8 * cur_addr[6  ]);
end : hrdata_reserved_16word_update

//=================================================
//	PROPERTY
//=================================================

//Check Reset value
property nreset_assert_value;
  @(posedge {base_clk}) {base_rst} == 0 |=> (HTRANS == 0) && (HREADYOUT == 1);
endproperty

//HRDATA Reserved check
property byte_hrdata_reserved_chk;
  @(posedge {base_clk})
  disable iff(HRESP)
  (HREADYOUT == 1) && (HTRANS inside {{2,3}}) && (HSIZE == 0) && (HWRITE == 0) |=> ##[0:$] HREADYOUT == 1 ##0 hrdata_reserved_byte == 0;
endproperty

property halfword_hrdata_reserved_chk;
  @(posedge {base_clk})
  disable iff(HRESP)
  (HREADYOUT == 1) && (HTRANS inside {{2,3}}) && (HSIZE == 1) && (HWRITE == 0) |=> ##[0:$] hrdata_reserved_halfword == 0 ##0 HREADYOUT == 1;
endproperty

property word_hrdata_reserved_chk;
  @(posedge {base_clk})
  disable iff(HRESP)
  (HREADYOUT == 1) && (HTRANS inside {{2,3}}) && (HSIZE == 2) && (HWRITE == 0) |=> ##[0:$] hrdata_reserved_word == 0 ##0 HREADYOUT == 1;
endproperty

property two_word_hrdata_reserved_chk;
  @(posedge {base_clk})
  disable iff(HRESP)
  (HREADYOUT == 1) && (HTRANS inside {{2,3}}) && (HSIZE == 3) && (HWRITE == 0) |=> ##[0:$] hrdata_reserved_2word == 0 ##0 HREADYOUT == 1;
endproperty

property four_word_hrdata_reserved_chk;
  @(posedge {base_clk})
  disable iff(HRESP)
  (HREADYOUT == 1) && (HTRANS inside {{2,3}}) && (HSIZE == 4) && (HWRITE == 0) |=> ##[0:$] hrdata_reserved_4word == 0 ##0 HREADYOUT == 1;
endproperty

property eight_word_hrdata_reserved_chk;
  @(posedge {base_clk})
  disable iff(HRESP)
  (HREADYOUT == 1) && (HTRANS inside {{2,3}}) && (HSIZE == 5) && (HWRITE == 0) |=> ##[0:$] hrdata_reserved_8word == 0 ##0 HREADYOUT == 1;
endproperty

property sixteen_word_hrdata_reserved_chk;
  @(posedge {base_clk})
  disable iff(HRESP)
  (HREADYOUT == 1) && (HTRANS inside {{2,3}}) && (HSIZE == 6) && (HWRITE == 0) |=> ##[0:$] hrdata_reserved_16word == 0 ##0 HREADYOUT == 1;
endproperty

property hresp_error_chk;
  @(posedge {base_clk})
  ($rose(HRESP == 1)) |-> ( ##0 (HRESP == 1) && (HREADYOUT == 0)
                           ##1 ((HRESP == 1) && (HREADYOUT == 1)));
endproperty

//=================================================
//	ASSERT
//=================================================

assert property (nreset_assert_value)
  else $error("Invalid Value When nreset is assert");

assert property (byte_hrdata_reserved_chk)
  else $error("Invalid byte HRDATA padding");

assert property (halfword_hrdata_reserved_chk)
  else $error("Invalid halfword HRDATA padding");

assert property (word_hrdata_reserved_chk)
  else $error("Invalid word HRDATA padding");

assert property (two_word_hrdata_reserved_chk)
  else $error("Invalid 2word HRDATA padding");

assert property (four_word_hrdata_reserved_chk)
  else $error("Invalid 4word HRDATA padding");

assert property (eight_word_hrdata_reserved_chk)
  else $error("Invalid 8word HRDATA padding");

assert property (sixteen_word_hrdata_reserved_chk)
  else $error("Invalid 16word HRDATA padding");

assert property (hresp_error_chk)
  else $error("Invalid HREADYOUT When HRESP == ERROR ");

endinterface
"""

        # ===== assertion_intf_inst.sv 생성 =====
        inst_lines: List[str] = []
        inst_lines.append("`include \"uvm_macros.svh\"")
        inst_lines.append("import uvm_pkg::*;")
        inst_lines.append("")
        inst_lines.append("assertion_intf u_assertion_intf();")
        inst_lines.append("")
        inst_lines.append(f"assign u_assertion_intf.{base_clk} = top.dut.{base_clk};")
        inst_lines.append(f"assign u_assertion_intf.{base_rst} = top.dut.{base_rst};")
        inst_lines.append(f"assign u_assertion_intf.HADDR = top.dut.{haddr};")
        inst_lines.append(f"assign u_assertion_intf.HBURST = top.dut.{hburst};")
        inst_lines.append(f"assign u_assertion_intf.HSIZE = top.dut.{hsize};")
        inst_lines.append(f"assign u_assertion_intf.HTRANS = top.dut.{htrans};")
        inst_lines.append(f"assign u_assertion_intf.HWRITE = top.dut.{hwrite};")
        inst_lines.append(f"assign u_assertion_intf.HWDATA = top.dut.{hwdata};")
        inst_lines.append(f"assign u_assertion_intf.HPROT = top.dut.{hprot};")
        inst_lines.append(f"assign u_assertion_intf.HRDATA = top.dut.{hrdata};")
        inst_lines.append(f"assign u_assertion_intf.HREADY = top.dut.{hready};")
        inst_lines.append(f"assign u_assertion_intf.HRESP = top.dut.{hresp};")
        inst_text = "\n".join(inst_lines) + "\n"

        return [sv_text, inst_text]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        """JSON 출력 (필요시)"""
        return parsed
