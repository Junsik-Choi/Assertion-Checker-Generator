from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json
import os
import re

from openpyxl import load_workbook
from .base import BaseAssertionPlugin
from .registry import register

ALLOWED_TYPES = ("counter",)

def _pick_one(title: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    print(title, flush=True)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Enter custom")
    while True:
        try:
            s = input("Select > ").strip()
        except EOFError:
            return options[0][1] if options else ""
        if allow_custom and s == "0":
            try:
                return input("Enter > ").strip()
            except EOFError:
                return ""
        if s.isdigit():
            i = int(s)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.", flush=True)

def _scan(title: str) -> str:
    print(title, flush=True)
    while True:
        try:
            return input("Enter > ").strip()
        except EOFError:
            return ""

def find_cell(ws, value: str) -> Tuple[Optional[int], Optional[int]]:
    tgt = value.strip().lower()
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if str(c.value).strip().lower() == tgt:
                return c.row, c.column
    return None, None

def _get_sheet_ci(wb, want_name: str = "Counter", create: bool = False):
    target = (want_name or "").strip().lower()
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == target:
            return wb[nm]
    if create:
        return wb.create_sheet(title=want_name)
    raise KeyError(f"Worksheet {want_name} does not exist.")

def _ensure_counter_layout(ws) -> Tuple[int, int, int]:
    c_row, c_col = find_cell(ws, "Counter")
    if c_row is None:
        c_row, c_col = 5, 2
        ws.cell(row=c_row, column=c_col, value="Counter")
        ws.cell(row=c_row - 3, column=c_col, value="Base Clock")
        ws.cell(row=c_row - 2, column=c_col, value="Base Reset")
        ws.cell(row=c_row + 1, column=c_col, value="Count Value")
        ws.cell(row=c_row + 2, column=c_col, value="Target")
        ws.cell(row=c_row + 1, column=c_col + 1, value="Logic Condition")
        ws.cell(row=c_row + 2, column=c_col + 1, value="Plus Condition")
        ws.cell(row=c_row + 2, column=c_col + 2, value="Reset Condition")
        ws.cell(row=c_row + 1, column=c_col + 3, value="Assertion Condition")
        ws.cell(row=c_row + 2, column=c_col + 3, value="Trigger Condition")
        ws.cell(row=c_row + 2, column=c_col + 4, value="Expect Count Value")
    min_col = c_col
    for mr in ws.merged_cells.ranges:
        if ws.cell(row=c_row, column=c_col).coordinate in mr:
            min_col = mr.min_col
            break
    target_row = c_row + 2
    target_col = min_col
    if not ws.cell(row=target_row, column=target_col).value:
        ws.cell(row=target_row, column=target_col, value="Target")
    if not ws.cell(row=target_row, column=target_col + 1).value:
        ws.cell(row=target_row, column=target_col + 1, value="Plus Condition")
    if not ws.cell(row=target_row, column=target_col + 2).value:
        ws.cell(row=target_row, column=target_col + 2, value="Reset Condition")
    if not ws.cell(row=target_row, column=target_col + 3).value:
        ws.cell(row=target_row, column=target_col + 3, value="Trigger Condition")
    if not ws.cell(row=target_row, column=target_col + 4).value:
        ws.cell(row=target_row, column=target_col + 4, value="Expect Count Value")
    data_row = target_row + 1
    return c_row, target_col, data_row

def parse_counter_block_for_row(ws, row: int, target_col: int) -> Dict[str, Any]:
    clk_row, clk_col = find_cell(ws, "Base Clock")
    rst_row, rst_col = find_cell(ws, "Base Reset")
    base_clk = ws.cell(row=clk_row, column=clk_col + 1).value if clk_row else None
    base_rst = ws.cell(row=rst_row, column=rst_col + 1).value if rst_row else None
    target = ws.cell(row=row, column=target_col).value
    plus_con = ws.cell(row=row, column=target_col + 1).value
    reset_con = ws.cell(row=row, column=target_col + 2).value
    trigger_con = ws.cell(row=row, column=target_col + 3).value
    exp_cnt_val = ws.cell(row=row, column=target_col + 4).value
    return {
        "Base Clock": str(base_clk).strip() if base_clk else "",
        "Reset": str(base_rst).strip() if base_rst else "",
        "Target": str(target).strip() if target else "",
        "Plus Condition": str(plus_con).strip() if plus_con else "",
        "Reset Condition": str(reset_con).strip() if reset_con else "",
        "Trigger Condition": str(trigger_con).strip() if trigger_con else "",
        "Expect Count Value": str(exp_cnt_val).strip() if exp_cnt_val else "",
    }

def _auto_pick_clk_rst(mod: Dict[str, Any]) -> Tuple[str, str]:
    clk = ""
    rst = ""
    clocks = mod.get("clocks") or []
    resets = mod.get("resets") or []
    if clocks:
        clk = clocks[0].get("name") or ""
    if resets:
        rst = resets[0].get("name") or ""
    if not clk:
        for it in mod.get("inputs", []):
            n = (it.get("name") or "").lower()
            if "clk" in n or n.endswith("clock"):
                clk = it.get("name") or ""
                break
    if not rst:
        for it in mod.get("inputs", []):
            n = (it.get("name") or "").lower()
            if "rst" in n or "reset" in n:
                rst = it.get("name") or ""
                break
    return clk, rst

# -----------------------------
# Is Port helpers
# -----------------------------
def _all_module_port_names(mod: Dict[str, Any]) -> List[str]:
    names = []
    for key in ("inputs", "outputs", "inouts", "clocks", "resets", "ports"):
        for p in mod.get(key, []) or []:
            n = (p.get("name") or "").strip()
            if n:
                names.append(n)
    # 중복 제거(순서 보존)
    seen = set(); out = []
    for n in names:
        if n not in seen:
            seen.add(n); out.append(n)
    return out

def _collect_expr_ports(mod: Dict[str, Any], expr: str) -> List[Tuple[str, str]]:
    """
    expr 안에서 모듈 포트 이름이 실제로 등장하는 것만 추출.
    """
    if not expr or not mod:
        return []
    all_names = _all_module_port_names(mod)
    # 포트명이 특수문자 포함 가능성 대비해서 re.escape 사용, 단어 경계 \b로 안전 매칭
    if not all_names:
        return []
    pat = r'\b(' + '|'.join(re.escape(n) for n in sorted(all_names, key=len, reverse=True)) + r')\b'
    # 중복 제거(순서 보존)
    seen = set(); found = []
    for m in re.finditer(pat, expr):
        n = m.group(1)
        if n not in seen:
            seen.add(n); found.append((n, _port_width_token(mod, n)))
    return found

# -----------------------------
# Width helpers
# -----------------------------
def _normalize_range_token(token: str) -> str:
    """Return normalized [msb:lsb] token; default to [0:0] for 1-bit/unknown."""
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    # numeric like '32' or '1'
    try:
        n = int(t, 10)
        if n >= 1:
            return f"[{n-1}:0]"
        return "[0:0]"
    except Exception:
        # unknown format -> treat as 1-bit
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    """Find port by name and return width token '[msb:lsb]' (defaults to [0:0])."""
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    # 1) search a broader 'ports' list first if present
    candidates = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in candidates:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            # Typical fields
            for key in ("packed_range", "range", "packed", "decl"):
                pr = it.get(key)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            # Width as integer/string
            for key in ("width", "bit_width", "width_bits"):
                w = it.get(key)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            # Separate ends
            for left_key, right_key in (("msb", "lsb"), ("left", "right")):
                msb = it.get(left_key)
                lsb = it.get(right_key)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            # Fallback 1-bit
            return "[0:0]"
    return "[0:0]"

def _fmt_input_decl(sig: str, width_tok: str) -> str:
    """Format input logic with width; defaults to [0:0] if empty."""
    width_tok = (width_tok or "").strip() or "[0:0]"
    return f"input logic {width_tok} {sig}"

def _update_counter_sheet(ws, cnt_cfg: Dict[str, str], module_info: Dict[str, Any]) -> int:
    """
    Counter 시트에 값을 기록하고, 기록한 '행 번호'를 반환한다.
    기존 데이터 아래의 첫 빈 행(Type 셀이 비어있는 곳)에 기록.
    """
    c_row, target_col, data_row = _ensure_counter_layout(ws)
    # 아래로 내려가며 Type 셀이 비어있는 첫 행을 찾음
    write_row = data_row
    while True:
        val = ws.cell(row=write_row, column=target_col).value
        if val is None or str(val).strip() == "":
            break
        write_row += 1
    # 선택 정보 기록
    ws.cell(row=write_row, column=target_col,     value=(cnt_cfg.get("target", "") or "").strip())
    ws.cell(row=write_row, column=target_col + 1, value=(cnt_cfg.get("plus_con", "") or "").strip())
    ws.cell(row=write_row, column=target_col + 2, value=(cnt_cfg.get("reset_con", "") or "").strip())
    ws.cell(row=write_row, column=target_col + 3, value=(cnt_cfg.get("trigger_con", "") or "").strip())
    ws.cell(row=write_row, column=target_col + 4, value=(cnt_cfg.get("exp_cnt_val", "") or "").strip())

    # Base Clock / Base Reset 라벨 및 값 정리
    clk_label = find_cell(ws, "Base Clock")
    rst_label = find_cell(ws, "Base Reset")
    if clk_label[0] is None:
        ws.cell(row=c_row - 3, column=target_col, value="Base Clock"); clk_label = (c_row - 3, target_col)
    if rst_label[0] is None:
        ws.cell(row=c_row - 2, column=target_col, value="Base Reset"); rst_label = (c_row - 2, target_col)
    clk_name, rst_name = _auto_pick_clk_rst(module_info)
    if clk_name:
        ws.cell(row=clk_label[0], column=clk_label[1] + 1, value=clk_name)
    if rst_name:
        ws.cell(row=rst_label[0], column=rst_label[1] + 1, value=rst_name)
    return write_row

def generate_verilog(info: Dict[str, Any]) -> str:
    clk         = info["Base Clock"]
    rst         = info["Reset"]
    plus_con    = info["Plus Condition"]
    reset_con   = info["Reset Condition"]
    trigger_con = info["Trigger Condition"]
    exp_cnt_val = info["Expect Count Value"]
    cnt         = info["Target"]

    w_clk           = info.get("Base Clock Width", "") or "[0:0]"
    w_rst           = info.get("Reset Width", "") or "[0:0]"

    logic_list = []
    declared = set()
    if clk : logic_list.append(f"logic {w_clk} {clk};"); declared.add(clk)
    if rst : logic_list.append(f"logic {w_rst} {rst};"); declared.add(rst)

    for name, w in (info.get("Plus Condition Ports") or []):
        if name not in declared:
            width = w or "[0:0]"
            logic_list.append(f"logic {width} {name};"); declared.add(name)

    for name, w in (info.get("Reset Condition Ports") or []):
        if name not in declared:
            width = w or "[0:0]"
            logic_list.append(f"logic {width} {name};"); declared.add(name)

    for name, w in (info.get("Trigger Condition Ports") or []):
        if name not in declared:
            width = w or "[0:0]"
            logic_list.append(f"logic {width} {name};"); declared.add(name)

    for name, w in (info.get("Expect Count Value Ports") or []):
        if name not in declared:
            width = w or "[0:0]"
            logic_list.append(f"logic {width} {name};"); declared.add(name)

    logics = "\n".join(logic_list)

    header = '`include "uvm_macros.svh"\nimport uvm_pkg::*;\n\n'
    return header + f"""interface assertion_intf();

{logics}

reg [31:0] {cnt};

always @(posedge {clk} or negedge {rst}) begin
    if(!{rst}) begin
        {cnt} <= 0;
    end
    else if({reset_con}) begin
        {cnt} <= 0;
    end
    else if({plus_con}) begin
        {cnt} <= {cnt}+1;
    end
    else begin
        {cnt} <= {cnt};
    end
end

property p_counter_check;
    @(posedge {clk}) disable iff(!{rst})
    {trigger_con} |-> ({cnt} == {exp_cnt_val});
endproperty

assert property (p_counter_check) else `uvm_error("ASSERTION", "Counter check failed")

endinterface
"""

def generate_inst_verilog(info: Dict[str, Any]) -> str:
    clk         = info["Base Clock"]
    rst         = info["Reset"]

    assign_list = []
    assigned = set()
    if clk : assign_list.append(f"assign u_assertion_intf.{clk} = top.dut.{clk};"); assigned.add(clk)
    if rst : assign_list.append(f"assign u_assertion_intf.{rst} = top.dut.{rst};"); assigned.add(rst)

    for name, w in (info.get("Plus Condition Ports") or []):
        if name not in assigned:
            assign_list.append(f"assign u_assertion_intf.{name} = top.dut.{name};"); assigned.add(name)

    for name, w in (info.get("Reset Condition Ports") or []):
        if name not in assigned:
            assign_list.append(f"assign u_assertion_intf.{name} = top.dut.{name};"); assigned.add(name)

    for name, w in (info.get("Trigger Condition Ports") or []):
        if name not in assigned:
            assign_list.append(f"assign u_assertion_intf.{name} = top.dut.{name};"); assigned.add(name)

    for name, w in (info.get("Expect Count Value Ports") or []):
        if name not in assigned:
            assign_list.append(f"assign u_assertion_intf.{name} = top.dut.{name};"); assigned.add(name)
    assigns = "\n".join(assign_list)

    header = '`include "uvm_macros.svh"\nimport uvm_pkg::*;\n\n'

    # inst 파일은 모듈 래퍼 없이 인스턴스와 assign만 생성
    return header + (
        f"assertion_intf u_assertion_intf();\n\n"
        f"{assigns}\n"
    )

def _get_forced_type() -> Optional[str]:
    t = (os.environ.get("ASSERTION_FORCE_TYPE") or "").strip().lower()
    return t if t in ALLOWED_TYPES else None

@register
class CounterPlugin(BaseAssertionPlugin):
    plugin_name = "counter"
    sheet_name = "Counter"

    def _load_module_define(self, xls_path: Path) -> Dict[str, Any]:
        try:
            md = xls_path.parent / "module_define.json"
            if md.exists():
                return json.loads(md.read_text(encoding="utf-8"))
            # fallback: assertion_inputs.json (same schema subset)
            ai = xls_path.parent / "assertion_inputs.json"
            if ai.exists():
                return json.loads(ai.read_text(encoding="utf-8"))
        except Exception:
            pass
        return {}

    def _interactive_collect(self, mod: Dict[str, Any]) -> Dict[str, str]:
        in_names = [(f"in : {p.get('name')}", p.get("name") or "") for p in (mod.get("inputs") or []) if p.get("name")]
        out_names = [(f"out: {p.get('name')}", p.get("name") or "") for p in (mod.get("outputs") or []) if p.get("name")]
        sig_opts = in_names + out_names or [("manual input", "")]
        target = "<Target>"
        reset_con = "<Reset Condition>"
        plus_con = "<Plus Condition>"
        trigger_con = "<Trigger Condition>"
        exp_cnt_val = "<Expect Counter Value>"
        while True:
            print("\n==================== Enter Items ====================")
            print(f"always @(posedge clk or negedge rstn) begin")
            print(f"    if(!rstn) begin")
            print(f"        [1]{target} <= 0;")
            print(f"    end")
            print(f"    else if([2]{reset_con}) begin")
            print(f"        [1]{target} <= 0;")
            print(f"    end")
            print(f"    else if([3]{plus_con}) begin")
            print(f"        [1]{target} <= [1]{target}+1;")
            print(f"    end")
            print(f"    else begin")
            print(f"        [1]{target} <= [1]{target};")
            print(f"    end")
            print(f"end")
            print(f"")
            print(f"property p_counter_check;")
            print(f"    @(posedge clk) disable iff(!rstn)")
            print(f"    [4]{trigger_con} |-> ([1]{target} == [5]{exp_cnt_val});")
            print(f"endproperty")
            print("=========================================================")
            print("Select item number to edit")
            print("Press Enter twice to confirm all")
            choice = input("> ").strip()
            if choice == "":
                missing = []
                if target       in ("", "<Target>"):                missing.append("[1]")
                if reset_con    in ("", "<Reset Condition>"):       missing.append("[2]")
                if plus_con     in ("", "<Plus Condition>"):        missing.append("[3]")
                if trigger_con  in ("", "<Trigger Condition>"):     missing.append("[4]")
                if exp_cnt_val  in ("", "<Expect Counter Value>"):  missing.append("[5]")
                if missing: print(f"{','.join(missing)} has NOT been entered yet.")
                print("Press Enter again to confirm, or select item number to edit")
                choice = input("> ").strip()
                if choice == "":
                    break
            if choice == "1":
                target = _scan("Enter Target")
            elif choice == "2":
                reset_con = _pick_one("Select Reset Condition", sig_opts, allow_custom=True)
            elif choice == "3":
                plus_con = _pick_one("Select Plus Condition", sig_opts, allow_custom=True)
            elif choice == "4":
                trigger_con = _pick_one("Select Trigger Condition", sig_opts, allow_custom=True)
            elif choice == "5":
                exp_cnt_val = _pick_one("Select Expect Counter Value", sig_opts, allow_custom=True)
            else:
                print("Invalid selection. Try again.")
                continue
        return {"target": target, "plus_con": plus_con, "reset_con": reset_con, "trigger_con": trigger_con, "exp_cnt_val": exp_cnt_val}

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        # 1) 플러그인 선택 직후: 타입/신호 선택 프롬프트
        mod = self._load_module_define(Path(xls_path))
        cnt_cfg = self._interactive_collect(mod)

        # 2) Excel에 기록
        wb_w = load_workbook(xls_path)
        try:
            ws_w = _get_sheet_ci(wb_w, self.sheet_name, create=False)
        except KeyError:
            ws_w = _get_sheet_ci(wb_w, self.sheet_name, create=True)
        write_row = _update_counter_sheet(ws_w, cnt_cfg, mod)
        wb_w.save(xls_path)

        # 3) data_only로 재오픈하여 '방금 기록한 행'만 파싱
        wb = load_workbook(xls_path, data_only=True)
        try:
            ws = _get_sheet_ci(wb, self.sheet_name, create=False)
        except KeyError:
            return {"blocks": []}
        _, target_col, _ = _ensure_counter_layout(ws)
        info = parse_counter_block_for_row(ws, write_row, target_col)
        # attach bit widths from module define
        info["Base Clock Width"] = _port_width_token(mod, info.get("Base Clock", ""))
        info["Reset Width"] = _port_width_token(mod, info.get("Reset", ""))
        info["Plus Condition Width"] = _port_width_token(mod, info.get("Plus Condition", ""))
        info["Reset Condition Width"] = _port_width_token(mod, info.get("Reset Condition", ""))
        info["Trigger Condition Width"] = _port_width_token(mod, info.get("Trigger Condition", ""))
        info["Expect Count Value Width"] = _port_width_token(mod, info.get("Expect Count Value", ""))
        # attach 'ports in custom' from module define
        info["Plus Condition Ports"] = _collect_expr_ports(mod, info.get("Plus Condition", ""))
        info["Reset Condition Ports"] = _collect_expr_ports(mod, info.get("Reset Condition", ""))
        info["Trigger Condition Ports"] = _collect_expr_ports(mod, info.get("Trigger Condition", ""))
        info["Expect Count Value Ports"] = _collect_expr_ports(mod, info.get("Expect Count Value", ""))
        blocks: List[Dict[str, Any]] = []
        ct = "counter"
        if ct in ALLOWED_TYPES and info.get("Target") and info.get("Plus Condition") and info.get("Trigger Condition") and info.get("Expect Count Value"):
            blocks.append(info)

        # 선택된 타입만 남기기
        forced = _get_forced_type()
        if forced:
            blocks = [b for b in blocks if "counter" == forced]
        return {"blocks": blocks}

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
        # 파일 저장은 assertion_builder.py에서 처리. 여기서는 문자열만 반환.
        sv_parts: List[str] = []
        inst_parts: List[str] = []
        forced = _get_forced_type()
        for info in parsed.get("blocks", []):
            if forced and ("counter" != forced):
                continue
            sv_parts.append(generate_verilog(info))
            inst_parts.append(generate_inst_verilog(info))
        combined_sv = "\n\n".join([s.strip() for s in sv_parts if str(s).strip()]) + ("\n" if sv_parts else "")
        combined_inst = "\n\n".join([s.strip() for s in inst_parts if str(s).strip()]) + ("\n" if inst_parts else "")
        return [combined_sv, combined_inst]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        return parsed

# Optional: standalone script entry point for legacy usage
def main(xlsx_path: str) -> None:
    # 간단 실행 테스트: 시트 준비만 수행
    wb = load_workbook(xlsx_path)
    try:
        ws = _get_sheet_ci(wb, "Counter", create=False)
    except KeyError:
        ws = _get_sheet_ci(wb, "Counter", create=True)
    _ensure_counter_layout(ws)
    wb.save(xlsx_path)