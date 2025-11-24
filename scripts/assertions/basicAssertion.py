from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import re

import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from .base import BaseAssertionPlugin
from .registry import register

# -----------------------------
# Local helpers (self-contained)
# -----------------------------
def _get_sheet_ci(wb, name: str, create: bool = True):
    tgt = name.strip().lower()
    for s in wb.sheetnames:
        if str(s).strip().lower() == tgt:
            return wb[s]
    if not create:
        raise KeyError(name)
    return wb.create_sheet(title=name)

def _find_cell(ws, key: str) -> Tuple[Optional[int], Optional[int]]:
    key_l = str(key).strip().lower()
    for r in range(1, (ws.max_row or 1) + 1):
        for c in range(1, (ws.max_column or 1) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip().lower() == key_l:
                return r, c
    return None, None

def _read_json(p: Path) -> Optional[Dict[str, Any]]:
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None

def _load_module_define(xls_path: Path) -> Dict[str, Any]:
    root = xls_path.parent
    md = _read_json(root / "module_define.json")
    if md:
        return md
    ai = _read_json(root / "assertion_inputs.json")
    if ai:
        return {
            "module": ai.get("module") or "",
            "clocks": ai.get("clocks") or [],
            "resets": ai.get("resets") or [],
            "inputs": ai.get("inputs") or [],
            "outputs": ai.get("outputs") or [],
            "inouts": ai.get("inouts") or [],
            "ports": ai.get("ports") or [],
        }
    return {}

def _scan(title: str) -> str:
    print(title, flush=True)
    while True:
        try:
            s = input("Enter > ").strip()
            return s.replace("\\n", "\n")
        except EOFError:
            return ""

def _scan_wrap(title: str) -> str:
    print(title, flush=True)
    print("(Multiple lines supported. Press Enter twice to finish.)", flush=True)

    lines = []
    empty_count = 0

    while True:
        try:
            s = input()
        except EOFError:
            break
        if s == "":
            empty_count += 1
            if empty_count == 2:
                lines.pop()
                break
            lines.append("")
            continue
        empty_count = 0
        lines.append(s.replace("\\n", "\n"))
    return "\n".join(lines)

def _normalize_range_token(token: Any) -> str:
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    try:
        n = int(t, 10)
        return f"[{n-1}:0]" if n >= 1 else "[0:0]"
    except Exception:
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    cands = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in cands:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            for k in ("packed_range", "range", "packed", "decl"):
                pr = it.get(k)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            for k in ("width", "bit_width", "width_bits"):
                w = it.get(k)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            for lk, rk in (("msb", "lsb"), ("left", "right")):
                msb = it.get(lk); lsb = it.get(rk)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            return "[0:0]"
    return "[0:0]"

def _fmt_logic_decl(sig: str, width_tok: str) -> str:
    tok = (width_tok or "").strip() or "[0:0]"
    return f"logic {tok} {sig};"

def _read_base_clk_rst(ws) -> Tuple[str, str]:
    cr, cc = _find_cell(ws, "Base Clock")
    rr, rc = _find_cell(ws, "Base Reset")
    clk = ws.cell(row=cr, column=cc + 1).value if cr else ""
    rst = ws.cell(row=rr, column=rc + 1).value if rr else ""
    return (str(clk).strip() if clk else ""), (str(rst).strip() if rst else "")

# 수식/시트 참조를 실제 값으로 역참조
def _resolve_ref(wb, raw: Any) -> str:
    s = "" if raw is None else str(raw).strip()
    if not s:
        return ""
    expr = s[1:].strip() if s.startswith("=") else s
    if "!" not in expr:
        return s
    sheet_name, addr = expr.split("!", 1)
    sheet_name = sheet_name.strip().strip("'").strip('"')
    try:
        ws2 = wb[sheet_name]
        v = ws2[addr].value
        return str(v).strip() if v is not None else ""
    except Exception:
        return s

def _read_label_right_resolved(wb, ws, label: str) -> str:
    r, c = _find_cell(ws, label)
    if r is None:
        return ""
    raw = ws.cell(row=r, column=c + 1).value
    val = _resolve_ref(wb, raw)
    return val if val != "" else (str(raw).strip() if raw is not None else "")

def _ensure_label(ws, label: str) -> Tuple[int, int]:
    r, c = _find_cell(ws, label)
    if r is not None:
        return r, c
    # 헤더 기준으로 첫 빈 라벨 슬롯에 생성
    hdr_r, hdr_c = _find_cell(ws, "basicAssertion")
    if hdr_r is None:
        hdr_r, hdr_c = 1, 1
        cell = ws.cell(row=hdr_r, column=hdr_c, value="basicAssertion")
        cell.alignment = Alignment(wrap_text=True)
    r, c = hdr_r + 1, hdr_c
    while ws.cell(row=r, column=c).value not in (None, ""):
        r += 1
    cell = ws.cell(row=r, column=c, value=label)
    cell.alignment = Alignment(wrap_text=True)
    return r, c

def _append_label_below(ws, label: str, value: str) -> None:
    r, c = _ensure_label(ws, label)
    rr = r + 1
    while ws.cell(row=rr, column=c).value not in (None, ""):
        rr += 1
    cell = ws.cell(row=rr, column=c, value=value)
    cell.alignment = Alignment(wrap_text=True)

def _read_column_values(ws, label: str) -> List[str]:
    vals: List[str] = []
    r, c = _find_cell(ws, label)
    if r is None:
        return vals
    rr = r + 1
    while True:
        v = ws.cell(row=rr, column=c).value
        if v is None or str(v).strip() == "":
            break
        vals.append(str(v).strip())
        rr += 1
    return vals

def _sv_header() -> str:
    return '`include "uvm_macros.svh"\nimport uvm_pkg::*;\n'

def _nz(s: Optional[str], placeholder: str) -> str:
    return (s or "").strip() or placeholder

# -----------------------------
# Is Port helpers
# -----------------------------
def _all_module_port_names(mod: Dict[str, Any]) -> List[str]:
    names = []
    for key in ("inputs", "outputs", "inouts", "clocks", "resets", "ports"):
        for p in mod.get(key, []) or []:
            n = (p.get("name") or "").strip()
            if n:
                names.append(n)
    # 중복 제거(순서 보존)
    seen = set(); out = []
    for n in names:
        if n not in seen:
            seen.add(n); out.append(n)
    return out

def _collect_expr_ports(mod: Dict[str, Any], expr: str) -> List[str]:
    """
    expr 안에서 모듈 포트 이름이 실제로 등장하는 것만 추출.
    """
    if not expr or not mod:
        return []
    all_names = _all_module_port_names(mod)
    # 포트명이 특수문자 포함 가능성 대비해서 re.escape 사용, 단어 경계 \b로 안전 매칭
    if not all_names:
        return []
    pat = r'\b(' + '|'.join(re.escape(n) for n in sorted(all_names, key=len, reverse=True)) + r')\b'
    # 중복 제거(순서 보존)
    seen = set(); found = []
    for m in re.finditer(pat, expr):
        n = m.group(1)
        if n not in seen:
            seen.add(n); found.append(n)
    return found

# ---------------- SV builders (multi-sets) ----------------

def _build_basicassertion_sv_multi(base_clk: str, base_rst: str,
                                   prop_sets: List[Dict[str, str]],
                                   seq_sets: List[Dict[str, str]],
                                   unique_ports: List[str],
                                   width_map: Dict[str, str]) -> str:
    lines: List[str] = []
    lines.append(_sv_header())
    lines.append("interface assertion_intf();")
    lines.append("")

    # 포트: Base + 유니크 DUT 포트들
    if base_clk:
        lines.append(_fmt_logic_decl(base_clk, width_map.get(base_clk, "[0:0]")))
    if base_rst:
        lines.append(_fmt_logic_decl(base_rst, width_map.get(base_rst, "[0:0]")))
    for p in unique_ports:
        if p in (base_clk, base_rst):
            continue
        lines.append(_fmt_logic_decl(p, width_map.get(p, "[0:0]")))
    lines.append("")
    # 각 세트별 property/assert
    for idx, st in enumerate(seq_sets, start=1):
        user_seq_nm = st.get("User Sequence Name", "")
        seq_bs_clk_con = st.get("Sequence Base Clock Condition", "")
        user_seq = st.get("User Sequence", "").replace("\n","\n    ")
        lines.append(f"sequence {user_seq_nm}();")
        lines.append(f"    @({seq_bs_clk_con})")
        lines.append(f"    {user_seq}")
        lines.append("endsequence")
        lines.append("")
    for idx, st in enumerate(prop_sets, start=1):
        user_prop_nm = st.get("User Property Name", "")
        prop_bs_clk_con = st.get("Property Base Clock Condition", "")
        dis_con = st.get("Disable Condition", "")
        trig_con = st.get("Trigger Condition", "")
        user_res = st.get("User Result", "").replace("\n","\n    ")
        lines.append(f"property {user_prop_nm}();")
        lines.append(f"    @({prop_bs_clk_con}) disable iff({dis_con})")
        lines.append(f"    {trig_con}")
        lines.append(f"    |-> {user_res};")
        lines.append("endproperty")
        lines.append("")
        lines.append(f'assert property ({user_prop_nm}) else $error("failed at %t", $time);')
        lines.append("")
    lines.append("endinterface")
    lines.append("")
    return "\n".join(lines)

def _build_basicassertion_inst_sv_multi(base_clk: str, base_rst: str,
                                        unique_ports: List[str]) -> str:
    lines: List[str] = []
    lines.append(_sv_header())
    lines.append("assertion_intf u_assertion_intf();")
    lines.append("")
    # base + unique 포트 assign
    ports = []
    if base_clk: ports.append(base_clk)
    if base_rst: ports.append(base_rst)
    ports.extend([p for p in unique_ports if p not in (base_clk, base_rst)])
    for p in ports:
        lines.append(f"assign u_assertion_intf.{p}  = top.dut.{p};")
    lines.append("")
    return "\n".join(lines)

# -----------------------------
# Plugin
# -----------------------------
@register
class BasicAssertionPlugin(BaseAssertionPlugin):
    plugin_name = "basicAssertion"
    sheet_name = "basicAssertion"

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        # 0) Load module info
        mod = _load_module_define(Path(xls_path))

        # 1) Open sheet and ensure header exists
        wb_w = load_workbook(xls_path)
        try:
            ws_w = _get_sheet_ci(wb_w, self.sheet_name, create=False)
        except KeyError:
            ws_w = _get_sheet_ci(wb_w, self.sheet_name, create=True)
        # Ensure header label
        if _find_cell(ws_w, "basicAssertion")[0] is None:
            cell = ws_w.cell(row=1, column=1, value="basicAssertion")
            cell.alignment = Alignment(wrap_text=True)

        # 2) Base Clock/Reset from Define sheet (resolve references, no prompts)
        try:
            ws_define = _get_sheet_ci(wb_w, "Define", create=False)
            clk = _read_label_right_resolved(wb_w, ws_define, "Base Clock")
            rst = _read_label_right_resolved(wb_w, ws_define, "Base Reset")
        except KeyError:
            clk = ""
            rst = ""

        # 3) 세트 입력 루프
        prop_sets: List[Dict[str, str]] = []
        seq_sets: List[Dict[str, str]] = []
        while True:
            print("Property or Sequece")
            print("[1] Property")
            print("[2] Sequence")
            while True:
                sel = input("> ").strip()
                if sel == "1":
                    user_prop_nm = "<User Property Name>"
                    prop_bs_clk_con = "<Property Base Clock Condition>"
                    dis_con = "<Disable Condition>"
                    trig_con = "<Trigger Condition>"
                    user_res = "<User Result>"
                    while True:
                        print("\n==================== Enter Property ====================")
                        print(f"property [1]{user_prop_nm}();")
                        print(f"    @([2]{prop_bs_clk_con}) disable iff([3]{dis_con})")
                        print(f"    [4]{trig_con}")
                        formatted_res = user_res.replace('\n', '\n    ')
                        print(f"    [5]{formatted_res};")
                        print("endproperty")
                        print("=========================================================")
                        print("Select item number to edit")
                        print("Press Enter twice to confirm all")
                        choice = input("> ").strip()
                        if choice == "":
                            missing = []
                            if user_prop_nm     in ("", "<User Property Name>"):            missing.append("[1]")
                            if prop_bs_clk_con  in ("", "<Property Base Clock Condition>"): missing.append("[2]")
                            if dis_con          in ("", "<Disable Condition>"):             missing.append("[3]")
                            if trig_con         in ("", "<Trigger Condition>"):             missing.append("[4]")
                            if user_res         in ("", "<User Result>"):                   missing.append("[5]")
                            if missing: print(f"{','.join(missing)} has NOT been entered yet.")
                            print("Press Enter again to confirm, or select item number to edit")
                            choice = input("> ").strip()
                            if choice == "":
                                break
                        if choice == "1":
                            user_prop_nm = _scan("Enter User Property Name")
                        elif choice == "2":
                            prop_bs_clk_con = _scan("Enter Property Base Clock Condition")
                        elif choice == "3":
                            dis_con = _scan("Enter Disable Condition")
                        elif choice == "4":
                            trig_con = _scan("Enter Trigger Condition")
                        elif choice == "5":
                            user_res = _scan_wrap("Enter User Result")
                        else:
                            print("Invalid selection. Try again.")
                            continue
                    _append_label_below(ws_w, "User Property Name", user_prop_nm)
                    _append_label_below(ws_w, "Property Base Clock Condition", prop_bs_clk_con)
                    _append_label_below(ws_w, "Disable Condition", dis_con)
                    _append_label_below(ws_w, "Trigger Condition", trig_con)
                    _append_label_below(ws_w, "User Result", user_res)
                    prop_sets.append({"User Property Name": user_prop_nm, "Property Base Clock Condition": prop_bs_clk_con, "Disable Condition": dis_con, "Trigger Condition": trig_con, "User Result": user_res})
                    break
                if sel == "2":
                    user_seq_nm = "<User Sequence Name>"
                    seq_bs_clk_con = "<Sequence Base Clock Condition>"
                    user_seq = "<User Sequence>"
                    while True:
                        print("\n==================== Enter Sequence ====================")
                        print(f"sequence [1]{user_seq_nm}();")
                        print(f"    @([2]{seq_bs_clk_con})")
                        formatted_seq = user_seq.replace('\n', '\n    ')
                        print(f"    [3]{formatted_seq}")
                        print("endsequence")
                        print("=========================================================")
                        print("Select item number to edit")
                        print("Press Enter twice to confirm all")
                        choice = input("> ").strip()
                        if choice == "":
                            missing = []
                            if user_seq_nm      in ("", "<User Sequence Name>"):            missing.append("[1]")
                            if seq_bs_clk_con   in ("", "<Sequence Base Clock Condition>"): missing.append("[2]")
                            if user_seq         in ("", "<User Sequence>"):                 missing.append("[3]")
                            if missing: print(f"{','.join(missing)} has NOT been entered yet.")
                            print("Press Enter again to confirm, or select item number to edit")
                            choice = input("> ").strip()
                            if choice == "":
                                break
                        if choice == "1":
                            user_seq_nm = _scan("Enter User Sequence Name")
                        elif choice == "2":
                            seq_bs_clk_con = _scan("Enter Sequence Base Clock Condition")
                        elif choice == "3":
                            user_seq = _scan_wrap("Enter User Sequence")
                        else:
                            print("Invalid selection. Try again.")
                            continue
                    _append_label_below(ws_w, "User Sequence Name", user_seq_nm)
                    _append_label_below(ws_w, "Sequence Base Clock Condition", seq_bs_clk_con)
                    _append_label_below(ws_w, "User Sequence", user_seq)
                    seq_sets.append({"User Sequence Name": user_seq_nm, "Sequence Base Clock Condition": seq_bs_clk_con, "User Sequence": user_seq})
                    break
                print("Invalid selection. Try again.")

            # 추가 여부 프롬프트
            print(f"\nCurrent number of Sequence = {len(seq_sets)}, Property = {len(prop_sets)}\n")
            print("MORE?")
            print("[1] Yes")
            print("[2] No")
            while True:
                sel = input("> ").strip()
                if sel == "1":
                    break
                if sel == "2":
                    break
                print("Invalid selection. Try again.")
            if sel == "2":
                break

        # 저장
        wb_w.save(xls_path)

        # 4) Reopen and read all values (sheet)
        wb = load_workbook(xls_path, data_only=True)
        ws = _get_sheet_ci(wb, self.sheet_name, create=False)
        clk_do, rst_do = _read_base_clk_rst(ws)
        
        # Define 시트에서 Base Clock/Reset 읽기 (우선순위 높음)
        try:
            ws_define = _get_sheet_ci(wb, "Define", create=False)
            clk_define = _read_label_right_resolved(wb, ws_define, "Base Clock")
            rst_define = _read_label_right_resolved(wb, ws_define, "Base Reset")
            clk_do = clk_define or clk_do
            rst_do = rst_define or rst_do
        except KeyError:
            pass
        
        clk = clk_do or clk
        rst = rst_do or rst

        # 시트에서 누적 입력 다시 수집(보정)
        user_prop_nm_list = _read_column_values(ws, "User Property Name")
        prop_bs_clk_con_list = _read_column_values(ws, "Property Base Clock Condition")
        dis_con_list = _read_column_values(ws, "Disable Condition")
        trig_con_list = _read_column_values(ws, "Trigger Condition")
        user_res_list = _read_column_values(ws, "User Result")
        user_seq_nm_list = _read_column_values(ws, "User Sequence Name")
        seq_bs_clk_con_list = _read_column_values(ws, "Sequence Base Clock Condition")
        user_seq_list = _read_column_values(ws, "User Sequence")

        prop_n = max(len(prop_bs_clk_con_list), len(dis_con_list), len(trig_con_list), len(user_res_list))
        seq_n = max(len(user_seq_nm_list), len(seq_bs_clk_con_list), len(user_seq_list))
        prop_sets_final: List[Dict[str, str]] = []
        seq_sets_final: List[Dict[str, str]] = []
        for i in range(prop_n):
            user_prop_nm_i = user_prop_nm_list[i] if i < len(user_prop_nm_list) else ""
            prop_bs_clk_con_i = prop_bs_clk_con_list[i] if i < len(prop_bs_clk_con_list) else ""
            dis_con_i = dis_con_list[i] if i < len(dis_con_list) else ""
            trig_con_i = trig_con_list[i] if i < len(trig_con_list) else ""
            user_res_i = user_res_list[i] if i < len(user_res_list) else ""
            prop_sets_final.append({"User Property Name": user_prop_nm_i, "Property Base Clock Condition": prop_bs_clk_con_i, "Disable Condition": dis_con_i, "Trigger Condition": trig_con_i, "User Result": user_res_i})
        for i in range(seq_n):
            user_seq_nm_i = user_seq_nm_list[i] if i < len(user_seq_nm_list) else ""
            seq_bs_clk_con_i = seq_bs_clk_con_list[i] if i < len(seq_bs_clk_con_list) else ""
            user_seq_i = user_seq_list[i] if i < len(user_seq_list) else ""
            seq_sets_final.append({"User Sequence Name": user_seq_nm_i, "Sequence Base Clock Condition": seq_bs_clk_con_i, "User Sequence": user_seq_i})

        # 유니크 DUT 포트 수집(모듈 포트 선언용)
        unique_ports: List[str] = []
        def _add_port(name: str):
            for nm in _collect_expr_ports(mod, name or ""):
                if nm and nm not in unique_ports:
                    unique_ports.append(nm)
        _add_port(clk)
        _add_port(rst)
        for st in prop_sets_final:
            _add_port(st["Property Base Clock Condition"])
            _add_port(st["Disable Condition"])
            _add_port(st["Trigger Condition"])
            _add_port(st["User Result"])
        for st in seq_sets_final:
            _add_port(st["Sequence Base Clock Condition"])
            _add_port(st["User Sequence"])

        # width map
        width_map: Dict[str, str] = {}
        for nm in unique_ports:
            width_map[nm] = _port_width_token(mod, nm)

        # parsed 구조
        parsed: Dict[str, Any] = {
            "Base Clock": clk or "",
            "Base Reset": rst or "",
            "prop_sets": prop_sets,  # 이번 실행에서 입력한 세트만 사용
            "seq_sets": seq_sets,  # 이번 실행에서 입력한 세트만 사용
            "unique_ports": unique_ports,
            "width_map": width_map,
        }
        return parsed

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
         base_clk = parsed.get("Base Clock", "")
         base_rst = parsed.get("Base Reset", "")
         prop_sets: List[Dict[str, str]] = parsed.get("prop_sets", []) or []
         seq_sets: List[Dict[str, str]] = parsed.get("seq_sets", []) or []
         unique_ports: List[str] = parsed.get("unique_ports", []) or []
         width_map: Dict[str, str] = parsed.get("width_map", {}) or {}

         # 플레이스홀더 보정(비어 있을 때)
         base_clk_p = _nz(base_clk, "UNDEF_CLK")
         base_rst_p = _nz(base_rst, "UNDEF_RST")

         sv = _build_basicassertion_sv_multi(base_clk_p, base_rst_p, prop_sets, seq_sets, unique_ports, width_map)
         inst_sv = _build_basicassertion_inst_sv_multi(base_clk_p, base_rst_p, unique_ports)

         # 파일 저장은 assertion_builder.py에서 수행
         return [sv, inst_sv]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        return parsed