from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import re

import json
from openpyxl import load_workbook

from .base import BaseAssertionPlugin
from .registry import register

# -----------------------------
# Local helpers (self-contained)
# -----------------------------
def _get_sheet_ci(wb, name: str, create: bool = True):
    tgt = name.strip().lower()
    for s in wb.sheetnames:
        if str(s).strip().lower() == tgt:
            return wb[s]
    if not create:
        raise KeyError(name)
    return wb.create_sheet(title=name)

def _find_cell(ws, key: str) -> Tuple[Optional[int], Optional[int]]:
    key_l = str(key).strip().lower()
    for r in range(1, (ws.max_row or 1) + 1):
        for c in range(1, (ws.max_column or 1) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip().lower() == key_l:
                return r, c
    return None, None

def _read_json(p: Path) -> Optional[Dict[str, Any]]:
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None

def _load_module_define(xls_path: Path) -> Dict[str, Any]:
    root = xls_path.parent
    md = _read_json(root / "module_define.json")
    if md:
        return md
    ai = _read_json(root / "assertion_inputs.json")
    if ai:
        return {
            "module": ai.get("module") or "",
            "clocks": ai.get("clocks") or [],
            "resets": ai.get("resets") or [],
            "inputs": ai.get("inputs") or [],
            "outputs": ai.get("outputs") or [],
            "inouts": ai.get("inouts") or [],
            "ports": ai.get("ports") or [],
        }
    return {}

def _signal_options(mod: Dict[str, Any]) -> List[Tuple[str, str]]:
    opts: List[Tuple[str, str]] = []
    for it in (mod.get("inputs") or []):
        n = it.get("name")
        if n:
            opts.append((n, n))
    # dedup
    seen, uniq = set(), []
    for label, val in opts:
        if val in seen:
            continue
        seen.add(val); uniq.append((label, val))
    return uniq

def _pick_one(question: str, options: List[Tuple[str, str]], allow_custom: bool = False) -> str:
    print(question)
    for i, (label, _) in enumerate(options, start=1):
        print(f"  [{i}] {label}")
    if allow_custom:
        print("  [0] Custom Value")
    while True:
        sel = input("> ").strip()
        if allow_custom and sel == "0":
            return input("Enter value: ").strip()
        if sel.isdigit():
            i = int(sel)
            if 1 <= i <= len(options):
                return options[i - 1][1]
        print("Invalid selection. Try again.")

def _normalize_range_token(token: Any) -> str:
    if token is None:
        return "[0:0]"
    t = str(token).strip().replace(" ", "")
    if not t:
        return "[0:0]"
    if t.startswith("[") and t.endswith("]"):
        return t
    try:
        n = int(t, 10)
        return f"[{n-1}:0]" if n >= 1 else "[0:0]"
    except Exception:
        return "[0:0]"

def _port_width_token(mod: Dict[str, Any], name: str) -> str:
    if not name or not mod:
        return "[0:0]"
    want = (name or "").strip()
    cands = [
        mod.get("ports") or [],
        mod.get("inputs") or [],
        mod.get("outputs") or [],
        mod.get("inouts") or [],
        mod.get("clocks") or [],
        mod.get("resets") or [],
    ]
    for arr in cands:
        for it in arr:
            if (it.get("name") or "") != want:
                continue
            for k in ("packed_range", "range", "packed", "decl"):
                pr = it.get(k)
                if pr is not None and str(pr).strip():
                    return _normalize_range_token(pr)
            for k in ("width", "bit_width", "width_bits"):
                w = it.get(k)
                if w is not None and str(w).strip():
                    return _normalize_range_token(str(w))
            for lk, rk in (("msb", "lsb"), ("left", "right")):
                msb = it.get(lk); lsb = it.get(rk)
                if msb is not None and lsb is not None:
                    try:
                        return f"[{int(msb)}:{int(lsb)}]"
                    except Exception:
                        return f"[{msb}:{lsb}]"
            return "[0:0]"
    return "[0:0]"

def _fmt_input_decl(sig: str, width_tok: str) -> str:
    tok = (width_tok or "").strip() or "[0:0]"
    return f"input logic {tok} {sig}"

def _ensure_define_base_clk_rst(wb, module_info: Dict[str, Any]) -> None:
    ws = None
    for nm in wb.sheetnames:
        if str(nm).strip().lower() == "define":
            ws = wb[nm]; break
    if ws is None:
        ws = wb.create_sheet(title="Define")
    clk_r, clk_c = _find_cell(ws, "Base Clock")
    rst_r, rst_c = _find_cell(ws, "Base Reset")
    if clk_r is None:
        clk_r, clk_c = 2, 1
        ws.cell(row=clk_r, column=clk_c, value="Base Clock")
    if rst_r is None:
        rst_r, rst_c = 3, 1
        ws.cell(row=rst_r, column=rst_c, value="Base Reset")
    # try auto-fill from module info
    clk = ""
    for it in (module_info.get("clocks") or []):
        n = it.get("name");
        if n: clk = n; break
    rst = ""
    for it in (module_info.get("resets") or []):
        n = it.get("name");
        if n: rst = n; break
    if clk and not (ws.cell(row=clk_r, column=clk_c + 1).value):
        ws.cell(row=clk_r, column=clk_c + 1, value=clk)
    if rst and not (ws.cell(row=rst_r, column=rst_c + 1).value):
        ws.cell(row=rst_r, column=rst_c + 1, value=rst)

def _read_base_clk_rst(ws) -> Tuple[str, str]:
    cr, cc = _find_cell(ws, "Base Clock")
    rr, rc = _find_cell(ws, "Base Reset")
    clk = ws.cell(row=cr, column=cc + 1).value if cr else ""
    rst = ws.cell(row=rr, column=rc + 1).value if rr else ""
    return (str(clk).strip() if clk else ""), (str(rst).strip() if rst else "")

# 수식/시트 참조를 실제 값으로 역참조
def _resolve_ref(wb, raw: Any) -> str:
    s = "" if raw is None else str(raw).strip()
    if not s:
        return ""
    expr = s[1:].strip() if s.startswith("=") else s
    if "!" not in expr:
        return s
    sheet_name, addr = expr.split("!", 1)
    sheet_name = sheet_name.strip().strip("'").strip('"')
    try:
        ws2 = wb[sheet_name]
        v = ws2[addr].value
        return str(v).strip() if v is not None else ""
    except Exception:
        return s

def _read_label_right_resolved(wb, ws, label: str) -> str:
    r, c = _find_cell(ws, label)
    if r is None:
        return ""
    raw = ws.cell(row=r, column=c + 1).value
    val = _resolve_ref(wb, raw)
    return val if val != "" else (str(raw).strip() if raw is not None else "")

def _read_label_below_resolved(wb, ws, label: str) -> str:
    r, c = _find_cell(ws, label)
    if r is None:
        return ""
    raw = ws.cell(row=r + 1, column=c).value
    val = _resolve_ref(wb, raw)
    return val if val != "" else (str(raw).strip() if raw is not None else "")

# 아래/우측 라벨 유틸 + append 지원
def _read_label_below(ws, label: str) -> str:
    r, c = _find_cell(ws, label)
    if r is None:
        return ""
    v = ws.cell(row=r + 1, column=c).value
    return str(v).strip() if v is not None else ""

def _ensure_label(ws, label: str) -> Tuple[int, int]:
    r, c = _find_cell(ws, label)
    if r is not None:
        return r, c
    # 헤더 기준으로 첫 빈 라벨 슬롯에 생성
    hdr_r, hdr_c = _find_cell(ws, "DelayCondition")
    if hdr_r is None:
        hdr_r, hdr_c = 1, 1
        ws.cell(row=hdr_r, column=hdr_c, value="DelayCondition")
    r, c = hdr_r + 1, hdr_c
    while ws.cell(row=r, column=c).value not in (None, ""):
        r += 1
    ws.cell(row=r, column=c, value=label)
    return r, c

def _append_label_below(ws, label: str, value: str) -> None:
    r, c = _ensure_label(ws, label)
    rr = r + 1
    while ws.cell(row=rr, column=c).value not in (None, ""):
        rr += 1
    ws.cell(row=rr, column=c, value=value)

def _read_column_values(ws, label: str) -> List[str]:
    vals: List[str] = []
    r, c = _find_cell(ws, label)
    if r is None:
        return vals
    rr = r + 1
    while True:
        v = ws.cell(row=rr, column=c).value
        if v is None or str(v).strip() == "":
            break
        vals.append(str(v).strip())
        rr += 1
    return vals

def _ensure_dc_layout(ws) -> Tuple[int, Dict[str, int], int]:
    # Header
    h_r, h_c = _find_cell(ws, "DelayCondition")
    if h_r is None:
        h_r, h_c = 1, 1
        ws.cell(row=h_r, column=h_c, value="DelayCondition")
    # Labels (legacy table layout, kept for compatibility)
    labels = ["Condition", "Target", "Delay_Min", "Delay_Max"]
    lab_row = h_r + 1
    col_map: Dict[str, int] = {}
    for i, lab in enumerate(labels):
        c = h_c + i
        if not ws.cell(row=lab_row, column=c).value:
            ws.cell(row=lab_row, column=c, value=lab)
        col_map[lab] = c
    return h_r, col_map, lab_row + 1

def _update_delay_sheet(ws, cfg: Dict[str, Any]) -> int:
    h_r, cols, data_start = _ensure_dc_layout(ws)
    # find first empty row at Condition col
    r = data_start
    while True:
        v = ws.cell(row=r, column=cols["Condition"]).value
        if v is None or str(v).strip() == "":
            break
        r += 1
    ws.cell(row=r, column=cols["Condition"], value=cfg.get("Condition", ""))
    ws.cell(row=r, column=cols["Target"],    value=cfg.get("Target", ""))
    ws.cell(row=r, column=cols["Delay_Min"], value=cfg.get("Delay_Min", "1"))
    ws.cell(row=r, column=cols["Delay_Max"], value=cfg.get("Delay_Max", "1"))
    # Ensure Base Clock/Reset labels exist below header
    clk_r, clk_c = _find_cell(ws, "Base Clock")
    rst_r, rst_c = _find_cell(ws, "Base Reset")
    if clk_r is None:
        clk_r, clk_c = h_r + 3, 1
        ws.cell(row=clk_r, column=clk_c, value="Base Clock")
    if rst_r is None:
        rst_r, rst_c = h_r + 4, 1
        ws.cell(row=rst_r, column=rst_c, value="Base Reset")
    return r

def _sv_header() -> str:
    return '`include "uvm_macros.svh"\nimport uvm_pkg::*;\n\n'

def _nz(s: Optional[str], placeholder: str) -> str:
    return (s or "").strip() or placeholder

def _is_simple_ident(s: str) -> bool:
    return bool(re.match(r"^[A-Za-z_][A-Za-z0-9_$]*$", (s or "").strip()))

# -----------------------------
# Is Port helpers
# -----------------------------
def _all_module_port_names(mod: Dict[str, Any]) -> List[str]:
    names = []
    for key in ("inputs", "outputs", "inouts", "clocks", "resets", "ports"):
        for p in mod.get(key, []) or []:
            n = (p.get("name") or "").strip()
            if n:
                names.append(n)
    # 중복 제거(순서 보존)
    seen = set(); out = []
    for n in names:
        if n not in seen:
            seen.add(n); out.append(n)
    return out

def _collect_expr_ports(mod: Dict[str, Any], expr: str) -> List[str]:
    """
    expr 안에서 모듈 포트 이름이 실제로 등장하는 것만 추출.
    """
    if not expr or not mod:
        return []
    all_names = _all_module_port_names(mod)
    # 포트명이 특수문자 포함 가능성 대비해서 re.escape 사용, 단어 경계 \b로 안전 매칭
    if not all_names:
        return []
    pat = r'\b(' + '|'.join(re.escape(n) for n in sorted(all_names, key=len, reverse=True)) + r')\b'
    # 중복 제거(순서 보존)
    seen = set(); found = []
    for m in re.finditer(pat, expr):
        n = m.group(1)
        if n not in seen:
            seen.add(n); found.append(n)
    return found

# ---------------- SV builders (multi-sets) ----------------

def _build_delaycondition_sv_multi(base_clk: str, base_rst: str,
                                   sets: List[Dict[str, str]],
                                   unique_ports: List[str],
                                   width_map: Dict[str, str]) -> str:
    lines: List[str] = []
    lines.append(_sv_header())
    lines.append("module assertion_delayCondition")
    lines.append("(")
    # 포트: Base + 유니크 DUT 포트들
    ports: List[str] = []
    if base_clk:
        ports.append(_fmt_input_decl(base_clk, width_map.get(base_clk, "[0:0]")))
    if base_rst:
        ports.append(_fmt_input_decl(base_rst, width_map.get(base_rst, "[0:0]")))
    for p in unique_ports:
        if p in (base_clk, base_rst):
            continue
        ports.append(_fmt_input_decl(p, width_map.get(p, "[0:0]")))
    # emit with commas
    for i, decl in enumerate(ports):
        comma = "," if i < len(ports) - 1 else ","
        lines.append(f"    {decl}{comma}")
    lines.append(");")
    lines.append("")
    # 각 세트별 property/assert
    for idx, st in enumerate(sets, start=1):
        trig = st.get("Trigger", "")
        res  = st.get("Result", "")
        d1   = st.get("Delay1", "1")
        d2   = st.get("Delay2", d1 or "1")
        lines.append(f"property p_delayCondition_check{idx}(trigger, result);")
        lines.append(f"    @(posedge {base_clk}) disable iff(!{base_rst})")
        lines.append(f"    $rose(trigger) |-> ##[{d1} : {d2}] $rose(result);")
        lines.append("endproperty")
        lines.append("")
        lines.append(f'assert property (p_delayCondition_check{idx}({trig}, {res})) else $error("failed at %t", $time);')
        lines.append("")
    lines.append("endmodule")
    lines.append("")
    return "\n".join(lines)

def _build_delaycondition_inst_sv_multi(base_clk: str, base_rst: str,
                                        unique_ports: List[str]) -> str:
    lines: List[str] = []
    lines.append(_sv_header())
    lines.append("assertion_delayCondition u_assertion_delayCondition();")
    lines.append("")
    # base + unique 포트 assign
    ports = []
    if base_clk: ports.append(base_clk)
    if base_rst: ports.append(base_rst)
    ports.extend([p for p in unique_ports if p not in (base_clk, base_rst)])
    for p in ports:
        lines.append(f"assign u_assertion_delayCondition.{p}  = top.dut.{p};")
    lines.append("")
    return "\n".join(lines)

# -----------------------------
# Plugin
# -----------------------------
@register
class DelayConditionPlugin(BaseAssertionPlugin):
    plugin_name = "delayCondition"
    sheet_name = "DelayCondition"

    def parse(self, xls_path: Path) -> Dict[str, Any]:
        # 0) Load module info
        mod = _load_module_define(Path(xls_path))

        # 1) Open sheet and ensure header exists
        wb_w = load_workbook(xls_path)
        try:
            ws_w = _get_sheet_ci(wb_w, self.sheet_name, create=False)
        except KeyError:
            ws_w = _get_sheet_ci(wb_w, self.sheet_name, create=True)
        # Ensure header label
        if _find_cell(ws_w, "DelayCondition")[0] is None:
            ws_w.cell(row=1, column=1, value="DelayCondition")

        # 2) Base Clock/Reset from label-right (resolve references, no prompts)
        clk = _read_label_right_resolved(wb_w, ws_w, "Base Clock")
        rst = _read_label_right_resolved(wb_w, ws_w, "Base Reset")

        # 3) 세트 입력 루프
        sets: List[Dict[str, str]] = []
        while True:
            # Trigger
            trig = _pick_one("Select Trigger signal", _signal_options(mod), allow_custom=True)
            _append_label_below(ws_w, "Trigger", trig)
            # Delay1/Delay2
            def _ask_int(prompt: str, default: str) -> str:
                while True:
                    s = input(f"{prompt} (integer, default {default}): ").strip() or default
                    if s.isdigit():
                        return s
                    print("Invalid integer. Try again.")
            d1 = _ask_int("Enter Delay1", "1")
            d2 = _ask_int("Enter Delay2", d1)
            _append_label_below(ws_w, "Delay1", d1)
            _append_label_below(ws_w, "Delay2", d2)
            # Result
            res = _pick_one("Select Result signal", _signal_options(mod), allow_custom=True)
            _append_label_below(ws_w, "Result", res)

            sets.append({"Trigger": trig, "Delay1": d1, "Delay2": d2, "Result": res})

            # 카운트/추가 여부 프롬프트
            print(f"\nCurrent number of generated Assertions = {len(sets)}\n")
            print("Would you like to generate additional Assertions?")
            print("[1] Yes")
            print("[2] No")
            while True:
                sel = input("> ").strip()
                if sel == "1":
                    break
                if sel == "2":
                    break
                print("Invalid selection. Try again.")
            if sel == "2":
                break

        # 저장
        wb_w.save(xls_path)

        # 4) Reopen and read all values (sheet)
        wb = load_workbook(xls_path, data_only=True)
        ws = _get_sheet_ci(wb, self.sheet_name, create=False)
        clk_do, rst_do = _read_base_clk_rst(ws)
        if not clk_do:
            clk_do = _read_label_right_resolved(wb_w, ws_w, "Base Clock")
        if not rst_do:
            rst_do = _read_label_right_resolved(wb_w, ws_w, "Base Reset")
        clk = clk_do or clk
        rst = rst_do or rst

        # 시트에서 누적 입력 다시 수집(보정)
        trig_list = _read_column_values(ws, "Trigger")
        d1_list   = _read_column_values(ws, "Delay1")
        d2_list   = _read_column_values(ws, "Delay2")
        res_list  = _read_column_values(ws, "Result")
        n = max(len(trig_list), len(d1_list), len(d2_list), len(res_list))
        sets_final: List[Dict[str, str]] = []
        for i in range(n):
            trig_i = trig_list[i] if i < len(trig_list) else ""
            d1_i   = d1_list[i] if i < len(d1_list) else "1"
            d2_i   = d2_list[i] if i < len(d2_list) else (d1_i or "1")
            res_i  = res_list[i] if i < len(res_list) else ""
            sets_final.append({"Trigger": trig_i, "Delay1": d1_i, "Delay2": d2_i, "Result": res_i})

        # 유니크 DUT 포트 수집(모듈 포트 선언용)
        unique_ports: List[str] = []
        def _add_port(name: str):
            for nm in _collect_expr_ports(mod, name or ""):
                if nm and nm not in unique_ports:
                    unique_ports.append(nm)
        _add_port(clk)
        _add_port(rst)
        for st in sets_final:
            _add_port(st["Trigger"])
            _add_port(st["Result"])

        # width map
        width_map: Dict[str, str] = {}
        for nm in unique_ports:
            width_map[nm] = _port_width_token(mod, nm)

        # parsed 구조
        parsed: Dict[str, Any] = {
            "Base Clock": clk or "",
            "Base Reset": rst or "",
            "sets": sets,  # 이번 실행에서 입력한 세트만 사용
            "unique_ports": unique_ports,
            "width_map": width_map,
        }
        return parsed

    def generate_sv(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> List[str]:
         base_clk = parsed.get("Base Clock", "")
         base_rst = parsed.get("Base Reset", "")
         sets: List[Dict[str, str]] = parsed.get("sets", []) or []
         unique_ports: List[str] = parsed.get("unique_ports", []) or []
         width_map: Dict[str, str] = parsed.get("width_map", {}) or {}

         # 플레이스홀더 보정(비어 있을 때)
         base_clk_p = _nz(base_clk, "UNDEF_CLK")
         base_rst_p = _nz(base_rst, "UNDEF_RST")

         sv = _build_delaycondition_sv_multi(base_clk_p, base_rst_p, sets, unique_ports, width_map)
         inst_sv = _build_delaycondition_inst_sv_multi(base_clk_p, base_rst_p, unique_ports)

         # 파일 저장은 assertion_builder.py에서 수행
         return [sv, inst_sv]

    def emit_json(self, parsed: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        return parsed