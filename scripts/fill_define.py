#!/usr/bin/env python3
import sys
import json
import re
import logging
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from copy import copy
from openpyxl.cell.cell import MergedCell

# -------------------- 전역 로거 --------------------
logger = None  # Will be initialized in main()

# -------------------- 로거 --------------------
def setup_logger(log_path: str = "fill_define.log"):
    global logger
    logger = logging.getLogger("fill_define")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fmt = logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s", "%H:%M:%S")

    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger

# -------------------- 서식 복사 헬퍼 --------------------
def copy_cell_format(source_cell, target_cell):
    """Copy formatting from source cell to target cell (keeps styles intact)"""
    if source_cell is None or target_cell is None:
        return
    
    try:
        # Copy font
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        
        # Copy border
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        
        # Copy fill
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        
        # Copy number format
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        
        # Copy alignment
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        
        # Copy protection
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
    except Exception:
        pass  # Silently ignore format copy errors

def copy_row_format(ws, source_row, target_row, start_col=1, end_col=None, logger=None):
    """Copy formatting from source row to target row"""
    if end_col is None:
        end_col = ws.max_column
    
    if logger:
        logger.debug(f"copy_row_format: source={source_row}, target={target_row}, cols={start_col}-{end_col}")
    
    for col in range(start_col, end_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        
        # Debug: check if source has any formatting
        if logger and col == start_col:
            has_border = source_cell.border and any([
                source_cell.border.left and source_cell.border.left.style,
                source_cell.border.right and source_cell.border.right.style,
                source_cell.border.top and source_cell.border.top.style,
                source_cell.border.bottom and source_cell.border.bottom.style
            ])
            logger.debug(f"  Source cell ({source_row},{col}) has_border={has_border}, font={source_cell.font.name if source_cell.font else 'None'}")
        
        copy_cell_format(source_cell, target_cell)

def set_cell_value_with_template(ws, row, col, value, template_row=None):
    """
    Set cell value while preserving or copying format.
    If template_row is provided, copy format from that row's same column.
    """
    target_cell = ws.cell(row=row, column=col)
    
    # Copy format from template if provided
    if template_row:
        template_cell = ws.cell(row=template_row, column=col)
        copy_cell_format(template_cell, target_cell)
    
    # Set value
    target_cell.value = value

# -------------------- 진행 표시 --------------------
class Progress:
    def __init__(self, total: int, bar_width: int = 28):
        self.total = max(1, int(total))
        self.count = 0
        self.bar_width = bar_width

    def step(self, msg: str):
        self.count += 1
        pct = int(self.count * 100 / self.total)
        filled = int(self.bar_width * pct / 100)
        bar = "#" * filled + "-" * (self.bar_width - filled)
        print(f"[{bar}] {pct:3d}% - {msg}")

# -------------------- 파일 백업 --------------------
# Backup Excel 기능 제거됨 (2025-10-14)

# -------------------- 엑셀 헬퍼 --------------------
def find_define_sheet(wb):
    for name in wb.sheetnames:
        if str(name).strip().lower() == "define":
            return wb[name]
    return wb.create_sheet("define")

def find_label_cell(ws, label: str):
    target = label.strip().casefold()
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.strip().casefold() == target:
                return cell
    return None

def ensure_base_label(ws, label: str):
    cell = find_label_cell(ws, label)
    if cell:
        return cell
    # 없으면 아래쪽에 라벨 생성(서식은 템플릿 없으면 기본)
    row = ws.max_row + 2
    ws.cell(row=row, column=1).value = label
    return ws.cell(row=row, column=1)

def right_cell(ws, cell):
    return ws.cell(row=cell.row, column=cell.column + 1)

def find_io_header(ws):
    # 한 줄에 Inputs/Bits, Outputs/Bits, Parameters/Bits가 있는 헤더 탐지
    for row in ws.iter_rows():
        header_map = {}
        for cell in row:
            if isinstance(cell.value, str):
                key = cell.value.strip().casefold()
                if key in ("inputs", "outputs", "parameters"):
                    header_map[key] = cell.column
        if header_map.get("inputs") and header_map.get("outputs") and header_map.get("parameters"):
            in_c = header_map["inputs"]
            out_c = header_map["outputs"]
            par_c = header_map["parameters"]
            return {
                "row": row[0].row,
                "inputs_col": in_c, "inputs_bits_col": in_c + 1,
                "outputs_col": out_c, "outputs_bits_col": out_c + 1,
                "params_col": par_c, "params_bits_col": par_c + 1
            }
    return None


def find_signal_assignments_header(ws):
    """
    Find 'Signal Assignments' header row with Name, Equation, Bits columns.
    Returns dict with row and column info, or None.
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "signal" in cell.value.strip().casefold() and "assignment" in cell.value.strip().casefold():
                # Found "Signal Assignments", now find Name/Equation/Bits in next row or same row
                header_row = cell.row
                logger.debug(f"Found 'Signal Assignments' at row {header_row}, col {cell.column}")
                # Check next row for Name/Equation/Bits
                next_row = list(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1))
                if next_row:
                    name_col = None
                    equation_col = None
                    bits_col = None
                    for c in next_row[0]:
                        if isinstance(c.value, str):
                            val = c.value.strip().casefold()
                            if val == "name":
                                # Check if this Name is close to "Signal Assignments"
                                # Signal Assignments is at cell.column, we want Name in same region
                                if c.column >= cell.column - 1:
                                    name_col = c.column
                            elif val == "equation":
                                equation_col = c.column
                            elif val == "bits":
                                # Only accept Bits near Equation (avoid IO section Bits)
                                if equation_col and c.column == equation_col + 1:
                                    bits_col = c.column
                    logger.debug(f"  Found columns: name={name_col}, equation={equation_col}, bits={bits_col}")
                    if name_col and equation_col:
                        return {
                            "header_row": header_row,
                            "data_row": header_row + 1,
                            "name_col": name_col,
                            "equation_col": equation_col,
                            "bits_col": bits_col or (equation_col + 1)
                        }
    logger.warning("Signal Assignments header NOT found in Define sheet")
    return None

def find_defines_header(ws):
    """
    Find 'Defines' header row with Name, Value columns.
    Returns dict with row and column info, or None.
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().casefold() == "defines":
                # Found "Defines", now find Name/Value in next row
                header_row = cell.row
                logger.debug(f"Found 'Defines' at row {header_row}, col {cell.column}")
                # Check next row for Name/Value
                next_row = list(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1))
                if next_row:
                    name_col = None
                    value_col = None
                    for c in next_row[0]:
                        if isinstance(c.value, str):
                            val = c.value.strip().casefold()
                            if val == "name":
                                name_col = c.column
                            elif val == "value":
                                value_col = c.column
                    logger.debug(f"  Found columns: name={name_col}, value={value_col}")
                    if name_col and value_col:
                        return {
                            "header_row": header_row,
                            "data_row": header_row + 1,
                            "name_col": name_col,
                            "value_col": value_col
                        }
    logger.debug("Defines header NOT found in Define sheet")
    return None

def create_io_header(ws):
    # 템플릿 없으면 간단 헤더 생성(서식 없음)
    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1).value = "Inputs"
    ws.cell(row=start_row, column=2).value = "Bits"
    ws.cell(row=start_row, column=4).value = "Outputs"
    ws.cell(row=start_row, column=5).value = "Bits"
    ws.cell(row=start_row, column=7).value = "Parameters"
    ws.cell(row=start_row, column=8).value = "Bits"
    return {
        "row": start_row,
        "inputs_col": 1, "inputs_bits_col": 2,
        "outputs_col": 4, "outputs_bits_col": 5,
        "params_col": 7, "params_bits_col": 8
    }

def next_empty_row(ws, col, start_row):
    r = max(1, start_row)
    maxr = ws.max_row + 2
    while r <= maxr:
        if ws.cell(row=r, column=col).value in (None, ""):
            return r
        r += 1
    return maxr

# ---------- 스타일 복사(서식 보존) ----------
def copy_cell_style(dst, src):
    # src가 셀이고 서식이 없다면 스킵
    if not src or not getattr(src, "has_style", False):
        return
    # 병합 셀이면 스킵(혹은 병합 범위 좌상단 셀로 대체)
    if isinstance(src, MergedCell):
        return
    try:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        # number_format은 문자열이라 복사 불필요
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    except TypeError:
        # 일부 버전/상황에서 StyleProxy가 섞여 들어오면 여기로 옴
        # 서식 복사는 생략하고 값만 쓰도록 조용히 통과
        pass


def clone_row_style(ws, template_row: int, target_row: int, cols):
    # 템플릿 행의 각 셀 스타일을 안전하게 복사
    for c in cols:
        tcell = ws.cell(row=target_row, column=c)
        scell = ws.cell(row=template_row, column=c)
        copy_cell_style(tcell, scell)

# ---------- 폭 변환([msb:lsb] -> 비트수 정수 텍스트) ----------
def bits_text(w):
    try:
        if not isinstance(w, str):
            return "1"
        s = w.strip()
        if s == "" or s.lower() == "none":
            return "1"
        m = re.match(r'^\[\s*(\d+)\s*:\s*(\d+)\s*\]$', s)
        if m:
            msb = int(m.group(1))
            lsb = int(m.group(2))
            return str(abs(msb - lsb) + 1)
        m = re.match(r'^\[\s*\d+\s*\]$', s)
        if m:
            return "1"
        if s.isdigit():
            return s
        # 수식/매크로는 원문 유지
        return s
    except Exception:
        return "1"

def set_cell_value_merged_safe(ws, row, col, value):
    """
    병합 셀 범위 내라면 좌상단 셀에만 값을 넣고, 나머지는 무시
    """
    for merged_range in ws.merged_cells.ranges:
        # merged_range가 문자열이 아닐 경우 무시
        try:
            if (row, col) in merged_range:
                min_row, min_col = merged_range.min_row, merged_range.min_col
                if (row, col) == (min_row, min_col):
                    ws.cell(row=row, column=col).value = value
                return
        except TypeError:
            continue
    ws.cell(row=row, column=col).value = value

def clear_base(ws):
    for label in ("Target Name", "Target Path", "Base Clock", "Base Reset"):
        c = find_label_cell(ws, label)
        if c:
            rc = right_cell(ws, c)
            rc.value = None  # 값만 지움(서식 유지)

def clear_defines(ws):
    """Clear all data in Defines section (keep formatting)"""
    hdr = find_defines_header(ws)
    if not hdr:
        logger.debug("Defines header not found, nothing to clear")
        return
    
    start = hdr["data_row"] + 1  # Start after header row
    cols = [hdr["name_col"], hdr["value_col"]]
    
    end = last_used_row_in_cols(ws, cols, start)
    if end < start:
        logger.debug("No data to clear in Defines")
        return
    
    logger.debug(f"Clearing Defines from row {start} to {end}")
    for r in range(start, end + 1):
        for c in cols:
            ws.cell(row=r, column=c).value = None  # 값만 지움(서식 유지)

def clear_io(ws):
    hdr = find_io_header(ws)
    if not hdr:
        return
    start = hdr["row"] + 1
    cols = [
        hdr["inputs_col"], hdr["inputs_bits_col"],
        hdr["outputs_col"], hdr["outputs_bits_col"],
        hdr["params_col"], hdr["params_bits_col"]
    ]
    end = last_used_row_in_cols(ws, cols, start)
    if end < start:
        return
    for r in range(start, end + 1):
        for c in cols:
            ws.cell(row=r, column=c).value = None  # 값만 지움(서식 유지)

def clear_signal_assignments(ws):
    """Clear all data in Signal Assignments section (keep formatting)"""
    hdr = find_signal_assignments_header(ws)
    if not hdr:
        logger.debug("Signal Assignments header not found, nothing to clear")
        return
    
    start = hdr["data_row"] + 1  # Start after header row
    cols = [hdr["name_col"], hdr["equation_col"], hdr["bits_col"]]
    
    end = last_used_row_in_cols(ws, cols, start)
    if end < start:
        logger.debug("No data to clear in Signal Assignments")
        return
    
    logger.debug(f"Clearing Signal Assignments from row {start} to {end}")
    for r in range(start, end + 1):
        for c in cols:
            ws.cell(row=r, column=c).value = None  # 값만 지움(서식 유지)

def last_used_row_in_cols(ws, cols, start_row):
    """
    주어진 열들 중에서 start_row부터 마지막으로 값이 있는 행 번호를 반환
    """
    max_row = ws.max_row
    last = start_row - 1
    for r in range(start_row, max_row + 1):
        for c in cols:
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                last = r
    return last

# -------------------- 메인 --------------------
def main():
    if len(sys.argv) != 3:
        print("사용법: python fill_define.py <엑셀파일> <JSON파일>")
        sys.exit(1)

    excel_path = sys.argv[1]
    json_path = sys.argv[2]
    setup_logger("fill_define.log")  # Initialize global logger

    # JSON 로드
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    module = data.get("module") or ""
    rtl_file_path = data.get("rtl_file_path") or ""  # Full path for internal use
    rtl_hierarchy = data.get("rtl_hierarchy") or ""  # Hierarchy path for display (EDA/RTL/file.v)
    top_path = data.get("top_path") or ""
    paths = data.get("paths") or []
    clocks = data.get("clocks") or []
    resets = data.get("resets") or []
    params = data.get("parameters") or []
    inputs = data.get("inputs") or []
    outputs = data.get("outputs") or []
    conditions = data.get("conditions") or []  # Add condition signals
    
    logger.info(f"Loaded data: module={module}, inputs={len(inputs)}, outputs={len(outputs)}, conditions={len(conditions)}")
    if conditions:
        logger.info(f"Conditions to process: {[c.get('name','?') for c in conditions]}")

    # Target Path: Extract parent directory path (everything except the module filename)
    # E.g., if rtl_file_path is "EDA/RTL/blur_scaler.v", show "EDA/RTL"
    if rtl_file_path:
        # Remove the module filename (last component) to get parent path
        parent_path = str(Path(rtl_file_path).parent)
        target_path_str = parent_path if parent_path != "." else module
    elif rtl_hierarchy:
        # If no full path, use hierarchy if available
        target_path_str = rtl_hierarchy
    elif module:
        target_path_str = module  # Use just module name, not full file path
    elif top_path:
        target_path_str = ",".join(f"{top_path}.{p}" if p else top_path for p in paths) if paths else top_path
    else:
        target_path_str = ",".join(p for p in paths)

    # 워크북 로드
    wb = load_workbook(excel_path)
    ws = find_define_sheet(wb)

    tasks = []

    # 1) 클리어
    def run_clear():
        clear_base(ws)
        clear_io(ws)
        clear_defines(ws)  # Clear Defines section
        clear_signal_assignments(ws)  # Also clear Signal Assignments section
        logger.debug("기존 값 클리어(서식 유지): Base, IO, Defines, Signal Assignments")
    tasks.append((run_clear, "기존 값 클리어"))

    # 2) IO 헤더 없으면 생성
    if not find_io_header(ws):
        def run_header():
            create_io_header(ws)
            logger.debug("IO 헤더 생성")
        tasks.append((run_header, "IO 헤더 생성"))

    # 3) Base 재기입(항상 덮어씀)
    def run_base_target_name():
        c = ensure_base_label(ws, "Target Name")
        set_cell_value_merged_safe(ws, c.row, c.column + 1, module)
        logger.debug("Base 채움: Target Name -> %s", module)
    tasks.append((run_base_target_name, "Base: Target Name"))

    def run_base_target_path():
        c = ensure_base_label(ws, "Target Path")
        set_cell_value_merged_safe(ws, c.row, c.column + 1, target_path_str)
        logger.debug("Base 채움: Target Path -> %s", target_path_str)
    tasks.append((run_base_target_path, "Base: Target Path"))

    def run_base_clock():
        c = ensure_base_label(ws, "Base Clock")
        clk = ",".join(x.get("name", "") for x in clocks if x.get("name"))
        set_cell_value_merged_safe(ws, c.row, c.column + 1, clk)
        logger.debug("Base 채움: Base Clock -> %s", clk)
    tasks.append((run_base_clock, "Base: Base Clock"))

    def run_base_reset():
        c = ensure_base_label(ws, "Base Reset")
        rst = ",".join(x.get("name", "") for x in resets if x.get("name"))
        set_cell_value_merged_safe(ws, c.row, c.column + 1, rst)
        logger.debug("Base 채움: Base Reset -> %s", rst)
    tasks.append((run_base_reset, "Base: Base Reset"))
    
    # Fill Defines section with parameters
    def run_defines():
        hdr = find_defines_header(ws)
        if hdr and params:
            row = hdr["data_row"] + 1
            for param in params:
                param_name = param.get("name", "")
                param_value = param.get("default", "")
                if param_name:
                    ws.cell(row=row, column=hdr["name_col"]).value = param_name
                    ws.cell(row=row, column=hdr["value_col"]).value = param_value
                    row += 1
            logger.debug("Defines 채움: %d parameters", len(params))
        else:
            if not hdr:
                logger.debug("Defines header not found, skipping")
            if not params:
                logger.debug("No parameters to fill")
    tasks.append((run_defines, "Defines: Parameters"))

    # 4) Inputs
    for it in inputs:
        name = it.get("name")
        if not name:
            continue
        bits = bits_text(it.get("width"))
        def run_in(n=name, b=bits):
            hdr = find_io_header(ws) or create_io_header(ws)
            template_row = hdr["row"] + 1  # Template row for format
            row = next_empty_row(ws, hdr["inputs_col"], hdr["row"] + 1)
            logger.debug(f"Input: {n} -> writing to row={row}, template_row={template_row}")
            # Copy template row format to preserve borders/fonts/colors
            copy_row_format(ws, template_row, row, start_col=hdr["inputs_col"], end_col=hdr["inputs_bits_col"], logger=logger)
            set_cell_value_merged_safe(ws, row, hdr["inputs_col"], n)
            set_cell_value_merged_safe(ws, row, hdr["inputs_bits_col"], b)
            logger.debug("Input 추가: %s (%s)", n, b)
        tasks.append((run_in, f"Inputs: {name}"))

    # 5) Outputs
    for it in outputs:
        name = it.get("name")
        if not name:
            continue
        bits = bits_text(it.get("width"))
        def run_out(n=name, b=bits):
            hdr = find_io_header(ws) or create_io_header(ws)
            template_row = hdr["row"] + 1  # Template row for format
            row = next_empty_row(ws, hdr["outputs_col"], hdr["row"] + 1)
            logger.debug(f"Output: {n} -> writing to row={row}, template_row={template_row}")
            # Copy template row format to preserve borders/fonts/colors
            copy_row_format(ws, template_row, row, start_col=hdr["outputs_col"], end_col=hdr["outputs_bits_col"], logger=logger)
            set_cell_value_merged_safe(ws, row, hdr["outputs_col"], n)
            set_cell_value_merged_safe(ws, row, hdr["outputs_bits_col"], b)
            logger.debug("Output 추가: %s (%s)", n, b)
        tasks.append((run_out, f"Outputs: {name}"))

    # 6) Parameters
    for it in params:
        if isinstance(it, dict):
            pname = it.get("name") or ""
            pbits = bits_text(it.get("width") or it.get("bits") or "")
        else:
            pname = str(it)
            pbits = ""
        if not pname:
            continue
        def run_pa(n=pname, b=pbits):
            hdr = find_io_header(ws) or create_io_header(ws)
            template_row = hdr["row"] + 1  # Template row for format
            row = next_empty_row(ws, hdr["params_col"], hdr["row"] + 1)
            logger.debug(f"Parameter: {n} -> writing to row={row}, template_row={template_row}")
            # Copy template row format to preserve borders/fonts/colors
            copy_row_format(ws, template_row, row, start_col=hdr["params_col"], end_col=hdr["params_bits_col"], logger=logger)
            set_cell_value_merged_safe(ws, row, hdr["params_col"], n)
            set_cell_value_merged_safe(ws, row, hdr["params_bits_col"], b)
            logger.debug("Parameter 추가: %s (%s)", n, b)
        tasks.append((run_pa, f"Parameters: {pname}"))

    # 7) Condition Signals (Signal Assignments)
    for cond in conditions:
        if isinstance(cond, dict):
            cname = cond.get("name", "")
            cexpr = cond.get("expr", "")
            cwidth = cond.get("width") or cond.get("bits") or 1  # Support both 'width' and 'bits'
        else:
            continue
        if not cname or not cexpr:
            continue
        cbits = str(cwidth) if cwidth else "1"  # Always show bits, default to "1"
        def run_cond(n=cname, e=cexpr, b=cbits):
            hdr = find_signal_assignments_header(ws)
            if not hdr:
                # Signal Assignments header not found, skip
                logger.warning("Signal Assignments header not found, skipping condition: %s", n)
                return
            template_row = hdr["data_row"] + 1  # Template row for format
            row = next_empty_row(ws, hdr["name_col"], hdr["data_row"] + 1)
            logger.debug(f"Condition: {n} -> writing to row={row}, template_row={template_row}")
            # Copy template row format to preserve borders/fonts/colors
            copy_row_format(ws, template_row, row, start_col=hdr["name_col"], end_col=hdr["bits_col"], logger=logger)
            set_cell_value_merged_safe(ws, row, hdr["name_col"], n)
            set_cell_value_merged_safe(ws, row, hdr["equation_col"], e)
            set_cell_value_merged_safe(ws, row, hdr["bits_col"], b)  # Always write bits (default "1")
            logger.debug("Condition Signal 추가: %s = %s (%s bits)", n, e, b)
        tasks.append((run_cond, f"Condition: {cname}"))

    # 8) 저장
    def run_save():
        wb.save(excel_path)
        logger.debug("저장 완료: %s", excel_path)
    tasks.append((run_save, "엑셀 저장"))

    # 실행 + 진행률
    prog = Progress(total=len(tasks))
    logger.info("작업 시작 (총 %d 단계)", len(tasks))
    for runner, desc in tasks:
        runner()
        prog.step(desc)
    logger.info("모든 작업 완료")

if __name__ == "__main__":
    main()