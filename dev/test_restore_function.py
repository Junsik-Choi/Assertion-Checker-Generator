#!/usr/bin/env python3
"""Test the fixed _restore_assertions_from_excel function"""
import sys
from pathlib import Path
import shutil

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from openpyxl import load_workbook
from cli_tui import _restore_assertions_from_excel, AppState
from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional

# Create a test Excel with known data
test_dir = Path("out/sessions/test_restore")
test_dir.mkdir(parents=True, exist_ok=True)

# Copy the Excel file from final_test
src_excel = Path("out/sessions/final_test/test.xlsx")
if not src_excel.exists():
    print("ERROR: Need final_test/test.xlsx. Run final_comprehensive_test.py first.")
    sys.exit(1)

test_excel = test_dir / "test.xlsx"
shutil.copy2(src_excel, test_excel)

print("=" * 80)
print("TEST: _restore_assertions_from_excel")
print("=" * 80)
print(f"\nTest Excel: {test_excel}\n")

# First, check what's in the Excel file
print("[1] Excel File Contents")
print("-" * 80)
wb = load_workbook(str(test_excel))

# Check Counter
for name in wb.sheetnames:
    if name.lower() == 'counter':
        ws = wb[name]
        print(f"\nCounter Sheet ('{name}'):")
        for row in range(8, min(11, ws.max_row + 1)):
            target = ws.cell(row, 2).value
            plus = ws.cell(row, 3).value
            if target:
                print(f"  Row {row}: {target} | {plus}")

# Check Handshake
for name in wb.sheetnames:
    if name.lower() == 'handshake':
        ws = wb[name]
        print(f"\nHandshake Sheet ('{name}'):")
        for row in range(7, min(10, ws.max_row + 1)):
            ptype = ws.cell(row, 3).value
            sender = ws.cell(row, 4).value
            if ptype:
                print(f"  Row {row}: {ptype} | {sender}")

# Check PulseWidth
for name in wb.sheetnames:
    if name.lower() == 'pulsewidth':
        ws = wb[name]
        print(f"\nPulseWidth Sheet ('{name}'):")
        for row in range(7, min(10, ws.max_row + 1)):
            ptype = ws.cell(row, 3).value
            signal = ws.cell(row, 5).value
            min_w = ws.cell(row, 6).value
            if ptype and signal:
                print(f"  Row {row}: {ptype} | {signal} | min:{min_w}")

wb.close()

# Now test the restore function
print("\n[2] Testing _restore_assertions_from_excel")
print("-" * 80)

# Create mock state
@dataclass
class MockModuleInfo:
    module: str = "test"
    module_hierarchy: str = ""

@dataclass
class MockState:
    session_excel_path: Optional[Path] = test_excel
    assertions: List[Dict[str, Any]] = field(default_factory=list)
    module_info: MockModuleInfo = field(default_factory=MockModuleInfo)

state = MockState()

# Call the restore function
_restore_assertions_from_excel(state)

print(f"\nRestored {len(state.assertions)} assertions:\n")

# Group by type
by_type = {}
for asrt in state.assertions:
    atype = asrt.get('type', 'unknown')
    if atype not in by_type:
        by_type[atype] = []
    by_type[atype].append(asrt)

for atype, asrts in by_type.items():
    print(f"{atype.upper()} ({len(asrts)} assertions):")
    for asrt in asrts:
        data = asrt.get('data', {})
        if atype == 'counter':
            print(f"  - {data.get('target', '?')}: plus={data.get('plus_con', '?')}")
        elif atype == 'handshake':
            print(f"  - {data.get('phase_type', '?')}: {data.get('sender', '?')} / {data.get('receiver', '?')}")
        elif atype == 'pulseWidth':
            print(f"  - {data.get('target_signal', '?')}: min={data.get('min_width', '?')} max={data.get('max_width', '?')}")

# Verify
print("\n[3] Verification")
print("-" * 80)

expected = {
    'counter': 3,
    'handshake': 3,
    'pulseWidth': 3
}

all_pass = True
for atype, expected_count in expected.items():
    actual_count = len(by_type.get(atype, []))
    if actual_count == expected_count:
        print(f"✅ {atype}: {actual_count}/{expected_count}")
    else:
        print(f"❌ {atype}: {actual_count}/{expected_count} (expected {expected_count})")
        all_pass = False

print("\n" + "=" * 80)
if all_pass:
    print("✅ RESTORE TEST PASSED")
    print("\nAll assertions were correctly restored from Excel!")
else:
    print("❌ RESTORE TEST FAILED")
    print("\nSome assertions were not restored correctly.")
print("=" * 80)

sys.exit(0 if all_pass else 1)
