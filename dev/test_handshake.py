#!/usr/bin/env python3
"""Test handshake writing with fresh Excel."""
import sys
from pathlib import Path
import shutil

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from cli_tui import _write_assertion_to_excel
from openpyxl import load_workbook

# Create fresh test Excel
sessions = Path("out/sessions")
test_folder = sessions / "test_handshake_fresh"
test_folder.mkdir(parents=True, exist_ok=True)

# Copy reference Excel
ref_excel = Path("Data/Assertion_TF.xlsx")
test_excel = test_folder / "test.xlsx"
shutil.copy2(ref_excel, test_excel)

print(f"Testing handshake with: {test_excel}\n")

# Write two handshake assertions
print("Writing assertion 1...")
_write_assertion_to_excel(
    str(test_excel), 
    'handshake', 
    {'phase_type': 'ready_valid', 'sender': 'o_valid', 'receiver': 'i_ready'}, 
    None
)

print("Writing assertion 2...")
_write_assertion_to_excel(
    str(test_excel), 
    'handshake', 
    {'phase_type': '4phase', 'sender': 'req_sig', 'receiver': 'ack_sig'}, 
    None
)

# Verify
wb = load_workbook(str(test_excel))
hs_sheet = None
for name in wb.sheetnames:
    if name.lower() == 'handshake':
        hs_sheet = name
        break

if hs_sheet:
    ws = wb[hs_sheet]
    r7_type = ws.cell(7, 3).value
    r7_sender = ws.cell(7, 4).value
    r8_type = ws.cell(8, 3).value
    r8_sender = ws.cell(8, 4).value
    
    print(f"\nRow 7: {r7_type} / {r7_sender}")
    print(f"Row 8: {r8_type} / {r8_sender}")
    
    if r7_type == 'ready_valid' and r8_type == '4phase':
        print("\n✅ Both handshake assertions written correctly!")
    else:
        print(f"\n❌ Failed. Expected 'ready_valid' and '4phase'")
        print(f"   Got: '{r7_type}' and '{r8_type}'")
else:
    print("❌ Handshake sheet not found")

wb.close()
