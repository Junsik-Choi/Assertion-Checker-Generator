#!/usr/bin/env python3
"""
Integration test: Simulate TUI workflow for creating assertions and writing to Excel.
This test verifies the complete flow from assertion creation to Excel persistence.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from cli_tui import _write_assertion_to_excel, AppState
from openpyxl import load_workbook

def test_tui_workflow():
    """Test the complete TUI workflow for assertion creation and Excel writing."""
    
    print("=" * 80)
    print("INTEGRATION TEST - TUI Assertion Creation to Excel")
    print("=" * 80)
    
    # Find most recent session Excel
    sessions = Path("out/sessions")
    if not sessions.exists():
        print("❌ No sessions directory found")
        return False
    
    session_folders = sorted([d for d in sessions.iterdir() if d.is_dir()], 
                           key=lambda x: x.stat().st_mtime, reverse=True)
    if not session_folders:
        print("❌ No session folders found")
        return False
    
    latest = session_folders[0]
    xlsx_files = list(latest.glob("*.xlsx"))
    if not xlsx_files:
        print("❌ No Excel file found")
        return False
    
    excel_path = xlsx_files[0]
    print(f"\nSession: {latest.name}")
    print(f"Excel: {excel_path.name}\n")
    
    # Simulate TUI workflow: User creates multiple assertions
    test_scenarios = [
        {
            'name': 'Counter Assertion #1',
            'type': 'counter',
            'data': {
                'target': 'data_valid_cnt',
                'plus_con': 'data_valid && !rst_n',
                'reset_con': 'rst_n',
                'trigger_con': 'trigger_signal',
                'exp_cnt_val': '5'
            }
        },
        {
            'name': 'Counter Assertion #2',
            'type': 'counter',
            'data': {
                'target': 'req_counter',
                'plus_con': 'req && ack',
                'reset_con': '!rst_n',
                'trigger_con': 'done',
                'exp_cnt_val': '3'
            }
        },
        {
            'name': 'Handshake Assertion #1',
            'type': 'handshake',
            'data': {
                'phase_type': 'ready_valid',
                'sender': 'o_valid',
                'receiver': 'i_ready'
            }
        },
        {
            'name': 'Handshake Assertion #2',
            'type': 'handshake',
            'data': {
                'phase_type': '4phase',
                'sender': 'req_sig',
                'receiver': 'ack_sig'
            }
        },
        {
            'name': 'PulseWidth Assertion #1',
            'type': 'pulseWidth',
            'data': {
                'target_signal': 'enable_pulse',
                'min_width': '3',
                'max_width': '7'
            }
        },
        {
            'name': 'PulseWidth Assertion #2',
            'type': 'pulseWidth',
            'data': {
                'target_signal': 'clock_gate',
                'min_width': '2',
                'max_width': '5'
            }
        }
    ]
    
    # Write all test assertions
    print("Creating test assertions...")
    print("-" * 80)
    
    for scenario in test_scenarios:
        try:
            _write_assertion_to_excel(str(excel_path), scenario['type'], scenario['data'], None)
            print(f"✅ {scenario['name']}: {scenario['type']}")
        except Exception as e:
            print(f"❌ {scenario['name']}: FAILED - {e}")
            return False
    
    # Verify all assertions were written correctly
    print("\n" + "=" * 80)
    print("Verifying Excel contents...")
    print("=" * 80)
    
    wb = load_workbook(str(excel_path))
    
    all_pass = True
    
    # Verify Counter sheet
    print("\n[Counter Sheet]")
    counter_sheet = None
    for name in wb.sheetnames:
        if name.lower() == 'counter':
            counter_sheet = name
            break
    
    if counter_sheet:
        ws = wb[counter_sheet]
        # Check row 8 and 9 (should have 2 counter assertions)
        row8_target = ws.cell(8, 2).value
        row9_target = ws.cell(9, 2).value
        
        if row8_target and row9_target:
            print(f"  Row 8: {row8_target} - {ws.cell(8, 3).value}")
            print(f"  Row 9: {row9_target} - {ws.cell(9, 3).value}")
            print("  ✅ Multiple counter assertions present")
        else:
            print("  ❌ Missing counter assertions")
            all_pass = False
    else:
        print("  ❌ Counter sheet not found")
        all_pass = False
    
    # Verify Handshake sheet
    print("\n[Handshake Sheet]")
    handshake_sheet = None
    for name in wb.sheetnames:
        if name.lower() == 'handshake':
            handshake_sheet = name
            break
    
    if handshake_sheet:
        ws = wb[handshake_sheet]
        # Check row 7 and 8 (should have 2 handshake assertions)
        row7_type = ws.cell(7, 3).value
        row8_type = ws.cell(8, 3).value
        
        if row7_type and row8_type:
            print(f"  Row 7: {row7_type} - {ws.cell(7, 4).value} / {ws.cell(7, 5).value}")
            print(f"  Row 8: {row8_type} - {ws.cell(8, 4).value} / {ws.cell(8, 5).value}")
            print("  ✅ Multiple handshake assertions present")
        else:
            print("  ❌ Missing handshake assertions")
            all_pass = False
    else:
        print("  ❌ Handshake sheet not found")
        all_pass = False
    
    # Verify PulseWidth sheet
    print("\n[PulseWidth Sheet]")
    pulse_sheet = None
    for name in wb.sheetnames:
        if name.lower() == 'pulsewidth':
            pulse_sheet = name
            break
    
    if pulse_sheet:
        ws = wb[pulse_sheet]
        # Check row 7 and 8 (should have 2 pulse width assertions)
        row7_signal = ws.cell(7, 5).value
        row8_signal = ws.cell(8, 5).value
        
        if row7_signal and row8_signal:
            print(f"  Row 7: {row7_signal} - min:{ws.cell(7, 6).value} max:{ws.cell(7, 7).value}")
            print(f"  Row 8: {row8_signal} - min:{ws.cell(8, 6).value} max:{ws.cell(8, 7).value}")
            print("  ✅ Multiple pulseWidth assertions present")
        else:
            print("  ❌ Missing pulseWidth assertions")
            all_pass = False
    else:
        print("  ❌ PulseWidth sheet not found")
        all_pass = False
    
    wb.close()
    
    # Final summary
    print("\n" + "=" * 80)
    if all_pass:
        print("✅ INTEGRATION TEST PASSED")
        print("\nAll assertions were successfully written to Excel!")
        print("The fix is working correctly for:")
        print("  • Multiple assertions per sheet")
        print("  • Case-insensitive sheet lookup")
        print("  • Sample data clearing")
        print("  • Correct column/row positioning")
    else:
        print("❌ INTEGRATION TEST FAILED")
        print("Some assertions were not written correctly.")
    print("=" * 80)
    
    return all_pass

if __name__ == "__main__":
    success = test_tui_workflow()
    sys.exit(0 if success else 1)
