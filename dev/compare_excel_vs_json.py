#!/usr/bin/env python3
"""
Quick check: What's in the Excel vs session.json?
Please CLOSE blur_scaler.xlsx before running this.
"""
import sys
import json
from pathlib import Path
from openpyxl import load_workbook

session_dir = Path("out/sessions/blur_scaler-20251110_100715")
excel_path = session_dir / "blur_scaler.xlsx"
json_path = session_dir / "session.json"

print("=" * 80)
print("COMPARISON: Excel vs session.json")
print("=" * 80)

# Check session.json
print("\n[1] session.json says:")
print("-" * 80)
with open(json_path, 'r', encoding='utf-8') as f:
    data = json.load(f)
    
for asrt in data.get('assertions', []):
    atype = asrt.get('type')
    adata = asrt.get('data', {})
    
    if atype == 'pulseWidth':
        print(f"  PulseWidth: signal={adata.get('target_signal')}, min={adata.get('min_width')}, max={adata.get('max_width')}")
    elif atype == 'handshake':
        print(f"  Handshake: type={adata.get('phase_type')}, sender={adata.get('sender')}, receiver={adata.get('receiver')}")
    elif atype == 'counter':
        print(f"  Counter: target={adata.get('target')}, plus={adata.get('plus_con')}")

# Check Excel
print("\n[2] Excel file says:")
print("-" * 80)

try:
    wb = load_workbook(str(excel_path), read_only=True)
    
    # Check pulseWidth sheet
    for name in wb.sheetnames:
        if name.lower() == 'pulsewidth':
            ws = wb[name]
            print(f"\nPulseWidth sheet ('{name}'):")
            for row in range(7, min(12, ws.max_row + 1)):
                pulse_type = ws.cell(row, 3).value
                signal = ws.cell(row, 5).value
                min_w = ws.cell(row, 6).value
                max_w = ws.cell(row, 7).value
                
                # Skip sample data
                if signal and signal != 'target_pulse':
                    print(f"  Row {row}: type={pulse_type}, signal={signal}, min={min_w}, max={max_w}")
    
    # Check handshake sheet
    for name in wb.sheetnames:
        if name.lower() == 'handshake':
            ws = wb[name]
            print(f"\nHandshake sheet ('{name}'):")
            for row in range(7, min(12, ws.max_row + 1)):
                phase = ws.cell(row, 3).value
                sender = ws.cell(row, 4).value
                receiver = ws.cell(row, 5).value
                
                # Skip sample data
                if phase and phase not in ('valid', 'req', 'ack'):
                    print(f"  Row {row}: type={phase}, sender={sender}, receiver={receiver}")
    
    # Check Counter sheet
    for name in wb.sheetnames:
        if name.lower() == 'counter':
            ws = wb[name]
            print(f"\nCounter sheet ('{name}'):")
            for row in range(8, min(12, ws.max_row + 1)):
                target = ws.cell(row, 2).value
                plus = ws.cell(row, 3).value
                
                # Skip sample data
                if target and target != 'target_counter':
                    print(f"  Row {row}: target={target}, plus={plus}")
    
    wb.close()
    
    print("\n" + "=" * 80)
    print("✅ CHECK COMPLETE")
    print("\nIf Excel shows different data than session.json,")
    print("then the restore function will now read the EXCEL data correctly.")
    print("=" * 80)

except PermissionError:
    print("\n❌ ERROR: Excel file is locked!")
    print("\nPlease CLOSE blur_scaler.xlsx in Excel and run again.")
    print("=" * 80)
    sys.exit(1)
