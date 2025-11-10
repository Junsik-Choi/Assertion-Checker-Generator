#!/usr/bin/env python3
"""Test restore with actual user session (blur_scaler-20251110_100715)"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from openpyxl import load_workbook
from cli_tui import _restore_assertions_from_excel
from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional

session_dir = Path("out/sessions/blur_scaler-20251110_100715")
excel_path = session_dir / "blur_scaler.xlsx"

if not excel_path.exists():
    print(f"ERROR: Excel file not found: {excel_path}")
    sys.exit(1)

print("=" * 80)
print("TEST: Restore from actual session (blur_scaler)")
print("=" * 80)
print(f"\nSession: {session_dir}")
print(f"Excel: {excel_path}\n")

# Check Excel contents
print("[1] Excel File Contents")
print("-" * 80)

wb = load_workbook(str(excel_path))

# Check all sheets
for name in wb.sheetnames:
    name_lower = name.lower()
    
    if name_lower == 'counter':
        ws = wb[name]
        print(f"\n{name} Sheet:")
        for row in range(8, min(12, ws.max_row + 1)):
            target = ws.cell(row, 2).value
            plus = ws.cell(row, 3).value
            if target and target != 'target_counter':
                print(f"  Row {row}: target={target}, plus={plus}")
    
    elif name_lower == 'handshake':
        ws = wb[name]
        print(f"\n{name} Sheet:")
        for row in range(7, min(11, ws.max_row + 1)):
            ptype = ws.cell(row, 3).value
            sender = ws.cell(row, 4).value
            receiver = ws.cell(row, 5).value
            if ptype and ptype not in ('valid', 'req', 'ack'):
                print(f"  Row {row}: type={ptype}, sender={sender}, receiver={receiver}")
    
    elif name_lower == 'pulsewidth':
        ws = wb[name]
        print(f"\n{name} Sheet:")
        for row in range(7, min(11, ws.max_row + 1)):
            ptype = ws.cell(row, 3).value
            signal = ws.cell(row, 5).value
            min_w = ws.cell(row, 6).value
            max_w = ws.cell(row, 7).value
            if signal and signal != 'target_pulse':
                print(f"  Row {row}: type={ptype}, signal={signal}, min={min_w}, max={max_w}")

wb.close()

# Now test restore
print("\n[2] Testing _restore_assertions_from_excel")
print("-" * 80)

@dataclass
class MockModuleInfo:
    module: str = "blur_scaler"
    module_hierarchy: str = ""

@dataclass
class MockState:
    session_excel_path: Optional[Path] = excel_path
    assertions: List[Dict[str, Any]] = field(default_factory=list)
    module_info: MockModuleInfo = field(default_factory=MockModuleInfo)

state = MockState()
_restore_assertions_from_excel(state)

print(f"\nRestored {len(state.assertions)} assertions:\n")

if len(state.assertions) == 0:
    print("⚠️  No assertions restored!")
else:
    for i, asrt in enumerate(state.assertions, 1):
        atype = asrt.get('type', 'unknown')
        data = asrt.get('data', {})
        
        print(f"{i}. Type: {atype}")
        if atype == 'counter':
            print(f"   Target: {data.get('target', '?')}")
            print(f"   Plus: {data.get('plus_con', '?')}")
        elif atype == 'handshake':
            print(f"   Phase: {data.get('phase_type', '?')}")
            print(f"   Sender: {data.get('sender', '?')}")
            print(f"   Receiver: {data.get('receiver', '?')}")
        elif atype == 'pulseWidth':
            print(f"   Signal: {data.get('target_signal', '?')}")
            print(f"   Min: {data.get('min_width', '?')}, Max: {data.get('max_width', '?')}")
        print()

print("=" * 80)
print("✅ TEST COMPLETE")
print("=" * 80)
