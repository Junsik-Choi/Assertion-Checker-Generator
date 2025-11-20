"""
Test file generation with new Excel parsing logic - matches wizard mode exactly
"""
import sys
from pathlib import Path

# Add scripts directory to path
scripts_dir = Path(__file__).parent.parent / "scripts"
sys.path.insert(0, str(scripts_dir))

from openpyxl import load_workbook
from assertions import get_registered_plugins
from assertions.handshake import (
    find_cell, _ensure_handshake_layout, 
    parse_handshake_block_for_row, _port_width_token,
    ALLOWED_TYPES, HandshakePlugin
)

def test_parse_excel():
    """Test NEW Excel parsing logic - exactly like wizard mode"""
    excel_path = Path(r"c:\Users\JunsChoi\OneDrive - HARMAN\문서\TF자료\Assertion TF\Assertion Script\out\sessions\blur_scaler-20251120_190954\blur_scaler.xlsx")
    
    print(f"Loading Excel: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    # Test NEW handshake parsing using official functions
    print("\n=== Testing NEW Handshake Parsing (Wizard-style) ===")
    ws = wb['handshake']
    
    try:
        _, type_col, data_row = _ensure_handshake_layout(ws)
        print(f"Layout: type_col={type_col}, data_row={data_row}")
    except Exception as e:
        print(f"ERROR: {e}")
        return
    
    # Parse all rows with data
    blocks = []
    for row_idx in range(data_row, min(20, ws.max_row + 1)):
        type_val = ws.cell(row=row_idx, column=type_col).value
        if not type_val or str(type_val).strip() == "":
            continue
        
        print(f"\nParsing row {row_idx}...")
        info = parse_handshake_block_for_row(ws, row_idx, type_col)
        print(f"  Type: {info.get('phase_type')}")
        print(f"  Sender: {info.get('Sender')}")
        print(f"  Receiver: {info.get('Receiver')}")
        print(f"  Base Clock: {info.get('Base Clock')}")
        print(f"  Reset: {info.get('Reset')}")
        
        # Add width info (use empty dict for module since we don't have real module_info)
        mod_dict = {"clocks": [], "resets": [], "inputs": [], "outputs": [], "inouts": []}
        info["Base Clock Width"] = _port_width_token(mod_dict, info.get("Base Clock", ""))
        info["Reset Width"] = _port_width_token(mod_dict, info.get("Reset", ""))
        info["Sender Width"] = _port_width_token(mod_dict, info.get("Sender", ""))
        info["Receiver Width"] = _port_width_token(mod_dict, info.get("Receiver", ""))
        
        # Validate
        pt = (info.get("phase_type", "") or "").lower()
        if pt in ALLOWED_TYPES and info.get("Sender") and info.get("Receiver"):
            blocks.append(info)
            print(f"  ✓ Valid block added")
        else:
            print(f"  ✗ Skipped (invalid)")
    
    print(f"\n=== Parsed {len(blocks)} blocks total ===")
    
    if blocks:
        # Test plugin generate
        print("\n=== Testing Plugin Generate with Parsed Data ===")
        plugin = HandshakePlugin()
        parsed_data = {"blocks": blocks}
        
        context = {
            "module_info": {},
            "define_excel_path": str(excel_path),
            "output_dir": "",
            "session_dir": "",
            "config": {}
        }
        
        result = plugin.generate_sv(parsed_data, context)
        print(f"Plugin returned {len(result)} outputs")
        
        if len(result) >= 1:
            print(f"\n=== Interface SV (first 800 chars) ===")
            print(result[0][:800])
        if len(result) >= 2:
            print(f"\n=== Instance SV (first 800 chars) ===")
            print(result[1][:800])
    
    wb.close()

if __name__ == "__main__":
    test_parse_excel()
