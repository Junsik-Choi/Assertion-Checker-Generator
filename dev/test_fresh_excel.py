#!/usr/bin/env python3
"""
Test with fresh Excel file to verify sample data clearing and multiple writes.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from cli_tui import _write_assertion_to_excel
from openpyxl import load_workbook

def test_fresh_excel():
    """Test with a fresh Excel file that has sample data."""
    
    # Find the test session we just created
    sessions = Path("out/sessions")
    test_session = None
    for folder in sorted(sessions.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if folder.name.startswith("test_session-"):
            test_session = folder
            break
    
    if not test_session:
        print("❌ Test session not found")
        return False
    
    excel_path = test_session / "test_module.xlsx"
    if not excel_path.exists():
        print("❌ Test Excel not found")
        return False
    
    print("=" * 80)
    print("FRESH EXCEL TEST - Sample Data Clearing & Multiple Writes")
    print("=" * 80)
    print(f"\nTest Excel: {excel_path}\n")
    
    # Phase 1: Check initial state (should have sample data)
    print("[Phase 1] Checking initial state...")
    print("-" * 80)
    wb = load_workbook(str(excel_path))
    
    # Check Counter sheet for sample data
    counter_sheet = None
    for name in wb.sheetnames:
        if name.lower() == 'counter':
            counter_sheet = name
            break
    
    if counter_sheet:
        ws = wb[counter_sheet]
        initial_val = ws.cell(8, 2).value
        print(f"Counter row 8: {initial_val}")
        if initial_val == 'cnt':
            print("✅ Sample data present (cnt)")
        else:
            print(f"⚠️  Unexpected value: {initial_val}")
    wb.close()
    
    # Phase 2: Write first assertion (should clear sample data)
    print("\n[Phase 2] Writing first assertion...")
    print("-" * 80)
    
    assertion1 = {
        'target': 'first_counter',
        'plus_con': 'enable1',
        'reset_con': 'rst1',
        'trigger_con': 'trig1',
        'exp_cnt_val': '1'
    }
    
    _write_assertion_to_excel(str(excel_path), 'counter', assertion1, None)
    
    wb = load_workbook(str(excel_path))
    ws = wb[counter_sheet]
    row8_val = ws.cell(8, 2).value
    print(f"Counter row 8 after first write: {row8_val}")
    if row8_val == 'first_counter':
        print("✅ First assertion written, sample data cleared")
    else:
        print(f"❌ Expected 'first_counter', got: {row8_val}")
        wb.close()
        return False
    wb.close()
    
    # Phase 3: Write second assertion (should append, not clear)
    print("\n[Phase 3] Writing second assertion...")
    print("-" * 80)
    
    assertion2 = {
        'target': 'second_counter',
        'plus_con': 'enable2',
        'reset_con': 'rst2',
        'trigger_con': 'trig2',
        'exp_cnt_val': '2'
    }
    
    _write_assertion_to_excel(str(excel_path), 'counter', assertion2, None)
    
    wb = load_workbook(str(excel_path))
    ws = wb[counter_sheet]
    row8_val = ws.cell(8, 2).value
    row9_val = ws.cell(9, 2).value
    print(f"Counter row 8: {row8_val}")
    print(f"Counter row 9: {row9_val}")
    
    if row8_val == 'first_counter' and row9_val == 'second_counter':
        print("✅ Second assertion appended correctly")
    else:
        print(f"❌ Expected 'first_counter' and 'second_counter'")
        print(f"   Got: '{row8_val}' and '{row9_val}'")
        wb.close()
        return False
    wb.close()
    
    # Phase 4: Write third assertion (should continue appending)
    print("\n[Phase 4] Writing third assertion...")
    print("-" * 80)
    
    assertion3 = {
        'target': 'third_counter',
        'plus_con': 'enable3',
        'reset_con': 'rst3',
        'trigger_con': 'trig3',
        'exp_cnt_val': '3'
    }
    
    _write_assertion_to_excel(str(excel_path), 'counter', assertion3, None)
    
    wb = load_workbook(str(excel_path))
    ws = wb[counter_sheet]
    row8_val = ws.cell(8, 2).value
    row9_val = ws.cell(9, 2).value
    row10_val = ws.cell(10, 2).value
    
    print(f"Counter row  8: {row8_val}")
    print(f"Counter row  9: {row9_val}")
    print(f"Counter row 10: {row10_val}")
    
    if (row8_val == 'first_counter' and 
        row9_val == 'second_counter' and 
        row10_val == 'third_counter'):
        print("✅ Third assertion appended correctly")
    else:
        print(f"❌ Assertion ordering incorrect")
        wb.close()
        return False
    wb.close()
    
    # Final summary
    print("\n" + "=" * 80)
    print("✅ ALL TESTS PASSED!")
    print("\nVerified:")
    print("  1. ✅ Sample data cleared on first write")
    print("  2. ✅ Subsequent writes append (don't clear)")
    print("  3. ✅ Multiple assertions maintained correctly")
    print("=" * 80)
    
    return True

if __name__ == "__main__":
    success = test_fresh_excel()
    sys.exit(0 if success else 1)
