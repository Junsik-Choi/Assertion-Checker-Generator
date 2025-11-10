#!/usr/bin/env python3
"""Check Counter sheet structure in detail"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from openpyxl import load_workbook

def check_counter_sheet():
    sessions = Path("out/sessions")
    session_folders = sorted([d for d in sessions.iterdir() if d.is_dir()], 
                           key=lambda x: x.stat().st_mtime, reverse=True)
    latest = session_folders[0]
    xlsx_files = list(latest.glob("*.xlsx"))
    excel_path = xlsx_files[0]
    
    print(f"Excel: {excel_path}")
    print("=" * 80)
    
    wb = load_workbook(str(excel_path))
    
    # Check Counter sheet
    ws = wb['Counter']
    print("\nCounter Sheet - ALL rows:")
    print("=" * 80)
    
    for row in range(1, min(20, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(12, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val is None:
                row_data.append("_")
            else:
                val_str = str(val)[:40]
                row_data.append(val_str)
        
        # Only show rows with some content
        if any(v != "_" for v in row_data):
            print(f"R{row:2}: {' | '.join(row_data)}")
    
    # Check handshake sheet (lowercase)
    print("\n" + "=" * 80)
    print("handshake sheet (lowercase):")
    print("=" * 80)
    if 'handshake' in wb.sheetnames:
        ws = wb['handshake']
        for row in range(1, min(10, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(8, ws.max_column + 1)):
                val = ws.cell(row, col).value
                if val is None:
                    row_data.append("_")
                else:
                    row_data.append(str(val)[:40])
            if any(v != "_" for v in row_data):
                print(f"R{row:2}: {' | '.join(row_data)}")
    
    # Check pulseWidth sheet (lowercase)
    print("\n" + "=" * 80)
    print("pulseWidth sheet (lowercase):")
    print("=" * 80)
    if 'pulseWidth' in wb.sheetnames:
        ws = wb['pulseWidth']
        for row in range(1, min(10, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(8, ws.max_column + 1)):
                val = ws.cell(row, col).value
                if val is None:
                    row_data.append("_")
                else:
                    row_data.append(str(val)[:40])
            if any(v != "_" for v in row_data):
                print(f"R{row:2}: {' | '.join(row_data)}")
    
    wb.close()

if __name__ == "__main__":
    check_counter_sheet()
