#!/usr/bin/env python3
"""Check what _restore_assertions_from_excel is actually reading"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from openpyxl import load_workbook

# Check the blur_scaler session
excel_path = Path("out/sessions/blur_scaler-20251110_100715/blur_scaler.xlsx")

if excel_path.exists():
    print(f"Reading: {excel_path}\n")
    
    wb = load_workbook(str(excel_path))
    
    # Check for PulseWidth sheet (case-insensitive)
    pulse_sheet = None
    for name in wb.sheetnames:
        if name.lower() == 'pulsewidth':
            pulse_sheet = name
            print(f"Found sheet: '{pulse_sheet}'")
            break
    
    if pulse_sheet:
        ws = wb[pulse_sheet]
        
        print(f"\n=== Current code reads (WRONG) ===")
        print("Starting from row 2, column 1:")
        for row_idx in range(2, min(10, ws.max_row + 1)):
            col1 = ws.cell(row=row_idx, column=1).value
            col3 = ws.cell(row=row_idx, column=3).value
            col4 = ws.cell(row=row_idx, column=4).value
            print(f"  Row {row_idx}: col1={col1}, col3={col3}, col4={col4}")
            if not col1:
                break
        
        print(f"\n=== Should read (CORRECT) ===")
        print("Starting from row 7, columns 3-7:")
        for row_idx in range(7, min(12, ws.max_row + 1)):
            col3 = ws.cell(row=row_idx, column=3).value  # Type
            col4 = ws.cell(row=row_idx, column=4).value  # Count_Trigger
            col5 = ws.cell(row=row_idx, column=5).value  # Target_Pulse
            col6 = ws.cell(row=row_idx, column=6).value  # Min
            col7 = ws.cell(row=row_idx, column=7).value  # Max
            print(f"  Row {row_idx}: type={col3}, trigger={col4}, signal={col5}, min={col6}, max={col7}")
            if not col3:
                break
    
    wb.close()
else:
    print(f"Excel file not found: {excel_path}")
