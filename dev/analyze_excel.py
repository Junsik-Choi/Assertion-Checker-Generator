#!/usr/bin/env python3
"""Analyze Excel structure to debug Signal Assignments issue"""
import sys
from pathlib import Path
from openpyxl import load_workbook

excel_path = Path("out/sessions/sfr_cap-20251014_191534/sfr_cap.xlsx")
if not excel_path.exists():
    print(f"Excel not found: {excel_path}")
    sys.exit(1)

wb = load_workbook(str(excel_path))
print(f"✓ Loaded workbook: {excel_path}")
print(f"✓ Sheets: {wb.sheetnames}\n")

# Find Define sheet
define_ws = None
for name in wb.sheetnames:
    if name.strip().lower() == "define":
        define_ws = wb[name]
        break

if not define_ws:
    print("❌ 'Define' sheet NOT found!")
    sys.exit(1)

print(f"✓ Found 'Define' sheet\n")
print("=" * 80)
print("ANALYZING SHEET STRUCTURE (first 50 rows):")
print("=" * 80)

# Print all non-empty cells in first 50 rows
for row_idx in range(1, min(51, define_ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(15, define_ws.max_column + 1)):
        cell = define_ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            row_data.append(f"Col{col_idx}='{cell.value}'")
    if row_data:
        print(f"Row {row_idx:2d}: {', '.join(row_data)}")

print("\n" + "=" * 80)
print("SEARCHING FOR 'Signal Assignments' HEADER:")
print("=" * 80)

found_signal_assignments = False
for row in define_ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str):
            val_lower = cell.value.strip().lower()
            if "signal" in val_lower and "assignment" in val_lower:
                print(f"\n✓ FOUND at Row {cell.row}, Col {cell.column}: '{cell.value}'")
                found_signal_assignments = True
                
                # Check next row for Name/Equation/Bits
                print(f"\n  Checking Row {cell.row + 1} for column headers:")
                next_row = list(define_ws.iter_rows(min_row=cell.row + 1, max_row=cell.row + 1))
                if next_row:
                    for c in next_row[0]:
                        if c.value:
                            print(f"    Col {c.column}: '{c.value}'")
                break
    if found_signal_assignments:
        break

if not found_signal_assignments:
    print("\n❌ 'Signal Assignments' header NOT FOUND in Define sheet!")
    print("\n💡 Possible reasons:")
    print("   1. Reference Excel doesn't have 'Signal Assignments' section")
    print("   2. Header text is different (e.g., 'Signal Assignment' without 's')")
    print("   3. Section exists but with different wording")

wb.close()
print("\n" + "=" * 80)
