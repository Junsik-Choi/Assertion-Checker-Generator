#!/usr/bin/env python3
"""
Excel 파일의 시트 이름을 확인하는 스크립트
"""
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    
    excel_path = Path("Data/Assertion_TF.xlsx")
    if not excel_path.exists():
        print(f"Excel 파일을 찾을 수 없습니다: {excel_path}")
        sys.exit(1)
    
    print(f"Excel 파일 확인 중: {excel_path}")
    print("-" * 60)
    
    try:
        # read_only 모드로 열기 (파일이 열려있어도 읽을 수 있음)
        wb = load_workbook(str(excel_path), read_only=True, data_only=True)
        
        print(f"\n✓ 총 {len(wb.sheetnames)}개의 시트 발견:\n")
        
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  [{i}] '{sheet_name}'")
        
        wb.close()
        
        print("\n" + "-" * 60)
        print("\n대소문자 구분 확인:")
        
        # 플러그인에서 찾는 시트명
        expected_sheets = {
            "Counter": "Counter 플러그인",
            "counter": "counter (소문자)",
            "handshake": "Handshake 플러그인",
            "Handshake": "Handshake (대문자)",
        }
        
        wb = load_workbook(str(excel_path), read_only=True)
        sheet_names_lower = {s.lower(): s for s in wb.sheetnames}
        
        for expected, desc in expected_sheets.items():
            if expected in wb.sheetnames:
                print(f"  ✓ '{expected}' - 정확히 일치 ({desc})")
            elif expected.lower() in sheet_names_lower:
                actual = sheet_names_lower[expected.lower()]
                print(f"  ~ '{expected}' → 실제: '{actual}' ({desc})")
            else:
                print(f"  ✗ '{expected}' - 찾을 수 없음 ({desc})")
        
        wb.close()
        
    except PermissionError:
        print("\n⚠ Excel 파일이 다른 프로그램에서 열려있습니다.")
        print("  파일을 닫고 다시 시도해주세요.")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ 오류 발생: {e}")
        sys.exit(1)
        
except ImportError:
    print("openpyxl 패키지가 설치되어 있지 않습니다.")
    print("설치: pip install openpyxl")
    sys.exit(1)

print("\n완료!")
