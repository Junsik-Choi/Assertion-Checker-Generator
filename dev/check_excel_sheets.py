#!/usr/bin/env python3
"""Check Excel sheets in the current session"""
import sys
from pathlib import Path

# Add scripts directory to path
sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from openpyxl import load_workbook

def check_excel():
    # Find most recent session Excel
    sessions = Path("out/sessions")
    if not sessions.exists():
        print("Sessions directory not found")
        return
    
    session_folders = sorted([d for d in sessions.iterdir() if d.is_dir()], 
                           key=lambda x: x.stat().st_mtime, reverse=True)
    if not session_folders:
        print("No session folders found")
        return
    
    latest = session_folders[0]
    xlsx_files = list(latest.glob("*.xlsx"))
    if not xlsx_files:
        print("No Excel file found")
        return
    
    excel_path = xlsx_files[0]
    print(f"Excel: {excel_path}")
    print("=" * 80)
    
    wb = load_workbook(str(excel_path))
    print(f"\nAll Sheets: {wb.sheetnames}")
    print("=" * 80)
    
    # Check each assertion sheet
    for sheet_name in ['Counter', 'Handshake', 'PulseWidth']:
        print(f"\n{'='*80}")
        print(f"{sheet_name} Sheet")
        print("=" * 80)
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
            
            # Show first 6 rows
            for row in range(1, min(7, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(10, ws.max_column + 1)):
                    val = ws.cell(row, col).value
                    if val is None:
                        row_data.append("")
                    else:
                        row_data.append(str(val)[:30])  # Limit length
                print(f"Row {row}: {row_data}")
        else:
            print(f"*** {sheet_name} sheet NOT FOUND ***")
    
    wb.close()

if __name__ == "__main__":
    check_excel()
