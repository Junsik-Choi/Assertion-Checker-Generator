#!/usr/bin/env python3
"""
Test pulseWidth wizard improvements:
1. pulse_type selection (hpulse/vpulse)
2. base_clock from state.clocks for hpulse
3. trigger_signal from signals for vpulse
4. Parameter support in min/max width
5. show_if conditional fields work correctly
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

print("=" * 80)
print("PULSEWIDTH WIZARD IMPROVEMENTS TEST")
print("=" * 80)

# Test 1: Field definitions
print("\n[1] Testing pulseWidth field definitions...")
print("-" * 80)

from cli_tui import _get_plugin_fields

fields = _get_plugin_fields('pulseWidth')
print(f"Total fields defined: {len(fields)}")

for i, field in enumerate(fields, 1):
    name = field.get('name')
    ftype = field.get('type')
    step = field.get('step')
    title = field.get('title')
    show_if = field.get('show_if')
    
    print(f"\n{i}. {name} (step {step})")
    print(f"   Type: {ftype}")
    print(f"   Title: {title}")
    if show_if:
        print(f"   Show if: {show_if}")
    
    if name == 'pulse_type':
        print(f"   Options: {field.get('options')}")
        assert ftype == 'choice', "pulse_type should be choice type"
        assert 'hpulse' in field.get('options', []), "hpulse should be in options"
        assert 'vpulse' in field.get('options', []), "vpulse should be in options"
    elif name == 'base_clock':
        assert show_if == {'pulse_type': 'hpulse'}, "base_clock should show only for hpulse"
        assert ftype == 'choice', "base_clock should be choice type"
    elif name == 'trigger_signal':
        assert show_if == {'pulse_type': 'vpulse'}, "trigger_signal should show only for vpulse"
        assert ftype == 'signal', "trigger_signal should be signal type"

print("\n✅ Field definitions look good!")

# Test 2: show_if logic
print("\n[2] Testing show_if logic...")
print("-" * 80)

from cli_tui import _should_show_field, _get_visible_fields

# Test hpulse scenario
hpulse_data = {'pulse_type': 'hpulse'}
visible_hpulse = _get_visible_fields(fields, hpulse_data)
visible_names_h = [f['name'] for f in visible_hpulse]

print(f"When pulse_type='hpulse', visible fields: {visible_names_h}")
assert 'base_clock' in visible_names_h, "base_clock should be visible for hpulse"
assert 'trigger_signal' not in visible_names_h, "trigger_signal should NOT be visible for hpulse"

# Test vpulse scenario
vpulse_data = {'pulse_type': 'vpulse'}
visible_vpulse = _get_visible_fields(fields, vpulse_data)
visible_names_v = [f['name'] for f in visible_vpulse]

print(f"When pulse_type='vpulse', visible fields: {visible_names_v}")
assert 'trigger_signal' in visible_names_v, "trigger_signal should be visible for vpulse"
assert 'base_clock' not in visible_names_v, "base_clock should NOT be visible for vpulse"

print("\n✅ show_if logic working correctly!")

# Test 3: Preview with new fields
print("\n[3] Testing preview generation...")
print("-" * 80)

from cli_tui import _generate_assertion_preview, AppState
from dataclasses import dataclass, field as dc_field
from typing import List, Dict, Any

@dataclass
class MockModuleInfo:
    clocks: List[Dict] = dc_field(default_factory=lambda: [{'name': 'I_CLK'}])
    resets: List[Dict] = dc_field(default_factory=lambda: [{'name': 'I_RSTN'}])
    parameters: List[Dict] = dc_field(default_factory=lambda: [
        {'name': 'DATA_WIDTH', 'default': '8'},
        {'name': 'PARAM_WIDTH', 'default': '11'}
    ])

@dataclass
class MockState:
    module_info: MockModuleInfo = dc_field(default_factory=MockModuleInfo)

# Test hpulse preview
hpulse_preview_data = {
    'pulse_type': 'hpulse',
    'base_clock': 'I_CLK',
    'target_signal': 'o_hsync',
    'min_width': '10',
    'max_width': '20'
}

state = MockState()
preview_lines = _generate_assertion_preview('pulseWidth', hpulse_preview_data, state)

print("\nhpulse Preview:")
for line in preview_lines[:15]:
    print(f"  {line}")

# Check preview contains key information
preview_text = '\n'.join(preview_lines)
assert 'hpulse' in preview_text, "Preview should show pulse type"
assert 'I_CLK' in preview_text, "Preview should show base clock"
assert 'o_hsync' in preview_text, "Preview should show target signal"
assert 'Available Parameters' in preview_text, "Preview should show available parameters"

print("\n✅ Preview generation working!")

# Test 4: Excel writing/reading
print("\n[4] Testing Excel write/read cycle...")
print("-" * 80)

from openpyxl import Workbook
import tempfile
import os

# Create temporary Excel file
temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
os.close(temp_fd)

try:
    # Create workbook with pulseWidth sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'pulseWidth'
    
    # Write headers (Row 6)
    ws.cell(6, 3, 'Type')
    ws.cell(6, 4, 'Count_Trigger')
    ws.cell(6, 5, 'Target_Pulse')
    ws.cell(6, 6, 'Expected_Min_Value')
    ws.cell(6, 7, 'Expected_Max_Value')
    
    wb.save(temp_path)
    wb.close()
    
    # Test writing hpulse assertion
    from cli_tui import _write_assertion_to_excel
    
    hpulse_data = {
        'pulse_type': 'hpulse',
        'base_clock': 'I_CLK',
        'target_signal': 'o_hsync',
        'min_width': '10',
        'max_width': '20'
    }
    
    _write_assertion_to_excel(temp_path, 'pulseWidth', hpulse_data, state)
    
    # Test writing vpulse assertion
    vpulse_data = {
        'pulse_type': 'vpulse',
        'trigger_signal': 'i_trigger',
        'target_signal': 'o_data_valid',
        'min_width': 'DATA_WIDTH',
        'max_width': 'PARAM_WIDTH'
    }
    
    _write_assertion_to_excel(temp_path, 'pulseWidth', vpulse_data, state)
    
    # Read back and verify
    from openpyxl import load_workbook
    wb = load_workbook(temp_path)
    ws = wb['pulseWidth']
    
    # Check row 7 (first data row - hpulse)
    row7_type = ws.cell(7, 3).value
    row7_trigger = ws.cell(7, 4).value
    row7_signal = ws.cell(7, 5).value
    row7_min = ws.cell(7, 6).value
    row7_max = ws.cell(7, 7).value
    
    print(f"\nRow 7 (hpulse):")
    print(f"  Type: {row7_type}")
    print(f"  Count_Trigger: {row7_trigger}")
    print(f"  Target: {row7_signal}")
    print(f"  Min: {row7_min}")
    print(f"  Max: {row7_max}")
    
    assert row7_type == 'hpulse', f"Expected hpulse, got {row7_type}"
    assert row7_trigger == 'I_CLK', f"Expected I_CLK, got {row7_trigger}"
    assert row7_signal == 'o_hsync', f"Expected o_hsync, got {row7_signal}"
    
    # Check row 8 (second data row - vpulse)
    row8_type = ws.cell(8, 3).value
    row8_trigger = ws.cell(8, 4).value
    row8_signal = ws.cell(8, 5).value
    row8_min = ws.cell(8, 6).value
    row8_max = ws.cell(8, 7).value
    
    print(f"\nRow 8 (vpulse):")
    print(f"  Type: {row8_type}")
    print(f"  Count_Trigger: {row8_trigger}")
    print(f"  Target: {row8_signal}")
    print(f"  Min: {row8_min}")
    print(f"  Max: {row8_max}")
    
    assert row8_type == 'vpulse', f"Expected vpulse, got {row8_type}"
    assert row8_trigger == 'i_trigger', f"Expected i_trigger, got {row8_trigger}"
    assert row8_signal == 'o_data_valid', f"Expected o_data_valid, got {row8_signal}"
    assert row8_min == 'DATA_WIDTH', f"Expected DATA_WIDTH, got {row8_min}"
    
    wb.close()
    
    print("\n✅ Excel write cycle working correctly!")
    
    # Test restore
    print("\n[5] Testing Excel restore...")
    print("-" * 80)
    
    from cli_tui import _restore_assertions_from_excel
    from dataclasses import field as dc_field
    
    @dataclass
    class RestoreState:
        session_excel_path: Path = Path(temp_path)
        assertions: List[Dict] = dc_field(default_factory=list)
        module_info: MockModuleInfo = dc_field(default_factory=MockModuleInfo)
    
    restore_state = RestoreState()
    _restore_assertions_from_excel(restore_state)
    
    print(f"\nRestored {len(restore_state.assertions)} assertions:")
    
    for i, asrt in enumerate(restore_state.assertions, 1):
        atype = asrt.get('type')
        data = asrt.get('data', {})
        print(f"\n{i}. Type: {atype}")
        for key, val in data.items():
            print(f"   {key}: {val}")
        
        if i == 1:
            assert data.get('pulse_type') == 'hpulse', "First assertion should be hpulse"
            assert data.get('base_clock') == 'I_CLK', "Should have base_clock"
            assert 'trigger_signal' not in data or data.get('trigger_signal') == '', "Should not have trigger_signal"
        elif i == 2:
            assert data.get('pulse_type') == 'vpulse', "Second assertion should be vpulse"
            assert data.get('trigger_signal') == 'i_trigger', "Should have trigger_signal"
            assert data.get('min_width') == 'DATA_WIDTH', "Should preserve parameter name"
    
    print("\n✅ Excel restore working correctly!")
    
finally:
    # Cleanup
    try:
        os.unlink(temp_path)
    except:
        pass

print("\n" + "=" * 80)
print("✅ ALL TESTS PASSED!")
print("\nSummary:")
print("  ✅ Field definitions correct (pulse_type, base_clock, trigger_signal)")
print("  ✅ show_if conditional logic working")
print("  ✅ Preview generation includes new fields")
print("  ✅ Excel write saves correct columns")
print("  ✅ Excel restore reads all fields correctly")
print("  ✅ Parameter names preserved (no validation error)")
print("=" * 80)
