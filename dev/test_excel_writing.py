#!/usr/bin/env python3
"""
Test script to verify Excel writing works correctly.
Creates test assertions for all types and checks if they are written to Excel.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from openpyxl import load_workbook

def verify_excel_writing():
    """Verify that assertions can be written to Excel properly."""
    
    # Find most recent session Excel
    sessions = Path("out/sessions")
    if not sessions.exists():
        print("❌ Sessions directory not found")
        return False
    
    session_folders = sorted([d for d in sessions.iterdir() if d.is_dir()], 
                           key=lambda x: x.stat().st_mtime, reverse=True)
    if not session_folders:
        print("❌ No session folders found")
        return False
    
    latest = session_folders[0]
    xlsx_files = list(latest.glob("*.xlsx"))
    if not xlsx_files:
        print("❌ No Excel file found")
        return False
    
    excel_path = xlsx_files[0]
    print(f"Testing Excel: {excel_path}")
    print("=" * 80)
    
    # Test writing to Counter sheet
    print("\n[1] Testing Counter sheet write...")
    try:
        from cli_tui import _write_assertion_to_excel
        
        # Create test data for counter
        counter_data = {
            'target': 'test_counter',
            'plus_con': 'test_enable',
            'reset_con': 'rst_n',
            'trigger_con': 'test_trigger',
            'exp_cnt_val': '10'
        }
        
        _write_assertion_to_excel(str(excel_path), 'counter', counter_data, None)
        
        # Verify it was written
        wb = load_workbook(str(excel_path))
        sheet_name = None
        for name in wb.sheetnames:
            if name.lower() == 'counter':
                sheet_name = name
                break
        
        if sheet_name:
            ws = wb[sheet_name]
            # Check row 8 (first data row after header)
            target_val = ws.cell(8, 2).value
            plus_val = ws.cell(8, 3).value
            
            if target_val == 'test_counter' and plus_val == 'test_enable':
                print("✅ Counter assertion written successfully!")
            else:
                print(f"❌ Counter assertion not found. Got: target={target_val}, plus={plus_val}")
                wb.close()
                return False
        else:
            print("❌ Counter sheet not found")
            wb.close()
            return False
        
        wb.close()
    except Exception as e:
        print(f"❌ Counter write failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Test writing to Handshake sheet
    print("\n[2] Testing Handshake sheet write...")
    try:
        handshake_data = {
            'phase_type': 'ready_valid',
            'sender': 'test_valid',
            'receiver': 'test_ready'
        }
        
        _write_assertion_to_excel(str(excel_path), 'handshake', handshake_data, None)
        
        # Verify it was written
        wb = load_workbook(str(excel_path))
        sheet_name = None
        for name in wb.sheetnames:
            if name.lower() == 'handshake':
                sheet_name = name
                break
        
        if sheet_name:
            ws = wb[sheet_name]
            # Check row 7 (first data row after header)
            type_val = ws.cell(7, 3).value
            sender_val = ws.cell(7, 4).value
            
            if type_val == 'ready_valid' and sender_val == 'test_valid':
                print("✅ Handshake assertion written successfully!")
            else:
                print(f"❌ Handshake assertion not found. Got: type={type_val}, sender={sender_val}")
                wb.close()
                return False
        else:
            print("❌ Handshake sheet not found")
            wb.close()
            return False
        
        wb.close()
    except Exception as e:
        print(f"❌ Handshake write failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Test writing to PulseWidth sheet
    print("\n[3] Testing PulseWidth sheet write...")
    try:
        pulse_data = {
            'target_signal': 'test_pulse',
            'min_width': '5',
            'max_width': '10'
        }
        
        _write_assertion_to_excel(str(excel_path), 'pulseWidth', pulse_data, None)
        
        # Verify it was written
        wb = load_workbook(str(excel_path))
        sheet_name = None
        for name in wb.sheetnames:
            if name.lower() == 'pulsewidth':
                sheet_name = name
                break
        
        if sheet_name:
            ws = wb[sheet_name]
            # Check row 7 (first data row after header)
            type_val = ws.cell(7, 3).value
            signal_val = ws.cell(7, 5).value
            min_val = ws.cell(7, 6).value
            
            if signal_val == 'test_pulse' and min_val == '5':
                print("✅ PulseWidth assertion written successfully!")
            else:
                print(f"❌ PulseWidth assertion not found. Got: signal={signal_val}, min={min_val}")
                wb.close()
                return False
        else:
            print("❌ PulseWidth sheet not found")
            wb.close()
            return False
        
        wb.close()
    except Exception as e:
        print(f"❌ PulseWidth write failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    print("\n" + "=" * 80)
    print("✅ ALL TESTS PASSED! Excel writing is working correctly.")
    print("=" * 80)
    return True

if __name__ == "__main__":
    success = verify_excel_writing()
    sys.exit(0 if success else 1)
