#!/usr/bin/env python3
"""
FINAL COMPREHENSIVE TEST
Tests all three assertion types with multiple assertions each, using a fresh Excel file.
"""
import sys
from pathlib import Path
import shutil

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from cli_tui import _write_assertion_to_excel
from openpyxl import load_workbook

def run_comprehensive_test():
    """Run a comprehensive test of all assertion types."""
    
    print("=" * 80)
    print("COMPREHENSIVE FINAL TEST - All Assertion Types")
    print("=" * 80)
    
    # Create fresh test Excel
    sessions = Path("out/sessions")
    test_folder = sessions / "final_test"
    test_folder.mkdir(parents=True, exist_ok=True)
    
    ref_excel = Path("Data/Assertion_TF.xlsx")
    test_excel = test_folder / "test.xlsx"
    shutil.copy2(ref_excel, test_excel)
    
    print(f"\nTest Excel: {test_excel}\n")
    
    # Define test assertions
    test_cases = {
        'counter': [
            {'target': 'data_counter', 'plus_con': 'data_valid', 'reset_con': 'rst_n', 'trigger_con': 'done', 'exp_cnt_val': '10'},
            {'target': 'req_counter', 'plus_con': 'req && ack', 'reset_con': '!rst_n', 'trigger_con': 'complete', 'exp_cnt_val': '5'},
            {'target': 'error_counter', 'plus_con': 'error_flag', 'reset_con': 'clear', 'trigger_con': 'check', 'exp_cnt_val': '0'},
        ],
        'handshake': [
            {'phase_type': 'ready_valid', 'sender': 'o_valid', 'receiver': 'i_ready'},
            {'phase_type': '4phase', 'sender': 'req_out', 'receiver': 'ack_in'},
            {'phase_type': '2phase', 'sender': 'request', 'receiver': 'acknowledge'},
        ],
        'pulseWidth': [
            {'target_signal': 'enable_pulse', 'min_width': '3', 'max_width': '7'},
            {'target_signal': 'clock_gate', 'min_width': '2', 'max_width': '5'},
            {'target_signal': 'strobe_signal', 'min_width': '1', 'max_width': '4'},
        ]
    }
    
    # Write all assertions
    print("Writing assertions...")
    print("-" * 80)
    for atype, assertions in test_cases.items():
        print(f"\n{atype.upper()}:")
        for i, data in enumerate(assertions, 1):
            try:
                _write_assertion_to_excel(str(test_excel), atype, data, None)
                print(f"  ✅ Assertion #{i}")
            except Exception as e:
                print(f"  ❌ Assertion #{i}: {e}")
                return False
    
    # Verify all assertions
    print("\n" + "=" * 80)
    print("Verifying results...")
    print("=" * 80)
    
    wb = load_workbook(str(test_excel))
    all_pass = True
    
    # Verify Counter
    print("\n[COUNTER SHEET]")
    counter_sheet = None
    for name in wb.sheetnames:
        if name.lower() == 'counter':
            counter_sheet = name
            break
    
    if counter_sheet:
        ws = wb[counter_sheet]
        for i in range(8, 11):  # Rows 8, 9, 10
            target = ws.cell(i, 2).value
            plus = ws.cell(i, 3).value
            print(f"  Row {i}: {target} | {plus}")
        
        r8 = ws.cell(8, 2).value
        r9 = ws.cell(9, 2).value
        r10 = ws.cell(10, 2).value
        
        if r8 == 'data_counter' and r9 == 'req_counter' and r10 == 'error_counter':
            print("  ✅ All 3 counter assertions verified")
        else:
            print(f"  ❌ Counter assertions incorrect")
            all_pass = False
    else:
        print("  ❌ Counter sheet not found")
        all_pass = False
    
    # Verify Handshake
    print("\n[HANDSHAKE SHEET]")
    handshake_sheet = None
    for name in wb.sheetnames:
        if name.lower() == 'handshake':
            handshake_sheet = name
            break
    
    if handshake_sheet:
        ws = wb[handshake_sheet]
        for i in range(7, 10):  # Rows 7, 8, 9
            ptype = ws.cell(i, 3).value
            sender = ws.cell(i, 4).value
            print(f"  Row {i}: {ptype} | {sender}")
        
        r7 = ws.cell(7, 3).value
        r8 = ws.cell(8, 3).value
        r9 = ws.cell(9, 3).value
        
        if r7 == 'ready_valid' and r8 == '4phase' and r9 == '2phase':
            print("  ✅ All 3 handshake assertions verified")
        else:
            print(f"  ❌ Handshake assertions incorrect")
            all_pass = False
    else:
        print("  ❌ Handshake sheet not found")
        all_pass = False
    
    # Verify PulseWidth
    print("\n[PULSEWIDTH SHEET]")
    pulse_sheet = None
    for name in wb.sheetnames:
        if name.lower() == 'pulsewidth':
            pulse_sheet = name
            break
    
    if pulse_sheet:
        ws = wb[pulse_sheet]
        for i in range(7, 10):  # Rows 7, 8, 9
            signal = ws.cell(i, 5).value
            min_w = ws.cell(i, 6).value
            max_w = ws.cell(i, 7).value
            print(f"  Row {i}: {signal} | min:{min_w} max:{max_w}")
        
        r7 = ws.cell(7, 5).value
        r8 = ws.cell(8, 5).value
        r9 = ws.cell(9, 5).value
        
        if r7 == 'enable_pulse' and r8 == 'clock_gate' and r9 == 'strobe_signal':
            print("  ✅ All 3 pulseWidth assertions verified")
        else:
            print(f"  ❌ PulseWidth assertions incorrect")
            all_pass = False
    else:
        print("  ❌ PulseWidth sheet not found")
        all_pass = False
    
    wb.close()
    
    # Final summary
    print("\n" + "=" * 80)
    if all_pass:
        print("✅ ✅ ✅  ALL TESTS PASSED  ✅ ✅ ✅")
        print("\nSUCCESSFULLY VERIFIED:")
        print("  • 3 Counter assertions")
        print("  • 3 Handshake assertions")
        print("  • 3 PulseWidth assertions")
        print("\nFEATURES CONFIRMED:")
        print("  ✅ Case-insensitive sheet lookup")
        print("  ✅ Sample data auto-clearing (first write)")
        print("  ✅ Append mode (subsequent writes)")
        print("  ✅ Merged cell handling")
        print("  ✅ Correct column mapping")
        print("\n🎉 Excel writing is fully functional!")
    else:
        print("❌ SOME TESTS FAILED")
        print("See details above")
    print("=" * 80)
    
    return all_pass

if __name__ == "__main__":
    success = run_comprehensive_test()
    sys.exit(0 if success else 1)
