#!/usr/bin/env python3
"""Test script to verify Excel Define sheet fix"""
import sys
from pathlib import Path
from openpyxl import load_workbook

excel_path = Path(r"c:\Users\JunsChoi\OneDrive - HARMAN\문서\TF자료\Assertion TF\Assertion Script\out\sessions\out_sync_gen-20251106_163952\out_sync_gen.xlsx")

print("Loading Excel file...")
wb = load_workbook(excel_path, data_only=True)
ws = wb["Define"]

print("\n=== Current Define Sheet ===")
for i in range(1, 7):
    row_data = [cell.value for cell in ws[i][:3]]
    print(f"Row {i}: {row_data}")

# Find Target Path value
target_path_value = None
for row in ws.iter_rows():
    for cell in row:
        if cell.value and str(cell.value).strip().lower() == "target path":
            value_cell = ws.cell(row=cell.row, column=cell.column + 1)
            target_path_value = str(value_cell.value or "").strip()
            print(f"\nFound Target Path at Row {cell.row}, Col {cell.column}")
            print(f"Current value: {repr(target_path_value)}")
            break
    if target_path_value:
        break

wb.close()

# Check if it needs fixing
if target_path_value:
    has_backslash = '\\' in target_path_value
    is_absolute = target_path_value.startswith('/')
    needs_fix = has_backslash or is_absolute
    
    print(f"\nAnalysis:")
    print(f"  Has backslash: {has_backslash}")
    print(f"  Starts with /: {is_absolute}")
    print(f"  Needs fixing: {needs_fix}")
    
    if needs_fix:
        print(f"\n✗ Excel Define sheet has file path format")
        print(f"  Should be converted to hierarchy format (forward slash)")
    else:
        print(f"\n✓ Excel Define sheet already has correct hierarchy format")
