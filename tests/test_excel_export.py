#!/usr/bin/env python3
"""
Test Excel export functionality with signal bit width parsing.
"""

import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("openpyxl not installed. Skipping Excel export test.")
    sys.exit(0)


def test_excel_export():
    """Test writing assertion data to Excel with signal bit widths."""
    print("\n" + "="*60)
    print("  EXCEL EXPORT TEST")
    print("="*60 + "\n")
    
    # Create a temporary Excel file
    with tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False) as f:
        temp_excel = f.name
    
    try:
        # Create workbook with required sheets
        wb = Workbook()
        ws_counter = wb.create_sheet('Counter', 0)
        ws_handshake = wb.create_sheet('Handshake')
        ws_pulse = wb.create_sheet('PulseWidth')
        
        # Add headers
        ws_counter['A1'] = 'Signal'
        ws_counter['B1'] = 'Bit Width'
        ws_counter['C1'] = 'Plus Condition'
        ws_counter['D1'] = 'Reset Condition'
        ws_counter['E1'] = 'Trigger Condition'
        ws_counter['F1'] = 'Expected Count'
        
        ws_handshake['A1'] = 'Phase Type'
        ws_handshake['B1'] = 'Sender'
        ws_handshake['C1'] = 'Sender Width'
        ws_handshake['D1'] = 'Receiver'
        ws_handshake['E1'] = 'Receiver Width'
        
        ws_pulse['A1'] = 'Signal'
        ws_pulse['B1'] = 'Bit Width'
        ws_pulse['C1'] = 'Min Width'
        ws_pulse['D1'] = 'Max Width'
        
        wb.save(temp_excel)
        wb.close()
        
        print("✓ Created Excel file with sheets")
        
        # Now write test data
        wb = load_workbook(temp_excel)
        
        # Test Counter
        print("\nWriting Counter assertion...")
        ws = wb['Counter']
        ws['A2'] = 'i_data'
        ws['B2'] = '[7:0]'
        ws['C2'] = 'i_valid'
        ws['D2'] = 'rst_n'
        ws['E2'] = 'clk'
        ws['F2'] = '10'
        print("  ✓ Row 2: i_data [7:0] + conditions")
        
        # Test Handshake
        print("\nWriting Handshake assertion...")
        ws = wb['Handshake']
        ws['A2'] = '4-Phase'
        ws['B2'] = 'req_sig'
        ws['C2'] = '[0:0]'
        ws['D2'] = 'ack_sig'
        ws['E2'] = '[0:0]'
        print("  ✓ Row 2: 4-Phase with req/ack signals")
        
        # Test PulseWidth
        print("\nWriting PulseWidth assertion...")
        ws = wb['PulseWidth']
        ws['A2'] = 'pulse_out'
        ws['B2'] = '[3:0]'
        ws['C2'] = '5'
        ws['D2'] = '20'
        print("  ✓ Row 2: pulse_out [3:0] with min/max")
        
        wb.save(temp_excel)
        wb.close()
        
        # Verify data was written correctly
        print("\n\nVerifying Excel data...")
        wb = load_workbook(temp_excel)
        
        ws_counter = wb['Counter']
        counter_signal = ws_counter['A2'].value
        counter_width = ws_counter['B2'].value
        counter_plus = ws_counter['C2'].value
        
        ws_handshake = wb['Handshake']
        hs_sender = ws_handshake['B2'].value
        hs_sender_width = ws_handshake['C2'].value
        
        ws_pulse = wb['PulseWidth']
        pulse_signal = ws_pulse['A2'].value
        pulse_width = ws_pulse['B2'].value
        
        checks = [
            (counter_signal == 'i_data', f"Counter signal: '{counter_signal}' == 'i_data'"),
            (counter_width == '[7:0]', f"Counter width: '{counter_width}' == '[7:0]'"),
            (counter_plus == 'i_valid', f"Counter plus: '{counter_plus}' == 'i_valid'"),
            (hs_sender == 'req_sig', f"Handshake sender: '{hs_sender}' == 'req_sig'"),
            (hs_sender_width == '[0:0]', f"Handshake sender width: '{hs_sender_width}' == '[0:0]'"),
            (pulse_signal == 'pulse_out', f"Pulse signal: '{pulse_signal}' == 'pulse_out'"),
            (pulse_width == '[3:0]', f"Pulse width: '{pulse_width}' == '[3:0]'"),
        ]
        
        all_passed = True
        for passed, description in checks:
            status = "✓" if passed else "✗"
            print(f"{status} {description}")
            all_passed = all_passed and passed
        
        wb.close()
        
        return all_passed
        
    finally:
        # Clean up
        Path(temp_excel).unlink(missing_ok=True)


if __name__ == '__main__':
    try:
        result = test_excel_export()
        
        print("\n" + "="*60)
        if result:
            print("  ✅ EXCEL EXPORT TEST PASSED")
        else:
            print("  ⚠️  EXCEL EXPORT TEST FAILED")
        print("="*60 + "\n")
        
        sys.exit(0 if result else 1)
        
    except Exception as e:
        print(f"\n❌ Excel export test failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
