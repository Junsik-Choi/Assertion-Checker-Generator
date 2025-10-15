#!/usr/bin/env python3
"""Test find_signal_assignments_header() function"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent / "scripts"))

from openpyxl import load_workbook

# Copy the function to test
def find_signal_assignments_header(ws):
    """
    Find 'Signal Assignments' header row with Name, Equation, Bits columns.
    Returns dict with row and column info, or None.
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "signal" in cell.value.strip().casefold() and "assignment" in cell.value.strip().casefold():
                # Found "Signal Assignments", now find Name/Equation/Bits in next row or same row
                header_row = cell.row
                print(f"✓ Found 'Signal Assignments' at row {header_row}, col {cell.column}")
                
                # Check next row for Name/Equation/Bits
                next_row = list(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1))
                if next_row:
                    name_col = None
                    equation_col = None
                    bits_col = None
                    
                    print(f"\nSearching for Name/Equation/Bits in row {header_row + 1}:")
                    for c in next_row[0]:
                        if isinstance(c.value, str):
                            val = c.value.strip().casefold()
                            print(f"  Col {c.column}: '{c.value}' (casefold='{val}')")
                            if val == "name":
                                # Check if this Name is close to "Signal Assignments"
                                # Signal Assignments is at cell.column (12)
                                # We want Name that's in the same region (col >= 11)
                                if c.column >= cell.column - 1:
                                    name_col = c.column
                                    print(f"    -> Matched as NAME (col {c.column})")
                            elif val == "equation":
                                equation_col = c.column
                                print(f"    -> Matched as EQUATION (col {c.column})")
                            elif val == "bits":
                                # Only accept Bits near Equation
                                if equation_col and c.column == equation_col + 1:
                                    bits_col = c.column
                                    print(f"    -> Matched as BITS (col {c.column})")
                    
                    if name_col and equation_col:
                        result = {
                            "header_row": header_row,
                            "data_row": header_row + 1,
                            "name_col": name_col,
                            "equation_col": equation_col,
                            "bits_col": bits_col or (equation_col + 1)
                        }
                        print(f"\n✓ RESULT: {result}")
                        return result
                    else:
                        print(f"\n❌ Found Signal Assignments but missing required columns:")
                        print(f"   name_col={name_col}, equation_col={equation_col}, bits_col={bits_col}")
    
    print("\n❌ 'Signal Assignments' header NOT found")
    return None

# Test
excel_path = Path("out/sessions/sfr_cap-20251014_191534/sfr_cap.xlsx")
wb = load_workbook(str(excel_path))
ws = None
for name in wb.sheetnames:
    if name.strip().lower() == "define":
        ws = wb[name]
        break

if ws:
    print("=" * 80)
    print("TESTING find_signal_assignments_header()")
    print("=" * 80)
    result = find_signal_assignments_header(ws)
    
    if result:
        print("\n" + "=" * 80)
        print("✓ SUCCESS - Header found correctly")
        print("=" * 80)
    else:
        print("\n" + "=" * 80)
        print("❌ FAILED - Header not found")
        print("=" * 80)
else:
    print("❌ Define sheet not found")

wb.close()
