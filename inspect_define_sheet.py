#!/usr/bin/env python3
"""
Check Define sheet for hierarchy, clocks, resets
"""
from openpyxl import load_workbook
from pathlib import Path

excel_path = Path(__file__).parent / "out/sessions/blur_scaler-20251203_141618/blur_scaler.xlsx"

print("=" * 70)
print("EXCEL DEFINE SHEET INSPECTION")
print("=" * 70)

if excel_path.exists():
    wb = load_workbook(str(excel_path))
    
    if 'Define' in wb.sheetnames:
        ws = wb['Define']
        print(f"\nDefine sheet")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        print("\nFirst 50 rows:")
        for row_idx in range(1, min(51, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(8, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(val)[:30] if val else "")
            # Only print non-empty rows
            if any(row_data):
                print(f"Row {row_idx}: {row_data}")
    else:
        print("No Define sheet")
    
    wb.close()
else:
    print(f"Excel not found: {excel_path}")

print("\n" + "=" * 70)
